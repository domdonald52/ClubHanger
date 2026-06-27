"""
Import aircraft from a filled-in ClubHangar aircraft template (.xlsx or .csv).

    python manage.py import_aircraft aircraft.xlsx --club wac --dry-run
    python manage.py import_aircraft aircraft.xlsx --club wac

Header matching is case-insensitive and tolerates minor wording variations.
Re-running the same file is safe — existing registrations are skipped unless
--update is passed.

No notifications or signals are fired. Safe to run before go-live.
"""
import csv
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import Aircraft, AircraftType, Club


HEADER_MAP = {
    'registration': 'registration', 'rego': 'registration', 'reg': 'registration',
    'type': 'type_name', 'aircraft type': 'type_name', 'make/model': 'type_name',
    'icao': 'icao_designator', 'icao designator': 'icao_designator',
    'serial': 'serial_number', 'serial number': 'serial_number', 's/n': 'serial_number',
    'engines': 'engine_count', 'engine count': 'engine_count', 'engine_count': 'engine_count',
    'seats': 'seats',
    'hobbs': 'records_hobbs', 'records hobbs': 'records_hobbs',
    'tacho': 'records_tacho', 'records tacho': 'records_tacho',
    'airswitch': 'records_airswitch', 'records airswitch': 'records_airswitch',
    'time method': 'total_time_method', 'total time method': 'total_time_method',
    'hobbs initial': 'hobbs_initial', 'hobbs start': 'hobbs_initial',
    'tacho initial': 'tacho_initial', 'tacho start': 'tacho_initial',
    'maint source': 'maint_time_source', 'maint time source': 'maint_time_source',
    'maint hours': 'maint_hours_initial', 'maint hours initial': 'maint_hours_initial',
    'fuel l/hr': 'fuel_consumption_per_hour', 'fuel litres/hr': 'fuel_consumption_per_hour',
    'fuel consumption': 'fuel_consumption_per_hour',
    'status': 'status',
    'available': 'is_available_for_hire', 'available for hire': 'is_available_for_hire',
    'leased': 'is_leased', 'is leased': 'is_leased',
}


def _norm(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower().rstrip('*').strip())


def _parse_bool(value, default=False):
    if value in (None, ''):
        return default
    return str(value).strip().upper() in ('TRUE', 'YES', '1', 'Y')


def _parse_decimal(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return None


class Command(BaseCommand):
    help = "Import aircraft from a filled-in aircraft template (.xlsx or .csv)."

    def add_arguments(self, parser):
        parser.add_argument('file', help="Path to the .xlsx or .csv file.")
        parser.add_argument('--club', required=True, help="Target club slug.")
        parser.add_argument('--sheet', default='Aircraft',
                            help="Worksheet name for .xlsx files (default: Aircraft).")
        parser.add_argument('--dry-run', action='store_true',
                            help="Validate and report without saving anything.")
        parser.add_argument('--update', action='store_true',
                            help="Update existing aircraft instead of skipping them.")

    def handle(self, *args, **options):
        try:
            club = Club.objects.get(slug=options['club'])
        except Club.DoesNotExist:
            raise CommandError(f"No club with slug '{options['club']}'.")

        dry_run = options['dry_run']
        update = options['update']

        header, rows = self._read(options['file'], options['sheet'])
        col = self._map_columns(header)

        if 'registration' not in col:
            raise CommandError("File must have a 'Registration' column.")

        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        errors = []

        with transaction.atomic():
            for row_no, values in rows:
                def cell(field):
                    i = col.get(field)
                    if i is None or i >= len(values):
                        return ''
                    v = values[i]
                    return '' if v is None else str(v).strip()

                reg = cell('registration').upper()
                if not reg:
                    continue
                # Skip template example rows
                if 'example' in reg.lower() or reg.startswith('ZK-EG'):
                    continue
                # Skip blank spacer rows
                if not any(str(v or '').strip() for v in values):
                    continue

                # Resolve or create AircraftType
                type_name = cell('type_name')
                aircraft_type = None
                if type_name:
                    aircraft_type, _ = AircraftType.objects.get_or_create(
                        club=club,
                        name=type_name,
                        defaults={'icao_designator': cell('icao_designator')},
                    )

                # Build field dict
                engine_count_raw = cell('engine_count')
                seats_raw = cell('seats')

                fields = {
                    'aircraft_type': aircraft_type,
                    'serial_number': cell('serial_number'),
                    'engine_count': int(engine_count_raw) if engine_count_raw.isdigit() else 1,
                    'seats': int(seats_raw) if seats_raw.isdigit() else 2,
                    'records_hobbs': _parse_bool(cell('records_hobbs'), default=True),
                    'records_tacho': _parse_bool(cell('records_tacho'), default=False),
                    'records_airswitch': _parse_bool(cell('records_airswitch'), default=False),
                    'total_time_method': cell('total_time_method') or 'hobbs',
                    'hobbs_initial': _parse_decimal(cell('hobbs_initial')),
                    'tacho_initial': _parse_decimal(cell('tacho_initial')),
                    'maint_time_source': cell('maint_time_source') or 'hobbs',
                    'maint_hours_initial': _parse_decimal(cell('maint_hours_initial')),
                    'fuel_consumption_per_hour': _parse_decimal(cell('fuel_consumption_per_hour')) or Decimal('0'),
                    'status': cell('status') or 'online',
                    'is_available_for_hire': _parse_bool(cell('is_available_for_hire'), default=True),
                    'is_leased': _parse_bool(cell('is_leased'), default=False),
                }

                existing = Aircraft.objects.filter(club=club, registration=reg).first()

                if existing:
                    if update:
                        if not dry_run:
                            for k, v in fields.items():
                                setattr(existing, k, v)
                            existing.save()
                        stats['updated'] += 1
                        self.stdout.write(f"  {'[DRY RUN] ' if dry_run else ''}Updated: {reg}")
                    else:
                        stats['skipped'] += 1
                        self.stdout.write(f"  Skipped (exists): {reg}")
                else:
                    if not dry_run:
                        Aircraft.objects.create(club=club, registration=reg, **fields)
                    stats['created'] += 1
                    self.stdout.write(f"  {'[DRY RUN] ' if dry_run else ''}Created: {reg}")

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write('')
        head = "DRY RUN — nothing was saved" if dry_run else "Import complete"
        self.stdout.write(self.style.MIGRATE_HEADING(head))
        self.stdout.write(
            f"  Created: {stats['created']}\n"
            f"  Updated: {stats['updated']}\n"
            f"  Skipped: {stats['skipped']}"
        )
        if errors:
            self.stdout.write(self.style.ERROR(f"\n  Errors ({len(errors)}):"))
            for row_no, msg in errors:
                self.stdout.write(f"    row {row_no}: {msg}")

    # ── File reading ──────────────────────────────────────────────────────────

    def _read(self, path, sheet_name):
        if path.lower().endswith(('.xlsx', '.xlsm')):
            return self._read_xlsx(path, sheet_name)
        return self._read_csv(path)

    def _read_xlsx(self, path, sheet_name):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_name.strip().lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active
        raw = [(i, list(r)) for i, r in enumerate(ws.iter_rows(values_only=True), start=1)]
        return self._split_header(raw)

    def _read_csv(self, path):
        try:
            with open(path, newline='', encoding='utf-8-sig') as fh:
                raw = [(i, row) for i, row in enumerate(csv.reader(fh), start=1)]
            return self._split_header(raw)
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")

    def _split_header(self, raw):
        for idx, (row_no, values) in enumerate(raw):
            norms = {_norm(v) for v in values}
            if 'registration' in norms or 'rego' in norms or 'reg' in norms:
                return list(values), raw[idx + 1:]
        raise CommandError("Could not find a header row containing a 'Registration' column.")

    def _map_columns(self, header):
        col = {}
        for i, h in enumerate(header):
            field = HEADER_MAP.get(_norm(h))
            if field and field not in col:
                col[field] = i
        return col
