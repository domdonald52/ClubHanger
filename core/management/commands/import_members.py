"""
Import members from a filled-in ClubHangar import template (the 'Members'
sheet) or a CSV with the same headers. Built for the Paper Aviator migration,
where the same shaped file is re-imported many times before go-live.

    python manage.py import_members members.xlsx --club migration-test --dry-run
    python manage.py import_members members.xlsx --club migration-test

Design notes (decisions pinned with the product owner):
  * Re-runnable / idempotent: members are upserted on EMAIL (the natural key),
    so importing the same file twice updates rather than duplicates.
  * --dry-run validates and reports inside a transaction, then rolls back.
  * The whole import is one atomic transaction: a hard error rolls the lot back.
  * Row-level problems are collected and reported together, not fail-fast.
  * Roles: everyone is assigned the club's default Member role; instructor/admin
    roles are set by hand afterwards. The template 'Role' column is ignored.
  * No email in the file -> a placeholder first.last@migrated.invalid is synthed
    and flagged for cleanup. No welcome/invite emails are ever sent on import.
  * Opening balance, Medical/BFR expiry columns are ignored here — they belong to
    the later financial and credentials passes.
"""
import csv
import re
import secrets
from datetime import date, datetime

from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import Account, Club, ClubMember, Role


# Normalised header -> ClubMember/User field. Headers are lower-cased, the
# trailing '*' (required marker) stripped, and internal whitespace collapsed.
HEADER_MAP = {
    'last name': 'last_name', 'surname': 'last_name',
    'first name': 'first_name', 'firstname': 'first_name', 'given name': 'first_name',
    'email': 'email', 'email address': 'email',
    'caa number': 'caa_number', 'caa': 'caa_number',
    'mobile': 'phone_mobile', 'mobile phone': 'phone_mobile', 'phone mobile': 'phone_mobile',
    'home phone': 'phone_home', 'phone home': 'phone_home',
    'date of birth': 'date_of_birth', 'dob': 'date_of_birth',
    'standing': 'standing', 'membership standing': 'standing',
    'subscription expires': 'subscription_expires',
    'subscription expiry': 'subscription_expires', 'sub expiry': 'subscription_expires',
}

# Columns we knowingly ignore in this pass (reported once so the operator knows).
IGNORED_HEADERS = {'role', 'medical expiry', 'bfr expiry',
                   'opening balance ($)', 'opening balance'}

DATE_FIELDS = {'date_of_birth', 'subscription_expires'}

# Map loose PA standing wording onto ClubMember.STANDING_CHOICES.
STANDING_MAP = {
    'active': 'active', 'current': 'active', 'financial': 'active', 'life': 'active',
    'pending': 'pending', 'suspended': 'suspended', 'lapsed': 'lapsed',
    'resigned': 'resigned', 'transferred': 'transferred',
    'non-member': 'non_member', 'non member': 'non_member', 'nonmember': 'non_member',
}
VALID_STANDINGS = {c[0] for c in ClubMember.STANDING_CHOICES}


def _norm_header(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower().rstrip('*').strip())


def _parse_date(value):
    """Return a date, or None for blanks. Raises ValueError on bad input."""
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unrecognised date '{text}' (use YYYY-MM-DD)")


class Command(BaseCommand):
    help = "Import members from a filled-in 'Members' template (.xlsx) or CSV."

    def add_arguments(self, parser):
        parser.add_argument('file', help="Path to the .xlsx or .csv file.")
        parser.add_argument('--club', required=True, help="Target club slug.")
        parser.add_argument('--sheet', default='Members',
                            help="Worksheet name for .xlsx files (default: Members).")
        parser.add_argument('--dry-run', action='store_true',
                            help="Validate and report, then roll back without saving.")

    def handle(self, *args, **options):
        try:
            club = Club.objects.get(slug=options['club'])
        except Club.DoesNotExist:
            raise CommandError(f"No club with slug '{options['club']}'. "
                               f"Run setup_test_club first.")

        member_role = (Role.objects.filter(club=club, system_role_type=Role.SYSTEM_MEMBER).first()
                       or Role.objects.filter(club=club, name__iexact='member').first())
        if member_role is None:
            raise CommandError(
                f"Club '{club.slug}' has no Member role. "
                f"Run: manage.py setup_defaults --club {club.slug}")

        header, rows = self._read(options['file'], options['sheet'])
        col = self._map_columns(header)
        if 'last_name' not in col or 'email' not in col:
            raise CommandError("File needs at least 'Last name' and 'Email' columns.")

        ignored_present = sorted(IGNORED_HEADERS & {_norm_header(h) for h in header})
        if ignored_present:
            self.stdout.write(self.style.NOTICE(
                "Ignoring columns (handled in later passes): " + ', '.join(ignored_present)))

        parsed, errors = self._parse_rows(rows, col)

        User = get_user_model()
        stats = {'created': 0, 'updated': 0, 'users_created': 0, 'accounts_created': 0}
        placeholders = []

        with transaction.atomic():
            for row_no, rec in parsed:
                email = rec['email']
                if not email:
                    email = self._placeholder_email(User, rec['first_name'], rec['last_name'])
                    placeholders.append((row_no, f"{rec['first_name']} {rec['last_name']}", email))

                user = User.objects.filter(email__iexact=email).first()
                if user is None:
                    user = User.objects.create(
                        username=self._unique_username(User, email),
                        email=email,
                        first_name=rec['first_name'],
                        last_name=rec['last_name'],
                        password=make_password(secrets.token_hex(20)),  # unusable until go-live
                    )
                    stats['users_created'] += 1
                else:
                    changed = []
                    if rec['first_name'] and user.first_name != rec['first_name']:
                        user.first_name = rec['first_name']; changed.append('first_name')
                    if rec['last_name'] and user.last_name != rec['last_name']:
                        user.last_name = rec['last_name']; changed.append('last_name')
                    if changed:
                        user.save(update_fields=changed)

                defaults = {'role': member_role}
                for field in ('caa_number', 'phone_mobile', 'phone_home',
                              'date_of_birth', 'standing', 'subscription_expires'):
                    if field in rec:
                        defaults[field] = rec[field]
                cm, made = ClubMember.objects.update_or_create(
                    club=club, user=user, defaults=defaults)
                stats['created' if made else 'updated'] += 1

                _, acct_made = Account.objects.get_or_create(club_member=cm)
                if acct_made:
                    stats['accounts_created'] += 1

            if options['dry_run']:
                transaction.set_rollback(True)

        self._report(options['dry_run'], stats, placeholders, errors, len(parsed))

    # ── File reading ──────────────────────────────────────────────────────────
    def _read(self, path, sheet_name):
        """Return (header_list, list_of_(row_number, values))."""
        if path.lower().endswith(('.xlsx', '.xlsm')):
            return self._read_xlsx(path, sheet_name)
        return self._read_csv(path)

    def _read_xlsx(self, path, sheet_name):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_name.strip().lower():
                ws = wb[name]; break
        if ws is None:
            ws = wb.active
        raw = [(i, list(r)) for i, r in enumerate(ws.iter_rows(values_only=True), start=1)]
        return self._split_header(raw)

    def _read_csv(self, path):
        with open(path, newline='', encoding='utf-8-sig') as fh:
            raw = [(i, row) for i, row in enumerate(csv.reader(fh), start=1)]
        return self._split_header(raw)

    def _split_header(self, raw):
        """Find the header row (first one containing 'email'), return it + the rest."""
        for idx, (row_no, values) in enumerate(raw):
            norms = {_norm_header(v) for v in values}
            if 'email' in norms:
                return list(values), raw[idx + 1:]
        raise CommandError("Could not find a header row containing an 'Email' column.")

    def _map_columns(self, header):
        col = {}
        for i, h in enumerate(header):
            field = HEADER_MAP.get(_norm_header(h))
            if field and field not in col:
                col[field] = i
        return col

    # ── Row parsing ───────────────────────────────────────────────────────────
    def _parse_rows(self, rows, col):
        parsed, errors = [], []
        for row_no, values in rows:
            def cell(field):
                i = col.get(field)
                if i is None or i >= len(values):
                    return ''
                v = values[i]
                return '' if v is None else (v if isinstance(v, (date, datetime)) else str(v).strip())

            last = str(cell('last_name')).strip()
            first = str(cell('first_name')).strip()
            email = str(cell('email')).strip().lower()

            if not any([last, first, email]):
                continue  # blank spacer row
            if email.endswith('@example.com') or last.lower().startswith('e.g.') \
                    or first.lower().startswith('e.g.'):
                continue  # template example / notes row
            if not last or not first:
                errors.append((row_no, "missing first or last name"))
                continue

            rec = {'first_name': first, 'last_name': last, 'email': email}
            try:
                for field in ('caa_number', 'phone_mobile', 'phone_home'):
                    if field in col:
                        rec[field] = str(cell(field)).strip()
                for field in DATE_FIELDS:
                    if field in col:
                        rec[field] = _parse_date(cell(field))
                if 'standing' in col:
                    rec['standing'] = self._map_standing(cell(field='standing'), row_no, errors)
            except ValueError as exc:
                errors.append((row_no, str(exc)))
                continue
            parsed.append((row_no, rec))
        return parsed, errors

    def _map_standing(self, value, row_no, errors):
        text = str(value or '').strip().lower()
        if not text:
            return 'active'
        if text in STANDING_MAP:
            return STANDING_MAP[text]
        if text in VALID_STANDINGS:
            return text
        errors.append((row_no, f"unknown standing '{value}' — defaulted to active"))
        return 'active'

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _placeholder_email(self, User, first, last):
        base = re.sub(r'[^a-z0-9.]', '', f"{first}.{last}".lower()).strip('.') or 'member'
        candidate = f"{base}@migrated.invalid"
        n = 2
        while User.objects.filter(email__iexact=candidate).exists():
            candidate = f"{base}{n}@migrated.invalid"; n += 1
        return candidate

    def _unique_username(self, User, email):
        base = re.sub(r'[^A-Za-z0-9.@+_-]', '', email.split('@')[0])[:30] or 'member'
        username, n = base, 1
        while User.objects.filter(username=username).exists():
            suffix = str(n)
            username = base[:30 - len(suffix)] + suffix
            n += 1
        return username

    # ── Reporting ─────────────────────────────────────────────────────────────
    def _report(self, dry_run, stats, placeholders, errors, total):
        self.stdout.write('')
        head = "DRY RUN — nothing was saved" if dry_run else "Import complete"
        self.stdout.write(self.style.MIGRATE_HEADING(head))
        self.stdout.write(
            f"  Rows processed:    {total}\n"
            f"  Members created:   {stats['created']}\n"
            f"  Members updated:   {stats['updated']}\n"
            f"  New user logins:   {stats['users_created']}\n"
            f"  Accounts created:  {stats['accounts_created']}")

        if placeholders:
            self.stdout.write(self.style.WARNING(
                f"\n  Placeholder emails synthed ({len(placeholders)}) — fix before go-live:"))
            for row_no, who, email in placeholders:
                self.stdout.write(f"    row {row_no}: {who} -> {email}")

        if errors:
            self.stdout.write(self.style.ERROR(f"\n  Rows with problems ({len(errors)}):"))
            for row_no, msg in errors:
                self.stdout.write(f"    row {row_no}: {msg}")
        else:
            self.stdout.write(self.style.SUCCESS("\n  No row errors."))
