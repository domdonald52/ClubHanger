import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import transaction
from datetime import datetime, timedelta, time, date

logger = logging.getLogger('clubhangar.audit')
from .models import (Club, ClubMember, Booking, Aircraft, AircraftType, Role, FlightType, BlockOutType,
                     SlotWatch, InstructorGrade, AircraftSurchargeType,
                     Aerodrome, FuelSurchargeRate, Invoice, InvoiceLineItem,
                     FlightCompletion, AircraftMaintenanceItem, ChargeRate, FlightChargeItem,
                     FlightLandingEntry, AccountTransaction, ClubConfig,
                     OccurrenceReport, OccurrenceType, OccurrenceAction, OccurrenceAuditEntry,
                     ContactType, MembershipHistoryEntry, VoucherType,
                     create_maint_log_entry, FlyingBudget, FeedbackMessage)
from .availability import find_available_slots, get_date_range
from .services import booking_service
from .services import availability_service
from .services import charging_service
from .services import qualification_service
from .permissions import require_staff, require_admin, require_manage, require_staff_api, require_admin_api


def _aware(dt):
    """Make a naive datetime timezone-aware in the active timezone."""
    if dt is not None and timezone.is_naive(dt):
        return timezone.make_aware(dt)
    return dt


def _audit(booking, user, event_type, notes='', field_name='', old_value='', new_value=''):
    """Thin wrapper — logic lives in booking_service.audit."""
    booking_service.audit(booking, user, event_type, notes=notes,
                          field_name=field_name, old_value=old_value, new_value=new_value)


def _blockout_check(club, aircraft, instructor, start_dt, end_dt, actor, override, exclude_booking_id=None):
    """Thin wrapper — logic lives in booking_service.check_blockout."""
    return booking_service.check_blockout(
        club, aircraft, instructor, start_dt, end_dt, actor, override,
        exclude_booking_id=exclude_booking_id
    )


@login_required
def index(request):
    memberships = ClubMember.objects.filter(user=request.user).select_related('club', 'role').order_by('club__name')
    if not memberships.exists():
        return render(request, 'core/no_access.html')
    if memberships.count() == 1:
        m = memberships.first()
        if m.is_staff:
            return redirect('core:gantt_day', club_slug=m.club.slug)
        return redirect('core:app_home', club_slug=m.club.slug)
    return render(request, 'core/club_select.html', {'memberships': memberships})


def get_config(club):
    """Return the club's config, creating a default if missing."""
    from .models import ClubConfig
    config, _ = ClubConfig.objects.get_or_create(club=club)
    return config


@login_required
def gantt_day(request, club_slug, year=None, month=None, day=None):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    config = get_config(club)

    if not year:
        today = timezone.localdate()
        year, month, day = today.year, today.month, today.day
    
    selected_date = datetime(year, month, day).date()
    day_start = _aware(datetime.combine(selected_date, config.operating_hours_start))
    day_end = _aware(datetime.combine(selected_date, config.operating_hours_end))
    slot_minutes = config.time_slot_interval
    
    # All instructors including inactive — ghost rows keep conflicted bookings visible.
    instructors = ClubMember.objects.filter(
        club=club, is_on_instructor_roster=True
    ).select_related('user').order_by('standing', 'user__last_name')
    
    aircraft_list = Aircraft.objects.filter(club=club, status='online').order_by('registration')
    all_aircraft = Aircraft.objects.filter(club=club).order_by('registration')
    
    # Get all bookings for day
    bookings = Booking.objects.filter(
        club=club,
        scheduled_start__gte=day_start,
        scheduled_start__lt=day_end
    ).exclude(status='cancelled').select_related('member__user', 'aircraft', 'instructor', 'confirmed_by', 'flight_type', 'flight_completion', 'client')
    from django.db.models import Count as _GCount
    bookings = bookings.annotate(watcher_count=_GCount('watchers'))

    # Pixel geometry for absolute-positioned pills
    total_minutes = int((day_end - day_start).total_seconds() // 60)
    _explicit_zoom = request.GET.get('zoom', '')
    if _explicit_zoom:
        px_per_min = float(_explicit_zoom)
    else:
        # Auto-fit to viewport if browser has sent us the width cookie
        _vw = int(request.COOKIES.get('vw', 0) or 0)
        if _vw > 0 and total_minutes > 0:
            # Reserve ~175px for row labels and ~50px for padding/scrollbar gutter
            _avail = max(400, _vw - 225)
            px_per_min = round(max(0.8, min(3.0, _avail / total_minutes)), 3)
        else:
            px_per_min = 2.0
    # Atypical-hours boundaries in pixels from day_start (for calendar shading)
    _slot_win_s, _slot_win_e = config.slot_window()
    typ_start_dt = _aware(datetime.combine(selected_date, _slot_win_s))
    typ_end_dt = _aware(datetime.combine(selected_date, _slot_win_e))
    track_width = int(total_minutes * px_per_min)
    typical_start_px = max(0, int((typ_start_dt - day_start).total_seconds() / 60 * px_per_min))
    typical_end_px = min(track_width, int((typ_end_dt - day_start).total_seconds() / 60 * px_per_min))

    # Column ticks (one per slot interval) for the header + gridlines
    ticks = []
    t = day_start
    while t <= day_end:
        offset = int((t - day_start).total_seconds() // 60 * px_per_min)
        ticks.append({'label': t.strftime('%H:%M'), 'left': offset})
        t += timedelta(minutes=slot_minutes)

    PILL_GAP = 2  # px gap either side so adjacent pills don't butt up

    def _check_live_conflict(b):
        """Live check for all conflict types. Returns (has_conflict, reason_str, issue_types_list)."""
        issues = []

        if not b.blockout_override:
            if b.blockout_conflict:
                issues.append(('blockout', b.blockout_conflict_reason or 'Block-out conflict'))
            else:
                for bo in day_blockouts:
                    if not bo.affects_booking(b):
                        continue
                    iv = bo.interval_on(selected_date)
                    if iv and iv[0] < b.scheduled_end and iv[1] > b.scheduled_start:
                        tname = bo.blockout_type.name if bo.blockout_type else (bo.label or 'block-out')
                        issues.append(('blockout', f"Overlaps {tname}"))
                        break

        if b.member and not b.member.is_current:
            issues.append(('member', f"Member {b.member.get_standing_display()}"))

        if b.aircraft and b.aircraft.status == 'retired':
            issues.append(('aircraft', f"Aircraft {b.aircraft.registration} is retired"))

        if b.instructor_id:
            instr_cm = next((i for i in instructors if i.user_id == b.instructor_id), None)
            if instr_cm:
                roster = _av_cache.get(instr_cm.id)
                # None = no schedule declared = always available (per the model).
                # Only flag when the instructor HAS windows and none apply (False).
                if roster is False:
                    issues.append(('instructor_roster', 'Instructor off roster'))

        if not issues:
            return False, '', []
        return True, '; '.join(r for _, r in issues), [t for t, _ in issues]

    def booking_geometry(b):
        # Completed bookings that returned early: display ends at arrived_at, not scheduled_end
        _arrived = getattr(b, 'arrived_at', None)
        effective_end = (
            _arrived if (b.status == 'completed' and _arrived and _arrived < b.scheduled_end)
            else b.scheduled_end
        )
        start_min = max(0, int((b.scheduled_start - day_start).total_seconds() // 60))
        end_min = min(total_minutes, int((effective_end - day_start).total_seconds() // 60))
        dur_min = max(slot_minutes, end_min - start_min)
        left = int(start_min * px_per_min)
        width = int(dur_min * px_per_min)
        local_start = timezone.localtime(b.scheduled_start)
        local_end = timezone.localtime(effective_end)
        desc = getattr(b, 'description', '') or getattr(b, 'notes', '') or ''
        member_name = ''
        if b.member and b.member.user:
            member_name = f"{b.member.user.first_name} {b.member.user.last_name}".strip()
        in_conflict, conflict_reason, issue_types = _check_live_conflict(b)
        _decl_pending = (
            getattr(b.flight_type, 'requires_declaration', False) and
            b.status in ('pending', 'confirmed') and
            not (hasattr(b, 'declaration') and not b.declaration.is_draft)
        )
        return {
            'id': b.id,
            'left': left + PILL_GAP,
            'width': max(width - PILL_GAP * 2, 4),
            'status': b.status,
            'conflict_kind': 'conflict' if in_conflict else ('declaration' if _decl_pending else ''),
            'is_dual': bool(b.instructor_id),
            'member_name': member_name or 'Unknown',
            'member_user_id': b.member.user_id if b.member else '',
            'description': desc,
            'aircraft_id': b.aircraft_id,
            'aircraft_reg': b.aircraft.registration if b.aircraft else '',
            'instructor_id': b.instructor_id,
            'instructor_name': (f"{b.instructor.first_name} {b.instructor.last_name}".strip()
                                if b.instructor else ''),
            'start_label': local_start.strftime('%H:%M'),
            'end_label': local_end.strftime('%H:%M'),
            'start_iso': b.scheduled_start.isoformat(),
            'duration_min': dur_min,
            'conflict': in_conflict,
            'override': b.blockout_override,
            'conflict_reason': conflict_reason,
            'issue_types': ','.join(issue_types),
            'flight_type_id': b.flight_type_id or '',
            'flight_type_name': b.flight_type.name if b.flight_type else '',
            'member_not_current': b.member is not None and not b.member.is_current,
            'member_standing': b.member.standing if b.member else '',
            'records_hobbs':      b.aircraft.records_hobbs      if b.aircraft else False,
            'records_tacho':      b.aircraft.records_tacho      if b.aircraft else False,
            'records_airswitch':  b.aircraft.records_airswitch  if b.aircraft else False,
            'total_time_method':  b.aircraft.total_time_method  if b.aircraft else '',
            'maint_time_source':  b.aircraft.maint_time_source  if b.aircraft else '',
            'client_id':   b.client_id or '',
            'client_name': b.client.name if b.client else '',
            'billed_to':   b.billed_to or '',
            'paid': (getattr(getattr(b, 'flight_completion', None), 'paid_at', None) is not None),
            'decl_pending': _decl_pending,
            'decl_url': (
                f'/manage/{club_slug}/bookings/{b.id}/declaration/'
                if getattr(b.flight_type, 'requires_declaration', False) and b.status in ('pending', 'confirmed')
                else ''
            ),
            'watcher_count': getattr(b, 'watcher_count', 0),
        }

    # Apply filters — suppress when arriving via booking deep-link (?book=1) because
    # that URL shares 'aircraft' and 'instructor' params with the modal pre-fill,
    # which would otherwise silently filter out all other bookings from the view.
    is_deeplink = request.GET.get('book') == '1'
    show_pending = request.GET.get('pending_only') == 'on'
    filter_instructor = None if is_deeplink else request.GET.get('instructor')
    filter_aircraft = None if is_deeplink else request.GET.get('aircraft')

    # Block-outs for this day, as geometry bands per resource
    from .models import BlockOut
    day_blockouts = [
        bo for bo in BlockOut.objects.filter(club=club).prefetch_related('aircraft', 'instructors', 'blockout_type')
        if bo.applies_on(selected_date)
    ]

    def band_geometry(bo):
        if bo.all_day or not (bo.start_time and bo.end_time):
            bo_start, bo_end = day_start, day_end
        else:
            bo_start = _aware(datetime.combine(selected_date, bo.start_time))
            bo_end = _aware(datetime.combine(selected_date, bo.end_time))
        # clamp to operating window
        bo_start = max(bo_start, day_start)
        bo_end = min(bo_end, day_end)
        if bo_end <= bo_start:
            return None
        left = int((bo_start - day_start).total_seconds() // 60 * px_per_min)
        width = int((bo_end - bo_start).total_seconds() // 60 * px_per_min)
        is_hard = (not bo.blockout_type) or bo.blockout_type.effective_is_hard
        return {
            'left': left, 'width': width,
            'label': (bo.blockout_type.name if bo.blockout_type else (bo.label or 'Blocked')),
            'color': (bo.blockout_type.color if bo.blockout_type else '#9aa3ad'),
            'start_label': bo_start.strftime('%H:%M'),
            'end_label': bo_end.strftime('%H:%M'),
            'is_hard': is_hard,
        }

    def bands_for_aircraft(ac):
        out = []
        for bo in day_blockouts:
            if bo.affects_aircraft(ac):
                g = band_geometry(bo)
                if g:
                    out.append(g)
        return out

    def bands_for_instructor(user):
        out = []
        for bo in day_blockouts:
            if bo.affects_instructor(user):
                g = band_geometry(bo)
                if g:
                    out.append(g)
        return out

    # Include inactive instructors so ghost rows can show conflicted bookings.
    # Re-query without is_active filter (the base query already has all instructors).
    from .models import InstructorAvailability
    _av_cache = {}  # club_member_id → bool | None
    for instr in instructors:
        windows = list(InstructorAvailability.objects.filter(club_member=instr))
        if not windows:
            _av_cache[instr.id] = None  # no schedule declared — treated as always available
        else:
            _av_cache[instr.id] = any(w.applies_on(selected_date) for w in windows)

    # Build grid data
    instructor_rows = []
    for instr in instructors:
        instr_bookings = bookings.filter(instructor=instr.user)
        if show_pending:
            instr_bookings = instr_bookings.filter(status='pending')
        if filter_aircraft:
            instr_bookings = instr_bookings.filter(aircraft_id=filter_aircraft)

        on_roster = _av_cache.get(instr.id)   # None / True / False
        has_bookings = instr_bookings.exists()
        # Operationally active: not resigned (standing may be suspended/lapsed but
        # could still have bookings to honour; resigned is definitive departure)
        is_active = instr.standing not in ('resigned',)

        # None means no availability schedule declared → treat as always available
        normal_show = is_active and on_roster is not False
        ghost = (not normal_show) and has_bookings
        if not normal_show and not ghost:
            continue

        if instr.standing in ('resigned', 'lapsed'):
            ghost_reason = 'inactive'
        elif on_roster is False:
            ghost_reason = 'off_roster'
        else:
            ghost_reason = None

        bands = bands_for_instructor(instr.user)
        instructor_rows.append({
            'type': 'instructor',
            'label': f"{instr.user.first_name} {instr.user.last_name}".strip() or instr.user.username,
            'row_key': f"instructor:{instr.user.id}",
            'resource_id': instr.user.id,
            'detail_id': instr.id,
            'pills': [booking_geometry(b) for b in instr_bookings],
            'bands': bands,
            'is_current_user': instr.user == request.user,
            'on_roster': on_roster,
            'ghost': ghost,
            'ghost_reason': ghost_reason,  # 'inactive' | 'off_roster' | None
            'has_hard_blockout': any(b['is_hard'] for b in bands),
            'has_soft_blockout': any(not b['is_hard'] for b in bands),
        })

    # Ghost rows for users who had the instructor role removed but still have bookings today
    shown_instructor_ids = {row['resource_id'] for row in instructor_rows}
    ex_instructor_ids = {
        b.instructor_id for b in bookings
        if b.instructor_id and b.instructor_id not in shown_instructor_ids
    }
    if ex_instructor_ids:
        from .models import User as _User
        ex_users = {u.id: u for u in _User.objects.filter(id__in=ex_instructor_ids)}
        for user_id, user in ex_users.items():
            ex_bookings = bookings.filter(instructor_id=user_id)
            if show_pending:
                ex_bookings = ex_bookings.filter(status='pending')
            if filter_aircraft:
                ex_bookings = ex_bookings.filter(aircraft_id=filter_aircraft)
            if not ex_bookings.exists():
                continue
            instructor_rows.append({
                'type': 'instructor',
                'label': f"{user.first_name} {user.last_name}".strip() or user.username,
                'row_key': f"instructor:{user.id}",
                'resource_id': user.id,
                'detail_id': None,
                'pills': [booking_geometry(b) for b in ex_bookings],
                'bands': [],
                'is_current_user': user == request.user,
                'on_roster': None,
                'ghost': True,
                'ghost_reason': 'role_changed',
                'has_hard_blockout': False,
                'has_soft_blockout': False,
            })

    aircraft_rows = []
    for ac in all_aircraft:
        is_online = ac.status == 'online'
        ac_bookings = bookings.filter(aircraft=ac)
        if show_pending:
            ac_bookings = ac_bookings.filter(status='pending')
        if filter_instructor:
            ac_bookings = ac_bookings.filter(instructor_id=filter_instructor)

        has_bookings = ac_bookings.exists()
        if not is_online and not has_bookings:
            continue  # retired with nothing booked today: omit entirely

        ghost = not is_online
        ac_bands = bands_for_aircraft(ac) if is_online else []
        aircraft_rows.append({
            'type': 'aircraft',
            'label': f"{ac.registration} ({ac.aircraft_type.name if ac.aircraft_type_id else '?'})",
            'row_key': f"aircraft:{ac.id}",
            'resource_id': ac.id,
            'detail_id': ac.id,
            'pills': [booking_geometry(b) for b in ac_bookings],
            'bands': ac_bands,
            'has_blockout': bool(ac_bands),
            'ghost': ghost,
            'ghost_reason': 'retired' if ghost else None,
        })

    # Row-label width: fit the longest instructor/aircraft label (7px/char approx at .76rem)
    all_labels = [r['label'] for r in instructor_rows + aircraft_rows]
    label_w = max(140, min(220, max((len(l) * 7 + 18 for l in all_labels), default=140)))

    # Navigation
    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)
    today = timezone.localdate()

    _members_qs = (ClubMember.objects
                   .filter(club=club, user__isnull=False)
                   .select_related('user')
                   .prefetch_related('account')
                   .order_by('user__last_name', 'user__first_name'))
    members_data = []
    for m in _members_qs:
        try:
            acct_warning = m.account.has_warning
        except Exception:
            acct_warning = False
        # Badge shown next to the member name in the booking modal
        if m.is_current:
            badge = 'current'
        elif m.standing == 'non_member':
            badge = 'non_member'
        else:
            badge = 'lapsed'   # suspended / lapsed / resigned / pending
        members_data.append({
            'id': m.user.id,
            'name': f"{m.user.first_name} {m.user.last_name}".strip() or m.user.username,
            'badge': badge,
            'acct_warning': acct_warning,
        })
    aircraft_data = [
        {'id': a.id, 'reg': a.registration, 'type': a.aircraft_type.name if a.aircraft_type_id else ''} for a in aircraft_list
    ]
    instructors_data = [
        {
            'id': i.user.id,
            'name': f"{i.user.first_name} {i.user.last_name}".strip(),
            'on_roster': _av_cache.get(i.id),  # True / False / None (None = no schedule = always available)
        }
        for i in instructors
    ]
    flight_types_data = [
        {'id': ft.id, 'name': ft.name, 'code': ft.code, 'is_solo': ft.is_solo, 'for_contacts': ft.for_contacts}
        for ft in FlightType.objects.filter(club=club)
    ]
    blockout_types_data = [
        {'id': bt.id, 'name': bt.name, 'target': bt.target}
        for bt in BlockOutType.objects.filter(club=club)
    ]

    # Now-line: pixel offset of current time — always show on today, clamped to track edges
    now_px = None
    if selected_date == today:
        now_dt = timezone.now()
        raw_px = (now_dt - day_start).total_seconds() / 60 * px_per_min
        now_px = max(0, min(track_width - 2, int(raw_px)))

    zoom_param = request.GET.get('zoom', '')
    can_manage = club_member.is_instructor or club_member.is_admin

    # Contacts for trial-flight / non-member client booking (staff only)
    if can_manage:
        from .models import Contact as _Contact
        contacts_data = [
            {
                'id': c.id,
                'name': c.name,
                'type': c.contact_type.name if c.contact_type_id else '',
            }
            for c in _Contact.objects.filter(club=club, converted_to_member__isnull=True)
            .select_related('contact_type')
            .order_by('name')
        ]
        contact_types_data = [
            {'id': ct.id, 'name': ct.name}
            for ct in ContactType.objects.filter(club=club, is_active=True).order_by('name')
        ]
    else:
        contacts_data = []
        contact_types_data = []
    club_phone = getattr(config, 'billing_phone', '') or ''

    # Instructor credential issues for the selected date (shown to staff only)
    instructor_cred_issues = (
        _instructor_cred_issues(club, selected_date) if can_manage else []
    )

    # METAR/TAF — fetched server-side so the pill and TAF are in the HTML on first paint,
    # eliminating the layout shift caused by async JS injection.
    wx_ctx = {}
    try:
        from django.conf import settings as _s
        _api_key = getattr(_s, 'AVWX_API_KEY', '')
        _home_ae = Aerodrome.objects.filter(club=club, is_home=True).first()
        if _api_key and _home_ae and getattr(_home_ae, 'icao_code', ''):
            _raw_m = _fetch_avwx(_home_ae.icao_code, 'metar', _api_key)
            if 'raw' in _raw_m:
                _m = _parse_avwx_metar(_raw_m)
                _raw_t = _fetch_avwx(_home_ae.icao_code, 'taf', _api_key)
                _age = _m['age_minutes']
                _ageStr = (
                    '' if _age is None else
                    f'{_age}min' if _age < 60 else
                    f'{_age//60}h{_age%60}m' if _age % 60 else f'{_age//60}h'
                )
                _wind = ''
                if _m['wind_dir_repr'] and _m['wind_speed']:
                    _wind = f"{_m['wind_dir_repr']}/{_m['wind_speed']}kt"
                elif _m['wind_speed']:
                    _wind = f"{_m['wind_speed']}kt"
                wx_ctx = {
                    'icao': _home_ae.icao_code,
                    'flight_rules': _m['flight_rules'],
                    'fr_cls': (_m['flight_rules'] or '').lower(),
                    'wind': _wind,
                    'qnh': _m['qnh_repr'] or '',
                    'age_str': _ageStr,
                    'stale': _age is not None and _age >= 60,
                    'very_stale': _age is not None and _age >= 120,
                    'raw': _m['raw'],
                    'taf_raw': _raw_t.get('raw', '') if 'raw' in _raw_t else '',
                }
    except Exception:
        pass

    # Active announcements (pinned first) — includes general + event-specific for today
    from .models import Announcement as _Ann
    from django.db.models import Q as _Q2
    announcements_today = list(_Ann.objects.filter(club=club).filter(
        _Q2(expires_at__isnull=True) | _Q2(expires_at__gte=selected_date)
    ).order_by('-is_pinned', 'type', 'created_at'))

    context = {
        'club': club,
        'club_member': club_member,
        'is_instructor': club_member.is_instructor,
        'can_book': True,  # all authenticated club members may create bookings
        'can_manage': can_manage,
        'current_user_id': request.user.id,
        'selected_date': selected_date,
        'today': today,
        'prev_date': prev_date,
        'next_date': next_date,
        'instructor_rows': instructor_rows,
        'aircraft_rows': aircraft_rows,
        'ticks': ticks,
        'track_width': track_width,
        'px_per_min': px_per_min,
        'slot_minutes': slot_minutes,
        'instructors': instructors,
        'aircraft_list': aircraft_list,
        'default_duration': config.default_booking_duration,
        'day_start_iso': day_start.isoformat(),
        'typical_hours_start': _slot_win_s.strftime('%H:%M'),
        'typical_hours_end': _slot_win_e.strftime('%H:%M'),
        'typical_start_px': typical_start_px,
        'typical_end_px': typical_end_px,
        'total_minutes': total_minutes,
        'zoom_param': zoom_param,
        'members_json': members_data,
        'aircraft_json': aircraft_data,
        'instructors_json': instructors_data,
        'flight_types_json': flight_types_data,
        'blockout_types_json': blockout_types_data,
        'contacts_json': contacts_data,
        'contact_types_json': contact_types_data,
        'watched_ids': list(
            SlotWatch.objects.filter(club_member=club_member)
            .values_list('booking_id', flat=True)
        ),
        'now_px': now_px,
        'label_w': label_w,
        'instructor_cred_issues': instructor_cred_issues,
        'announcements_today': announcements_today,
        'club_phone': club_phone,
        'wx': wx_ctx,
    }

    return render(request, 'core/gantt_day.html', context)


@login_required
@require_POST
@transaction.atomic
def reschedule_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if not (club_member.is_admin or club_member.is_instructor):
        return JsonResponse({'error': 'Only instructors and admins can reschedule'}, status=403)

    new_start_str = request.POST.get('new_start')
    if not new_start_str:
        return JsonResponse({'error': 'Missing new_start'}, status=400)

    try:
        new_start_dt  = _aware(datetime.fromisoformat(new_start_str))
        duration_param = request.POST.get('duration')
        duration = int(duration_param) if duration_param else int(
            (booking.scheduled_end - booking.scheduled_start).total_seconds() / 60
        )

        aircraft_id   = request.POST.get('aircraft_id')
        instructor_id = request.POST.get('instructor_id')
        override      = request.POST.get('override') in ('1', 'true', 'on')

        aircraft = None
        if aircraft_id:
            aircraft = Aircraft.objects.filter(id=aircraft_id, club=club).first()
            if not aircraft:
                return JsonResponse({'error': 'Aircraft not found'}, status=404)

        instructor = None
        if instructor_id:
            from .models import User as _User
            instructor = _User.objects.filter(id=instructor_id).first()

        result = booking_service.reschedule(
            booking, club_member, new_start_dt, duration,
            aircraft=aircraft, instructor=instructor, override=override
        )
        if not result.ok:
            status_code = 409 if result.data.get('blockout') else 400
            return JsonResponse({'error': result.error, **result.data}, status=status_code)
        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def _credential_checks(booking):
    """
    Return a list of {label, status ('ok'|'warn'|'info'), detail} for the
    booking's member + flight type. Used at confirmation time.
    """
    from datetime import date as _d
    from django.db.models import Q as _Q

    member = booking.member
    ft = booking.flight_type
    today = _d.today()
    creds = member.user.credentials.all()

    LICENCE_CODES = ('ppl', 'cpl', 'atpl', 'instr_c', 'instr_b', 'instr_a',
                     'instr_d', 'instr_e', 'examiner')
    MEDICAL_CODES = ('medical_c1', 'medical_c2', 'medical_c3', 'dlr9')

    def latest_valid(*codes):
        return (creds
                .filter(credential_type__code__in=codes)
                .filter(_Q(expiry_date__isnull=True) | _Q(expiry_date__gte=today))
                .order_by('-expiry_date')
                .first())

    checks = []

    # ── Medical ──────────────────────────────────────────────────────────────
    med = latest_valid(*MEDICAL_CODES)
    if not med:
        checks.append({'label': 'Medical certificate',
                       'status': 'warn',
                       'detail': 'No current medical certificate on record'})
    elif ft.is_solo and med.credential_type.code == 'medical_c3':
        checks.append({'label': 'Medical certificate',
                       'status': 'warn',
                       'detail': 'Class 3 medical only — Class 2 or better is required for private solo flying'})
    else:
        exp = f', valid to {med.expiry_date}' if med.expiry_date else ''
        checks.append({'label': 'Medical certificate', 'status': 'ok',
                       'detail': f'{med.display_name}{exp}'})

    if ft.is_solo:
        # ── Pilot licence ─────────────────────────────────────────────────────
        licence = latest_valid(*LICENCE_CODES)
        if not licence:
            checks.append({'label': 'Pilot licence',
                           'status': 'warn',
                           'detail': 'No PPL or higher licence on record'})
        else:
            checks.append({'label': 'Pilot licence', 'status': 'ok',
                           'detail': licence.display_name})

        # ── Flight Review (BFR) — every 24 months ────────────────────────────
        fr = latest_valid('fr')
        if not fr:
            checks.append({'label': 'Flight Review (BFR)',
                           'status': 'warn',
                           'detail': 'No current Flight Review on record'})
        else:
            exp = f', valid to {fr.expiry_date}' if fr.expiry_date else ''
            checks.append({'label': 'Flight Review (BFR)', 'status': 'ok',
                           'detail': f'Current{exp}'})

    # ── Age ───────────────────────────────────────────────────────────────────
    if ft.is_solo:
        if member.date_of_birth:
            age = (today - member.date_of_birth).days // 365
            min_age = 16 if ft.is_training else 17
            if age < min_age:
                checks.append({'label': 'Minimum age',
                               'status': 'warn',
                               'detail': f'Member is {age} years old — minimum is {min_age} for this flight type'})
            else:
                checks.append({'label': 'Minimum age', 'status': 'ok',
                               'detail': f'Age {age} — meets minimum of {min_age}'})
        else:
            checks.append({'label': 'Minimum age', 'status': 'info',
                           'detail': 'Date of birth not recorded — cannot verify minimum age requirement'})

    return checks


@login_required
def prev_readings_api(request, booking_id):
    """Return the last recorded meter end readings for the aircraft on this booking."""
    booking = get_object_or_404(Booking, id=booking_id)
    try:
        actor = ClubMember.objects.get(user=request.user, club=booking.club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if not (actor.is_admin or actor.is_instructor):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    from django.db.models import Q as _Q
    prev = (FlightCompletion.objects
            .filter(booking__aircraft=booking.aircraft, booking__club=booking.club)
            .exclude(booking=booking)
            .filter(_Q(hobbs_end__isnull=False) | _Q(tacho_end__isnull=False) | _Q(airswitch_end__isnull=False))
            .order_by('-booking__arrived_at', '-created_at')
            .first())
    ac = booking.aircraft
    def _val(prev_reading, initial):
        v = prev_reading if (prev and prev_reading is not None) else initial
        return float(v) if v is not None else None
    return JsonResponse({
        'hobbs_end':     _val(prev.hobbs_end     if prev else None, ac.hobbs_initial     if ac else None),
        'tacho_end':     _val(prev.tacho_end     if prev else None, ac.tacho_initial     if ac else None),
        'airswitch_end': _val(prev.airswitch_end if prev else None, ac.airswitch_initial if ac else None),
    })


@login_required
def credential_check_api(request, booking_id):
    """Pre-confirmation credential check for staff — member credentials + aircraft maintenance."""
    booking = get_object_or_404(Booking, id=booking_id)
    try:
        actor = ClubMember.objects.get(user=request.user, club=booking.club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if not (actor.is_admin or actor.is_instructor):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    checks = _credential_checks(booking)

    # Aircraft maintenance items with progress
    ac = booking.aircraft
    maint_items = list(AircraftMaintenanceItem.objects.filter(aircraft=ac).order_by('urgency', 'due_date'))
    latest_fc = (FlightCompletion.objects
                 .filter(booking__aircraft=ac)
                 .exclude(hobbs_end__isnull=True)
                 .order_by('-booking__arrived_at', '-created_at')
                 .values('hobbs_end', 'tacho_end').first())
    current_hobbs = float(latest_fc['hobbs_end']) if latest_fc and latest_fc['hobbs_end'] else None

    _today = timezone.localdate()
    maint_data = []
    for m in maint_items:
        progress_pct = None
        detail = ''
        if m.last_completed_date and m.due_date:
            total_days = max(1, (m.due_date - m.last_completed_date).days)
            elapsed = (_today - m.last_completed_date).days
            days_left = (m.due_date - _today).days
            progress_pct = min(100, max(0, round(elapsed / total_days * 100)))
            detail = f'{days_left}d remaining' if days_left >= 0 else f'{abs(days_left)}d overdue'
        elif m.last_completed_hours is not None and m.due_hours is not None and current_hobbs is not None:
            total_h = float(m.due_hours - m.last_completed_hours)
            hrs_left = float(m.due_hours) - current_hobbs
            if total_h > 0:
                progress_pct = min(100, max(0, round((current_hobbs - float(m.last_completed_hours)) / total_h * 100)))
            detail = f'{hrs_left:.1f}h remaining' if hrs_left >= 0 else f'{abs(hrs_left):.1f}h overdue'
        elif m.due_date:
            days_left = (m.due_date - _today).days
            detail = f'{days_left}d remaining' if days_left >= 0 else f'{abs(days_left)}d overdue'
        maint_data.append({
            'name': m.name,
            'urgency': m.urgency,
            'progress_pct': progress_pct,
            'detail': detail,
        })

    return JsonResponse({
        'member': booking.member.user.get_full_name(),
        'flight_type': booking.flight_type.name,
        'is_solo': booking.flight_type.is_solo,
        'checks': checks,
        'has_warnings': any(c['status'] == 'warn' for c in checks),
        'maintenance': maint_data,
        'has_maint_warnings': any(m['urgency'] in ('amber', 'red') for m in maint_data),
        'current_hobbs': current_hobbs,
        'aircraft_reg': ac.registration,
    })


@login_required
@require_POST
def confirm_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if not (club_member.is_admin or club_member.is_instructor):
        return JsonResponse({'error': 'Only instructors and admins can confirm'}, status=403)

    result = booking_service.confirm(booking, request.user)
    if not result.ok:
        return JsonResponse({'error': result.error}, status=400)
    return JsonResponse({'success': True})


@login_required
@require_POST
def depart_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    is_own = (booking.member.user == request.user)
    # Plain members may only depart their own private hire (solo, no instructor) flights
    is_own_solo = (is_own
                   and booking.flight_type and booking.flight_type.is_solo
                   and not booking.instructor_id)
    if not (actor.is_admin or actor.is_instructor or is_own_solo):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    no_decl_reason = request.POST.get('no_declaration_reason', '').strip()
    elig_override = request.POST.get('eligibility_override_reason', '').strip()
    # If member was changed in the dialog without saving first, apply it now
    member_user_id = request.POST.get('member_user_id', '').strip()
    if member_user_id and (actor.is_admin or actor.is_instructor):
        new_member = ClubMember.objects.filter(club=club, user_id=member_user_id).first()
        if new_member and new_member != booking.member:
            booking.member = new_member
            booking.save(update_fields=['member'])
    result = booking_service.depart(booking, request.user, no_decl_reason)
    if not result.ok:
        return JsonResponse({'error': result.error, **result.data}, status=400)
    if elig_override:
        from .models import BookingAuditLog
        BookingAuditLog.objects.create(
            booking=booking, user=request.user,
            event_type='field_changed',
            notes=f'Eligibility override: {elig_override}',
        )
    return JsonResponse({'success': True, 'status': 'departed'})


@login_required
@require_POST
@transaction.atomic
def checkin_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    if not (actor.is_admin or actor.is_instructor):
        return JsonResponse({'error': 'Instructors only'}, status=403)

    result = booking_service.check_in(
        booking, request.user,
        outcome=request.POST.get('outcome', 'completed'),
        outcome_notes=request.POST.get('outcome_notes', '').strip(),
        hobbs_start=request.POST.get('hobbs_start', '').strip() or None,
        hobbs_end=request.POST.get('hobbs_end', '').strip() or None,
        tacho_start=request.POST.get('tacho_start', '').strip() or None,
        tacho_end=request.POST.get('tacho_end', '').strip() or None,
        airswitch_start=request.POST.get('airswitch_start', '').strip() or None,
        airswitch_end=request.POST.get('airswitch_end', '').strip() or None,
    )
    if not result.ok:
        return JsonResponse({'error': result.error}, status=400)
    _special_fields = ['time_if_simulated', 'time_if_actual', 'time_night', 'time_low_flying', 'time_terrain_awareness']
    if any(request.POST.get(f, '').strip() for f in _special_fields):
        try:
            _fc = FlightCompletion.objects.get(booking=booking)
            _total = float(_fc.actual_flight_hours or 0)
            for _sf in _special_fields:
                _sv = request.POST.get(_sf, '').strip()
                if _sv:
                    _fval = float(_sv)
                    setattr(_fc, _sf, _fval if _fval <= _total else _total)
                else:
                    setattr(_fc, _sf, None)
            _fc.save(update_fields=_special_fields)
        except (FlightCompletion.DoesNotExist, ValueError):
            pass
    return JsonResponse({'success': True, 'status': 'completed',
                         'charges_url': result.data.get('charges_url', '')})


@login_required
@require_POST
def toggle_watch(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    try:
        member = ClubMember.objects.get(user=request.user, club=booking.club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not a member'}, status=403)
    if booking.member == member:
        return JsonResponse({'error': "Can't watch your own booking"}, status=400)
    watch, created = SlotWatch.objects.get_or_create(booking=booking, club_member=member)
    if not created:
        watch.delete()
        return JsonResponse({'watching': False})
    return JsonResponse({'watching': True})


@login_required
@require_POST
def reject_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    is_own = booking.member == club_member
    if not (club_member.is_admin or club_member.is_instructor or is_own):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    result = booking_service.cancel(
        booking, request.user,
        release_slot=request.POST.get('release') == '1',
        reason=request.POST.get('reason', ''),
        reason_other=request.POST.get('reason_other', ''),
    )
    if not result.ok:
        return JsonResponse({'error': result.error}, status=400)
    return JsonResponse({'success': True})


@login_required
@require_POST
@transaction.atomic
def create_booking(request):
    try:
        actor = ClubMember.objects.filter(user=request.user).first()
        if not actor:
            return JsonResponse({'error': 'Not a club member'}, status=403)

        # Use club_slug from POST (sent by Gantt JS) to ensure the right club is used
        # when the user is a member of multiple clubs.
        _club_slug = request.POST.get('club_slug', '').strip()
        if _club_slug:
            from .models import Club as _Club
            _slug_club = _Club.objects.filter(slug=_club_slug).first()
            if _slug_club:
                _slug_actor = ClubMember.objects.filter(user=request.user, club=_slug_club).first()
                if _slug_actor:
                    actor = _slug_actor

        club     = actor.club
        config   = get_config(club)
        aircraft_id   = request.POST.get('aircraft_id')
        start_time    = request.POST.get('start_time')
        duration      = int(request.POST.get('duration') or config.default_booking_duration)
        instructor_id = request.POST.get('instructor_id')
        member_id     = request.POST.get('member_id')
        description   = request.POST.get('description', '')
        override      = request.POST.get('override') in ('1', 'true', 'on')

        if not aircraft_id or not start_time:
            return JsonResponse({'error': 'Missing aircraft or start time'}, status=400)

        aircraft = Aircraft.objects.get(id=aircraft_id, club=club)
        start_dt = _aware(datetime.fromisoformat(start_time))
        end_dt   = start_dt + timedelta(minutes=duration)

        if start_dt < timezone.now():
            return JsonResponse({'error': 'Bookings cannot be made in the past'}, status=400)

        # Who is the booking FOR?
        booking_member = actor
        if member_id and (actor.is_admin or actor.is_instructor):
            from .models import User as _User
            target_user = _User.objects.filter(id=member_id).first()
            if target_user:
                booking_member = ClubMember.objects.filter(user=target_user, club=club).first() or actor

        instructor = None
        if instructor_id:
            from .models import User as _User
            instructor = _User.objects.filter(id=instructor_id).first()

        # Booking block check — applied to the member the booking is FOR
        from .permissions import check_booking_block
        bb_blocked, bb_msg = check_booking_block(booking_member, config)
        if bb_blocked:
            return JsonResponse({'error': bb_msg, 'booking_blocked': True}, status=403)

        # Flight type resolution
        flight_type_id = request.POST.get('flight_type_id')
        if flight_type_id:
            flight_type = FlightType.objects.filter(club=club, id=flight_type_id).first()
        elif not instructor:
            flight_type = (FlightType.objects.filter(club=club, is_solo=True, code='student_solo').first()
                           or FlightType.objects.filter(club=club, is_solo=True).first())
        else:
            flight_type = (FlightType.objects.filter(club=club, code='student_dual').first()
                           or FlightType.objects.filter(club=club, is_solo=False).first())
        if not flight_type:
            flight_type = FlightType.objects.filter(club=club).first()
        if not flight_type:
            return JsonResponse({'error': 'No flight types configured'}, status=400)

        result = booking_service.create(
            club, actor, aircraft, start_dt, end_dt, flight_type,
            instructor=instructor, booking_member=booking_member,
            description=description, override=override,
        )
        if not result.ok:
            status_code = 409 if result.data.get('blockout') else 400
            return JsonResponse({'error': result.error, **result.data}, status=status_code)

        # Attach non-member client (trial flight, Young Eagles, etc.) — staff only
        if actor.is_admin or actor.is_instructor:
            client_id = request.POST.get('client_id', '').strip()
            billed_to = request.POST.get('billed_to', '').strip()
            if client_id:
                from .models import Contact as _Contact
                client_obj = _Contact.objects.filter(id=client_id, club=club).first()
                if client_obj:
                    created_booking = Booking.objects.get(id=result.data['booking_id'])
                    created_booking.client = client_obj
                    created_booking.billed_to = billed_to
                    created_booking.save(update_fields=['client', 'billed_to'])

        return JsonResponse({'success': True, 'booking_id': result.data['booking_id']})

    except Aircraft.DoesNotExist:
        return JsonResponse({'error': 'Aircraft not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)
    except Exception as e:
        print(f"Error creating booking: {e}")
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@login_required
@require_POST
@transaction.atomic
def quick_create_contact(request):
    """AJAX: create a Contact from the calendar booking modal (staff/instructor only)."""
    actor = ClubMember.objects.filter(user=request.user).first()
    if not actor or not (actor.is_admin or actor.is_instructor):
        return JsonResponse({'error': 'Not authorized'}, status=403)
    club = actor.club
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    from .models import Contact as _Contact
    email    = request.POST.get('email', '').strip()
    phone    = request.POST.get('phone', '').strip()
    ctype_id = request.POST.get('contact_type', '').strip()
    contact_type = ContactType.objects.filter(id=ctype_id, club=club).first() if ctype_id else None
    contact = _Contact.objects.create(
        club=club, name=name, email=email, phone=phone,
        contact_type=contact_type, created_by=request.user,
    )
    return JsonResponse({
        'success': True,
        'id': contact.id,
        'name': contact.name,
        'type': contact.contact_type.name if contact.contact_type else '',
    })


@login_required
@require_POST
@transaction.atomic
def edit_booking(request, booking_id):
    """Full edit from the modal: member, aircraft, instructor, time, duration, description."""
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    is_own_pending = (booking.member == actor and booking.status == 'pending')
    if not (actor.is_admin or actor.is_instructor or is_own_pending):
        if booking.member == actor and booking.status == 'confirmed':
            club_phone = getattr(get_config(club), 'billing_phone', '') or ''
            contact_hint = f' Call the club on {club_phone}.' if club_phone else ''
            return JsonResponse({
                'error': f'This booking is confirmed — to make changes, please contact the club.{contact_hint}',
                'confirmed_contact': True,
            }, status=403)
        return JsonResponse({'error': 'Not authorized'}, status=403)

    try:
        config        = get_config(club)
        aircraft_id   = request.POST.get('aircraft_id')
        start_time    = request.POST.get('start_time')
        duration      = int(request.POST.get('duration') or config.default_booking_duration)
        instructor_id = request.POST.get('instructor_id')
        member_id     = request.POST.get('member_id')
        description   = request.POST.get('description', '')
        override      = request.POST.get('override') in ('1', 'true', 'on')

        aircraft = Aircraft.objects.get(id=aircraft_id, club=club) if aircraft_id else booking.aircraft
        start_dt = _aware(datetime.fromisoformat(start_time)) if start_time else booking.scheduled_start
        end_dt   = start_dt + timedelta(minutes=duration)

        instructor = None
        if instructor_id:
            from .models import User as _User
            instructor = _User.objects.filter(id=instructor_id).first()

        booking_member = None
        if member_id:
            from .models import User as _User
            tu = _User.objects.filter(id=member_id).first()
            if tu:
                booking_member = ClubMember.objects.filter(user=tu, club=club).first()

        flight_type = None
        flight_type_id = request.POST.get('flight_type_id')
        if flight_type_id:
            flight_type = FlightType.objects.filter(club=club, id=flight_type_id).first()

        result = booking_service.edit(
            booking, actor, aircraft, start_dt, end_dt,
            flight_type=flight_type, instructor=instructor,
            booking_member=booking_member, description=description, override=override,
        )
        if not result.ok:
            status_code = 409 if result.data.get('blockout') else 400
            return JsonResponse({'error': result.error, **result.data}, status=status_code)

        # Attach / clear non-member client — staff only
        if actor.is_admin or actor.is_instructor:
            client_id = request.POST.get('client_id', '').strip()
            billed_to = request.POST.get('billed_to', '').strip()
            from .models import Contact as _Contact
            if client_id:
                client_obj = _Contact.objects.filter(id=client_id, club=club).first()
                booking.client = client_obj
                booking.billed_to = billed_to if client_obj else ''
            else:
                booking.client = None
                booking.billed_to = ''
            booking.save(update_fields=['client', 'billed_to'])

        return JsonResponse({'success': True})

    except Aircraft.DoesNotExist:
        return JsonResponse({'error': 'Aircraft not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)


@login_required
def availability_search(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    
    config = get_config(club)
    instructors = ClubMember.objects.filter(club=club, is_on_instructor_roster=True).select_related('user')
    aircraft_list = Aircraft.objects.filter(club=club, status='online')
    aircraft_types = sorted(set(
        a.aircraft_type.name for a in aircraft_list if a.aircraft_type_id
    ))
    
    results = []
    search_performed = True
    filters_applied = {}
    result_count = 0
    calendar_days = []

    _p = request.GET if request.method == 'GET' else request.POST
    if True:  # always search; defaults give a useful first load
        # Default to 'today'; fall forward to this_week or next_week when today is nearly over
        _default_range = 'today'
        if not _p.get('range_type'):
            from datetime import datetime as _dt2
            _ops_end_dt = _aware(_dt2.combine(timezone.localdate(), config.operating_hours_end))
            _mins_left = (_ops_end_dt - timezone.now()).total_seconds() / 60
            if _mins_left < (config.default_booking_duration or 90):
                # If today is Sunday (end of week) jump straight to next week
                _default_range = 'next_week' if timezone.localdate().weekday() == 6 else 'this_week'
        range_type = _p.get('range_type', _default_range)
        aircraft_filter = _p.get('aircraft', '')
        aircraft_type_filter = _p.get('aircraft_type', '')
        instructor_filter = _p.get('instructor', '')
        booking_kind = _p.get('booking_kind', 'dual')
        duration = int(_p.get('duration') or config.default_booking_duration)

        # --- Reconcile aircraft type vs specific aircraft (specific wins) ---
        specific_aircraft = Aircraft.objects.filter(club=club, id=aircraft_filter).first() if aircraft_filter else None
        if specific_aircraft:
            # A specific tail implies its type; ignore any conflicting type filter.
            aircraft_type_filter = specific_aircraft.aircraft_type.name if specific_aircraft.aircraft_type_id else ''
        elif aircraft_type_filter:
            # Type chosen but no specific aircraft: leave as type-only filter.
            pass

        # --- Solo vs dual is the master switch ---
        is_solo = (booking_kind == 'solo')

        specific_instructor = None
        if not is_solo and instructor_filter and instructor_filter not in ('any', 'none', ''):
            from .models import User
            specific_instructor = User.objects.filter(id=instructor_filter).first()

        filters_applied = {
            'range_type': range_type,
            'aircraft': aircraft_filter,
            'aircraft_type': aircraft_type_filter,
            'instructor': instructor_filter,
            'booking_kind': booking_kind,
            'duration': duration,
        }

        date_start, date_end = get_date_range(range_type)
        # Never search past days — clamp start to today
        date_start = max(date_start, timezone.localdate())

        typ_start, typ_end = config.slot_window()
        _min_mins = config.time_slot_interval or 30

        def clip_and_mark(s, st_dt, en_dt):
            """Clip span to typical operating hours. Returns False if span is too short after clipping."""
            from datetime import time as _time
            day = st_dt.date()
            tz = st_dt.tzinfo
            from datetime import datetime as _dt
            typ_s = timezone.make_aware(_dt.combine(day, typ_start), tz) if tz else _dt.combine(day, typ_start)
            typ_e = timezone.make_aware(_dt.combine(day, typ_end), tz) if tz else _dt.combine(day, typ_end)
            clipped_start = max(st_dt, typ_s)
            clipped_end = min(en_dt, typ_e)
            if (clipped_end - clipped_start).total_seconds() / 60 < _min_mins:
                return False
            s['start'] = clipped_start
            s['end'] = clipped_end
            _mins = int((clipped_end - clipped_start).total_seconds() / 60)
            s['minutes'] = _mins
            s['start_label'] = clipped_start.strftime('%H:%M')
            s['end_label'] = clipped_end.strftime('%H:%M')
            _h, _m = divmod(_mins, 60)
            s['duration_label'] = (
                f'{_h}h {_m}m' if _h and _m else
                f'{_h}h'       if _h else
                f'{_m} min'
            )
            s['atypical'] = False
            s['start_iso'] = clipped_start.isoformat()
            return True

        by_day = {}
        _now = timezone.now()

        if is_solo:
            raw = availability_service.free_spans_solo(
                club, date_start, date_end,
                aircraft=specific_aircraft, aircraft_type=aircraft_type_filter or None,
                min_minutes=config.time_slot_interval,
            )
            slot_day = {}  # date -> {start_label -> slot_dict}
            for entry in raw:
                d = entry['date']; ac = entry['aircraft']
                type_name = ac.aircraft_type.name if ac.aircraft_type_id else ac.registration
                for s in entry['spans']:
                    if s['start'] <= _now:
                        continue
                    st = timezone.localtime(s['start']); en = timezone.localtime(s['end'])
                    if not clip_and_mark(s, st, en):
                        continue
                    if d not in slot_day:
                        slot_day[d] = {}
                    slot = slot_day[d].setdefault(s['start_label'], {
                        'start_label': s['start_label'], 'end_label': s['end_label'],
                        'start_iso': s['start_iso'], 'end_iso': s['end'].isoformat(),
                        'year': d.year, 'month': d.month, 'day': d.day,
                        'type_pills': {},
                    })
                    # Solo: one pill per tail (reg is the key) so each aircraft is bookable individually
                    if ac.registration not in slot['type_pills']:
                        slot['type_pills'][ac.registration] = {'reg': ac.registration, 'type_name': type_name, 'aircraft_id': ac.id}
            by_day = {
                d: {'slot_rows': sorted(
                    [{**{k: val for k, val in v.items() if k != 'type_pills'},
                      'type_pills': list(v['type_pills'].values())} for v in slots.values()],
                    key=lambda x: x['start_label']
                )}
                for d, slots in slot_day.items()
            }
        else:
            raw = availability_service.free_spans_dual(
                club, date_start, date_end,
                aircraft=specific_aircraft, aircraft_type=aircraft_type_filter or None,
                instructor=specific_instructor, min_minutes=config.time_slot_interval,
            )
            instr_day = {}  # date -> {instr_id -> {instructor_name, slot_map}}
            for entry in raw:
                d = entry['date']; ac = entry['aircraft']
                type_name = ac.aircraft_type.name if ac.aircraft_type_id else ac.registration
                for ir in entry['instructor_rows']:
                    instr = ir['instructor']
                    instr_name = f"{instr.first_name} {instr.last_name}".strip() or instr.username
                    for s in ir['spans']:
                        if s['start'] <= _now:
                            continue
                        st = timezone.localtime(s['start']); en = timezone.localtime(s['end'])
                        if not clip_and_mark(s, st, en):
                            continue
                        if d not in instr_day:
                            instr_day[d] = {}
                        if instr.id not in instr_day[d]:
                            instr_day[d][instr.id] = {
                                'instructor_name': instr_name, 'instructor_id': instr.id, 'slot_map': {},
                            }
                        slot = instr_day[d][instr.id]['slot_map'].setdefault(s['start_label'], {
                            'start_label': s['start_label'], 'end_label': s['end_label'],
                            'start_iso': s['start_iso'], 'end_iso': s['end'].isoformat(),
                            'year': d.year, 'month': d.month, 'day': d.day,
                            'type_pills': {},
                        })
                        if ac.registration not in slot['type_pills']:
                            slot['type_pills'][ac.registration] = {'reg': ac.registration, 'type_name': type_name, 'aircraft_id': ac.id}
            by_day = {}
            for d, instrs in instr_day.items():
                instr_rows = []
                for data in sorted(instrs.values(), key=lambda x: x['instructor_name']):
                    slot_rows = sorted(
                        [{**{k: val for k, val in v.items() if k != 'type_pills'},
                          'type_pills': list(v['type_pills'].values())} for v in data['slot_map'].values()],
                        key=lambda x: x['start_label']
                    )
                    instr_rows.append({
                        'instructor_name': data['instructor_name'],
                        'instructor_id': data['instructor_id'],
                        'slot_rows': slot_rows,
                    })
                by_day[d] = {'instructor_rows': instr_rows}

        from datetime import timedelta as _td2
        results = []
        for d in sorted(by_day.keys()):
            _mon = d - _td2(days=d.weekday())
            _sun = _mon + _td2(days=6)
            _wlabel = (
                f'{_mon.strftime("%-d")}–{_sun.strftime("%-d %b")}' if _mon.month == _sun.month
                else f'{_mon.strftime("%-d %b")}–{_sun.strftime("%-d %b")}'
            )
            day_data = by_day[d]
            results.append({
                'date': d,
                'weekday': d.strftime('%A'),
                'date_label': d.strftime('%a, %-d %b'),
                'is_weekend': d.weekday() >= 5,
                'year': d.year, 'month': d.month, 'day': d.day,
                'day_anchor': f'day-{d.year}-{d.month:02d}-{d.day:02d}',
                'iso_week_key': _mon.isoformat(),
                'week_label': _wlabel,
                'instructor_rows': day_data.get('instructor_rows', []),
                'slot_rows': day_data.get('slot_rows', []),
            })

        if is_solo:
            result_count = sum(len(day['slot_rows']) for day in results)
        else:
            result_count = sum(
                len({s['start_label'] for ir in day['instructor_rows'] for s in ir['slot_rows']})
                for day in results
            )

        # Build flat day list for the calendar band widget
        from datetime import timedelta as _td, date as _date
        _slot_dates = {d['date'] for d in results}
        _cal_days = []
        _cur = date_start
        _prev_month = None
        while _cur <= date_end:
            _cal_days.append({
                'day': _cur.day,
                'weekday_short': _cur.strftime('%a'),
                'month_short': _cur.strftime('%b'),
                'has_slots': _cur in _slot_dates,
                'anchor': f'day-{_cur.year}-{_cur.month:02d}-{_cur.day:02d}',
                'show_month': _cur.month != _prev_month,
            })
            _prev_month = _cur.month
            _cur += _td(days=1)
        calendar_days = _cal_days
    
    # Aircraft -> type map for the live JS reconciliation
    import json as _json
    aircraft_type_map = _json.dumps({str(a.id): (a.aircraft_type.name if a.aircraft_type_id else '') for a in aircraft_list})

    context = {
        'club': club,
        'club_member': club_member,
        'is_instructor': club_member.is_instructor,
        'instructors': instructors,
        'aircraft_list': aircraft_list,
        'aircraft_types': aircraft_types,
        'aircraft_type_map': aircraft_type_map,
        'duration_choices': config.duration_choices(),
        'default_duration': config.default_booking_duration,
        'results': results,
        'search_performed': search_performed,
        'filters_applied': filters_applied,
        'result_count': result_count,
        'calendar_days': calendar_days if search_performed else [],
    }
    
    return render(request, 'core/availability_search.html', context)


@login_required
def reschedule_options(request, booking_id):
    """Get available alternatives for rescheduling a booking."""
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    
    if not (club_member.is_admin or club_member.is_instructor):
        return JsonResponse({'error': 'Only instructors and admins can reschedule'}, status=403)
    
    # Instructors busy during this booking's window
    busy_instructors = Booking.objects.filter(
        club=club,
        instructor__isnull=False,
        scheduled_start__lt=booking.scheduled_end,
        scheduled_end__gt=booking.scheduled_start
    ).exclude(id=booking.id).values_list('instructor_id', flat=True)
    
    available_instructors = ClubMember.objects.filter(
        club=club,
        is_on_instructor_roster=True,
    ).exclude(user_id__in=busy_instructors).select_related('user')
    
    # Get available aircraft of same type
    busy_aircraft = Booking.objects.filter(
        club=club,
        scheduled_start__lt=booking.scheduled_end,
        scheduled_end__gt=booking.scheduled_start
    ).exclude(id=booking.id).values_list('aircraft_id', flat=True)
    
    available_aircraft = Aircraft.objects.filter(
        club=club,
        aircraft_type=booking.aircraft.aircraft_type_id,
        status='online'
    ).exclude(id__in=busy_aircraft)
    
    instructors_list = [
        {'id': i.user.id, 'name': f"{i.user.first_name} {i.user.last_name}"}
        for i in available_instructors
    ]
    
    aircraft_list = [
        {'id': a.id, 'registration': a.registration}
        for a in available_aircraft
    ]
    
    return JsonResponse({
        'instructors': instructors_list,
        'aircraft': aircraft_list,
        'current_instructor_id': booking.instructor_id if booking.instructor else None,
        'current_aircraft_id': booking.aircraft_id,
    })


@login_required
@require_POST
def update_booking(request, booking_id):
    """Update booking with new instructor or aircraft."""
    booking = get_object_or_404(Booking, id=booking_id)
    club = booking.club
    
    try:
        club_member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not authorized'}, status=403)
    
    if not (club_member.is_admin or club_member.is_instructor):
        return JsonResponse({'error': 'Only instructors and admins can update'}, status=403)
    
    instructor_id = request.POST.get('instructor_id')
    aircraft_id = request.POST.get('aircraft_id')
    
    if instructor_id:
        from .models import User
        booking.instructor = User.objects.filter(id=instructor_id).first()
    
    if aircraft_id:
        new_aircraft = Aircraft.objects.filter(id=aircraft_id, club=club).first()
        if new_aircraft:
            booking.aircraft = new_aircraft
    
    booking.save()
    return JsonResponse({'success': True})


def _budget_fy_year(config):
    """Return the integer year that the current FY starts."""
    today = timezone.localdate()
    return today.year if today.month >= config.fy_start_month else today.year - 1


def _budget_month_labels(config):
    """Return list of (month_int, 'Jan', year) for the 12 months of the current FY."""
    fy_year = _budget_fy_year(config)
    labels = []
    m, y = config.fy_start_month, fy_year
    for _ in range(12):
        labels.append((m, date(y, m, 1).strftime('%b'), y))
        m += 1
        if m > 12:
            m = 1; y += 1
    return labels


def _lapse_preview(club, config):
    """Return list of members who would be lapsed if auto-lapse ran now."""
    from datetime import timedelta as _td
    cutoff = timezone.localdate() - _td(days=config.lapse_grace_days)
    return list(
        ClubMember.objects.filter(
            club=club, standing='active',
            subscription_expires__lt=cutoff,
        ).select_related('user').order_by('subscription_expires')
    )


def _next_fy_end(config):
    """Return the next financial year end date for the club."""
    import calendar as _cal
    today = timezone.localdate()
    fy_start = config.fy_start_month
    fy_end_month = fy_start - 1 or 12
    fy_end_year = today.year if today.month < fy_start else today.year + 1
    if fy_end_month >= fy_start:
        fy_end_year -= 1
    fy_end_day = _cal.monthrange(fy_end_year, fy_end_month)[1]
    fy_end = date(fy_end_year, fy_end_month, fy_end_day)
    if fy_end < today:
        fy_end_year += 1
        fy_end_day = _cal.monthrange(fy_end_year, fy_end_month)[1]
        fy_end = date(fy_end_year, fy_end_month, fy_end_day)
    return fy_end


def _renewal_preview(club, config):
    """Active members eligible for bulk renewal: renewal_required role, fee set, no outstanding sub invoice."""
    already_invoiced = set(
        Invoice.objects.filter(
            club=club,
            description__startswith='Membership subscription',
        ).exclude(status__in=('paid', 'void')).values_list('member_id', flat=True)
    )
    return list(
        ClubMember.objects.filter(
            club=club,
            standing='active',
            role__renewal_required=True,
            role__annual_renewal_fee__isnull=False,
        ).exclude(id__in=already_invoiced)
        .select_related('role', 'user').order_by('user__last_name')
    )


def club_settings(request, club_slug, mode='settings'):
    """Admin-only settings page. mode='settings' shows general/billing/roles; mode='types' shows reference type lists."""
    club = get_object_or_404(Club, slug=club_slug)
    try:
        member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if not member.is_admin:
        return render(request, 'core/no_access.html', {'club': club}, status=403)

    config = get_config(club)
    saved = request.GET.get('saved') == '1'
    is_types = (mode == 'types')
    _redir_name = 'core:club_types' if is_types else 'core:club_settings'

    ft_error = None
    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'upload_logo':
            if request.FILES.get('logo'):
                config.logo = request.FILES['logo']
                config.save(update_fields=['logo'])
            elif request.POST.get('remove_logo'):
                config.logo.delete(save=True)
            from django.shortcuts import redirect as _redirect
            return _redirect(_redir_name, club_slug=club_slug)

        elif action == 'upload_app_banner':
            if request.FILES.get('app_banner'):
                config.app_banner = request.FILES['app_banner']
                config.save(update_fields=['app_banner'])
            elif request.POST.get('remove_app_banner'):
                config.app_banner.delete(save=True)
            from django.shortcuts import redirect as _redirect
            return _redirect(_redir_name, club_slug=club_slug)

        elif action == 'add_flight_type':
            ft_name = request.POST.get('ft_name', '').strip()
            ft_is_solo     = request.POST.get('ft_is_solo') == 'on'
            ft_is_training = request.POST.get('ft_is_training') == 'on'
            ft_is_trial    = request.POST.get('ft_is_trial') == 'on'
            if ft_name:
                from django.utils.text import slugify
                code = slugify(ft_name).replace('-', '_')[:20]
                if FlightType.objects.filter(club=club, code=code).exists():
                    code = code[:18] + '_2'
                ft_for_contacts = request.POST.get('ft_for_contacts') == 'on'
                FlightType.objects.create(
                    club=club, name=ft_name, code=code,
                    is_solo=ft_is_solo, is_training=ft_is_training, is_trial=ft_is_trial,
                    for_contacts=ft_for_contacts,
                )
            else:
                ft_error = "Name is required."

        elif action == 'set_flight_type_solo':
            ft_id = request.POST.get('ft_id')
            ft_is_solo = request.POST.get('ft_is_solo') == '1'
            ft = FlightType.objects.filter(club=club, id=ft_id).first()
            if ft:
                ft.is_solo = ft_is_solo
                ft.save(update_fields=['is_solo'])
            from django.shortcuts import redirect as _redirect
            return _redirect(_redir_name, club_slug=club_slug)

        elif action == 'set_flight_type_flag':
            ft_id = request.POST.get('ft_id')
            flag = request.POST.get('ft_flag')
            value = request.POST.get('ft_value') == '1'
            allowed = {'is_training', 'is_billable', 'requires_declaration', 'is_solo', 'is_trial', 'for_contacts'}
            ft = FlightType.objects.filter(club=club, id=ft_id).first()
            if ft and flag in allowed:
                setattr(ft, flag, value)
                ft.save(update_fields=[flag])
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'edit_flight_type':
            ft_id = request.POST.get('ft_id')
            ft = FlightType.objects.filter(club=club, id=ft_id).first()
            if ft:
                name = request.POST.get('ft_name', '').strip()
                if name:
                    ft.name = name
                    ft.save(update_fields=['name'])

        elif action == 'delete_flight_type':
            ft_id = request.POST.get('ft_id')
            ft = FlightType.objects.filter(club=club, id=ft_id).first()
            if ft:
                if Booking.objects.filter(club=club, flight_type=ft).exists():
                    ft_error = f"Cannot delete '{ft.name}' — it has existing bookings."
                else:
                    ft.delete()

        # ── BlockOutType management ──────────────────────────────────────
        elif action == 'add_blockout_type':
            bot_name = request.POST.get('bot_name', '').strip()
            bot_color = request.POST.get('bot_color', '#9aa3ad').strip()
            bot_target = request.POST.get('bot_target', 'instructor')
            bot_is_hard = request.POST.get('bot_is_hard') != '0'
            if bot_name:
                BlockOutType.objects.create(club=club, name=bot_name, color=bot_color,
                                            target=bot_target, is_hard=bot_is_hard)
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'set_blockout_type_hard':
            bot_id = request.POST.get('bot_id')
            is_hard = 'is_hard' in request.POST
            bt = BlockOutType.objects.filter(club=club, id=bot_id).first()
            if bt:
                bt.is_hard = is_hard
                bt.save(update_fields=['is_hard'])
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'edit_blockout_type':
            bot_id = request.POST.get('bot_id')
            bt = BlockOutType.objects.filter(club=club, id=bot_id).first()
            if bt:
                name = request.POST.get('bot_name', '').strip()
                color = request.POST.get('bot_color', '').strip()
                if name:
                    bt.name = name
                if color:
                    bt.color = color
                bt.save(update_fields=['name', 'color'])

        elif action == 'delete_blockout_type':
            bot_id = request.POST.get('bot_id')
            bt = BlockOutType.objects.filter(club=club, id=bot_id).first()
            if bt:
                bt.delete()
            return redirect(_redir_name, club_slug=club_slug)

        # ── Role management ──────────────────────────────────────────────────
        elif action == 'add_role':
            rname = request.POST.get('role_name', '').strip()
            if rname:
                Role.objects.get_or_create(club=club, name=rname)
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'delete_role':
            role = Role.objects.filter(club=club, id=request.POST.get('role_id')).first()
            if role and role.is_system_role:
                from django.contrib import messages as _msg
                _msg.error(request, f'"{role.name}" is a system role and cannot be deleted.')
            elif role:
                role.delete()
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'rename_role':
            role = Role.objects.filter(club=club, id=request.POST.get('role_id')).first()
            new_name = request.POST.get('role_name', '').strip()
            if role and new_name and new_name != role.name:
                if not Role.objects.filter(club=club, name=new_name).exclude(pk=role.pk).exists():
                    role.name = new_name
                    role.save(update_fields=['name'])
                else:
                    from django.contrib import messages as _msg
                    _msg.error(request, f'A role named "{new_name}" already exists.')
            return redirect(redirect(_redir_name, club_slug=club_slug).url + '?tab=roles&saved=1')

        elif action == 'set_role_fee':
            role = Role.objects.filter(club=club, id=request.POST.get('role_id')).first()
            if role:
                raw = request.POST.get('fee', '').strip()
                role.annual_renewal_fee = raw if raw else None
                role.save(update_fields=['annual_renewal_fee'])
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'set_role_permission':
            _BOOL_PERMS = {'can_access_manage', 'can_access_fleet', 'can_access_safety',
                           'can_access_settings', 'can_access_reports', 'is_superadmin', 'renewal_required'}
            role = Role.objects.filter(club=club, id=request.POST.get('role_id')).first()
            perm = request.POST.get('perm', '')
            # Admin system role permissions are always ALL — block changes
            if role and perm and role.system_role_type != 'admin':
                if perm == 'bookings_access':
                    val = request.POST.get('value', '')
                    if val in dict(Role.BOOKINGS_CHOICES):
                        role.bookings_access = val
                        role.save(update_fields=['bookings_access'])
                elif perm in _BOOL_PERMS:
                    setattr(role, perm, request.POST.get('value') == '1')
                    role.save(update_fields=[perm])
            return redirect(_redir_name, club_slug=club_slug)

        # ── Instructor grade management ──────────────────────────────────────
        elif action == 'add_instructor_grade':
            ig_name = request.POST.get('ig_name', '').strip()
            ig_rate = request.POST.get('ig_rate', '').strip()
            ig_order = request.POST.get('ig_order', '0').strip()
            if ig_name and ig_rate:
                try:
                    InstructorGrade.objects.get_or_create(
                        club=club, name=ig_name,
                        defaults={'hourly_rate': ig_rate,
                                  'display_order': int(ig_order) if ig_order.isdigit() else 0}
                    )
                except Exception:
                    pass
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'delete_instructor_grade':
            InstructorGrade.objects.filter(club=club, id=request.POST.get('ig_id')).delete()
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'edit_instructor_grade':
            ig = InstructorGrade.objects.filter(club=club, id=request.POST.get('ig_id')).first()
            if ig:
                name = request.POST.get('ig_name', '').strip()
                if name:
                    ig.name = name
                ig.hourly_rate = request.POST.get('ig_rate', ig.hourly_rate)
                order = request.POST.get('ig_order', '').strip()
                if order.isdigit():
                    ig.display_order = int(order)
                ig.save(update_fields=['name', 'hourly_rate', 'display_order'])
            return redirect(_redir_name, club_slug=club_slug)

        # ── Aircraft surcharge type management ───────────────────────────────
        elif action == 'add_surcharge_type':
            st_name = request.POST.get('st_name', '').strip()
            st_amount = request.POST.get('st_amount', '').strip()
            st_desc = request.POST.get('st_desc', '').strip()
            if st_name and st_amount:
                try:
                    AircraftSurchargeType.objects.get_or_create(
                        club=club, name=st_name,
                        defaults={'amount': st_amount, 'description': st_desc}
                    )
                except Exception:
                    pass
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'delete_surcharge_type':
            AircraftSurchargeType.objects.filter(club=club, id=request.POST.get('st_id')).delete()
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'edit_surcharge_type':
            st = AircraftSurchargeType.objects.filter(club=club, id=request.POST.get('st_id')).first()
            if st:
                name = request.POST.get('st_name', '').strip()
                if name:
                    st.name = name
                st.description = request.POST.get('st_desc', '').strip()
                st.amount = request.POST.get('st_amount', st.amount)
                st.save(update_fields=['name', 'description', 'amount'])
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'add_aircraft_type':
            at_name = request.POST.get('at_name', '').strip()
            at_icao = request.POST.get('at_icao', '').strip().upper()
            if at_name:
                AircraftType.objects.get_or_create(club=club, name=at_name,
                                                    defaults={'icao_designator': at_icao})
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'edit_aircraft_type':
            at = AircraftType.objects.filter(club=club, id=request.POST.get('at_id')).first()
            if at:
                name = request.POST.get('at_name', '').strip()
                if name:
                    at.name = name
                at.icao_designator = request.POST.get('at_icao', '').strip().upper()
                at.save(update_fields=['name', 'icao_designator'])
            return redirect(_redir_name, club_slug=club_slug)

        elif action == 'delete_aircraft_type':
            at = AircraftType.objects.filter(club=club, id=request.POST.get('at_id')).first()
            if at:
                if at.aircraft.exists():
                    ft_error = f"Cannot delete '{at.name}' — it is assigned to aircraft in the fleet."
                elif at.type_ratings.exists():
                    ft_error = f"Cannot delete '{at.name}' — it is referenced by member type ratings."
                else:
                    at.delete()
            if not ft_error:
                return redirect(_redir_name, club_slug=club_slug)

        elif action == 'add_occurrence_type':
            ot_name = request.POST.get('ot_name', '').strip()
            ot_desc = request.POST.get('ot_desc', '').strip()
            if ot_name:
                OccurrenceType.objects.get_or_create(club=club, name=ot_name, defaults={'description': ot_desc})
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=incident-types&saved=1")

        elif action == 'edit_occurrence_type':
            ot = OccurrenceType.objects.filter(club=club, id=request.POST.get('ot_id')).first()
            if ot:
                ot_name = request.POST.get('ot_name', '').strip()
                if ot_name:
                    ot.name = ot_name
                ot.description = request.POST.get('ot_desc', '').strip()
                ot.sort_order = int(request.POST.get('ot_sort', ot.sort_order) or ot.sort_order)
                ot.save()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=incident-types&saved=1")

        elif action == 'delete_occurrence_type':
            ot = OccurrenceType.objects.filter(club=club, id=request.POST.get('ot_id')).first()
            if ot and not ot.reports.exists():
                ot.delete()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=incident-types&saved=1")

        elif action == 'toggle_occurrence_type':
            ot = OccurrenceType.objects.filter(club=club, id=request.POST.get('ot_id')).first()
            if ot:
                ot.is_active = not ot.is_active
                ot.save(update_fields=['is_active'])
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=incident-types&saved=1")

        elif action == 'add_contact_type':
            ct_name = request.POST.get('ct_name', '').strip()
            if ct_name:
                ContactType.objects.get_or_create(club=club, name=ct_name)
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=contact-types&saved=1")

        elif action == 'edit_contact_type':
            ct = ContactType.objects.filter(club=club, id=request.POST.get('ct_id')).first()
            if ct:
                ct_name = request.POST.get('ct_name', '').strip()
                if ct_name: ct.name = ct_name
                ct.save()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=contact-types&saved=1")

        elif action == 'delete_contact_type':
            ct = ContactType.objects.filter(club=club, id=request.POST.get('ct_id')).first()
            if ct and not ct.contacts.exists():
                ct.delete()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=contact-types&saved=1")

        elif action == 'toggle_contact_type':
            ct = ContactType.objects.filter(club=club, id=request.POST.get('ct_id')).first()
            if ct:
                ct.is_active = not ct.is_active
                ct.save(update_fields=['is_active'])
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=contact-types&saved=1")

        elif action == 'add_voucher_type':
            vt_name = request.POST.get('vt_name', '').strip()
            vt_val  = request.POST.get('vt_value', '').strip()
            if vt_name and vt_val:
                try:
                    VoucherType.objects.get_or_create(
                        club=club, name=vt_name,
                        defaults={
                            'default_value': vt_val,
                            'description': request.POST.get('vt_desc', '').strip(),
                            'sort_order': VoucherType.objects.filter(club=club).count(),
                        }
                    )
                except Exception:
                    pass
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=voucher-types&saved=1")

        elif action == 'edit_voucher_type':
            vt = VoucherType.objects.filter(club=club, id=request.POST.get('vt_id')).first()
            if vt:
                vt.name        = request.POST.get('vt_name', vt.name).strip() or vt.name
                try: vt.default_value = float(request.POST.get('vt_value', vt.default_value))
                except (ValueError, TypeError): pass
                vt.description = request.POST.get('vt_desc', '').strip()
                vt.sort_order  = int(request.POST.get('vt_sort', vt.sort_order) or vt.sort_order)
                vt.save()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=voucher-types&saved=1")

        elif action == 'delete_voucher_type':
            vt = VoucherType.objects.filter(club=club, id=request.POST.get('vt_id')).first()
            if vt:
                vt.delete()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=voucher-types&saved=1")

        elif action == 'toggle_voucher_type':
            vt = VoucherType.objects.filter(club=club, id=request.POST.get('vt_id')).first()
            if vt:
                vt.is_active = not vt.is_active
                vt.save(update_fields=['is_active'])
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=voucher-types&saved=1")

        elif action == 'save_budget':
            import calendar as _cal
            fy_year = int(request.POST.get('fy_year', 0) or 0)
            if fy_year:
                ac_list = Aircraft.objects.filter(club=club).exclude(status='retired')
                for ac in ac_list:
                    for m in range(1, 13):
                        key = f'budget_{ac.id}_{m}'
                        raw = request.POST.get(key, '').strip()
                        try:
                            hrs = float(raw)
                        except (ValueError, TypeError):
                            continue
                        FlyingBudget.objects.update_or_create(
                            club=club, aircraft=ac, fy_year=fy_year, month=m,
                            defaults={'budgeted_hours': round(hrs, 1)},
                        )
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=budget&saved=1")

        elif action == 'run_auto_lapse':
            from datetime import timedelta as _td
            from .models import MembershipHistoryEntry as _MHE
            grace = config.lapse_grace_days
            _now = timezone.localdate()
            cutoff = _now - _td(days=grace)
            candidates = ClubMember.objects.filter(
                club=club, standing='active',
                subscription_expires__lt=cutoff,
            ).select_related('user')
            lapsed_names = []
            for m in candidates:
                days_over = (_now - m.subscription_expires).days
                m.standing    = 'lapsed'
                m.resigned_at = m.resigned_at or _now
                m.save(update_fields=['standing', 'resigned_at'])
                if m.user:
                    m.user.is_active = False
                    m.user.save(update_fields=['is_active'])
                _MHE.objects.create(
                    club_member=m, event_type='standing_change',
                    changed_by=request.user,
                    old_value='active', new_value='lapsed',
                    note=f'Auto-lapsed via Settings — subscription expired {m.subscription_expires} '
                         f'({days_over} days ago, grace period {grace} days)',
                )
                lapsed_names.append(m.user.get_full_name() if m.user else '—')
            msg = (f'Lapsed {len(lapsed_names)} member(s): {", ".join(lapsed_names)}.'
                   if lapsed_names else 'No members were due for lapsing.')
            from django.contrib import messages as _msgs
            _msgs.success(request, msg)
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=membership&saved=1")

        elif action == 'run_bulk_renewal':
            from datetime import timedelta as _td
            from django.db.models import F as _F
            from django.contrib import messages as _msgs
            fy_end = _next_fy_end(config)
            fy_label = f'FY{str(fy_end.year)[2:]}'
            candidates = _renewal_preview(club, config)
            today = timezone.localdate()
            created = 0
            for m in candidates:
                config.refresh_from_db()
                inv_number = config.invoice_number_next
                config.invoice_number_next = _F('invoice_number_next') + 1
                config.save(update_fields=['invoice_number_next'])
                due = today + _td(days=config.payment_terms_days or 14)
                desc = f'Membership subscription — {fy_label} ({m.role.name})'
                inv = Invoice.objects.create(
                    club=club, member=m,
                    invoice_number=inv_number,
                    issue_date=today, due_date=due,
                    description=desc,
                    gst_rate=config.gst_rate,
                    amount_paid=0,
                    status='sent',
                    sent_at=timezone.now(),
                    subscription_expiry_date=fy_end,
                    created_by=request.user,
                )
                InvoiceLineItem.objects.create(
                    invoice=inv,
                    description=desc,
                    quantity=1, unit='', rate=m.role.annual_renewal_fee, amount=m.role.annual_renewal_fee,
                )
                from .services import notification_service as _ns
                _ns.notify_invoice_issued(inv)
                created += 1
            msg = (f'{created} renewal invoice(s) sent for {fy_label}.'
                   if created else f'No members eligible for {fy_label} bulk renewal.')
            _msgs.success(request, msg)
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=membership&saved=1")

        elif action == 'save_membership_tab':
            try:
                fy = int(request.POST.get('fy_start_month', config.fy_start_month))
                if 1 <= fy <= 12:
                    config.fy_start_month = fy
            except (ValueError, TypeError):
                pass
            try:
                gd = int(request.POST.get('lapse_grace_days', config.lapse_grace_days))
                if gd >= 0:
                    config.lapse_grace_days = gd
            except (ValueError, TypeError):
                pass
            config.save()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=membership&saved=1")

        elif action == 'save_billing':
            for field in ['billing_name',
                          'gst_number', 'bank_name', 'bank_account', 'payment_terms_text',
                          'invoice_number_prefix']:
                setattr(config, field, request.POST.get(field, '').strip())
            try:
                config.gst_rate = float(request.POST.get('gst_rate', config.gst_rate))
            except (ValueError, TypeError):
                pass
            try:
                config.payment_terms_days = int(request.POST.get('payment_terms_days', config.payment_terms_days))
            except (ValueError, TypeError):
                pass
            try:
                next_num = int(request.POST.get('invoice_number_next', ''))
                if next_num > 0:
                    config.invoice_number_next = next_num
            except (ValueError, TypeError):
                pass
            try:
                fy = int(request.POST.get('fy_start_month', config.fy_start_month))
                if 1 <= fy <= 12:
                    config.fy_start_month = fy
            except (ValueError, TypeError):
                pass
            try:
                gd = int(request.POST.get('lapse_grace_days', config.lapse_grace_days))
                if gd >= 0:
                    config.lapse_grace_days = gd
            except (ValueError, TypeError):
                pass
            config.save()
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab=billing&saved=1")

        else:
            club_name = request.POST.get('club_name', '').strip()
            if club_name:
                club.name = club_name
                club.save(update_fields=['name'])
            for field in ['theme_banner', 'theme_primary', 'theme_accent',
                          'dual_accent',
                          'theme_weekend', 'theme_atypical']:
                val = request.POST.get(field, '').strip()
                if val:
                    setattr(config, field, val)
            _cc_raw = request.POST.get('chart_colors', '').strip()
            if _cc_raw:
                import json as _ccjson
                try:
                    _cc = _ccjson.loads(_cc_raw)
                    if isinstance(_cc, list):
                        config.chart_colors = [c for c in _cc if isinstance(c, str) and c.startswith('#')]
                except (ValueError, TypeError):
                    pass
            _fc = request.POST.get('font_choice', '').strip()
            if _fc in dict(config.FONT_CHOICES):
                config.font_choice = _fc
            config.compact_mode = request.POST.get('compact_mode') == 'on'
            oh_start = request.POST.get('operating_hours_start')
            oh_end = request.POST.get('operating_hours_end')
            if oh_start:
                config.operating_hours_start = oh_start
            if oh_end:
                config.operating_hours_end = oh_end
            config.booking_slots = request.POST.get('booking_slots', '').strip()
            config.billing_phone = request.POST.get('billing_phone', '').strip()
            config.billing_address = request.POST.get('billing_address', '').strip()
            config.billing_email = request.POST.get('billing_email', '').strip()
            # Booking blocks (bookings tab only)
            if request.POST.get('_tab') == 'bookings':
                config.booking_block_enabled = request.POST.get('booking_block_enabled') == 'on'
                config.booking_block_message = request.POST.get('booking_block_message', '').strip()
                for attr in ['booking_block_unpaid_flight_days', 'booking_block_invoice_days']:
                    raw = request.POST.get(attr, '').strip()
                    if raw == '':
                        setattr(config, attr, None)
                    else:
                        try:
                            v = int(raw)
                            if v >= 1:
                                setattr(config, attr, v)
                        except (ValueError, TypeError):
                            pass
            config.save()
            _tab = request.POST.get('_tab', 'general')
            return redirect(f"{redirect(_redir_name, club_slug=club_slug).url}?tab={_tab}&saved=1")

    color_fields = [
        ('theme_banner', 'Banner', config.theme_banner),
        ('theme_primary', 'Primary (buttons, links)', config.theme_primary),
        ('theme_accent', 'Accent', config.theme_accent),
        ('dual_accent', 'Dual-flight accent', config.dual_accent),
        ('theme_weekend', 'Weekend shade', config.theme_weekend),
        ('theme_atypical', 'Outside typical hours', config.theme_atypical),
    ]

    _BOOL_PERM_NAMES = ['can_access_manage', 'can_access_safety', 'can_access_fleet', 'can_access_reports', 'can_access_settings', 'is_superadmin', 'renewal_required']
    from django.db.models import Count as _Count, Case as _Case, When as _When, IntegerField as _IntF, Value as _Val
    _sys_order = _Case(
        _When(system_role_type='member',     then=_Val(0)),
        _When(system_role_type='instructor', then=_Val(1)),
        _When(system_role_type='admin',      then=_Val(2)),
        default=_Val(3), output_field=_IntF(),
    )
    roles = Role.objects.filter(club=club).annotate(
        member_count=_Count('clubmember'),
        sys_order=_sys_order,
    ).order_by('sys_order', 'name')
    for r in roles:
        r.perm_items = [(p, getattr(r, p)) for p in _BOOL_PERM_NAMES]
    from .models import ClubConfig as _CC2
    font_options = [
        {'value': val, 'label': lbl, 'stack': _CC2.FONT_STACKS[val][0]}
        for val, lbl in _CC2.FONT_CHOICES
    ]

    return render(request, 'core/club_settings.html', {
        'club': club,
        'config': config,
        'font_options': font_options,
        'color_fields': color_fields,
        'all_blockout_types': BlockOutType.objects.filter(club=club, target='all'),
        'instructor_blockout_types': BlockOutType.objects.filter(club=club, target='instructor'),
        'aircraft_blockout_types': BlockOutType.objects.filter(club=club, target='aircraft'),
        'flight_types': FlightType.objects.filter(club=club),
        'instructor_grades': InstructorGrade.objects.filter(club=club),
        'surcharge_types': AircraftSurchargeType.objects.filter(club=club),
        'aircraft_type_list': AircraftType.objects.filter(club=club),
        'occurrence_types': OccurrenceType.objects.filter(club=club),
        'contact_types_list': ContactType.objects.filter(club=club).order_by('name'),
        'voucher_types_list': VoucherType.objects.filter(club=club),
        'roles': roles,
        'saved': saved,
        'ft_error': ft_error,
        'is_types': is_types,
        'fy_month_choices': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'lapse_preview': _lapse_preview(club, config),
        'renewal_preview': _renewal_preview(club, config),
        'renewal_fy_end': _next_fy_end(config),
        'renewal_fy_label': f'FY{str(_next_fy_end(config).year)[2:]}',
        # Budget
        'budget_aircraft': Aircraft.objects.filter(club=club).exclude(status='retired').order_by('registration'),
        'budget_entries': {(b.aircraft_id, b.fy_year, b.month): b.budgeted_hours
                          for b in FlyingBudget.objects.filter(club=club)},
        'budget_fy_year': _budget_fy_year(config),
        'budget_month_labels': _budget_month_labels(config),
    })


@login_required
def manage_announcements(request, club_slug):
    from .models import Announcement
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'create':
            Announcement.objects.create(
                club=club,
                type=request.POST.get('type', 'announcement'),
                title=request.POST.get('title', '').strip(),
                body=request.POST.get('body', '').strip(),
                event_date=request.POST.get('event_date') or None,
                expires_at=request.POST.get('expires_at') or None,
                is_pinned=request.POST.get('is_pinned') == 'on',
                created_by=request.user,
            )
        elif action == 'edit':
            ann = get_object_or_404(Announcement, club=club, id=request.POST.get('ann_id'))
            ann.type = request.POST.get('type', ann.type)
            ann.title = request.POST.get('title', '').strip() or ann.title
            ann.body = request.POST.get('body', '').strip()
            ann.event_date = request.POST.get('event_date') or None
            ann.expires_at = request.POST.get('expires_at') or None
            ann.is_pinned = request.POST.get('is_pinned') == 'on'
            ann.save()
        elif action == 'delete':
            Announcement.objects.filter(club=club, id=request.POST.get('ann_id')).delete()
        return redirect(f"{request.path}?saved=1")

    from urllib.parse import urlencode as _ue
    saved    = request.GET.get('saved') == '1'
    sort     = request.GET.get('sort', 'date')
    sort_dir = request.GET.get('dir', 'asc')
    _ANN_SORT = {
        'type':    ('type', 'title'),
        'title':   ('title',),
        'date':    ('event_date', 'title'),
        'expires': ('expires_at', 'title'),
        'pinned':  ('-is_pinned', 'title'),
    }
    if sort not in _ANN_SORT:
        sort = 'date'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'asc'
    _ann_order = _ANN_SORT[sort]
    if sort_dir == 'desc':
        _ann_order = tuple(f[1:] if f.startswith('-') else '-'+f for f in _ann_order)
    announcements = Announcement.objects.filter(club=club).order_by(*_ann_order)
    _base_qs = _ue({k: v for k, v in request.GET.items() if k not in ('sort', 'dir') and v})
    return render(request, 'core/manage_announcements.html', {
        'club': club, 'club_member': actor,
        'announcements': announcements,
        'saved': saved,
        'sort': sort, 'sort_dir': sort_dir, 'base_qs': _base_qs,
        'type_choices': [('announcement','Announcement'),('info','Information'),
                         ('safety','Safety Notice'),('event','Event'),('flyaway','Fly-Away')],
    })


@login_required
def manage_bookings(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    from django.db.models import Q
    today = timezone.localdate()

    _conflict_q = (
        Q(blockout_conflict=True) |
        Q(member__standing__in=['suspended', 'lapsed', 'resigned']) |
        Q(member__standing='active', member__subscription_expires__lt=today) |
        Q(aircraft__status='retired')
    )

    # Pre-load instructor availability windows for off-roster conflict detection
    from .models import InstructorAvailability as _IA
    _roster_member_ids = set(
        ClubMember.objects.filter(club=club, is_on_instructor_roster=True).values_list('user_id', flat=True)
    )
    _av_windows = {}  # user_id -> list[InstructorAvailability]
    for av in _IA.objects.filter(club_member__club=club).select_related('club_member'):
        _av_windows.setdefault(av.club_member.user_id, []).append(av)

    def _instructor_off_roster(booking):
        uid = booking.instructor_id
        if not uid or uid not in _roster_member_ids:
            return False
        windows = _av_windows.get(uid, [])
        if not windows:
            return False  # no schedule declared = assumed available (per model)
        bdate = booking.scheduled_start.date()
        return not any(w.applies_on(bdate) for w in windows)

    if request.method == 'POST':
        action = request.POST.get('action', '')
        ids = [int(i) for i in request.POST.getlist('booking_ids') if i.isdigit()]
        qs = Booking.objects.filter(club=club, id__in=ids).exclude(status='cancelled')
        if action == 'confirm':
            qs.filter(status='pending').update(
                status='confirmed', confirmed_by=request.user, confirmed_at=timezone.now()
            )
        elif action == 'cancel':
            reason = request.POST.get('bulk_cancel_reason', 'no_longer_required')
            qs.update(status='cancelled', cancellation_reason=reason)
        elif action == 'move_aircraft':
            ac = Aircraft.objects.filter(club=club, id=request.POST.get('target_aircraft_id'), status='online').first()
            if ac:
                qs.update(aircraft=ac)
        elif action == 'move_instructor':
            from .models import User as _User
            instr = _User.objects.filter(id=request.POST.get('target_instructor_id')).first()
            if instr:
                qs.update(instructor=instr)
        return redirect(request.get_full_path())

    from datetime import timedelta
    from urllib.parse import urlencode

    tab              = request.GET.get('tab', 'active')
    f_aircraft       = request.GET.get('aircraft', '')
    f_instructor     = request.GET.get('instructor', '')
    f_status         = request.GET.get('status', '')
    f_date_from      = request.GET.get('date_from', '')
    f_date_to        = request.GET.get('date_to', '')
    show_all_history = request.GET.get('all_history') == '1'

    _base_qs = (Booking.objects
                .filter(club=club)
                .select_related('member__user', 'aircraft', 'instructor',
                                'flight_type', 'flight_completion'))

    # Aircraft + instructor filters apply to both sections
    if f_aircraft:
        _base_qs = _base_qs.filter(aircraft_id=f_aircraft)
    if f_instructor:
        _base_qs = _base_qs.filter(instructor_id=f_instructor)

    def conflict_reasons(b):
        r = []
        if b.id in _clashing_ids:
            # Distinguish which resource is double-booked for the reason label
            _ac_clash = (b.aircraft_id and Booking.objects
                .filter(club=club, aircraft_id=b.aircraft_id,
                        status__in=['pending','confirmed','departed'],
                        scheduled_start__lt=b.scheduled_end,
                        scheduled_end__gt=b.scheduled_start)
                .exclude(pk=b.pk).exists())
            _in_clash = (b.instructor_id and Booking.objects
                .filter(club=club, instructor_id=b.instructor_id,
                        status__in=['pending','confirmed','departed'],
                        scheduled_start__lt=b.scheduled_end,
                        scheduled_end__gt=b.scheduled_start)
                .exclude(pk=b.pk).exists())
            if _ac_clash: r.append('Aircraft double-booked')
            if _in_clash: r.append('Instructor double-booked')
        if b.blockout_conflict:
            r.append(b.blockout_conflict_reason or 'Block-out conflict')
        if b.member:
            if b.member.standing in ('suspended', 'lapsed', 'resigned'):
                r.append(f'Member {b.member.get_standing_display()}')
            elif b.member.standing == 'active' and b.member.subscription_expires and b.member.subscription_expires < today:
                r.append('Subscription expired')
        if b.aircraft and b.aircraft.status == 'retired':
            r.append('Aircraft retired')
        if _instructor_off_roster(b):
            r.append('Instructor off roster')
        _fc = getattr(b, 'flight_completion', None)
        if _fc and _fc.amount_paid and _fc.total_charge and _fc.amount_paid > _fc.total_charge:
            _over = _fc.amount_paid - _fc.total_charge
            r.append(f'Overpaid ${_over:.2f} — review and credit account')
        return r

    _STATUS_ORDER = {'completed': 0, 'departed': 1, 'confirmed': 2, 'pending': 3}

    # ── Booking-vs-booking aircraft clash detection ────────────────────────────
    from django.db.models import Q as _Q2, Exists as _Exists, OuterRef as _OuterRef
    _ac_clash_inner = (Booking.objects
        .filter(
            club=club,
            aircraft_id=_OuterRef('aircraft_id'),
            status__in=['pending', 'confirmed', 'departed'],
            scheduled_start__lt=_OuterRef('scheduled_end'),
            scheduled_end__gt=_OuterRef('scheduled_start'),
        )
        .exclude(pk=_OuterRef('pk')))
    _instr_clash_inner = (Booking.objects
        .filter(
            club=club,
            instructor_id=_OuterRef('instructor_id'),
            status__in=['pending', 'confirmed', 'departed'],
            scheduled_start__lt=_OuterRef('scheduled_end'),
            scheduled_end__gt=_OuterRef('scheduled_start'),
        )
        .exclude(pk=_OuterRef('pk')))
    _future_active = Booking.objects.filter(
        club=club, status__in=['pending', 'confirmed'],
        scheduled_start__date__gte=today)
    _clashing_ids = set(
        _future_active.filter(aircraft__isnull=False)
        .annotate(_has_clash=_Exists(_ac_clash_inner))
        .filter(_has_clash=True)
        .values_list('id', flat=True)
    ) | set(
        _future_active.filter(instructor__isnull=False)
        .annotate(_has_clash=_Exists(_instr_clash_inner))
        .filter(_has_clash=True)
        .values_list('id', flat=True)
    )

    # ── Needs-attention section (aircraft/instructor filters apply; status filter applies on active tab) ──
    from django.db.models import F as _F
    _near_cutoff = today + timedelta(days=7)
    _attn_qs = _base_qs.filter(
        _Q2(status='departed') |
        _Q2(status='completed', flight_completion__paid_at__isnull=True) |
        _Q2(status__in=['pending', 'confirmed'], scheduled_start__date__lte=_near_cutoff) |
        _Q2(id__in=_clashing_ids) |
        _Q2(status='completed', flight_completion__amount_paid__gt=_F('flight_completion__total_charge'))
    )
    if f_status:
        _attn_qs = _attn_qs.filter(status=f_status)
    _attn_list = list(_attn_qs.order_by('scheduled_start'))
    _attn_list.sort(key=lambda b: (_STATUS_ORDER.get(b.status, 9), b.scheduled_start))
    attention_data = [{'b': b, 'reasons': conflict_reasons(b)} for b in _attn_list]

    # ── Full bookings list (all filters apply) ──
    _bk_sort     = request.GET.get('sort', 'date')
    _bk_sort_dir = request.GET.get('dir', 'asc')
    _BK_SORT = {
        'date':       ('scheduled_start',),
        'member':     ('member__user__last_name', 'member__user__first_name', 'scheduled_start'),
        'aircraft':   ('aircraft__registration', 'scheduled_start'),
        'instructor': ('instructor__last_name', 'scheduled_start'),
        'status':     ('status', 'scheduled_start'),
    }
    if _bk_sort not in _BK_SORT:
        _bk_sort = 'date'
    if _bk_sort_dir not in ('asc', 'desc'):
        _bk_sort_dir = 'asc'
    _bk_order = _BK_SORT[_bk_sort]
    if _bk_sort_dir == 'desc':
        _bk_order = tuple('-'+f for f in _bk_order)

    # Apply the same sort to the active-tab attention list
    _ATTN_SORT_KEYS = {
        'member':     lambda d: (d['b'].member.user.last_name.lower() if d['b'].member and d['b'].member.user else '', d['b'].scheduled_start),
        'aircraft':   lambda d: (d['b'].aircraft.registration if d['b'].aircraft else '', d['b'].scheduled_start),
        'instructor': lambda d: (d['b'].instructor.last_name.lower() if d['b'].instructor else '', d['b'].scheduled_start),
        'date':       lambda d: d['b'].scheduled_start,
        'status':     lambda d: (d['b'].status, d['b'].scheduled_start),
    }
    if _bk_sort in _ATTN_SORT_KEYS:
        attention_data.sort(key=_ATTN_SORT_KEYS[_bk_sort], reverse=(_bk_sort_dir == 'desc'))

    _list_qs = _base_qs
    if f_status:
        _list_qs = _list_qs.filter(status=f_status)
    qs = _list_qs.exclude(status='cancelled').order_by(*_bk_order)

    using_default_window = False
    if not (f_date_from or f_date_to or f_status or show_all_history):
        from django.db.models import Q as _Q3
        cutoff = today - timedelta(days=30)
        qs = qs.filter(
            _Q3(status__in=['pending', 'confirmed', 'departed']) |
            _Q3(status__in=['completed', 'transferred'], scheduled_start__date__gte=cutoff)
        )
        using_default_window = True
    if f_date_from:
        try:
            qs = qs.filter(scheduled_start__date__gte=f_date_from)
        except Exception:
            pass
    if f_date_to:
        try:
            qs = qs.filter(scheduled_start__date__lte=f_date_to)
        except Exception:
            pass
    from django.core.paginator import Paginator as _Pag
    _all_bookings = [{'b': b, 'reasons': conflict_reasons(b)} for b in qs]
    _paginator     = _Pag(_all_bookings, 50)
    bookings_page  = _paginator.get_page(request.GET.get('page'))

    aircraft_list = Aircraft.objects.filter(club=club, status='online').order_by('registration')
    instructors = ClubMember.objects.filter(club=club, is_on_instructor_roster=True).select_related('user')
    members_qs = ClubMember.objects.filter(club=club).select_related('user').order_by('user__last_name')

    from urllib.parse import urlencode as _ue
    def _tab_url(t):
        p = {k: v for k, v in request.GET.items() if k not in ('tab', 'page') and v}
        p['tab'] = t
        return '?' + _ue(p)

    # Build a query-string for pagination links that preserves all filters except page
    _filter_qs = _ue({k: v for k, v in request.GET.items() if k != 'page' and v})
    _bk_base_qs = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir') and v})

    return render(request, 'core/manage_bookings.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'tab': tab,
        'url_tab_active': _tab_url('active'),
        'url_tab_all': _tab_url('all'),
        'attention_data': attention_data,
        'bookings_page': bookings_page,
        'f_aircraft': f_aircraft, 'f_instructor': f_instructor,
        'f_status': f_status, 'f_date_from': f_date_from, 'f_date_to': f_date_to,
        'aircraft_list': aircraft_list, 'instructors': instructors, 'members_qs': members_qs,
        'using_default_window': using_default_window,
        'filter_qs': _filter_qs,
        'sort': _bk_sort, 'sort_dir': _bk_sort_dir, 'base_qs': _bk_base_qs,
    })


@login_required
@transaction.atomic
def booking_detail(request, club_slug, booking_id):
    from .models import (FlightCompletion, FlightChargeItem, FlightLandingEntry,
                         AccountTransaction, FuelSurchargeRate, ChargeRate, Aerodrome)
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    booking = get_object_or_404(Booking, club=club, id=booking_id)
    is_own_booking = (booking.member == actor)
    # Non-staff can only view their own bookings (read-only + cancel)
    if not (actor.is_admin or actor.is_instructor or is_own_booking):
        return render(request, 'core/no_access.html', {'club': club}, status=403)

    error = None
    success = None

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'confirm' and booking.status == 'pending' and (actor.is_admin or actor.is_instructor):
            booking.status = 'confirmed'
            booking.confirmed_by = request.user
            booking.confirmed_at = timezone.now()
            booking.save()
            _audit(booking, request.user, 'confirmed')
            from .services import notification_service
            notification_service.notify_booking_confirmed(booking)
            from .email_notifications import booking_confirmed as _email_confirmed
            _email_confirmed(booking)
            success = 'Booking confirmed.'

        elif action == 'cancel_booking' and booking.status in ('pending', 'confirmed'):
            reason = request.POST.get('cancellation_reason', 'no_longer_required')
            reason_other = request.POST.get('cancellation_reason_other', '')
            result = booking_service.cancel(
                booking, request.user,
                reason=reason,
                reason_other=reason_other,
            )
            if result.ok:
                _audit(booking, request.user, 'cancelled')
                from .services import notification_service as _ns
                _ns.notify_booking_cancelled(booking)
                from .email_notifications import booking_cancelled as _email_cancelled
                _email_cancelled(booking, reason=reason)
                success = 'Booking cancelled.'
            else:
                error = result.error

        elif action == 'depart' and booking.status == 'confirmed' and (actor.is_admin or actor.is_instructor):
            no_decl_reason = request.POST.get('no_declaration_reason', '').strip()
            elig_override_reason = request.POST.get('eligibility_override_reason', '').strip()
            requires_decl = booking.flight_type.requires_declaration
            has_decl = hasattr(booking, 'declaration') and not booking.declaration.is_draft
            # Run eligibility to check for blocks (POST-side validation mirrors the GET-side display)
            _elig = qualification_service.check_eligibility(booking) if booking.member else None
            has_elig_blocks = _elig.has_blocks if _elig else False
            if requires_decl and not has_decl and not no_decl_reason:
                error = 'This flight type requires a departure declaration. Provide a reason to override.'
            elif has_elig_blocks and not elig_override_reason:
                error = 'One or more compliance checks failed. Provide an override reason to proceed.'
            else:
                booking.status = 'departed'
                booking.departed_at = timezone.now()
                if requires_decl and not has_decl:
                    booking.departed_without_declaration = True
                    booking.departed_without_declaration_reason = no_decl_reason
                # Snapshot fuel rate at departure
                fuel_rate = FuelSurchargeRate.current_rate(club, booking.aircraft)
                booking.save()
                # Store snapshot on a pending FlightCompletion stub
                FlightCompletion.objects.get_or_create(
                    booking=booking,
                    defaults={
                        'logged_by': request.user,
                        'fuel_surcharge_rate_snapshot': fuel_rate.rate if fuel_rate else None,
                        'departed_with_aircraft': booking.aircraft,
                        'departed_with_instructor': booking.instructor,
                    }
                )
                audit_notes = f'Compliance override: {elig_override_reason}' if elig_override_reason else None
                _audit(booking, request.user, 'departed', notes=audit_notes)
                success = 'Checked out.'

        elif action == 'undo_depart' and booking.status == 'departed' and actor.is_admin:
            fc = getattr(booking, 'flight_completion', None)
            # Delete the FC shell created at departure (fuel rate snapshot only, no charges yet)
            if fc and (fc.amount_paid or 0) > 0:
                error = 'Cannot undo departure — a payment is recorded against this flight.'
            elif fc and fc.charge_items.exists():
                error = 'Cannot undo departure — charge items exist. Use void check-in instead.'
            else:
                if fc:
                    fc.delete()
                booking.status = 'confirmed'
                booking.departed_at = None
                booking.departed_without_declaration = False
                booking.departed_without_declaration_reason = ''
                booking.save(update_fields=['status', 'departed_at',
                                            'departed_without_declaration',
                                            'departed_without_declaration_reason'])
                _audit(booking, request.user, 'undo_depart')
                success = 'Departure undone — flight is back to confirmed.'

        elif action == 'change_details' and booking.status in ('pending', 'confirmed') and actor.is_admin:
            new_instr_id = request.POST.get('instructor_id', '').strip()
            new_ac_id    = request.POST.get('aircraft_id', '').strip()
            new_desc     = request.POST.get('description', booking.description or '').strip()
            changed = []
            old_instructor = booking.instructor
            if new_instr_id == '__none__':
                booking.instructor = None
                changed.append('instructor')
            elif new_instr_id:
                from .models import User as _U
                new_instr = _U.objects.filter(id=new_instr_id).first()
                if new_instr and new_instr != booking.instructor:
                    booking.instructor = new_instr
                    changed.append('instructor')
            if new_ac_id:
                new_ac = Aircraft.objects.filter(club=club, id=new_ac_id).first()
                if new_ac and new_ac != booking.aircraft:
                    booking.aircraft = new_ac
                    changed.append('aircraft')
            if new_desc != (booking.description or ''):
                booking.description = new_desc
                changed.append('description')
            if changed:
                booking.save(update_fields=['instructor', 'aircraft', 'description'])
                _audit(booking, request.user, 'edited', notes=', '.join(changed) + ' updated')
                notif_changes = ', '.join(changed)
                from .services import notification_service as _ns
                _ns.notify_booking_amended(booking, changes=notif_changes, old_instructor=old_instructor)
                from .email_notifications import booking_amended as _email_amended
                _email_amended(booking, changes=notif_changes)
                success = 'Booking updated.'

        elif action == 'undo_confirm' and booking.status == 'confirmed' and actor.is_admin:
            booking.status = 'pending'
            booking.confirmed_by = None
            booking.confirmed_at = None
            booking.save(update_fields=['status', 'confirmed_by', 'confirmed_at'])
            _audit(booking, request.user, 'undo_confirm')
            from .services import notification_service as _ns
            _ns.notify_booking_unconfirmed(booking)
            from .email_notifications import booking_unconfirmed as _email_unconfirmed
            _email_unconfirmed(booking)
            success = 'Confirmation undone — flight is back to pending.'

        elif action == 'checkin' and booking.status == 'departed' and (actor.is_admin or actor.is_instructor):
            outcome = request.POST.get('outcome', 'completed')
            outcome_notes = request.POST.get('outcome_notes', '').strip()
            hobbs_start      = request.POST.get('hobbs_start', '').strip() or None
            hobbs_end        = request.POST.get('hobbs_end', '').strip() or None
            tacho_start      = request.POST.get('tacho_start', '').strip() or None
            tacho_end        = request.POST.get('tacho_end', '').strip() or None
            airswitch_start  = request.POST.get('airswitch_start', '').strip() or None
            airswitch_end    = request.POST.get('airswitch_end', '').strip() or None
            new_ft_id = request.POST.get('flight_type_id', '').strip()
            gap_explanation = request.POST.get('gap_explanation', '').strip()

            ac = booking.aircraft
            if ac.records_hobbs and (not hobbs_start or not hobbs_end):
                error = 'Hobbs start and end are required for this aircraft.'
            elif ac.records_tacho and (not tacho_start or not tacho_end):
                error = 'Tacho start and end are required for this aircraft.'
            elif ac.records_airswitch and (not airswitch_start or not airswitch_end):
                error = 'Air switch start and end are required for this aircraft.'
            elif hobbs_start and hobbs_end:
                try:
                    if float(hobbs_end) <= float(hobbs_start):
                        error = 'Hobbs end must be greater than start.'
                except ValueError:
                    error = 'Invalid Hobbs reading.'
            elif tacho_start and tacho_end:
                try:
                    if float(tacho_end) <= float(tacho_start):
                        error = 'Tacho end must be greater than start.'
                except ValueError:
                    error = 'Invalid Tacho reading.'
            elif airswitch_start and airswitch_end:
                try:
                    if float(airswitch_end) <= float(airswitch_start):
                        error = 'Air switch end must be greater than start.'
                except ValueError:
                    error = 'Invalid air switch reading.'

            # Gap detection: compare submitted start against the last recorded end for this aircraft
            if not error:
                from django.db.models import Q as _Q
                prev_fc = (FlightCompletion.objects
                           .filter(booking__aircraft=ac, booking__club=club)
                           .exclude(booking=booking)
                           .filter(_Q(hobbs_end__isnull=False) | _Q(tacho_end__isnull=False) | _Q(airswitch_end__isnull=False))
                           .order_by('-booking__arrived_at', '-created_at')
                           .first())
                _gap_detected = False
                _gap_label = ''
                if prev_fc:
                    try:
                        if hobbs_start and prev_fc.hobbs_end is not None:
                            gap_h = float(hobbs_start) - float(prev_fc.hobbs_end)
                            if gap_h > 0.05:
                                _gap_detected = True
                                _gap_label = (f'Hobbs start {hobbs_start} is {gap_h:.1f}h ahead of '
                                              f'last recorded end ({prev_fc.hobbs_end})')
                        if not _gap_detected and tacho_start and prev_fc.tacho_end is not None:
                            gap_t = float(tacho_start) - float(prev_fc.tacho_end)
                            if gap_t > 0.005:
                                _gap_detected = True
                                _gap_label = (f'Tacho start {tacho_start} is {gap_t:.2f} ahead of '
                                              f'last recorded end ({prev_fc.tacho_end})')
                        if not _gap_detected and airswitch_start and prev_fc.airswitch_end is not None:
                            gap_a = float(airswitch_start) - float(prev_fc.airswitch_end)
                            if gap_a > 0.05:
                                _gap_detected = True
                                _gap_label = (f'Air switch start {airswitch_start} is {gap_a:.1f}h ahead of '
                                              f'last recorded end ({prev_fc.airswitch_end})')
                    except (TypeError, ValueError):
                        pass
                if _gap_detected and not gap_explanation:
                    error = f'Meter gap detected — {_gap_label}. An explanation is required (see warning below).'

            # Validate split handover readings are between flight start and end
            if not error and request.POST.get('has_split') == '1':
                _sv_meth = ac.total_time_method
                _sv_start_str = {'tacho': tacho_start, 'hobbs': hobbs_start, 'airswitch': airswitch_start}.get(_sv_meth)
                _sv_end_str   = {'tacho': tacho_end,   'hobbs': hobbs_end,   'airswitch': airswitch_end}.get(_sv_meth)
                if _sv_start_str and _sv_end_str:
                    try:
                        _sv_s    = float(_sv_start_str)
                        _sv_e    = float(_sv_end_str)
                        _sv_prev = _sv_s
                        for _hn in range(1, 4):
                            _hv_str = request.POST.get(f'seg_{_sv_meth}_h{_hn}', '').strip()
                            if not _hv_str:
                                break
                            _hv = float(_hv_str)
                            if _hv <= _sv_prev:
                                error = f'Handover {_hn} reading ({_hv_str}) must be greater than the previous value ({_sv_prev}).'
                                break
                            if _hv >= _sv_e:
                                error = f'Handover {_hn} reading ({_hv_str}) must be less than the flight end ({_sv_end_str}).'
                                break
                            _sv_prev = _hv
                    except (ValueError, TypeError):
                        error = 'Invalid split handover reading.'

            if not error:
                fc, _ = FlightCompletion.objects.get_or_create(
                    booking=booking, defaults={'logged_by': request.user}
                )
                fc.outcome = outcome
                fc.outcome_notes = outcome_notes
                fc.hobbs_start     = hobbs_start
                fc.hobbs_end       = hobbs_end
                fc.tacho_start     = tacho_start
                fc.tacho_end       = tacho_end
                fc.airswitch_start = airswitch_start
                fc.airswitch_end   = airswitch_end
                fc.logged_by = request.user

                method = booking.aircraft.total_time_method
                try:
                    if method == 'hobbs' and hobbs_start and hobbs_end:
                        fc.actual_flight_hours = float(hobbs_end) - float(hobbs_start)
                    elif method == 'tacho' and tacho_start and tacho_end:
                        fc.actual_flight_hours = round(float(tacho_end) - float(tacho_start), 2)
                    elif method == 'airswitch' and airswitch_start and airswitch_end:
                        fc.actual_flight_hours = float(airswitch_end) - float(airswitch_start)
                except (ValueError, TypeError):
                    pass

                if booking.instructor:
                    instr_member = ClubMember.objects.filter(user=booking.instructor, club=club).first()
                    if instr_member and instr_member.instructor_grade:
                        fc.instructor_rate_snapshot = instr_member.instructor_grade.hourly_rate

                if new_ft_id:
                    new_ft = FlightType.objects.filter(club=club, id=new_ft_id).first()
                    if new_ft and new_ft != booking.flight_type:
                        fc.original_flight_type = booking.flight_type
                        booking.flight_type = new_ft
                        booking.save(update_fields=['flight_type'])

                if gap_explanation:
                    fc.meter_gap_note = gap_explanation
                fc.save()
                create_maint_log_entry(fc)
                for _mi in booking.aircraft.maintenance_items.all():
                    _mi.recalc_urgency()
                    _mi.save(update_fields=['urgency'])
                booking.status = 'completed'
                booking.arrived_at = timezone.now()
                booking.save(update_fields=['status', 'arrived_at'])

                # Split flight segments (up to 4 members)
                from .models import FlightSegment as _FS
                _segments = []
                if request.POST.get('has_split') == '1':
                    _meth = booking.aircraft.total_time_method

                    def _hpvals(n):
                        return (
                            request.POST.get(f'seg_hobbs_h{n}',     '').strip() or None,
                            request.POST.get(f'seg_tacho_h{n}',     '').strip() or None,
                            request.POST.get(f'seg_airswitch_h{n}', '').strip() or None,
                        )

                    _hp = [_hpvals(n) for n in range(1, 4)]   # (hobbs, tacho, air) for H1 H2 H3
                    _extra_ids = [request.POST.get(f'seg_member_{i}', '').strip() for i in range(2, 5)]
                    _extra = [ClubMember.objects.filter(club=club, id=mid).first() if mid else None
                              for mid in _extra_ids]

                    # Build member list and handover list, stopping at first incomplete pair
                    _seg_members   = [booking.member]
                    _seg_handovers = []   # list of (hobbs, tacho, air) tuples
                    _split_error   = None
                    for _hvtuple, _em in zip(_hp, _extra):
                        if any(_hvtuple) and _em:
                            _seg_handovers.append(_hvtuple)
                            _seg_members.append(_em)
                        elif any(_hvtuple) and not _em:
                            _split_error = 'A handover reading was provided but no member was selected for that segment.'
                            break
                        else:
                            break
                    if not _split_error and len(_seg_members) >= 2:
                        _seen_ids = set()
                        for _m in _seg_members:
                            if _m.id in _seen_ids:
                                _split_error = f'{_m.user.get_full_name()} appears more than once in the split — each pilot must be unique.'
                                break
                            _seen_ids.add(_m.id)
                    if _split_error:
                        error = _split_error

                    if not error and len(_seg_members) >= 2:
                        _fc_start = (fc.hobbs_start, fc.tacho_start, fc.airswitch_start)
                        _fc_end   = (fc.hobbs_end,   fc.tacho_end,   fc.airswitch_end)
                        for _i, _m in enumerate(_seg_members):
                            _seg_s = _seg_handovers[_i - 1] if _i > 0         else _fc_start
                            _seg_e = _seg_handovers[_i]     if _i < len(_seg_handovers) else _fc_end
                            hs, ts, as_ = _seg_s
                            he, te, ae  = _seg_e
                            _seg = _FS.objects.create(
                                flight_completion=fc, member=_m, sequence=_i + 1,
                                hobbs_start=hs, hobbs_end=he,
                                tacho_start=ts, tacho_end=te,
                                airswitch_start=as_, airswitch_end=ae,
                                hours=_calc_segment_hours(_meth, hs, he, ts, te, as_, ae),
                            )
                            _segments.append(_seg)

                hours = fc.actual_flight_hours
                if _segments:
                    _generate_segment_charges(fc, _segments, booking)
                else:
                    hire_rate = ChargeRate.objects.filter(
                        aircraft=booking.aircraft, flight_type=booking.flight_type,
                        time_method=booking.aircraft.total_time_method
                    ).first()
                    if hire_rate and hours:
                        FlightChargeItem.objects.get_or_create(
                            flight_completion=fc, item_type='hire',
                            defaults={'description': f'Aircraft hire — {booking.aircraft.registration}',
                                      'amount': round(float(hire_rate.amount) * float(hours), 2)}
                        )
                    if fc.fuel_surcharge_rate_snapshot and hours and not (hire_rate and hire_rate.includes_fuel) and fc.fuel_surcharge_rate_snapshot > 0:
                        FlightChargeItem.objects.get_or_create(
                            flight_completion=fc, item_type='fuel',
                            defaults={'description': 'Fuel charge',
                                      'amount': round(float(fc.fuel_surcharge_rate_snapshot) * float(hours), 2)}
                        )
                    if fc.instructor_rate_snapshot and hours and booking.instructor:
                        FlightChargeItem.objects.get_or_create(
                            flight_completion=fc, item_type='instructor',
                            defaults={'description': f'Instructor fee — {booking.instructor.get_full_name()}',
                                      'amount': round(float(fc.instructor_rate_snapshot) * float(hours), 2)}
                        )
                    for sc in booking.aircraft.surcharges.all():
                        FlightChargeItem.objects.get_or_create(
                            flight_completion=fc, item_type='surcharge',
                            defaults={'description': sc.name, 'amount': sc.amount}
                        )

                _update_total(fc)
                _audit(booking, request.user, 'completed')
                from .services import notification_service as _ns
                _ns.notify_flight_charged(fc)
                success = 'Flight checked in.'

        elif action == 'edit_checkin' and booking.status == 'completed' and actor.is_admin:
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                hobbs_start      = request.POST.get('hobbs_start', '').strip() or None
                hobbs_end        = request.POST.get('hobbs_end', '').strip() or None
                tacho_start      = request.POST.get('tacho_start', '').strip() or None
                tacho_end        = request.POST.get('tacho_end', '').strip() or None
                airswitch_start  = request.POST.get('airswitch_start', '').strip() or None
                airswitch_end    = request.POST.get('airswitch_end', '').strip() or None
                outcome          = request.POST.get('outcome', fc.outcome)
                outcome_notes    = request.POST.get('outcome_notes', fc.outcome_notes or '').strip()
                ac = booking.aircraft
                if ac.records_hobbs and (not hobbs_start or not hobbs_end):
                    error = 'Hobbs start and end are required for this aircraft.'
                if not error and ac.records_tacho and (not tacho_start or not tacho_end):
                    error = 'Tacho start and end are required for this aircraft.'
                if not error and ac.records_airswitch and (not airswitch_start or not airswitch_end):
                    error = 'Air switch start and end are required for this aircraft.'
                if not error:
                    try:
                        if hobbs_start and hobbs_end and float(hobbs_end) <= float(hobbs_start):
                            error = 'Hobbs end must be greater than start.'
                        if tacho_start and tacho_end and float(tacho_end) <= float(tacho_start):
                            error = 'Tacho end must be greater than start.'
                        if airswitch_start and airswitch_end and float(airswitch_end) <= float(airswitch_start):
                            error = 'Air switch end must be greater than start.'
                    except ValueError:
                        error = 'Invalid meter reading.'
                if not error:
                    fc.hobbs_start      = hobbs_start
                    fc.hobbs_end        = hobbs_end
                    fc.tacho_start      = tacho_start
                    fc.tacho_end        = tacho_end
                    fc.airswitch_start  = airswitch_start
                    fc.airswitch_end    = airswitch_end
                    fc.outcome          = outcome
                    fc.outcome_notes    = outcome_notes
                    _special_fields = ['time_if_simulated', 'time_if_actual', 'time_night', 'time_low_flying', 'time_terrain_awareness']
                    for _sf in _special_fields:
                        _sv = request.POST.get(_sf, '').strip()
                        setattr(fc, _sf, float(_sv) if _sv else None)
                    method = ac.total_time_method
                    try:
                        if method == 'hobbs' and hobbs_start and hobbs_end:
                            fc.actual_flight_hours = round(float(hobbs_end) - float(hobbs_start), 2)
                        elif method == 'tacho' and tacho_start and tacho_end:
                            fc.actual_flight_hours = round(float(tacho_end) - float(tacho_start), 2)
                        elif method == 'airswitch' and airswitch_start and airswitch_end:
                            fc.actual_flight_hours = round(float(airswitch_end) - float(airswitch_start), 2)
                    except (ValueError, TypeError):
                        pass
                    fc.save()
                    # Rebuild auto-generated charges; preserve manual/landing/one-off items.
                    # Clear segments — edit_checkin always produces non-segmented charges.
                    fc.segments.all().delete()
                    fc.charge_items.filter(
                        item_type__in=['hire', 'fuel', 'instructor', 'surcharge']
                    ).delete()
                    hours = fc.actual_flight_hours
                    hire_rate = ChargeRate.objects.filter(
                        aircraft=ac, flight_type=booking.flight_type,
                        time_method=ac.total_time_method
                    ).first()
                    if hire_rate and hours:
                        FlightChargeItem.objects.create(
                            flight_completion=fc, item_type='hire',
                            description=f'Aircraft hire — {ac.registration}',
                            amount=round(float(hire_rate.amount) * float(hours), 2),
                        )
                    if fc.fuel_surcharge_rate_snapshot and hours and not (hire_rate and hire_rate.includes_fuel) and fc.fuel_surcharge_rate_snapshot > 0:
                        FlightChargeItem.objects.create(
                            flight_completion=fc, item_type='fuel',
                            description='Fuel charge',
                            amount=round(float(fc.fuel_surcharge_rate_snapshot) * float(hours), 2),
                        )
                    if fc.instructor_rate_snapshot and hours and booking.instructor:
                        FlightChargeItem.objects.create(
                            flight_completion=fc, item_type='instructor',
                            description=f'Instructor fee — {booking.instructor.get_full_name()}',
                            amount=round(float(fc.instructor_rate_snapshot) * float(hours), 2),
                        )
                    for sc in ac.surcharges.all():
                        FlightChargeItem.objects.create(
                            flight_completion=fc, item_type='surcharge',
                            description=sc.name, amount=sc.amount,
                        )
                    _update_total(fc)
                    for _mi in booking.aircraft.maintenance_items.all():
                        _mi.recalc_urgency()
                        _mi.save(update_fields=['urgency'])
                    _audit(booking, request.user, 'edit_checkin')
                    success = 'Check-in details updated. Review charges and payment below.'

        elif action == 'set_client' and actor.can_access_manage:
            from .models import Contact as _Cont
            cid = request.POST.get('client_id', '').strip()
            booking.client = _Cont.objects.filter(id=cid, club=club).first() if cid else None
            booking.billed_to = request.POST.get('billed_to', '')
            booking.save(update_fields=['client', 'billed_to'])
            return redirect(request.path + ('?inline=1&saved=1' if is_inline else '?saved=1'))

        elif action == 'add_charge' and booking.status == 'completed' and (actor.is_admin or actor.is_instructor):
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                item_type = request.POST.get('item_type', 'one_off')
                description = request.POST.get('description', '').strip()
                amount = request.POST.get('amount', '').strip()
                ae_id  = request.POST.get('aerodrome_id', '')
                ft_id  = request.POST.get('fee_type_id', '')
                from .models import AerodromeFeeType
                ae = Aerodrome.objects.filter(club=club, id=ae_id).first() if ae_id else None
                fee_type = AerodromeFeeType.objects.filter(id=ft_id).first() if ft_id else None
                result = charging_service.add_charge(
                    fc, item_type, description, amount,
                    aerodrome=ae, fee_type=fee_type,
                    custom_icao=request.POST.get('custom_icao', '').strip(),
                    custom_name=request.POST.get('custom_name', '').strip(),
                    quantity=int(request.POST.get('quantity', 1) or 1),
                    unit_amount=request.POST.get('unit_amount', amount) or amount,
                )
                if result.ok:
                    _inv_count = fc.invoices.exclude(status='void').count()
                    success = 'Charge added.' + (
                        f' ⚠ This flight has {_inv_count} issued invoice(s) — '
                        'the invoice total no longer matches. Update or re-issue the invoice.'
                        if _inv_count else ''
                    )
                else:
                    error = result.error

        elif action == 'delete_charge' and booking.status == 'completed':
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                result = charging_service.delete_charge(fc, request.POST.get('item_id'))
                if result.ok:
                    _inv_count = fc.invoices.exclude(status='void').count()
                    success = 'Charge removed.' + (
                        f' ⚠ This flight has {_inv_count} issued invoice(s) — '
                        'the invoice total no longer matches. Update or re-issue the invoice.'
                        if _inv_count else ''
                    )

        elif action == 'confirm_payment' and booking.status == 'completed':
            # Quick single-payment record (default to booking member)
            fc = getattr(booking, 'flight_completion', None)
            if fc and not fc.is_paid:
                amount_str = request.POST.get('payment_amount', '').strip()
                pay_amount = amount_str if amount_str else str(fc.balance_owing)
                result = charging_service.record_payment(
                    fc, booking, request.user, pay_amount,
                    method=request.POST.get('payment_method', 'eftpos'),
                )
                if result.ok:
                    success = result.data['message']
                else:
                    error = result.error

        elif action == 'add_payee' and booking.status == 'completed':
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                from .models import ClubMember as _CM
                member_id = request.POST.get('member_id', '').strip()
                amount_str = request.POST.get('payment_amount', '').strip()
                method = request.POST.get('payment_method', 'eftpos')
                try:
                    payee = _CM.objects.get(id=member_id, club=club)
                except _CM.DoesNotExist:
                    error = 'Member not found'
                else:
                    result = charging_service.allocate_payment(
                        fc, booking, request.user, amount_str, method=method, member=payee
                    )
                    if result.ok:
                        success = 'Payee added.'
                    else:
                        error = result.error

        elif action == 'record_payee' and booking.status == 'completed':
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                _pid_for_check = request.POST.get('payment_id', '').strip()
                _fp_for_check = fc.payments.filter(id=_pid_for_check).first()
                _fc_inv = (fc.invoices.filter(member=_fp_for_check.member, status__in=['draft', 'sent']).first()
                           if _fp_for_check else None)
                if _fc_inv:
                    error = f'Payment is via invoice {_fc_inv.display_number}. Record payment against the invoice.'
                else:
                    from decimal import Decimal as _D, InvalidOperation as _IO
                    from django.db.models import Sum as _Sum
                    payment_id = request.POST.get('payment_id', '').strip()
                    new_amount_str = request.POST.get('payment_amount', '').strip()
                    new_method = request.POST.get('payment_method', '').strip()
                    # Update pending fp amount/method if the form submitted new values
                    if new_amount_str:
                        try:
                            from .models import FlightPayment as _FP
                            _new_amt = round(_D(new_amount_str), 2)
                            if _new_amt <= 0:
                                raise _IO('zero')
                            _fp_obj = fc.payments.get(id=payment_id, paid_at__isnull=True)
                            _other = fc.payments.exclude(id=payment_id).aggregate(t=_Sum('amount'))['t'] or _D('0')
                            if _other + _new_amt > _D(str(fc.total_charge or 0)):
                                _over = _other + _new_amt - _D(str(fc.total_charge or 0))
                                error = f'Amount ${_new_amt:.2f} would exceed the flight total by ${_over:.2f}.'
                            else:
                                _fp_obj.amount = _new_amt
                                if new_method:
                                    _fp_obj.method = new_method
                                _fp_obj.save(update_fields=['amount', 'method'])
                        except (_IO, Exception):
                            if not error:
                                error = 'Invalid amount.'
                    if not error:
                        # Check total allocations match flight total before recording
                        _all_alloc = fc.payments.aggregate(t=_Sum('amount'))['t'] or _D('0')
                        _fc_total = _D(str(fc.total_charge or 0))
                        if abs(_all_alloc - _fc_total) > _D('0.01'):
                            _diff = _fc_total - _all_alloc
                            error = (f'Allocations total ${_all_alloc:.2f} but flight total is ${_fc_total:.2f} '
                                     f'(${abs(_diff):.2f} {"unallocated" if _diff > 0 else "over-allocated"}). '
                                     f'Adjust allocations before recording payment.')
                    if not error:
                        result = charging_service.record_allocated_payment(
                            fc, booking, request.user, payment_id
                        )
                        if result.ok:
                            success = result.data['message']
                        else:
                            error = result.error

        elif action == 'record_new_payee' and booking.status == 'completed':
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                _mid_for_check = request.POST.get('member_id', '').strip()
                _fc_inv = fc.invoices.filter(member_id=_mid_for_check, status__in=['draft', 'sent']).first()
                if _fc_inv:
                    error = f'Payment is via invoice {_fc_inv.display_number}. Record payment against the invoice.'
                else:
                    from .models import ClubMember as _CM
                    member_id = request.POST.get('member_id', '').strip()
                    amount_str = request.POST.get('payment_amount', '').strip()
                    method = request.POST.get('payment_method', 'eftpos')
                    try:
                        payee = _CM.objects.get(id=member_id, club=club)
                    except _CM.DoesNotExist:
                        error = 'Member not found'
                    else:
                        result = charging_service.record_payment(
                            fc, booking, request.user, amount_str, method=method, member=payee
                        )
                        if result.ok:
                            success = result.data['message']
                        else:
                            error = result.error

        elif action == 'remove_payee' and booking.status == 'completed':
            fc = getattr(booking, 'flight_completion', None)
            if fc and actor.is_admin:
                payment_id = request.POST.get('payment_id', '').strip()
                result = charging_service.remove_payment_allocation(fc, payment_id)
                if result.ok:
                    success = 'Payee removed.'
                else:
                    error = result.error

        elif action == 'record_multi_payment' and booking.status == 'completed':
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                _fc_inv = fc.invoices.filter(member=booking.member, status__in=['draft', 'sent']).first()
                if _fc_inv:
                    error = f'Payment is via invoice {_fc_inv.display_number}. Record payment against the invoice.'
                else:
                    from decimal import Decimal as _D, InvalidOperation
                    method = request.POST.get('payment_method', 'eftpos')
                    received_str = request.POST.get('amount_received', '').strip()
                    try:
                        received = _D(received_str)
                    except InvalidOperation:
                        error = 'Invalid amount received'
                    else:
                        # Arrears clearance — separate from flight amounts
                        try:
                            arrears_clear_amt = _D(request.POST.get('arrears_amount', '0') or '0')
                        except InvalidOperation:
                            arrears_clear_amt = _D('0')
                        include_arrears = bool(
                            request.POST.get('include_arrears') and arrears_clear_amt > 0
                        )
                        received_for_flights = received - arrears_clear_amt if include_arrears else received

                        # Build ordered list: current flight first, then selected others by date
                        fc_amounts = [(fc, booking, fc.balance_owing)]
                        selected_ids = request.POST.getlist('other_fc_ids')
                        if selected_ids:
                            _other_fcs = (FlightCompletion.objects
                                          .filter(id__in=selected_ids,
                                                  booking__member=booking.member,
                                                  booking__club=club)
                                          .select_related('booking__aircraft', 'booking')
                                          .order_by('booking__scheduled_start'))
                            for ofc in _other_fcs:
                                amt_str = request.POST.get(f'other_amount_{ofc.id}', '').strip()
                                try:
                                    amt = _D(amt_str)
                                except InvalidOperation:
                                    amt = ofc.balance_owing
                                fc_amounts.append((ofc, ofc.booking, amt))
                        result = charging_service.record_multi_payment(
                            fc, booking, request.user, method, fc_amounts, received_for_flights
                        )
                        if result.ok:
                            if include_arrears:
                                from .models import AccountTransaction as _AT
                                try:
                                    _acct2 = booking.member.account
                                    _AT.objects.create(
                                        account=_acct2,
                                        transaction_type='deposit',
                                        direction='credit',
                                        amount=arrears_clear_amt,
                                        payment_method=method,
                                        description='Arrears clearance — collected with flight payment',
                                        created_by=request.user,
                                    )
                                    _acct2.apply_transaction(arrears_clear_amt, 'credit')
                                except Exception:
                                    pass
                            success = result.data['message']
                        else:
                            error = result.error

        elif action == 'void_checkin' and booking.status == 'completed' and actor.is_admin:
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                _active_invoices = fc.invoices.exclude(status='void')
                _inv_paid = _active_invoices.filter(amount_paid__gt=0).exists()
                if (fc.amount_paid and fc.amount_paid > 0) or _inv_paid:
                    error = ('Cannot void check-in — a payment has been recorded against this flight or one of its invoices. '
                             'Reverse all payments first, then void.')
                else:
                    # Void any outstanding invoices first
                    _voided_count = _active_invoices.count()
                    _active_invoices.update(status='void')
                    # Clear all charge items, segments, and meter data, reset status to departed
                    fc.charge_items.all().delete()
                    fc.segments.all().delete()
                    fc.hobbs_start = fc.hobbs_end = None
                    fc.tacho_start = fc.tacho_end = None
                    fc.airswitch_start = fc.airswitch_end = None
                    fc.actual_flight_hours = 0
                    fc.outcome = 'completed'
                    fc.outcome_notes = ''
                    fc.total_charge = 0
                    fc.meter_gap_note = ''
                    fc.save()
                    booking.status = 'departed'
                    booking.arrived_at = None
                    booking.save(update_fields=['status', 'arrived_at'])
                    _audit(booking, request.user, 'void_checkin')
                    if _voided_count:
                        success = (f'Check-in voided — flight is back to departed. '
                                   f'{_voided_count} invoice(s) voided and will need to be re-issued after re-check-in.')
                    else:
                        success = 'Check-in voided — flight is back to departed.'

        elif action == 'reverse_payment' and booking.status == 'completed' and actor.is_admin:
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                payment_id = request.POST.get('payment_id') or None
                result = charging_service.reverse_payment(fc, booking, request.user, payment_id=payment_id)
                if result.ok:
                    success = result.data['message']
                else:
                    error = result.error

        elif action == 'record_refund' and booking.status == 'completed' and actor.is_admin:
            fc = getattr(booking, 'flight_completion', None)
            if fc:
                result = charging_service.record_refund(
                    fc, booking, request.user,
                    amount=request.POST.get('refund_amount', '').strip(),
                    method=request.POST.get('refund_method', 'eftpos'),
                )
                if result.ok:
                    success = result.data['message']
                else:
                    error = result.error

        is_inline = request.POST.get('inline') == '1' or request.GET.get('inline') == '1'
        if error:
            pass  # fall through to re-render with error
        elif success and not error:
            if is_inline:
                return redirect(f'{request.path}?inline=1&saved=1')
            from django.urls import reverse as _rev
            return redirect(_rev('core:booking_detail', kwargs={'club_slug': club_slug, 'booking_id': booking_id}) + '?saved=1')

    is_inline = request.GET.get('inline') == '1'

    # GET — build context
    fc = getattr(booking, 'flight_completion', None)
    charge_items = fc.charge_items.select_related('segment').all() if fc else []

    # Split-flight segments with charge items pre-grouped
    fc_segments = []
    if fc:
        _segs = list(fc.segments.select_related('member__user').order_by('sequence'))
        if _segs:
            from collections import defaultdict as _dd
            _by_seg = _dd(list)
            for _ci in charge_items:
                if _ci.segment_id:
                    _by_seg[_ci.segment_id].append(_ci)
            from decimal import Decimal as _D
            for _s in _segs:
                _s.segment_charges = _by_seg.get(_s.id, [])
                _s.charge_total = sum(ci.amount for ci in _s.segment_charges)
            fc_segments = _segs
            charge_items = [_ci for _ci in charge_items if not _ci.segment_id]
            # Compute suggested payment per segment: segment charges + proportional share
            # of any non-segment charges (landing fees, one-offs, etc.).
            _total_seg = sum(_s.charge_total for _s in fc_segments)
            _non_seg = _D(str(fc.total_charge or 0)) - _total_seg
            _n = len(fc_segments)
            for _s in fc_segments:
                _ratio = (_s.charge_total / _total_seg) if _total_seg else (_D('1') / _n)
                _s.suggested_payment = (_s.charge_total + (_non_seg * _ratio)).quantize(_D('0.01'))
            # Last segment absorbs any rounding remainder
            if fc_segments:
                _rounding = _D(str(fc.total_charge or 0)) - sum(_s.suggested_payment for _s in fc_segments)
                fc_segments[-1].suggested_payment += _rounding
    total = fc.total_charge if fc else 0
    balance_owing = fc.balance_owing if fc else 0
    from decimal import Decimal as _D
    overpayment = max(_D('0'), (_D(str(fc.amount_paid or 0)) - _D(str(fc.total_charge or 0)))) if fc else _D('0')
    fc_payments = list(fc.payments.select_related('member__user').order_by('created_at')) if fc else []
    _allocated_member_ids = {fp.member_id for fp in fc_payments}
    fc_segments_pending = [s for s in fc_segments if s.member_id not in _allocated_member_ids]
    fc_payments_pending = [fp for fp in fc_payments if not fp.paid_at]
    fc_payments_paid = [fp for fp in fc_payments if fp.paid_at]
    fc_payments_paid_total = sum(fp.amount for fp in fc_payments_paid)

    # Build per-member invoice lookup (ForeignKey now, one per member per FC)
    _fc_invoices = list(fc.invoices.all()) if fc else []
    fc_invoices = _fc_invoices
    _invoice_by_member = {inv.member_id: inv for inv in _fc_invoices}
    # Attach invoice to each FlightPayment for template convenience
    for _fp in fc_payments:
        _fp.flight_invoice = _invoice_by_member.get(_fp.member_id)

    # True if any active invoice has a payment recorded (blocks void_checkin)
    fc_any_invoice_paid = any(inv.amount_paid > 0 for inv in _fc_invoices if inv.status != 'void')

    # True if any payee with a non-zero payment has no invoice yet
    fc_has_uninvoiced_payees = any(
        fp.amount > 0 and not _invoice_by_member.get(fp.member_id)
        for fp in fc_payments
    ) if fc_payments else (fc and not _fc_invoices)
    club_members = list(ClubMember.objects.filter(club=club).exclude(standing='resigned').select_related('user').order_by('user__last_name', 'user__first_name'))
    from .models import Contact as _Cont
    contacts = list(_Cont.objects.filter(club=club, converted_to_member__isnull=True).order_by('name'))

    # Other unpaid/partially-paid flights for this member (for payment warning)
    if fc:
        _base = (FlightCompletion.objects
                 .filter(booking__member=booking.member, booking__club=club)
                 .exclude(booking=booking)
                 .select_related('booking__aircraft', 'booking'))
        other_unpaid_list = list(
            _base.filter(paid_at__isnull=True, total_charge__gt=0)
        )
        other_outstanding_list = list(
            _base.filter(paid_at__isnull=False).extra(where=['amount_paid < total_charge'])
        )
    else:
        other_unpaid_list = []
        other_outstanding_list = []

    other_unpaid_total = sum((_D(str(x.total_charge)) for x in other_unpaid_list), _D('0'))
    other_outstanding_total = sum((_D(str(x.balance_owing)) for x in other_outstanding_list), _D('0'))
    other_total = other_unpaid_total + other_outstanding_total
    other_total_count = len(other_unpaid_list) + len(other_outstanding_list)
    aerodromes = Aerodrome.objects.filter(club=club, is_active=True).prefetch_related('fee_types')
    flight_types = FlightType.objects.filter(club=club)
    requires_decl = booking.flight_type.requires_declaration
    has_submitted_decl = (hasattr(booking, 'declaration') and
                          not booking.declaration.is_draft
                          if requires_decl else False)

    # Eligibility check — shown in the depart section so staff can see issues before letting a member fly
    eligibility = None
    if booking.status in ('confirmed', 'pending') and booking.member:
        eligibility = qualification_service.check_eligibility(booking)

    # Maintenance items for the aircraft — only AMBER/RED, shown in the same status panel
    departure_maint_items = []
    if booking.status in ('confirmed', 'pending') and booking.aircraft:
        _maint_qs = AircraftMaintenanceItem.objects.filter(
            aircraft=booking.aircraft,
            urgency__in=['amber', 'red'],
        ).order_by('urgency', 'due_date')
        _latest_hobbs = (FlightCompletion.objects
                         .filter(booking__aircraft=booking.aircraft)
                         .exclude(hobbs_end__isnull=True)
                         .order_by('-booking__arrived_at', '-created_at')
                         .values('hobbs_end').first())
        _cur_hobbs = float(_latest_hobbs['hobbs_end']) if _latest_hobbs else None
        _today_d = timezone.localdate()
        for _mi in _maint_qs:
            _detail = ''
            if _mi.last_completed_date and _mi.due_date:
                _dl = (_mi.due_date - _today_d).days
                _detail = f'{_dl}d remaining' if _dl >= 0 else f'{abs(_dl)}d overdue'
            elif _mi.last_completed_hours is not None and _mi.due_hours is not None and _cur_hobbs is not None:
                _hl = float(_mi.due_hours) - _cur_hobbs
                _detail = f'{_hl:.1f}h remaining' if _hl >= 0 else f'{abs(_hl):.1f}h overdue'
            elif _mi.due_date:
                _dl = (_mi.due_date - _today_d).days
                _detail = f'{_dl}d remaining' if _dl >= 0 else f'{abs(_dl)}d overdue'
            departure_maint_items.append({'name': _mi.name, 'urgency': _mi.urgency, 'detail': _detail})

    # Previous meter readings for this aircraft — used by the gap-detection JS in the check-in form
    from django.db.models import Q as _Q
    _prev_fc = (FlightCompletion.objects
                .filter(booking__aircraft=booking.aircraft, booking__club=club)
                .exclude(booking=booking)
                .filter(_Q(hobbs_end__isnull=False) | _Q(tacho_end__isnull=False) | _Q(airswitch_end__isnull=False))
                .order_by('-booking__arrived_at', '-created_at')
                .first()) if booking.status == 'departed' else None
    prev_hobbs_end     = float(_prev_fc.hobbs_end)     if _prev_fc and _prev_fc.hobbs_end     is not None else None
    prev_tacho_end     = float(_prev_fc.tacho_end)     if _prev_fc and _prev_fc.tacho_end     is not None else None
    prev_airswitch_end = float(_prev_fc.airswitch_end) if _prev_fc and _prev_fc.airswitch_end is not None else None

    # Rate data for the check-in JS charge preview
    import json as _json
    checkin_rates_json = 'null'
    if booking.status == 'departed':
        _hire = ChargeRate.objects.filter(
            aircraft=booking.aircraft, flight_type=booking.flight_type,
            time_method=booking.aircraft.total_time_method
        ).first()
        _instr_rate = None
        if booking.instructor:
            _im = ClubMember.objects.filter(user=booking.instructor, club=club).first()
            if _im and _im.instructor_grade:
                _instr_rate = float(_im.instructor_grade.hourly_rate)
        _fuel_snap = float(fc.fuel_surcharge_rate_snapshot) if fc and fc.fuel_surcharge_rate_snapshot else None
        _surcharge_list = [
            {'name': sc.name, 'amount': float(sc.amount)}
            for sc in booking.aircraft.surcharges.all()
        ]
        checkin_rates_json = _json.dumps({
            'time_method':    booking.aircraft.total_time_method,
            'hire_rate':      float(_hire.amount) if _hire else None,
            'includes_fuel':  _hire.includes_fuel if _hire else False,
            'fuel_rate':      _fuel_snap,
            'instructor_rate': _instr_rate,
            'surcharges':     _surcharge_list,
        })

    try:
        _acct = booking.member.account
    except Exception:
        _acct = None
    _arrears_clearable = (
        _acct and _acct.balance < 0 and _acct.has_warning
    )

    # Instructor availability at this booking's time slot (two queries, not N+1)
    _active_statuses = ['confirmed', 'departed']
    _confirmed_conflict_ids = set(
        Booking.objects.filter(
            club=club,
            status__in=_active_statuses,
            instructor__isnull=False,
            scheduled_start__lt=booking.scheduled_end,
            scheduled_end__gt=booking.scheduled_start,
        ).exclude(pk=booking.pk).values_list('instructor_id', flat=True)
    )
    _pending_conflict_ids = set(
        Booking.objects.filter(
            club=club,
            status='pending',
            instructor__isnull=False,
            scheduled_start__lt=booking.scheduled_end,
            scheduled_end__gt=booking.scheduled_start,
        ).exclude(pk=booking.pk).values_list('instructor_id', flat=True)
    ) - _confirmed_conflict_ids

    def _instr_avail(member):
        uid = member.user_id
        if uid in _confirmed_conflict_ids:
            return 'red'
        if uid in _pending_conflict_ids:
            return 'orange'
        return 'green'

    _rostered = list(
        ClubMember.objects.filter(club=club, is_on_instructor_roster=True)
        .select_related('user').order_by('user__last_name')
    )
    rostered_instructors = [(_m, _instr_avail(_m)) for _m in _rostered]

    ctx = {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'booking': booking,
        'member_account': _acct,
        'arrears_clearable': _arrears_clearable,
        'fc': fc,
        'fc_segments': fc_segments,
        'fc_segments_pending': fc_segments_pending,
        'charge_items': charge_items,
        'contacts': contacts,
        'total': total,
        'balance_owing': balance_owing,
        'overpayment': overpayment,
        'fc_payments': fc_payments,
        'fc_payments_pending': fc_payments_pending,
        'fc_payments_paid': fc_payments_paid,
        'fc_payments_paid_total': fc_payments_paid_total,
        'fc_invoices': fc_invoices,
        'fc_has_uninvoiced_payees': fc_has_uninvoiced_payees,
        'fc_any_invoice_paid': fc_any_invoice_paid,
        'club_members': club_members,
        'other_unpaid_list': other_unpaid_list,
        'other_outstanding_list': other_outstanding_list,
        'other_unpaid_total': other_unpaid_total,
        'other_outstanding_total': other_outstanding_total,
        'other_total': other_total,
        'other_total_count': other_total_count,
        'aerodromes': aerodromes,
        'flight_types': flight_types,
        'requires_decl': requires_decl,
        'has_submitted_decl': has_submitted_decl,
        'eligibility': eligibility,
        'departure_maint_items': departure_maint_items,
        'error': error,
        'success': success,
        'prev_hobbs_end': prev_hobbs_end,
        'prev_tacho_end': prev_tacho_end,
        'prev_airswitch_end': prev_airswitch_end,
        'checkin_rates_json': checkin_rates_json,
        'rostered_instructors': rostered_instructors,
        'online_aircraft': Aircraft.objects.filter(club=club, status='online').order_by('registration'),
        'base_template': 'core/base_inline.html' if is_inline else 'core/base.html',
        'inline_title': f'{booking.member.user.get_full_name()} · {booking.aircraft.registration}',
        'watchers': list(SlotWatch.objects.filter(booking=booking).select_related('club_member__user')) if actor.can_access_manage else [],
    }
    return render(request, 'core/booking_detail.html', ctx)


def _inline_redirect(request, view_name, saved=False, error='', **kwargs):
    """Redirect back to the same page, preserving ?inline=1, ?saved=1, ?err=..., ?back=, ?back_label="""
    from django.urls import reverse
    from urllib.parse import quote
    url = reverse(view_name, kwargs=kwargs)
    params = []
    if not saved and request.GET.get('inline') == '1':
        params.append('inline=1')
    if saved:
        params.append('saved=1')
    if error:
        params.append(f'err={quote(error)}')
    back = request.GET.get('back', '')
    if back:
        params.append(f'back={quote(back, safe="/")}')
    back_label = request.GET.get('back_label', '')
    if back_label:
        params.append(f'back_label={quote(back_label)}')
    if params:
        url += '?' + '&'.join(params)
    return redirect(url)


def _update_total(fc):
    """Thin wrapper — logic lives in booking_service.update_total."""
    booking_service.update_total(fc)


def _calc_segment_hours(method, hobbs_s, hobbs_e, tacho_s, tacho_e, air_s, air_e):
    from decimal import Decimal as _D
    def _d(v):
        try: return _D(str(v)) if v else None
        except Exception: return None
    try:
        if method == 'hobbs':
            s, e = _d(hobbs_s), _d(hobbs_e)
            if s and e and e > s: return round(e - s, 2)
        elif method == 'tacho':
            s, e = _d(tacho_s), _d(tacho_e)
            if s and e and e > s:
                return round(e - s, 2)
        elif method == 'airswitch':
            s, e = _d(air_s), _d(air_e)
            if s and e and e > s: return round(e - s, 2)
    except Exception:
        pass
    return _D('0')


def _generate_segment_charges(fc, segments, booking, config=None):
    from decimal import Decimal as _D
    from .models import FlightChargeItem as _FCI, ChargeRate
    ac = booking.aircraft
    hire_rate = ChargeRate.objects.filter(
        aircraft=ac, flight_type=booking.flight_type,
        time_method=ac.total_time_method
    ).first()
    total_hours = sum(s.hours for s in segments) or _D('0')

    for seg in segments:
        h = seg.hours
        if not h:
            continue
        if hire_rate:
            _FCI.objects.create(
                flight_completion=fc, segment=seg, item_type='hire',
                description=f'Aircraft hire — {ac.registration} ({seg.member.user.get_full_name()})',
                amount=round(float(hire_rate.amount) * float(h), 2),
            )
        if fc.fuel_surcharge_rate_snapshot and not (hire_rate and hire_rate.includes_fuel):
            _FCI.objects.create(
                flight_completion=fc, segment=seg, item_type='fuel',
                description=f'Fuel levy ({seg.member.user.get_full_name()})',
                amount=round(float(fc.fuel_surcharge_rate_snapshot) * float(h), 2),
            )

    # Instructor fee — applied to non-instructor segments, weighted by hours
    if fc.instructor_rate_snapshot and booking.instructor and total_hours:
        instr_user = booking.instructor
        student_segs = [s for s in segments if s.member.user_id != instr_user.pk]
        student_total = sum(s.hours for s in student_segs) or _D('0')
        if student_total:
            for seg in student_segs:
                amt = round(float(fc.instructor_rate_snapshot) * float(total_hours) * float(seg.hours) / float(student_total), 2)
                if amt > 0:
                    _FCI.objects.create(
                        flight_completion=fc, segment=seg, item_type='instructor',
                        description=f'Instructor fee — {booking.instructor.get_full_name()} ({seg.member.user.get_full_name()})',
                        amount=amt,
                    )

    # Flat surcharges — split proportionally, last segment absorbs rounding
    if total_hours:
        for sc in ac.surcharges.all():
            amounts = [round(float(sc.amount) * float(s.hours) / float(total_hours), 2) for s in segments]
            amounts[-1] = round(float(sc.amount) - sum(amounts[:-1]), 2)
            for seg, amt in zip(segments, amounts):
                if amt > 0:
                    _FCI.objects.create(
                        flight_completion=fc, segment=seg, item_type='surcharge',
                        description=f'{sc.name} ({seg.member.user.get_full_name()})',
                        amount=amt,
                    )


@login_required
def manage_member_detail(request, club_slug, member_id):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    member = get_object_or_404(ClubMember, club=club, id=member_id)

    if request.method == 'POST':
        action = request.POST.get('action', '')
        _saved = action in ('save_membership',)
        if action == 'avatar_upload':
            if request.FILES.get('avatar'):
                member.avatar = request.FILES['avatar']
                member.save(update_fields=['avatar'])
            return _inline_redirect(request, 'core:manage_member_detail', club_slug=club_slug, member_id=member_id)
        elif action == 'save_contact':
            u = member.user
            u.first_name = request.POST.get('first_name', '').strip()
            u.last_name = request.POST.get('last_name', '').strip()
            u.email = request.POST.get('email', '').strip()
            u.save(update_fields=['first_name', 'last_name', 'email'])
            member.phone_mobile = request.POST.get('phone_mobile', '').strip()
            member.phone_home = request.POST.get('phone_home', '').strip()
            member.phone_work = request.POST.get('phone_work', '').strip()
            member.address_line1 = request.POST.get('address_line1', '').strip()
            member.address_line2 = request.POST.get('address_line2', '').strip()
            member.suburb = request.POST.get('suburb', '').strip()
            member.postcode = request.POST.get('postcode', '').strip()
            dob = request.POST.get('date_of_birth', '').strip()
            member.date_of_birth = dob or None
            member.sex = request.POST.get('sex', '').strip()
            member.next_of_kin_name  = request.POST.get('next_of_kin_name', '').strip()
            member.next_of_kin_phone = request.POST.get('next_of_kin_phone', '').strip()
            member.save()
            return _inline_redirect(request, 'core:manage_member_detail',
                                    club_slug=club_slug, member_id=member_id, saved=True)
        elif action == 'save_membership' and actor.is_admin:
            _old_standing = member.standing
            _old_role = member.role
            _old_sub_exp = member.subscription_expires
            _old_cat = member.membership_category

            standing = request.POST.get('standing')
            if standing in dict(ClubMember.STANDING_CHOICES):
                member.standing = standing
                if standing in ('resigned', 'lapsed') and not member.resigned_at:
                    member.resigned_at = timezone.localdate()
            sub_exp = request.POST.get('subscription_expires')
            member.subscription_expires = sub_exp or None
            role_id = request.POST.get('role_id')
            new_role = Role.objects.filter(club=club, id=role_id).first() if role_id else None
            new_has_admin = request.POST.get('has_admin_access') == 'on'
            # Guard: don't remove the last admin
            from django.db.models import Q
            would_be_admin = new_has_admin or (new_role and new_role.effective_is_admin)
            if not would_be_admin:
                other_admins = ClubMember.objects.filter(club=club).exclude(id=member.id).filter(
                    Q(has_admin_access=True) |
                    Q(role__is_superadmin=True) |
                    Q(role__can_access_settings=True)
                )
                if not other_admins.exists():
                    return _inline_redirect(request, 'core:manage_member_detail',
                                            club_slug=club_slug, member_id=member_id,
                                            error='Cannot remove admin access — at least one admin must remain.')
            member.role = new_role
            member.has_admin_access = new_has_admin
            member.save()
            # Audit log
            _sc = dict(ClubMember.STANDING_CHOICES)
            if member.standing != _old_standing:
                MembershipHistoryEntry.objects.create(
                    club_member=member, event_type='standing_change', changed_by=request.user,
                    old_value=_sc.get(_old_standing, _old_standing),
                    new_value=_sc.get(member.standing, member.standing),
                )
                if member.user:
                    if member.standing in ('lapsed', 'resigned', 'suspended'):
                        member.user.is_active = False
                        member.user.save(update_fields=['is_active'])
                    elif member.standing == 'active':
                        member.user.is_active = True
                        member.user.save(update_fields=['is_active'])
            if member.role != _old_role:
                MembershipHistoryEntry.objects.create(
                    club_member=member, event_type='role_change', changed_by=request.user,
                    old_value=_old_role.name if _old_role else '—',
                    new_value=member.role.name if member.role else '—',
                )
            if member.subscription_expires != _old_sub_exp:
                MembershipHistoryEntry.objects.create(
                    club_member=member, event_type='subscription_renewed', changed_by=request.user,
                    old_value=str(_old_sub_exp) if _old_sub_exp else '—',
                    new_value=str(member.subscription_expires) if member.subscription_expires else '—',
                )
        elif action == 'save_notes' and actor.is_admin:
            pass  # notes field not yet on model — placeholder

        elif action in ('add_credential', 'edit_credential') and (actor.is_admin or actor.is_instructor):
            from .models import MemberCredential, CredentialType as _CT
            ct_id = request.POST.get('credential_type', '').strip()
            ct_obj = _CT.objects.filter(id=ct_id).first() if ct_id else None
            name = request.POST.get('cred_name', '').strip()
            cert_num = request.POST.get('certificate_number', '').strip()
            issue_str = request.POST.get('issue_date', '').strip() or None
            expiry_str = request.POST.get('expiry_date', '').strip() or None
            notes = request.POST.get('notes', '').strip()
            ac_type_id = request.POST.get('cred_aircraft_type_id', '').strip()
            ac_type_obj = (AircraftType.objects.filter(club=club, id=ac_type_id).first()
                           if ac_type_id and ct_obj and ct_obj.category == 'type_rating' else None)
            if action == 'add_credential' and ct_obj:
                cred = MemberCredential(
                    member=member.user, credential_type=ct_obj, name=name,
                    aircraft_type=ac_type_obj,
                    certificate_number=cert_num, issue_date=issue_str,
                    expiry_date=expiry_str, notes=notes, created_by=request.user,
                )
                if request.FILES.get('evidence'):
                    cred.evidence = request.FILES['evidence']
                cred.save()
            elif action == 'edit_credential':
                cred_id = request.POST.get('cred_id')
                cred = MemberCredential.objects.filter(member=member.user, id=cred_id).first()
                if cred:
                    if ct_obj:
                        cred.credential_type = ct_obj
                    cred.aircraft_type = ac_type_obj
                    cred.name = name
                    cred.certificate_number = cert_num
                    cred.issue_date = issue_str
                    cred.expiry_date = expiry_str
                    cred.notes = notes
                    if request.FILES.get('evidence'):
                        cred.evidence = request.FILES['evidence']
                    cred.save()

        elif action == 'delete_credential' and (actor.is_admin or actor.is_instructor):
            from .models import MemberCredential
            MemberCredential.objects.filter(member=member.user, id=request.POST.get('cred_id')).delete()

        elif action == 'set_credit_limit' and actor.is_admin:
            from .models import Account as _Account
            acct, _ = _Account.objects.get_or_create(club_member=member, defaults={'balance': 0})
            raw = request.POST.get('credit_limit', '').strip()
            if raw == '' or raw.lower() == 'exempt':
                acct.credit_limit = None
            else:
                try:
                    acct.credit_limit = float(raw)
                except ValueError:
                    pass
            acct.save(update_fields=['credit_limit'])

        elif action == 'account_topup' and actor.is_admin:
            from .models import AccountTransaction
            amount = request.POST.get('amount', '').strip()
            pay_method = request.POST.get('payment_method', 'bank_transfer')
            ref = request.POST.get('reference', '').strip()
            desc = request.POST.get('description', '').strip() or 'Account top-up'
            if amount:
                try:
                    acct, _ = member.account.__class__.objects.get_or_create(club_member=member, defaults={'balance': 0})
                    AccountTransaction.objects.create(
                        account=acct, transaction_type='top_up', direction='credit',
                        amount=amount, description=desc,
                        payment_method=pay_method, reference=ref, created_by=request.user,
                    )
                    acct.apply_transaction(amount, 'credit')
                except Exception:
                    pass

        elif action == 'account_adjustment' and actor.is_admin:
            from .models import AccountTransaction
            amount = request.POST.get('amount', '').strip()
            direction = request.POST.get('direction', 'credit')
            desc = request.POST.get('description', '').strip()
            if amount and desc:
                try:
                    acct = member.account
                    AccountTransaction.objects.create(
                        account=acct, transaction_type='adjustment', direction=direction,
                        amount=amount, description=desc, created_by=request.user,
                    )
                    acct.apply_transaction(amount, direction)
                except Exception:
                    pass

        elif action == 'account_transfer' and actor.is_admin:
            from .models import AccountTransaction
            amount = request.POST.get('amount', '').strip()
            dest_id = request.POST.get('dest_member_id', '').strip()
            desc = request.POST.get('description', '').strip() or 'Member-to-member transfer'
            dest = ClubMember.objects.filter(club=club, id=dest_id).first() if dest_id else None
            if amount and dest and dest != member:
                try:
                    src_acct = member.account
                    dst_acct = dest.account
                    AccountTransaction.objects.create(
                        account=src_acct, transaction_type='adjustment', direction='debit',
                        amount=amount, description=f'Transfer to {dest.user.get_full_name()} — {desc}',
                        created_by=request.user,
                    )
                    src_acct.apply_transaction(amount, 'debit')
                    AccountTransaction.objects.create(
                        account=dst_acct, transaction_type='adjustment', direction='credit',
                        amount=amount, description=f'Transfer from {member.user.get_full_name()} — {desc}',
                        created_by=request.user,
                    )
                    dst_acct.apply_transaction(amount, 'credit')
                except Exception:
                    pass

        elif action == 'add_frequent_passenger':
            from .models import FrequentPassenger
            fp_name  = request.POST.get('fp_name', '').strip()
            fp_phone = request.POST.get('fp_phone', '').strip()
            if fp_name:
                FrequentPassenger.objects.create(club_member=member, name=fp_name, phone=fp_phone)

        elif action == 'delete_frequent_passenger':
            from .models import FrequentPassenger
            FrequentPassenger.objects.filter(club_member=member, id=request.POST.get('fp_id')).delete()

        elif action == 'raise_subscription_invoice' and actor.is_admin:
            from decimal import Decimal, InvalidOperation
            from datetime import date as _date, timedelta as _td
            from django.db.models import F as _F
            from django.urls import reverse as _rev
            from django.contrib import messages as _msgs
            config = get_config(club)
            # Block if an unpaid subscription invoice already exists
            existing = Invoice.objects.filter(
                club=club, member=member,
                description__startswith='Membership subscription',
            ).exclude(status__in=['paid', 'void']).first()
            if existing:
                _msgs.error(request,
                    f'An unpaid subscription invoice already exists ({existing.display_number}). '
                    f'Pay or void it before raising another.')
                return _inline_redirect(request, 'core:manage_member_detail', club_slug=club_slug, member_id=member_id)
            raw_amount = request.POST.get('amount', '').strip()
            raw_expiry = request.POST.get('subscription_expiry_date', '').strip()
            notes = request.POST.get('notes', '').strip()
            try:
                amount = Decimal(raw_amount)
                assert amount > 0
            except (InvalidOperation, AssertionError):
                return _inline_redirect(request, 'core:manage_member_detail', club_slug=club_slug, member_id=member_id)
            today = timezone.localdate()
            try:
                expiry_date = date.fromisoformat(raw_expiry)
            except (ValueError, TypeError):
                expiry_date = None
            due = today + _td(days=config.payment_terms_days)
            ClubConfig = config.__class__
            ClubConfig.objects.filter(pk=config.pk).select_for_update().get()
            config.refresh_from_db()
            inv_number = config.invoice_number_next
            config.invoice_number_next = _F('invoice_number_next') + 1
            config.save(update_fields=['invoice_number_next'])
            fy_label = f'FY{str(expiry_date.year)[2:]}' if expiry_date else ''
            description = 'Membership subscription'
            if fy_label:
                description += f' — {fy_label}'
            if member.role:
                description += f' ({member.role.name})'
            inv = Invoice.objects.create(
                club=club, member=member,
                invoice_number=inv_number,
                issue_date=today, due_date=due,
                description=description,
                notes=notes,
                gst_rate=config.gst_rate,
                amount_paid=0,
                status='sent',
                sent_at=timezone.now(),
                subscription_expiry_date=expiry_date,
                created_by=request.user,
            )
            InvoiceLineItem.objects.create(
                invoice=inv,
                description=description,
                quantity=1, unit='', rate=amount, amount=amount,
            )
            from .services import notification_service as _ns
            _ns.notify_invoice_issued(inv)
            return _inline_redirect(request, 'core:manage_member_detail',
                                    club_slug=club_slug, member_id=member_id, saved=True)

        elif action in ('save_lesson_note', 'delete_lesson_note', 'email_lesson_note') and (actor.is_instructor or actor.is_admin):
            from .models import LessonNote as _LN
            from django.urls import reverse as _rev3
            _detail_base = _rev3('core:manage_member_detail', kwargs={'club_slug': club_slug, 'member_id': member_id})
            if action == 'save_lesson_note':
                _booking_id = request.POST.get('booking_id')
                _booking_obj = get_object_or_404(Booking, id=_booking_id, club=club, member=member)
                _note_id = request.POST.get('note_id')
                if _note_id:
                    _note = get_object_or_404(_LN, id=_note_id)
                else:
                    _note, _ = _LN.objects.get_or_create(booking=_booking_obj)
                if _booking_obj.instructor:
                    _note.author = _booking_obj.instructor
                _note.exercises_covered = request.POST.get('exercises_covered', '').strip()
                _note.debrief_notes     = request.POST.get('debrief_notes', '').strip()
                _note.next_lesson_plan  = request.POST.get('next_lesson_plan', '').strip()
                _note.save()
                return redirect(_detail_base + '?saved=1#training')
            elif action == 'delete_lesson_note':
                _note = get_object_or_404(_LN, id=request.POST.get('note_id'), booking__club=club)
                _note.delete()
                return redirect(_detail_base + '#training')
            elif action == 'email_lesson_note':
                _note = get_object_or_404(_LN, id=request.POST.get('note_id'), booking__club=club)
                from .email_notifications import lesson_note_emailed as _email_note
                _email_note(_note, club)
                return redirect(_detail_base + '?saved=1#training')

        return _inline_redirect(request, 'core:manage_member_detail', club_slug=club_slug, member_id=member_id, saved=_saved)

    from .models import MemberCredential, AccountTransaction, FrequentPassenger as _FP
    from django.db.models import Q as _Q
    _now = timezone.now()
    _PAST_STATUSES = ('completed', 'transferred')
    _ACTIVE_STATUSES = ('pending', 'confirmed', 'departed')
    upcoming_bookings = (Booking.objects
                         .filter(club=club, member=member, status__in=_ACTIVE_STATUSES,
                                 scheduled_start__gte=_now)
                         .select_related('aircraft', 'instructor', 'flight_type', 'flight_completion')
                         .order_by('scheduled_start')[:10])
    past_bookings = (Booking.objects
                     .filter(club=club, member=member)
                     .filter(_Q(status__in=_PAST_STATUSES) | _Q(scheduled_start__lt=_now))
                     .exclude(status='cancelled')
                     .select_related('aircraft', 'instructor', 'flight_type', 'flight_completion')
                     .order_by('-scheduled_start')[:20])
    credentials = (MemberCredential.objects
                   .filter(member=member.user)
                   .select_related('credential_type', 'aircraft_type')
                   .order_by('expiry_date'))
    roles = Role.objects.filter(club=club)

    try:
        account = member.account
        transactions = account.transactions.select_related('flight_completion__booking__aircraft').order_by('-created_at')[:30]
    except Exception:
        account = None
        transactions = []

    all_members = (ClubMember.objects.filter(club=club)
                   .exclude(id=member.id)
                   .select_related('user')
                   .order_by('user__last_name'))

    from .models import CredentialType
    credential_types = CredentialType.objects.filter(region='NZ-CAA').order_by('display_order')
    frequent_passengers = _FP.objects.filter(club_member=member).order_by('name')
    _is_inline = request.GET.get('inline') == '1'
    membership_history = member.membership_history.select_related('changed_by').order_by('-changed_at')
    # FY-end shortcut: compute the next FY end date for the subscription_expires helper
    import calendar as _cal
    _cfg = get_config(club)
    _fy_start = _cfg.fy_start_month
    _today = timezone.localdate()
    _fy_end_month = (_fy_start - 2) % 12 + 1
    _fy_end_year  = _today.year if _today.month < _fy_start else _today.year + 1
    if _fy_end_month >= _fy_start:
        _fy_end_year -= 1
    _fy_end_day = _cal.monthrange(_fy_end_year, _fy_end_month)[1]
    _next_fy_end = date(_fy_end_year, _fy_end_month, _fy_end_day)
    # If today is already past that FY end, move to the following year's FY end
    if _next_fy_end < _today:
        _fy_end_year += 1
        _fy_end_day = _cal.monthrange(_fy_end_year, _fy_end_month)[1]
        _next_fy_end = date(_fy_end_year, _fy_end_month, _fy_end_day)

    _fy_label = f'FY{str(_next_fy_end.year)[2:]}'
    # Pro-rata amount: fraction of FY remaining × annual fee
    _prorata_amount = None
    if member.role and member.role.annual_renewal_fee:
        _fy_start_year = _next_fy_end.year - 1 if _next_fy_end.month < _fy_start else _next_fy_end.year
        _fy_start_date = date(_fy_start_year, _fy_start, 1)
        _fy_total_days = (_next_fy_end - _fy_start_date).days + 1
        _days_remaining = (_next_fy_end - _today).days + 1
        from decimal import Decimal as _D, ROUND_HALF_UP
        _prorata_amount = (member.role.annual_renewal_fee * _D(str(_days_remaining / _fy_total_days))).quantize(_D('0.01'), rounding=ROUND_HALF_UP)
    _existing_sub_inv = Invoice.objects.filter(
        club=club, member=member,
        description__startswith='Membership subscription',
    ).exclude(status__in=['paid', 'void']).first()

    from .models import LessonNote as _LN
    lesson_notes = (_LN.objects.filter(booking__member=member, booking__club=club)
                    .select_related('booking__aircraft', 'booking__instructor', 'author')
                    .order_by('-booking__scheduled_start'))
    dual_bookings = (Booking.objects
                     .filter(member=member, club=club, status='completed', instructor__isnull=False)
                     .select_related('aircraft', 'instructor')
                     .order_by('-scheduled_start')[:50])
    _edit_note_id = request.GET.get('edit_note')
    edit_note = _LN.objects.filter(id=_edit_note_id, booking__club=club).first() if _edit_note_id else None

    return render(request, 'core/manage_member_detail.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'member': member, 'upcoming_bookings': upcoming_bookings, 'past_bookings': past_bookings,
        'credentials': credentials, 'account': account, 'transactions': transactions,
        'all_members': all_members, 'roles': roles,
        'standing_choices': ClubMember.STANDING_CHOICES,
        'credential_types': credential_types,
        'aircraft_type_list': AircraftType.objects.filter(club=club),
        'frequent_passengers': frequent_passengers,
        'membership_history': membership_history,
        'config': _cfg,
        'fy_end': _next_fy_end,
        'fy_label': _fy_label,
        'prorata_amount': _prorata_amount,
        'existing_sub_invoice': _existing_sub_inv,
        'base_template': 'core/base_inline.html' if _is_inline else 'core/base.html',
        'inline_title': member.user.get_full_name(),
        'lesson_notes': lesson_notes,
        'dual_bookings': dual_bookings,
        'edit_note': edit_note,
    })


@login_required
def lesson_note_print(request, club_slug, note_id):
    from .models import LessonNote
    club = get_object_or_404(Club, slug=club_slug)
    note = get_object_or_404(LessonNote, id=note_id, booking__club=club)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err
    config = get_config(club)
    fc = getattr(note.booking, 'flight_completion', None)
    return render(request, 'core/lesson_note_print.html', {
        'club': club, 'note': note, 'config': config, 'fc': fc,
    })


@login_required
def app_training(request, club_slug):
    from .models import LessonNote
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')
    notes = (LessonNote.objects
             .filter(booking__member=actor, booking__club=club)
             .select_related('booking__aircraft', 'booking__flight_completion', 'author')
             .order_by('-booking__scheduled_start'))
    return render(request, 'core/app/training.html', {
        'club': club, 'club_member': actor, 'notes': notes,
    })


def app_credentials(request, club_slug):
    from .models import MemberCredential
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')
    credentials = (MemberCredential.objects
                   .filter(member=actor.user)
                   .select_related('credential_type', 'aircraft_type')
                   .order_by('credential_type__display_order', 'expiry_date'))
    return render(request, 'core/app/credentials.html', {
        'club': club, 'club_member': actor,
        'credentials': credentials,
        'saved': request.GET.get('saved') == '1',
    })


@login_required
def app_flight_log(request, club_slug):
    from django.db.models import Sum, Q as _Q
    from django.core.paginator import Paginator
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')
    base_qs = (FlightCompletion.objects
               .filter(booking__member=actor, booking__club=club, booking__status='completed')
               .select_related('booking__aircraft', 'booking__instructor', 'booking__flight_type')
               .order_by('-booking__scheduled_start'))

    # Year filter
    year_param = request.GET.get('year', '')
    years = (base_qs.dates('booking__scheduled_start', 'year')
             .values_list('booking__scheduled_start__year', flat=True).distinct().order_by('-booking__scheduled_start__year'))
    years = sorted(set(base_qs.values_list('booking__scheduled_start__year', flat=True).distinct()), reverse=True)
    filtered_qs = base_qs.filter(booking__scheduled_start__year=year_param) if year_param else base_qs

    # Totals over the filtered set
    agg = filtered_qs.aggregate(
        total=Sum('actual_flight_hours'),
        dual=Sum('actual_flight_hours', filter=_Q(booking__instructor__isnull=False)),
        solo=Sum('actual_flight_hours', filter=_Q(booking__instructor__isnull=True)),
        if_sim=Sum('time_if_simulated'),
        if_act=Sum('time_if_actual'),
        night=Sum('time_night'),
    )
    total_hrs = agg['total'] or 0
    dual_hrs  = agg['dual']  or 0
    solo_hrs  = agg['solo']  or 0
    if_hrs    = (agg['if_sim'] or 0) + (agg['if_act'] or 0)
    night_hrs = agg['night'] or 0

    # Pagination
    paginator = Paginator(filtered_qs, 30)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'core/app/flight_log.html', {
        'club': club, 'club_member': actor,
        'page_obj': page_obj,
        'total_hrs': total_hrs,
        'dual_hrs': dual_hrs,
        'solo_hrs': solo_hrs,
        'if_hrs': if_hrs,
        'night_hrs': night_hrs,
        'years': years,
        'year_param': year_param,
    })


@login_required
def my_profile(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    if request.method == 'POST':
        action = request.POST.get('action', 'save_contact')
        _profile_url = redirect('core:my_profile', club_slug=club_slug).url
        if action == 'avatar_upload':
            if request.FILES.get('avatar'):
                member.avatar = request.FILES['avatar']
                member.save(update_fields=['avatar'])
            return redirect(_profile_url + '?saved=1')
        elif action == 'save_notifications':
            from .models import NotificationPreference
            pref, _ = NotificationPreference.objects.get_or_create(club_member=member)
            # Per-type alert toggles
            _toggle_fields = [
                'booking_confirmed', 'booking_cancelled', 'booking_reminder',
                'credential_expiring', 'subscription_expiring',
                'instructor_booking_urgent', 'instructor_booking_upcoming',
                'maintenance_alert', 'lapsed_credentials', 'slot_released',
                'payment_reminder', 'invoice_sent',
            ]
            for f in _toggle_fields:
                setattr(pref, f, request.POST.get(f) == 'on')
            # Slot-release filters
            pref.aircraft.set(request.POST.getlist('notify_aircraft'))
            pref.instructors.set(request.POST.getlist('notify_instructors'))
            raw_days = request.POST.get('max_days_ahead', '').strip()
            pref.max_days_ahead = int(raw_days) if raw_days.isdigit() else None
            pref.save()
            return redirect(_profile_url + '?saved=1#notifications')
        elif action == 'unwatch_slot':
            booking_id = request.POST.get('booking_id')
            if booking_id:
                SlotWatch.objects.filter(club_member=member, booking_id=booking_id).delete()
            return redirect(_profile_url + '?saved=1#notifications')
        elif action == 'delete_notifications':
            from .models import NotificationPreference
            NotificationPreference.objects.filter(club_member=member).delete()
            from django.contrib import messages as _msg
            _msg.success(request, 'Notification preferences reset.')
            return redirect(_profile_url + '#notifications')
        else:
            u = request.user
            u.first_name = request.POST.get('first_name', '').strip()
            u.last_name = request.POST.get('last_name', '').strip()
            u.email = request.POST.get('email', '').strip()
            u.save(update_fields=['first_name', 'last_name', 'email'])
            member.phone_mobile = request.POST.get('phone_mobile', '').strip()
            member.phone_home = request.POST.get('phone_home', '').strip()
            member.address_line1 = request.POST.get('address_line1', '').strip()
            member.suburb = request.POST.get('suburb', '').strip()
            member.postcode = request.POST.get('postcode', '').strip()
            dob = request.POST.get('date_of_birth', '').strip()
            member.date_of_birth = dob or None
            member.sex = request.POST.get('sex', '').strip()
            member.save()
            return redirect(_profile_url + '?saved=1')

    from django.db.models import Q as _Q
    _now = timezone.now()
    _PAST_STATUSES = ('completed', 'transferred')
    _ACTIVE_STATUSES = ('pending', 'confirmed', 'departed')
    upcoming = (Booking.objects
                .filter(club=club, member=member, status__in=_ACTIVE_STATUSES,
                        scheduled_start__gte=_now)
                .select_related('aircraft', 'instructor', 'flight_type', 'flight_completion')
                .order_by('scheduled_start')[:10])
    past = (Booking.objects
            .filter(club=club, member=member)
            .filter(_Q(status__in=_PAST_STATUSES) | _Q(scheduled_start__lt=_now))
            .exclude(status='cancelled')
            .select_related('aircraft', 'instructor', 'flight_type')
            .order_by('-scheduled_start')[:20])

    try:
        account = member.account
    except Exception:
        account = None

    from .models import NotificationPreference
    try:
        notification_pref = member.notification_prefs
    except NotificationPreference.DoesNotExist:
        notification_pref = None

    # Bookings where this user is the instructor (next 7 days)
    upcoming_as_instructor = None
    if member.is_instructor:
        upcoming_as_instructor = (Booking.objects
            .filter(club=club, instructor=request.user,
                    status__in=('pending', 'confirmed', 'departed'),
                    scheduled_start__gte=_now,
                    scheduled_start__lt=_now + timedelta(days=7))
            .select_related('aircraft', 'member__user', 'flight_type')
            .order_by('scheduled_start'))

    club_aircraft = Aircraft.objects.filter(club=club, status='online').order_by('registration')
    club_instructors = ClubMember.objects.filter(
        club=club, is_on_instructor_roster=True
    ).select_related('user').order_by('user__last_name')

    _raw_toggles = [
        ('booking_confirmed',           'Booking confirmed',                             True),
        ('booking_cancelled',           'Booking cancelled',                             True),
        ('booking_reminder',            'Booking reminder (day before)',                 True),
        ('credential_expiring',         'Credential expiring (medical, BFR, etc.)',      True),
        ('subscription_expiring',       'Subscription expiring',                         True),
        ('instructor_booking_urgent',   'New booking assigned — within 2 days (urgent)', True),
        ('instructor_booking_upcoming', 'New booking assigned — within 10 days',         True),
        ('maintenance_alert',           'Maintenance alert (amber/red items)',            True),
        ('lapsed_credentials',          'Member has lapsed credentials — flight today',   True),
        ('slot_released',               'Slot released by another member (opt-in)',       False),
        ('payment_reminder',            'Account balance reminder',                        True),
        ('invoice_sent',                'Invoice issued to me',                            True),
    ]
    _notif_toggles = [
        {'field': f, 'label': l,
         'enabled': getattr(notification_pref, f, default) if notification_pref else default}
        for f, l, default in _raw_toggles
    ]
    paid_flights = (FlightCompletion.objects
                    .filter(booking__member=member, booking__club=club)
                    .exclude(amount_paid=0)
                    .exclude(amount_paid__isnull=True)
                    .select_related('booking__aircraft', 'booking__flight_type')
                    .order_by('-booking__scheduled_start')[:50])

    my_actions = (OccurrenceAction.objects
                  .filter(assigned_to=member, status=OccurrenceAction.STATUS_OPEN)
                  .select_related('report__occurrence_type', 'report__club')
                  .order_by('due_date', 'created_at'))

    watched_slots = (SlotWatch.objects
                     .filter(club_member=member,
                             booking__status__in=('pending', 'confirmed'),
                             booking__scheduled_start__gte=timezone.now())
                     .select_related('booking__aircraft', 'booking__member__user',
                                     'booking__flight_type')
                     .order_by('booking__scheduled_start'))

    return render(request, 'core/my_profile.html', {
        'club': club, 'club_member': member, 'is_instructor': member.is_instructor,
        'member': member, 'upcoming': upcoming, 'past': past, 'account': account,
        'notification_pref': notification_pref,
        'notification_toggle_fields': _notif_toggles,
        'club_aircraft': club_aircraft,
        'club_instructors': club_instructors,
        'upcoming_as_instructor': upcoming_as_instructor,
        'paid_flights': paid_flights,
        'my_actions': my_actions,
        'watched_slots': watched_slots,
    })


def _recompute_conflicts_for_club(club):
    """After a blockout is created/edited/deleted, refresh conflict flags on all active bookings."""
    from .models import recompute_blockout_conflict as _rbc
    qs = (Booking.objects
          .filter(club=club, status__in=['pending', 'confirmed', 'departed'])
          .select_related('aircraft', 'instructor', 'flight_completion'))
    for b in qs:
        _rbc(b)


def _create_blockout_from_post(request, club, scope, aircraft=None, instructor_user=None):
    """Create a BlockOut from POST data. scope, aircraft/instructor_user pre-determined by caller."""
    from .models import BlockOut, BlockOutType
    bot_id = request.POST.get('bot_id')
    blockout_type = BlockOutType.objects.filter(club=club, id=bot_id).first() if bot_id else None
    label = request.POST.get('label', '').strip()
    recurrence = request.POST.get('recurrence', 'one_off')
    all_day = request.POST.get('all_day') in ('on', '1', 'true')
    date_str = request.POST.get('date', '')
    weekday_str = request.POST.get('weekday', '0')

    bo = BlockOut(club=club, blockout_type=blockout_type, label=label,
                  scope=scope, recurrence=recurrence, all_day=all_day,
                  created_by=request.user)
    if recurrence == 'one_off':
        from datetime import date as _date
        try:
            bo.date = _date.fromisoformat(date_str) if date_str else None
        except ValueError:
            bo.date = None
    elif recurrence == 'weekly':
        bo.weekday = int(weekday_str) if weekday_str.isdigit() else 0

    if not all_day:
        bo.start_time = request.POST.get('start_time') or None
        bo.end_time = request.POST.get('end_time') or None

    bo.save()

    if scope == 'aircraft' and aircraft:
        bo.aircraft.set([aircraft])
    elif scope == 'instructors' and instructor_user:
        bo.instructors.set([instructor_user])


@login_required
def manage_blockouts(request, club_slug):
    """All-resource block-outs (scope='all') and overview."""
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    from .models import BlockOut, BlockOutType
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'add_blockout':
            _create_blockout_from_post(request, club, scope='all')
            _recompute_conflicts_for_club(club)
        elif action == 'delete_blockout':
            bo_id = request.POST.get('bo_id')
            BlockOut.objects.filter(club=club, id=bo_id).delete()
            _recompute_conflicts_for_club(club)
        elif action == 'edit_blockout':
            bo_id = request.POST.get('bo_id')
            bo = BlockOut.objects.filter(club=club, id=bo_id, scope='all').first()
            if bo:
                bot_id = request.POST.get('bot_id')
                bo.blockout_type = BlockOutType.objects.filter(club=club, id=bot_id).first() if bot_id else None
                bo.label = request.POST.get('label', '').strip()
                bo.recurrence = request.POST.get('recurrence', 'one_off')
                bo.all_day = request.POST.get('all_day') in ('on', '1', 'true')
                from datetime import date as _date
                if bo.recurrence == 'one_off':
                    date_str = request.POST.get('date', '')
                    try:
                        bo.date = _date.fromisoformat(date_str) if date_str else None
                    except ValueError:
                        bo.date = None
                    bo.weekday = None
                elif bo.recurrence == 'weekly':
                    weekday_str = request.POST.get('weekday', '0')
                    bo.weekday = int(weekday_str) if weekday_str.isdigit() else 0
                    bo.date = None
                else:
                    bo.date = None
                    bo.weekday = None
                if not bo.all_day:
                    bo.start_time = request.POST.get('start_time') or None
                    bo.end_time = request.POST.get('end_time') or None
                else:
                    bo.start_time = None
                    bo.end_time = None
                bo.save()
                _recompute_conflicts_for_club(club)
        return redirect('core:manage_blockouts', club_slug=club_slug)

    from django.db.models import Q
    _today = timezone.localdate()
    _active_q = (
        Q(recurrence='one_off', date__gte=_today) |
        Q(recurrence__in=['weekly', 'daily'], active_until__isnull=True) |
        Q(recurrence__in=['weekly', 'daily'], active_until__gte=_today)
    )
    all_blockouts = (BlockOut.objects
                     .filter(club=club, scope='all')
                     .filter(_active_q)
                     .select_related('blockout_type')
                     .order_by('recurrence', 'date', 'weekday', 'start_time'))
    past_count = BlockOut.objects.filter(club=club, scope='all').exclude(_active_q).count()
    blockout_types = BlockOutType.objects.filter(club=club, target='all')
    return render(request, 'core/manage_blockouts.html', {
        'club': club,
        'club_member': actor,
        'is_instructor': actor.is_instructor,
        'blockouts': all_blockouts,
        'past_count': past_count,
        'blockout_types': blockout_types,
    })


@login_required
def manage_members(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    modal_error = modal_error_id = None
    modal_error_values = {}
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'add_member' and actor.is_admin:
            from django.contrib.auth import get_user_model as _get_user
            from django.contrib import messages as _messages
            _User = _get_user()
            send_invite = request.POST.get('send_invite') == 'on'
            first = request.POST.get('first_name', '').strip()
            last = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip().lower()
            password = request.POST.get('password', '').strip()
            if not email:
                modal_error = 'Email address is required.'
            elif not send_invite and not (first and last):
                modal_error = 'First name and last name are required.'
            elif _User.objects.filter(email=email).exists():
                modal_error = 'An account with that email address already exists.'
            elif send_invite:
                from .models import ClubInvite
                from .email_notifications import club_invite as _email_invite
                from datetime import timedelta as _td
                ClubInvite.objects.filter(club=club, email=email, accepted_at__isnull=True).delete()
                invite = ClubInvite.objects.create(
                    club=club, email=email,
                    invited_by=request.user,
                    expires_at=timezone.now() + _td(days=7),
                )
                _email_invite(invite)
                logger.info('invite_sent club=%s email=%s by=%s', club.slug, email, request.user.email)
                _messages.success(request, f'Invite sent to {email}.')
                return redirect('core:manage_members', club_slug=club_slug)
            else:
                import secrets as _secrets
                auto_generated = not password
                pw = password or _secrets.token_urlsafe(12)
                user = _User.objects.create_user(
                    username=email, email=email,
                    first_name=first, last_name=last, password=pw
                )
                new_cm = ClubMember.objects.create(club=club, user=user)
                MembershipHistoryEntry.objects.create(
                    club_member=new_cm, event_type='joined',
                    changed_by=request.user, new_value=email,
                )
                if auto_generated:
                    _messages.info(request,
                        f"Member created. Auto-generated password: {pw} — note this now and give it to {first}. It won't be shown again.")
                return redirect('core:manage_members', club_slug=club_slug)
            modal_error_id = 'add-member-modal'
            modal_error_values = {'first_name': first, 'last_name': last, 'email': email}
        else:
            cm_id = request.POST.get('cm_id')
            cm = ClubMember.objects.filter(club=club, id=cm_id).first() if cm_id else None
            if cm:
                if action == 'set_standing':
                    standing = request.POST.get('standing')
                    if standing in dict(ClubMember.STANDING_CHOICES):
                        old_standing = cm.standing
                        cm.standing = standing
                        if standing in ('resigned', 'lapsed') and not cm.resigned_at:
                            cm.resigned_at = timezone.localdate()
                        cm.save(update_fields=['standing', 'resigned_at'])
                        MembershipHistoryEntry.objects.create(
                            club_member=cm, event_type='standing_change',
                            changed_by=request.user,
                            old_value=dict(ClubMember.STANDING_CHOICES).get(old_standing, old_standing),
                            new_value=dict(ClubMember.STANDING_CHOICES).get(standing, standing),
                        )
                elif action == 'set_role':
                    role_id = request.POST.get('role_id')
                    new_role = Role.objects.filter(club=club, id=role_id).first() if role_id else None
                    # Guard: must always have at least one member on the admin system role
                    _admin_role = Role.objects.filter(club=club, system_role_type='admin').first()
                    if _admin_role and cm.role == _admin_role and new_role != _admin_role:
                        _remaining_admins = ClubMember.objects.filter(
                            club=club, role=_admin_role).exclude(id=cm.id).count()
                        if _remaining_admins == 0:
                            from django.contrib import messages as _msg
                            _msg.error(request, 'Cannot remove the last Administrator. Assign another member to Administrator first.')
                            return redirect('core:manage_members', club_slug=club_slug)
                    old_role = cm.role
                    cm.role = new_role
                    cm.save(update_fields=['role'])
                    MembershipHistoryEntry.objects.create(
                        club_member=cm, event_type='role_change',
                        changed_by=request.user,
                        old_value=old_role.name if old_role else '—',
                        new_value=new_role.name if new_role else '—',
                    )
            return redirect('core:manage_members', club_slug=club_slug)

    from django.core.paginator import Paginator as _Pag
    from urllib.parse import urlencode as _ue
    q           = request.GET.get('q', '').strip()
    f_standing  = request.GET.get('standing', '')
    f_role      = request.GET.get('role', '')
    debt_filter = request.GET.get('debt', '')
    sort        = request.GET.get('sort', 'name')
    sort_dir    = request.GET.get('dir', 'asc')

    _SORT_MAP = {
        'name':    ('user__last_name', 'user__first_name'),
        'standing': ('standing', 'user__last_name'),
        'role':    ('role__name', 'user__last_name'),
        'expires': ('subscription_expires', 'user__last_name'),
        'joined':  ('join_date', 'user__last_name'),
    }
    if sort not in _SORT_MAP:
        sort = 'name'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'asc'
    _order = _SORT_MAP[sort]
    if sort_dir == 'desc':
        _order = tuple('-' + f for f in _order)

    members = (ClubMember.objects
               .filter(club=club)
               .select_related('user', 'role', 'membership_category')
               .order_by(*_order))
    if q:
        from django.db.models import Q as _Q
        members = members.filter(
            _Q(user__first_name__icontains=q) |
            _Q(user__last_name__icontains=q) |
            _Q(user__email__icontains=q)
        )
    if f_standing:
        members = members.filter(standing=f_standing)
    if f_role == '__none__':
        members = members.filter(role__isnull=True)
    elif f_role:
        members = members.filter(role_id=f_role)
    f_exp_from = request.GET.get('exp_from', '').strip()
    f_exp_to   = request.GET.get('exp_to', '').strip()
    if debt_filter == '1':
        members = members.filter(account__balance__lt=0)
    if f_exp_from:
        try: members = members.filter(subscription_expires__gte=f_exp_from)
        except Exception: pass
    if f_exp_to:
        try: members = members.filter(subscription_expires__lte=f_exp_to)
        except Exception: pass

    _today        = timezone.localdate()
    _config       = get_config(club)
    _grace_days   = getattr(_config, 'lapse_grace_days', 60)
    _lapse_cutoff = _today - timedelta(days=_grace_days)
    _warn_cutoff  = _today + timedelta(days=30)

    _paginator   = _Pag(members, 25)
    members_page = _paginator.get_page(request.GET.get('page'))

    # base_qs preserves filters but not sort/page — template appends sort params per column
    _base_qs   = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir') and v})
    _filter_qs = _ue({k: v for k, v in request.GET.items() if k != 'page' and v})

    roles = Role.objects.filter(club=club).order_by('name')

    from .models import ClubInvite as _CI
    pending_invites = _CI.objects.filter(
        club=club, accepted_at__isnull=True, expires_at__gt=timezone.now()
    ).select_related('role', 'invited_by')

    return render(request, 'core/manage_members.html', {
        'club': club,
        'club_member': actor,
        'is_instructor': actor.is_instructor,
        'members_page': members_page,
        'total_count': _paginator.count,
        'roles': roles,
        'standing_choices': ClubMember.STANDING_CHOICES,
        'q': q,
        'f_standing': f_standing,
        'f_role': f_role,
        'f_exp_from': f_exp_from,
        'f_exp_to': f_exp_to,
        'sort': sort,
        'sort_dir': sort_dir,
        'base_qs': _base_qs,
        'modal_error': modal_error,
        'modal_error_id': modal_error_id,
        'modal_error_values': modal_error_values,
        'today_iso': timezone.localdate().isoformat(),
        'today': _today,
        'lapse_cutoff': _lapse_cutoff,
        'warn_cutoff': _warn_cutoff,
        'filter_qs': _filter_qs,
        'pending_invites': pending_invites,
    })


@login_required
def send_invite(request, club_slug):
    """Admin sends a single-use invite link to a prospective member's email address."""
    if request.method != 'POST':
        return redirect('core:manage_members', club_slug=club_slug)
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    from .models import ClubInvite, Role as _Role
    from django.contrib import messages as _msgs
    from datetime import timedelta

    email = request.POST.get('email', '').strip().lower()
    role_id = request.POST.get('role_id', '').strip()

    if not email:
        _msgs.error(request, 'Email address is required.')
        return redirect('core:manage_members', club_slug=club_slug)

    role = None
    if role_id:
        role = _Role.objects.filter(club=club, id=role_id).first()

    # Cancel any existing pending invite for this email so the new one supersedes it
    ClubInvite.objects.filter(club=club, email=email, accepted_at__isnull=True).delete()

    invite = ClubInvite.objects.create(
        club=club, email=email, role=role,
        invited_by=request.user,
        expires_at=timezone.now() + timedelta(days=7),
    )

    from .email_notifications import club_invite as _email_invite
    _email_invite(invite)

    _msgs.success(request, f'Invite sent to {email}.')
    return redirect('core:manage_members', club_slug=club_slug)


def accept_invite(request, token):
    """Public view — no login required. Recipient accepts a club invite."""
    from .models import ClubInvite
    invite = get_object_or_404(ClubInvite, token=token)
    club = invite.club

    def _render(state, error=None, form_values=None):
        return render(request, 'core/invite_accept.html', {
            'invite': invite, 'club': club,
            'state': state, 'error': error,
            'form_values': form_values or {},
        })

    if invite.is_accepted:
        return _render('already_accepted')
    if invite.is_expired:
        return _render('expired')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'join' and request.user.is_authenticated:
            cm, created = ClubMember.objects.get_or_create(
                club=club, user=request.user,
                defaults={'standing': 'pending'},
            )
            if not created and cm.standing == 'resigned':
                cm.standing = 'pending'
                cm.save(update_fields=['standing'])
            if invite.role and not cm.role:
                cm.role = invite.role
                cm.save(update_fields=['role'])
            invite.accepted_at = timezone.now()
            invite.club_member = cm
            invite.save(update_fields=['accepted_at', 'club_member'])
            MembershipHistoryEntry.objects.create(
                club_member=cm, event_type='joined',
                changed_by=request.user, new_value='via invite',
            )
            return redirect('core:app_home', club_slug=club.slug)

        elif action == 'create':
            from django.contrib.auth import get_user_model as _gum, login as _login
            _User = _gum()
            first = request.POST.get('first_name', '').strip()
            last  = request.POST.get('last_name', '').strip()
            pw    = request.POST.get('password', '').strip()
            pw2   = request.POST.get('password2', '').strip()
            fv    = {'first_name': first, 'last_name': last}

            if not (first and last and pw):
                return _render('create', 'All fields are required.', fv)
            if pw != pw2:
                return _render('create', 'Passwords do not match.', fv)
            if len(pw) < 8:
                return _render('create', 'Password must be at least 8 characters.', fv)
            if _User.objects.filter(email=invite.email).exists():
                return _render('create',
                    'An account with this email already exists. Use the login link below.', fv)

            user = _User.objects.create_user(
                username=invite.email, email=invite.email,
                first_name=first, last_name=last, password=pw,
            )
            cm = ClubMember.objects.create(
                club=club, user=user, standing='pending', role=invite.role,
            )
            invite.accepted_at = timezone.now()
            invite.club_member = cm
            invite.save(update_fields=['accepted_at', 'club_member'])
            MembershipHistoryEntry.objects.create(
                club_member=cm, event_type='joined',
                changed_by=user, new_value='via invite',
            )
            _login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect('core:app_home', club_slug=club.slug)

    # GET — determine which state to show
    if request.user.is_authenticated:
        try:
            existing_cm = ClubMember.objects.get(club=club, user=request.user)
            if existing_cm.standing not in ('resigned',):
                return _render('already_member')
        except ClubMember.DoesNotExist:
            pass
        return _render('join')
    else:
        return _render('create')


@login_required
def registrar_export(request, club_slug):
    """Point-in-time member register — returns HTML table or CSV download."""
    import csv as _csv
    from io import StringIO as _StringIO
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    as_of_str = request.GET.get('as_of', '').strip()
    fmt = request.GET.get('fmt', 'html')
    if fmt == 'csv':
        logger.info('export club=%s type=registrar as_of=%s by=%s', club.slug, as_of_str or 'today', request.user.email)

    from django.db.models import Q as _Q

    as_of = None
    if as_of_str:
        try:
            as_of = date.fromisoformat(as_of_str)
        except ValueError:
            pass
    if as_of is None:
        as_of = timezone.localdate()

    # Members whose membership was active at as_of date:
    # joined on or before as_of AND (no resigned_at OR resigned_at > as_of)
    qs = (ClubMember.objects
          .filter(club=club, join_date__lte=as_of)
          .filter(_Q(resigned_at__isnull=True) | _Q(resigned_at__gt=as_of))
          .select_related('role', 'user')
          .order_by('user__last_name', 'user__first_name'))

    fields = ['name', 'email', 'role', 'standing', 'join_date', 'resigned_at',
              'date_of_birth', 'address', 'phone_mobile']

    rows = []
    for m in qs:
        addr = ', '.join(filter(None, [m.address_line1, m.address_line2]))
        rows.append({
            'name': m.user.get_full_name(),
            'email': m.user.email,
            'role': m.role.name if m.role else '—',
            'standing': m.standing or '—',
            'join_date': str(m.join_date) if m.join_date else '—',
            'resigned_at': str(m.resigned_at) if m.resigned_at else '—',
            'date_of_birth': str(m.date_of_birth) if m.date_of_birth else '—',
            'address': addr or '—',
            'phone_mobile': m.phone_mobile or '—',
        })

    if fmt == 'csv':
        buf = _StringIO()
        w = _csv.DictWriter(buf, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
        resp = HttpResponse(buf.getvalue(), content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="members_{as_of}.csv"'
        return resp

    return JsonResponse({'as_of': str(as_of), 'count': len(rows), 'rows': rows})


@login_required
def manage_aircraft(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    from .models import AircraftStatus, BlockOut, BlockOutType
    retire_error = ''
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'set_status' and actor.is_admin:
            ac_id = request.POST.get('ac_id')
            status = request.POST.get('status')
            force = request.POST.get('force') == '1'
            ac = Aircraft.objects.filter(club=club, id=ac_id).first()
            if ac and status in [s.value for s in AircraftStatus]:
                if status == 'retired' and not force:
                    future_count = Booking.objects.filter(
                        club=club, aircraft=ac,
                        scheduled_start__gt=timezone.now(),
                    ).exclude(status='cancelled').count()
                    if future_count:
                        retire_error = (
                            f"{ac.registration} has "
                            f"{future_count} future booking{'s' if future_count != 1 else ''} — "
                            "they will appear on the Exceptions screen for reassignment."
                        )
                        # Store pending retire context for the confirm button
                        retire_error = {'msg': retire_error, 'ac_id': ac_id, 'count': future_count}
                        status = None  # don't save yet
                if status:
                    ac.status = status
                    ac.save(update_fields=['status'])
        elif action == 'add_aircraft' and actor.is_admin:
            reg = request.POST.get('registration', '').strip().upper()
            ac_type_id = request.POST.get('aircraft_type_id', '').strip()
            ac_type_obj = AircraftType.objects.filter(club=club, id=ac_type_id).first() if ac_type_id else None
            if reg:
                Aircraft.objects.get_or_create(
                    club=club, registration=reg,
                    defaults={'aircraft_type': ac_type_obj}
                )
        elif action == 'add_aircraft_blockout':
            ac_id = request.POST.get('ac_id')
            ac = Aircraft.objects.filter(club=club, id=ac_id).first()
            if ac:
                _create_blockout_from_post(request, club, scope='aircraft', aircraft=ac)
        elif action == 'delete_blockout':
            bo_id = request.POST.get('bo_id')
            BlockOut.objects.filter(club=club, id=bo_id).delete()
        if not retire_error:
            return redirect('core:manage_aircraft', club_slug=club_slug)

    online_aircraft  = Aircraft.objects.filter(club=club, status='online').order_by('registration')
    retired_aircraft = Aircraft.objects.filter(club=club, status='retired').order_by('registration')
    aircraft_list = list(online_aircraft) + list(retired_aircraft)
    # Attach aircraft-scoped block-outs to each aircraft (active/upcoming only)
    from .models import BlockOut
    from django.db.models import Q
    _today = timezone.localdate()
    _active_q = (
        Q(recurrence='one_off', date__gte=_today) |
        Q(recurrence__in=['weekly', 'daily'], active_until__isnull=True) |
        Q(recurrence__in=['weekly', 'daily'], active_until__gte=_today)
    )
    ac_blockouts_qs = (BlockOut.objects
                       .filter(club=club, scope='aircraft')
                       .filter(_active_q)
                       .prefetch_related('aircraft', 'blockout_type')
                       .order_by('recurrence', 'date', 'weekday', 'start_time'))
    bo_by_ac = {}
    for bo in ac_blockouts_qs:
        for a in bo.aircraft.all():
            bo_by_ac.setdefault(a.id, []).append(bo)

    past_ac_bos = (BlockOut.objects
                   .filter(club=club, scope='aircraft')
                   .exclude(_active_q)
                   .prefetch_related('aircraft'))
    past_bo_count_by_ac = {}
    for bo in past_ac_bos:
        for a in bo.aircraft.all():
            past_bo_count_by_ac[a.id] = past_bo_count_by_ac.get(a.id, 0) + 1

    for ac in aircraft_list:
        ac.bo_list = bo_by_ac.get(ac.id, [])
        ac.past_bo_count = past_bo_count_by_ac.get(ac.id, 0)

    aircraft_blockout_types = BlockOutType.objects.filter(club=club, target='aircraft')
    return render(request, 'core/manage_aircraft.html', {
        'club': club,
        'club_member': actor,
        'is_instructor': actor.is_instructor,
        'online_aircraft': online_aircraft,
        'retired_aircraft': retired_aircraft,
        'aircraft_list': aircraft_list,
        'aircraft_type_list': AircraftType.objects.filter(club=club),
        'status_choices': AircraftStatus.choices,
        'aircraft_blockout_types': aircraft_blockout_types,
        'retire_error': retire_error,
    })


@login_required
def manage_aircraft_detail(request, club_slug, aircraft_id):
    from .models import ChargeRate, FuelSurchargeRate, AircraftSurchargeType, AircraftMaintenanceItem, BlockOutType
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    ac = get_object_or_404(Aircraft, club=club, id=aircraft_id)

    if request.method == 'POST':
        action = request.POST.get('action', '')

        _saved = action in ('save_details', 'save_instruments')
        if action == 'save_details' and actor.is_admin:
            ac_type_id = request.POST.get('aircraft_type_id', '').strip()
            ac_type_obj = AircraftType.objects.filter(club=club, id=ac_type_id).first() if ac_type_id else None
            if ac_type_obj:
                ac.aircraft_type = ac_type_obj
            ac.serial_number = request.POST.get('serial_number', '').strip()
            seats = request.POST.get('seats', '')
            if seats.isdigit():
                ac.seats = int(seats)
            engines = request.POST.get('engine_count', '')
            if engines.isdigit():
                ac.engine_count = int(engines)
            ac.is_leased = request.POST.get('is_leased') == 'on'
            ac.is_available_for_hire = request.POST.get('is_available_for_hire') == 'on'
            ac.save()

        elif action == 'save_instruments' and actor.is_admin:
            ac.records_hobbs = request.POST.get('records_hobbs') == 'on'
            ac.records_tacho = request.POST.get('records_tacho') == 'on'
            ac.records_airswitch = request.POST.get('records_airswitch') == 'on'
            ttm = request.POST.get('total_time_method', '')
            if ttm in dict(Aircraft.TOTAL_TIME_METHOD_CHOICES):
                ac.total_time_method = ttm
            fuel_cph = request.POST.get('fuel_consumption_per_hour', '').strip()
            try:
                ac.fuel_consumption_per_hour = fuel_cph
            except Exception:
                pass
            hobbs_init     = request.POST.get('hobbs_initial', '').strip()
            tacho_init     = request.POST.get('tacho_initial', '').strip()
            airswitch_init = request.POST.get('airswitch_initial', '').strip()
            ac.hobbs_initial     = hobbs_init or None
            ac.tacho_initial     = tacho_init or None
            ac.airswitch_initial = airswitch_init or None
            mts = request.POST.get('maint_time_source', '')
            if mts in dict(Aircraft.MAINT_SOURCE_CHOICES):
                ac.maint_time_source = mts
            frac = request.POST.get('maint_time_fraction', '').strip()
            try:
                f = float(frac)
                if 0.5 <= f <= 1.0:
                    ac.maint_time_fraction = round(f, 2)
            except (ValueError, TypeError):
                pass
            mhi = request.POST.get('maint_hours_initial', '').strip()
            ac.maint_hours_initial = mhi or None
            ac.save()

        elif action == 'save_hire_rate' and actor.is_admin:
            ft_id = request.POST.get('ft_id')
            time_method = request.POST.get('time_method', 'hobbs')
            amount = request.POST.get('amount', '').strip()
            includes_fuel = request.POST.get('includes_fuel') == 'on'
            ft = FlightType.objects.filter(club=club, id=ft_id).first()
            if ft and amount:
                ChargeRate.objects.update_or_create(
                    aircraft=ac, flight_type=ft, time_method=time_method,
                    defaults={'club': club, 'amount': amount, 'includes_fuel': includes_fuel}
                )

        elif action == 'edit_hire_rate' and actor.is_admin:
            rate_id = request.POST.get('rate_id')
            rate = ChargeRate.objects.filter(club=club, aircraft=ac, id=rate_id).first()
            if rate:
                amount = request.POST.get('amount', '').strip()
                if amount:
                    rate.amount = amount
                rate.includes_fuel = request.POST.get('includes_fuel') == 'on'
                rate.save(update_fields=['amount', 'includes_fuel'])

        elif action == 'delete_hire_rate' and actor.is_admin:
            ChargeRate.objects.filter(club=club, aircraft=ac, id=request.POST.get('rate_id')).delete()

        elif action == 'add_fuel_rate' and actor.is_admin:
            rate = request.POST.get('fuel_rate', '').strip()
            effective_from = request.POST.get('effective_from', '').strip()
            notes = request.POST.get('notes', '').strip()
            if rate and effective_from:
                FuelSurchargeRate.objects.create(
                    club=club, aircraft=ac, rate=rate,
                    effective_from=effective_from, notes=notes
                )

        elif action == 'toggle_fuel_rate' and actor.is_admin:
            r = FuelSurchargeRate.objects.filter(club=club, aircraft=ac, id=request.POST.get('rate_id')).first()
            if r:
                r.is_active = not r.is_active
                r.save(update_fields=['is_active'])

        elif action == 'delete_fuel_rate' and actor.is_admin:
            FuelSurchargeRate.objects.filter(club=club, aircraft=ac, id=request.POST.get('rate_id')).delete()

        elif action == 'toggle_surcharge' and actor.is_admin:
            sc_id = request.POST.get('sc_id')
            sc = AircraftSurchargeType.objects.filter(club=club, id=sc_id).first()
            if sc:
                if ac.surcharges.filter(id=sc.id).exists():
                    ac.surcharges.remove(sc)
                else:
                    ac.surcharges.add(sc)

        elif action == 'save_surcharges' and actor.is_admin:
            selected_ids = [int(i) for i in request.POST.getlist('surcharge_ids') if i.isdigit()]
            ac.surcharges.set(AircraftSurchargeType.objects.filter(club=club, id__in=selected_ids))

        elif action == 'add_maintenance' and actor.is_admin:
            name = request.POST.get('maint_name', '').strip()
            if name:
                AircraftMaintenanceItem.objects.create(
                    aircraft=ac, name=name,
                    description=request.POST.get('maint_desc', '').strip(),
                    due_date=request.POST.get('due_date') or None,
                    due_hours=request.POST.get('due_hours') or None,
                    last_completed_date=request.POST.get('last_completed_date') or None,
                    last_completed_hours=request.POST.get('last_completed_hours') or None,
                    interval_days=request.POST.get('interval_days') or None,
                    interval_hours=request.POST.get('interval_hours') or None,
                    warn_days=request.POST.get('warn_days') or None,
                    alert_days=request.POST.get('alert_days') or None,
                    warn_hours=request.POST.get('warn_hours') or None,
                    alert_hours=request.POST.get('alert_hours') or None,
                )

        elif action == 'edit_maintenance' and actor.is_admin:
            maint_id = request.POST.get('maint_id')
            m = AircraftMaintenanceItem.objects.filter(aircraft=ac, id=maint_id).first()
            if m:
                name = request.POST.get('maint_name', '').strip()
                if name:
                    m.name = name
                m.description = request.POST.get('maint_desc', '').strip()
                m.due_date = request.POST.get('due_date') or None
                m.due_hours = request.POST.get('due_hours') or None
                m.last_completed_date = request.POST.get('last_completed_date') or None
                m.last_completed_hours = request.POST.get('last_completed_hours') or None
                m.interval_days = request.POST.get('interval_days') or None
                m.interval_hours = request.POST.get('interval_hours') or None
                m.warn_days = request.POST.get('warn_days') or None
                m.alert_days = request.POST.get('alert_days') or None
                m.warn_hours = request.POST.get('warn_hours') or None
                m.alert_hours = request.POST.get('alert_hours') or None
                m.save()

        elif action == 'mark_maintenance_done' and actor.is_admin:
            maint_id = request.POST.get('maint_id')
            m = AircraftMaintenanceItem.objects.filter(aircraft=ac, id=maint_id).first()
            if m:
                _today = timezone.localdate()
                m.last_completed_date = _today
                # Advance due_date by interval_days if configured
                if m.interval_days:
                    import datetime as _dt
                    m.due_date = _today + _dt.timedelta(days=int(m.interval_days))
                # Snapshot current Hobbs/tacho for hours-based items
                _last_log = ac.maint_log.order_by('-date', '-id').first()
                if _last_log and _last_log.hobbs_reading is not None:
                    m.last_completed_hours = _last_log.hobbs_reading
                    if m.interval_hours and m.last_completed_hours is not None:
                        m.due_hours = float(m.last_completed_hours) + float(m.interval_hours)
                m.save()
                m.recalc_urgency()
                m.save(update_fields=['urgency'])
                _saved = True

        elif action == 'delete_maintenance' and actor.is_admin:
            AircraftMaintenanceItem.objects.filter(aircraft=ac, id=request.POST.get('maint_id')).delete()

        elif action == 'add_manual_log_entry' and actor.is_admin:
            from .models import MaintenanceLogEntry as _MLE
            from decimal import Decimal as _D
            entry_date = request.POST.get('entry_date') or timezone.localdate().isoformat()
            notes = request.POST.get('notes', '').strip()
            hobbs = request.POST.get('hobbs_reading') or None
            tacho = request.POST.get('tacho_reading') or None
            airswitch = request.POST.get('airswitch_reading') or None
            try:
                maint_hrs = _D(request.POST.get('maint_hours', '0') or '0')
            except Exception:
                maint_hrs = _D('0')
            # Compute cumulative total from last log entry
            prev = _MLE.objects.filter(aircraft=ac).order_by('-date', '-id').first()
            prev_total = prev.maint_hours_total if prev else _D(str(ac.maint_hours_initial or 0))
            _MLE.objects.create(
                aircraft=ac,
                date=entry_date,
                hobbs_reading=hobbs,
                tacho_reading=tacho,
                airswitch_reading=airswitch,
                maint_hours_flight=maint_hrs,
                maint_hours_total=prev_total + maint_hrs,
                notes=notes,
            )
            _saved = True

        elif action == 'add_aircraft_blockout':
            _create_blockout_from_post(request, club, scope='aircraft', aircraft=ac)

        elif action == 'delete_blockout':
            from .models import BlockOut
            BlockOut.objects.filter(club=club, id=request.POST.get('bo_id')).delete()

        return _inline_redirect(request, 'core:manage_aircraft_detail', club_slug=club_slug, aircraft_id=aircraft_id, saved=_saved)

    from .models import BlockOut
    from django.db.models import Q
    _today = timezone.localdate()
    _active_q = (
        Q(recurrence='one_off', date__gte=_today) |
        Q(recurrence__in=['weekly', 'daily'], active_until__isnull=True) |
        Q(recurrence__in=['weekly', 'daily'], active_until__gte=_today)
    )
    blockouts = (BlockOut.objects
                 .filter(club=club, scope='aircraft')
                 .filter(_active_q)
                 .prefetch_related('aircraft', 'blockout_type')
                 .order_by('recurrence', 'date', 'weekday', 'start_time'))
    ac_blockouts = [bo for bo in blockouts if ac in bo.aircraft.all()]

    hire_rates = (ChargeRate.objects
                  .filter(aircraft=ac)
                  .select_related('flight_type')
                  .order_by('flight_type__name', 'time_method'))
    fuel_rates = FuelSurchargeRate.objects.filter(aircraft=ac).order_by('-effective_from')
    all_surcharge_types = AircraftSurchargeType.objects.filter(club=club)
    assigned_surcharge_ids = set(ac.surcharges.values_list('id', flat=True))
    maintenance_items = list(AircraftMaintenanceItem.objects.filter(aircraft=ac).order_by('urgency', 'due_date'))

    # Compute progress percentage for each maintenance item and attach as a dynamic attribute.
    # Priority: date-based interval if both last_completed_date and due_date are set; else hours-based.
    _latest_meters = (FlightCompletion.objects
                      .filter(booking__aircraft=ac, booking__club=club)
                      .exclude(hobbs_end__isnull=True)
                      .order_by('-booking__arrived_at', '-created_at')
                      .values('hobbs_end').first())
    _current_hobbs = float(_latest_meters['hobbs_end']) if _latest_meters else None
    _today = timezone.localdate()
    for _m in maintenance_items:
        _m.progress_pct = None
        if _m.last_completed_date and _m.due_date:
            _total = max(1, (_m.due_date - _m.last_completed_date).days)
            _elapsed = (_today - _m.last_completed_date).days
            _m.progress_pct = min(100, max(0, round(_elapsed / _total * 100)))
        elif _m.last_completed_hours is not None and _m.due_hours is not None and _current_hobbs is not None:
            _total_h = float(_m.due_hours - _m.last_completed_hours)
            if _total_h > 0:
                _elapsed_h = _current_hobbs - float(_m.last_completed_hours)
                _m.progress_pct = min(100, max(0, round(_elapsed_h / _total_h * 100)))
    flight_types = FlightType.objects.filter(club=club, is_billable=True)
    aircraft_blockout_types = BlockOutType.objects.filter(club=club, target='aircraft')

    from django.core.paginator import Paginator as _FHPag
    _fh_qs = (Booking.objects
              .filter(club=club, aircraft=ac)
              .exclude(status='cancelled')
              .select_related('member__user', 'instructor', 'flight_type', 'flight_completion')
              .order_by('-scheduled_start'))
    flight_history = _FHPag(_fh_qs, 25).get_page(request.GET.get('fh_page'))

    from .models import MaintenanceLogEntry
    maint_log = (MaintenanceLogEntry.objects
                 .filter(aircraft=ac)
                 .select_related('flight_completion__booking__member__user')
                 .order_by('-date', '-id')[:100])

    _at = ac.aircraft_type
    _designator = (_at.icao_designator or '').upper() if _at else ''
    _engine_count = ac.engine_count or 1
    if _engine_count >= 2:
        _ac_icon_type = 'twin'
    elif any(_designator.startswith(p) for p in ('C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9')):
        _ac_icon_type = 'high_wing'
    else:
        _ac_icon_type = 'low_wing'

    _is_inline = request.GET.get('inline') == '1'
    return render(request, 'core/manage_aircraft_detail.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'ac': ac,
        'ac_icon_type': _ac_icon_type,
        'hire_rates': hire_rates,
        'fuel_rates': fuel_rates,
        'all_surcharge_types': all_surcharge_types,
        'assigned_surcharge_ids': assigned_surcharge_ids,
        'maintenance_items': maintenance_items,
        'maint_log': maint_log,
        'flight_types': flight_types,
        'aircraft_blockout_types': aircraft_blockout_types,
        'ac_blockouts': ac_blockouts,
        'flight_history': flight_history,
        'aircraft_type_list': AircraftType.objects.filter(club=club),
        'base_template': 'core/base_inline.html' if _is_inline else 'core/base.html',
        'inline_title': ac.registration,
        'records_instruments': [
            ('records_hobbs',      'Hobbs meter',  ac.records_hobbs),
            ('records_tacho',      'Tachometer',   ac.records_tacho),
            ('records_airswitch',  'Air switch',   ac.records_airswitch),
        ],
    })


@login_required
def aircraft_maintenance_log(request, club_slug, aircraft_id):
    import csv as _csv
    from decimal import Decimal as _D
    from django.http import HttpResponse as _HR

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    ac = get_object_or_404(Aircraft, id=aircraft_id, club=club)

    entries_qs = (
        ac.maint_log
        .select_related('flight_completion__booking__member__user')
        .order_by('date', 'id')
    )

    rows = []
    prev_hobbs_end = None
    for e in entries_qs:
        fc = e.flight_completion
        hobbs_start = fc.hobbs_start if fc else None
        hobbs_end   = fc.hobbs_end   if fc else None
        tacho_start = fc.tacho_start if fc else None
        tacho_end   = fc.tacho_end   if fc else None
        as_start    = fc.airswitch_start if fc else None
        as_end      = fc.airswitch_end   if fc else None
        member_name = fc.booking.member.user.get_full_name() if (fc and fc.booking and fc.booking.member) else None

        gap = None
        if hobbs_start is not None and prev_hobbs_end is not None:
            g = _D(str(hobbs_start)) - _D(str(prev_hobbs_end))
            if abs(g) > _D('0.001'):
                gap = g

        # Flight hours from the billing instrument (hobbs/tacho/airswitch)
        flight_hrs = None
        ttm = ac.total_time_method
        if ttm == 'hobbs' and hobbs_start is not None and hobbs_end is not None:
            flight_hrs = _D(str(hobbs_end)) - _D(str(hobbs_start))
        elif ttm == 'tacho' and tacho_start is not None and tacho_end is not None:
            flight_hrs = _D(str(tacho_end)) - _D(str(tacho_start))
        elif ttm == 'airswitch' and as_start is not None and as_end is not None:
            flight_hrs = _D(str(as_end)) - _D(str(as_start))

        rows.append({
            'date': e.date,
            'member': member_name,
            'is_manual': fc is None,
            'notes': e.notes,
            'hobbs_start': hobbs_start,
            'hobbs_end': hobbs_end,
            'hobbs_gap': gap,
            'tacho_start': tacho_start,
            'tacho_end': tacho_end,
            'as_start': as_start,
            'as_end': as_end,
            'flight_hrs': flight_hrs,
            'maint_flight': e.maint_hours_flight,
            'maint_total': e.maint_hours_total,
        })
        prev_hobbs_end = e.hobbs_reading  # MaintenanceLogEntry stores end reading

    if request.GET.get('fmt') == 'csv':
        resp = _HR(content_type='text/csv; charset=utf-8-sig')
        resp['Content-Disposition'] = f'attachment; filename="{ac.registration}_maintenance_log.csv"'
        w = _csv.writer(resp)
        w.writerow(['Date', 'Member/Note', 'Hobbs start', 'Hobbs end', 'Hobbs gap',
                    'Tacho start', 'Tacho end', 'Air sw. start', 'Air sw. end',
                    'Flight hrs', 'Maint hrs (flight)', 'Maint hrs (total)'])
        for r in rows:
            w.writerow([
                r['date'],
                r['member'] or (f"Manual — {r['notes']}" if r['notes'] else 'Manual'),
                r['hobbs_start'] or '',
                r['hobbs_end'] or '',
                r['hobbs_gap'] or '',
                r['tacho_start'] or '',
                r['tacho_end'] or '',
                r['as_start'] or '',
                r['as_end'] or '',
                r['flight_hrs'] or '',
                r['maint_flight'],
                r['maint_total'],
            ])
        return resp

    return render(request, 'core/aircraft_maintenance_log.html', {
        'club': club, 'club_member': actor, 'ac': ac, 'rows': rows,
    })


@login_required
def manage_instructors(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    if request.method == 'POST' and actor.is_admin:
        action = request.POST.get('action', '')
        if action == 'add_instructor':
            member_id = request.POST.get('member_id', '').strip()
            member = ClubMember.objects.filter(club=club, id=member_id).first()
            if member and member.role and member.role.effective_is_instructor:
                member.is_on_instructor_roster = True
                member.save(update_fields=['is_on_instructor_roster'])
        elif action == 'remove_instructor':
            member_id = request.POST.get('member_id', '').strip()
            member = ClubMember.objects.filter(club=club, id=member_id).first()
            if member:
                future_bookings = Booking.objects.filter(
                    club=club, instructor=member.user,
                    scheduled_end__gte=timezone.now(),
                    status__in=['pending', 'confirmed'],
                )
                for b in future_bookings:
                    new_instr_id = request.POST.get(f'reassign_{b.id}', '').strip()
                    if new_instr_id:
                        new_instr = ClubMember.objects.filter(
                            club=club, id=new_instr_id, is_on_instructor_roster=True
                        ).first()
                        b.instructor = new_instr.user if new_instr else None
                    else:
                        b.instructor = None
                    b.save(update_fields=['instructor'])
                ClubMember.objects.filter(club=club, id=member_id).update(is_on_instructor_roster=False)
        return redirect('core:manage_instructors', club_slug=club_slug)

    from .models import InstructorAvailability
    from django.db.models import Q
    _today = timezone.localdate()
    _active_q = (
        Q(recurrence='one_off', date__gte=_today) |
        Q(recurrence='weekly', active_until__isnull=True) |
        Q(recurrence='weekly', active_until__gte=_today)
    )
    av_counts = {}
    for av in InstructorAvailability.objects.filter(club_member__club=club).filter(_active_q):
        av_counts[av.club_member_id] = av_counts.get(av.club_member_id, 0) + 1

    instructors = (ClubMember.objects
                   .filter(club=club, is_on_instructor_roster=True)
                   .select_related('user', 'instructor_grade', 'role')
                   .order_by('user__last_name'))
    for instr in instructors:
        instr.av_count = av_counts.get(instr.id, 0)
        instr.future_bookings = list(
            Booking.objects
            .filter(club=club, instructor=instr.user,
                    scheduled_end__gte=timezone.now(),
                    status__in=['pending', 'confirmed'])
            .select_related('member__user', 'aircraft')
            .order_by('scheduled_start')
        )

    # Members eligible to add: have instructor-type role, active standing, not already on roster
    roster_ids = set(instructors.values_list('id', flat=True))
    # Use system_role_type='instructor' if set, fall back to permission flags for legacy roles
    from django.db.models import Q as _Q
    _instr_q = _Q(role__system_role_type='instructor') | _Q(
        role__bookings_access='manage_all', role__can_access_manage=True
    )
    eligible_members = (ClubMember.objects
                        .filter(_instr_q, club=club, standing='current')
                        .exclude(id__in=roster_ids)
                        .select_related('user', 'role')
                        .order_by('user__last_name', 'user__first_name'))
    any_instructor_role = ClubMember.objects.filter(_instr_q, club=club).exists()
    all_on_roster = not eligible_members.exists() and any_instructor_role

    # ── Mini roster calendar (rolling 5-week window) ─────────────────────────
    from datetime import date as _date2, timedelta as _td
    _ROSTER_COLORS = [
        '#2563eb', '#16a34a', '#dc2626', '#d97706',
        '#7c3aed', '#0891b2', '#be185d', '#065f46',
        '#92400e', '#1e40af',
    ]

    # Start from Monday of the current week, show 5 full weeks
    _week_start = _today - _td(days=_today.weekday())
    _cal_start  = _week_start
    _cal_end    = _cal_start + _td(days=34)

    # Assign color + initials + availability windows to each instructor
    instr_list = list(instructors)
    for _i, _instr in enumerate(instr_list):
        _instr.cal_color    = _ROSTER_COLORS[_i % len(_ROSTER_COLORS)]
        _instr.cal_initials = (_instr.user.first_name[:1] + _instr.user.last_name[:1]).upper() or '?'
        _instr._av_wins     = list(InstructorAvailability.objects.filter(club_member=_instr))

    # Block-outs covering this window (scope='all' only for the overview)
    from .models import BlockOut as _BlockOut
    _raw_bos = list(_BlockOut.objects.filter(club=club, scope='all').filter(
        Q(recurrence='one_off',  date__range=(_cal_start, _cal_end)) |
        Q(recurrence='weekly') |
        Q(recurrence='daily',    date__isnull=True) |
        Q(recurrence='daily',    date__lte=_cal_end)
    ).select_related('blockout_type'))

    def _has_blockout(day):
        for _bo in _raw_bos:
            if _bo.recurrence == 'daily':
                return True
            if _bo.recurrence == 'weekly' and _bo.weekday == day.weekday():
                return True
            if _bo.recurrence == 'one_off' and _bo.date == day:
                return True
        return False

    def _on_roster_on(instr, day):
        if not instr._av_wins:
            return True  # no schedule = always available
        return any(w.applies_on(day) for w in instr._av_wins)

    # Build 5 weeks of rows (Mon-first, no padding cells)
    _cal_weeks = []
    _d = _cal_start
    while _d <= _cal_end:
        _row = []
        for _ in range(7):
            _row.append({
                'date':        _d,
                'is_today':    _d == _today,
                'is_past':     _d < _today,
                'is_weekend':  _d.weekday() >= 5,
                'instructors': [_instr for _instr in instr_list if _on_roster_on(_instr, _d)],
                'has_blockout': _has_blockout(_d),
            })
            _d += _td(days=1)
        _cal_weeks.append(_row)

    return render(request, 'core/manage_instructors.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'instructors': instructors,
        'eligible_members': eligible_members,
        'all_on_roster': all_on_roster,
        'cal_weeks':    _cal_weeks,
        'cal_range':    f'{_cal_start.strftime("%-d %b")} – {_cal_end.strftime("%-d %b %Y")}',
        'instr_list':   instr_list,
        'has_blockouts': any(_has_blockout(_cal_start + _td(days=i)) for i in range(35)),
    })


@login_required
def manage_instructor_detail(request, club_slug, member_id):
    from .models import InstructorAvailability, BlockOut, BlockOutType, MemberCredential
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    instr = get_object_or_404(ClubMember, club=club, id=member_id)

    if request.method == 'POST':
        action = request.POST.get('action', '')
        _saved = action in ('save_contact', 'save_details')

        if action == 'save_details' and actor.is_admin:
            # Grade is the only instructor-specific field; contact is edited on the member profile
            grade_id = request.POST.get('grade_id', '')
            instr.instructor_grade = InstructorGrade.objects.filter(club=club, id=grade_id).first() if grade_id else None
            instr.save(update_fields=['instructor_grade'])

        elif action == 'save_grade' and actor.is_admin:
            grade_id = request.POST.get('grade_id')
            instr.instructor_grade = InstructorGrade.objects.filter(club=club, id=grade_id).first() if grade_id else None
            instr.save(update_fields=['instructor_grade'])

        elif action == 'add_instructor_availability':
            recurrence = request.POST.get('av_recurrence', 'weekly')
            all_day = request.POST.get('av_all_day') == 'on'
            av = InstructorAvailability(club_member=instr, recurrence=recurrence, all_day=all_day)
            if recurrence == 'weekly':
                wd = request.POST.get('av_weekday', '0')
                av.weekday = int(wd) if wd.isdigit() else 0
            else:
                av.date = request.POST.get('av_date') or None
            if not all_day:
                av.start_time = request.POST.get('av_start_time') or None
                av.end_time = request.POST.get('av_end_time') or None
            av.active_from = request.POST.get('av_active_from') or None
            av.active_until = request.POST.get('av_active_until') or None
            av.notes = request.POST.get('av_notes', '').strip()
            av.save()

        elif action == 'delete_instructor_availability':
            InstructorAvailability.objects.filter(club_member=instr, id=request.POST.get('av_id')).delete()

        elif action == 'add_instructor_blockout':
            _create_blockout_from_post(request, club, scope='instructors', instructor_user=instr.user)

        elif action == 'delete_blockout':
            BlockOut.objects.filter(club=club, id=request.POST.get('bo_id')).delete()

        return _inline_redirect(request, 'core:manage_instructor_detail', club_slug=club_slug, member_id=member_id, saved=_saved)

    from django.db.models import Q
    _today = timezone.localdate()
    _active_q = (
        Q(recurrence='one_off', date__gte=_today) |
        Q(recurrence='weekly', active_until__isnull=True) |
        Q(recurrence='weekly', active_until__gte=_today)
    )
    av_windows = InstructorAvailability.objects.filter(club_member=instr).filter(_active_q).order_by('weekday', 'date')
    past_av_count = InstructorAvailability.objects.filter(club_member=instr).exclude(_active_q).count()

    _bo_active_q = (
        Q(recurrence='one_off', date__gte=_today) |
        Q(recurrence__in=['weekly', 'daily'], active_until__isnull=True) |
        Q(recurrence__in=['weekly', 'daily'], active_until__gte=_today)
    )
    all_instr_bos = (BlockOut.objects
                     .filter(club=club, scope='instructors')
                     .filter(_bo_active_q)
                     .prefetch_related('instructors', 'blockout_type'))
    instr_blockouts = [bo for bo in all_instr_bos if instr.user in bo.instructors.all()]

    credentials = (MemberCredential.objects
                   .filter(member=instr.user)
                   .select_related('credential_type', 'aircraft_type')
                   .order_by('credential_type__display_order', 'expiry_date'))
    instructor_grades = InstructorGrade.objects.filter(club=club).order_by('display_order')
    instructor_blockout_types = BlockOutType.objects.filter(club=club).exclude(target='aircraft')

    upcoming_bookings = (Booking.objects
                         .filter(club=club, instructor=instr.user)
                         .exclude(status='cancelled')
                         .filter(scheduled_start__gte=timezone.now())
                         .select_related('member__user', 'aircraft', 'flight_type')
                         .order_by('scheduled_start')[:10])

    _is_inline = request.GET.get('inline') == '1'
    return render(request, 'core/manage_instructor_detail.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'instr': instr,
        'av_windows': av_windows,
        'past_av_count': past_av_count,
        'instr_blockouts': instr_blockouts,
        'credentials': credentials,
        'instructor_grades': instructor_grades,
        'instructor_blockout_types': instructor_blockout_types,
        'upcoming_bookings': upcoming_bookings,
        'base_template': 'core/base_inline.html' if _is_inline else 'core/base.html',
        'inline_title': instr.user.get_full_name(),
    })


@login_required
def manage_contacts(request, club_slug):
    from .models import Contact as _C
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err

    if request.method == 'POST':
        name     = request.POST.get('name', '').strip()
        email    = request.POST.get('email', '').strip()
        phone    = request.POST.get('phone', '').strip()
        is_org   = request.POST.get('is_organisation') == '1'
        org      = request.POST.get('organisation', '').strip()
        ctype_id = request.POST.get('contact_type', '').strip()
        notes    = request.POST.get('notes', '').strip()
        if name:
            _C.objects.create(
                club=club, name=name, email=email, phone=phone,
                is_organisation=is_org, organisation=org,
                contact_type=ContactType.objects.filter(id=ctype_id, club=club).first() if ctype_id else None,
                notes=notes,
                created_by=request.user,
            )
        return redirect('core:manage_contacts', club_slug=club_slug)

    f_type = request.GET.get('type', '')
    f_q    = request.GET.get('q', '').strip()
    from django.db.models import Q as _Q, Count as _Count
    qs = _C.objects.filter(club=club)
    if f_type:
        qs = qs.filter(contact_type_id=f_type)
    if f_q:
        qs = qs.filter(_Q(name__icontains=f_q) | _Q(organisation__icontains=f_q) | _Q(email__icontains=f_q))
    from django.core.paginator import Paginator as _Pag
    from urllib.parse import urlencode as _ue
    sort     = request.GET.get('sort', 'name')
    sort_dir = request.GET.get('dir', 'asc')
    _CT_SORT = {
        'name':         ('name',),
        'type':         ('contact_type__name', 'name'),
        'organisation': ('organisation', 'name'),
    }
    if sort not in _CT_SORT:
        sort = 'name'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'asc'
    _ct_order = _CT_SORT[sort]
    if sort_dir == 'desc':
        _ct_order = tuple('-'+f for f in _ct_order)
    qs = qs.annotate(booking_count=_Count('bookings')).order_by(*_ct_order)
    _paginator    = _Pag(qs, 50)
    contacts_page = _paginator.get_page(request.GET.get('page'))
    _filter_qs    = _ue({k: v for k, v in request.GET.items() if k != 'page' and v})
    _base_qs      = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir') and v})

    return render(request, 'core/manage_contacts.html', {
        'club': club, 'club_member': actor,
        'contacts_page': contacts_page,
        'total_count': _paginator.count,
        'f_type': f_type, 'f_q': f_q,
        'contact_types': list(ContactType.objects.filter(club=club).order_by('name')),
        'filter_qs': _filter_qs,
        'sort': sort, 'sort_dir': sort_dir, 'base_qs': _base_qs,
    })


@login_required
def contact_detail(request, club_slug, contact_id):
    from .models import Contact as _C
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err

    contact = get_object_or_404(_C, id=contact_id, club=club)
    error = None

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'save':
            contact.name          = request.POST.get('name', contact.name).strip() or contact.name
            contact.email         = request.POST.get('email', '').strip()
            contact.phone         = request.POST.get('phone', '').strip()
            contact.is_organisation = request.POST.get('is_organisation') == '1'
            contact.organisation  = request.POST.get('organisation', '').strip()
            _ct_id = request.POST.get('contact_type', '').strip()
            contact.contact_type  = ContactType.objects.filter(id=_ct_id, club=club).first() if _ct_id else None
            contact.notes         = request.POST.get('notes', '').strip()
            contact.save()
            redir = request.path + '?saved=1'
            return redirect(redir)

        elif action == 'convert_to_member' and contact.can_convert:
            from django.contrib.auth import get_user_model as _gum
            from django.contrib.auth.hashers import make_password
            import secrets
            _User = _gum()
            email = request.POST.get('email', '').strip() or contact.email
            send_invite = request.POST.get('send_invite') == 'on'
            if not email:
                error = 'An email address is required to create a member account.'
            elif _User.objects.filter(email=email).exists():
                error = 'A user with this email already exists.'
            else:
                username = email.split('@')[0][:30]
                base_u = username
                i = 1
                while _User.objects.filter(username=username).exists():
                    username = f'{base_u}{i}'; i += 1
                names = contact.name.strip().split(None, 1)
                first = names[0]; last = names[1] if len(names) > 1 else ''
                user = _User.objects.create(
                    username=username, email=email,
                    first_name=first, last_name=last,
                    password=make_password(secrets.token_hex(20)),
                )
                member_role = ClubMember.objects.filter(
                    club=club, role__system_role_type='member'
                ).values_list('role_id', flat=True).first()
                new_member = ClubMember.objects.create(
                    club=club, user=user,
                    role_id=member_role,
                    standing='current',
                )
                contact.converted_to_member = new_member
                contact.save(update_fields=['converted_to_member'])
                if send_invite:
                    from .models import ClubInvite
                    from .email_notifications import club_invite as _email_invite
                    from datetime import timedelta as _td
                    ClubInvite.objects.filter(club=club, email=email, accepted_at__isnull=True).delete()
                    invite = ClubInvite.objects.create(
                        club=club, email=email,
                        invited_by=request.user,
                        expires_at=timezone.now() + _td(days=7),
                        club_member=new_member,
                    )
                    _email_invite(invite)
                    logger.info('invite_sent club=%s email=%s by=%s (contact conversion)', club.slug, email, request.user.email)
                return redirect('core:manage_member_detail', club_slug=club_slug, member_id=new_member.id)

    bookings = (contact.bookings
                .select_related('aircraft', 'flight_type', 'member__user')
                .order_by('-scheduled_start')[:20])

    is_inline = request.GET.get('inline') == '1'
    return render(request, 'core/contact_detail.html', {
        'club': club, 'club_member': actor,
        'contact': contact,
        'bookings': bookings,
        'error': error,
        'contact_types': list(ContactType.objects.filter(club=club).order_by('name')),
        'base_template': 'core/base_inline.html' if is_inline else 'core/base.html',
        'inline_title': f'Manage <span class="crumb-sep">›</span> Contacts <span class="crumb-sep">›</span> <span class="crumb-cur">{contact.name}</span>',
    })


@login_required
@require_POST
@transaction.atomic
def create_blockout(request):
    """Create a block-out from the calendar UI. Staff/admin only."""
    from .models import BlockOut, User as _User

    actor = ClubMember.objects.filter(user=request.user).first()
    if not actor or not (actor.is_admin or actor.is_instructor):
        return JsonResponse({'error': 'Not authorized'}, status=403)

    club = actor.club

    bot_id = request.POST.get('blockout_type_id')
    scope = request.POST.get('scope', 'all')
    label = request.POST.get('label', '').strip()
    recurrence = request.POST.get('recurrence', 'one_off')
    all_day = request.POST.get('all_day') in ('1', 'true', 'on')
    date_str = request.POST.get('date', '')
    weekday_str = request.POST.get('weekday', '')
    start_time_str = request.POST.get('start_time', '') or None
    end_time_str = request.POST.get('end_time', '') or None
    aircraft_ids = request.POST.getlist('aircraft_ids')
    instructor_ids = request.POST.getlist('instructor_ids')

    blockout_type = BlockOutType.objects.filter(club=club, id=bot_id).first() if bot_id else None

    bo = BlockOut(
        club=club, blockout_type=blockout_type, label=label,
        scope=scope, recurrence=recurrence, all_day=all_day,
        created_by=request.user,
    )

    if recurrence == 'one_off':
        try:
            from datetime import date as _date
            bo.date = _date.fromisoformat(date_str) if date_str else None
        except ValueError:
            return JsonResponse({'error': 'Invalid date'}, status=400)
    elif recurrence == 'weekly':
        bo.weekday = int(weekday_str) if weekday_str.isdigit() else 0

    if not all_day:
        bo.start_time = start_time_str
        bo.end_time = end_time_str

    bo.save()

    if scope == 'aircraft' and aircraft_ids:
        bo.aircraft.set(Aircraft.objects.filter(club=club, id__in=aircraft_ids))
    elif scope == 'instructors' and instructor_ids:
        bo.instructors.set(_User.objects.filter(id__in=instructor_ids))

    _recompute_conflicts_for_club(club)
    return JsonResponse({'success': True, 'id': bo.id})


@login_required
def manage_charges(request, club_slug):
    from .models import FlightCompletion, AccountTransaction
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    from django.core.paginator import Paginator
    from urllib.parse import urlencode as _ue

    tab = request.GET.get('tab', 'unpaid')
    if tab not in ('unpaid', 'transactions'):
        tab = 'unpaid'

    # Unpaid flights table sort
    usort     = request.GET.get('usort', 'date')
    usort_dir = request.GET.get('udir', 'desc')
    _UP_SORT = {
        'member':   ('booking__member__user__last_name', 'booking__member__user__first_name'),
        'aircraft': ('booking__aircraft__registration',),
        'date':     ('booking__scheduled_start',),
        'type':     ('booking__flight_type__name',),
    }
    if usort not in _UP_SORT:
        usort = 'date'
    if usort_dir not in ('asc', 'desc'):
        usort_dir = 'desc'
    _up_order = _UP_SORT[usort]
    if usort_dir == 'desc':
        _up_order = tuple('-'+f for f in _up_order)
    unpaid = (FlightCompletion.objects
              .filter(booking__club=club, paid_at__isnull=True)
              .select_related('booking__member__user', 'booking__aircraft', 'booking__flight_type',
                              'booking__instructor')
              .order_by(*_up_order))

    # Transactions table sort
    sort     = request.GET.get('sort', 'date')
    sort_dir = request.GET.get('dir', 'desc')
    _TX_SORT = {
        'member': ('account__club_member__user__last_name', 'account__club_member__user__first_name'),
        'type':   ('transaction_type', '-created_at'),
        'date':   ('created_at',),
        'amount': ('amount', '-created_at'),
    }
    if sort not in _TX_SORT:
        sort = 'date'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'desc'
    _tx_order = _TX_SORT[sort]
    if sort_dir == 'desc':
        _tx_order = tuple(f[1:] if f.startswith('-') else '-'+f for f in _tx_order)
    tx_qs = (AccountTransaction.objects
             .filter(account__club_member__club=club)
             .select_related('account__club_member__user', 'flight_completion__booking__aircraft')
             .order_by(*_tx_order))
    tx_page = Paginator(tx_qs, 50).get_page(request.GET.get('page'))
    _base_qs   = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir', 'usort', 'udir') and v})
    _filter_qs = _ue({k: v for k, v in request.GET.items() if k != 'page' and v})

    return render(request, 'core/manage_charges.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'tab': tab,
        'unpaid': unpaid,
        'tx_page': tx_page,
        'sort': sort, 'sort_dir': sort_dir,
        'usort': usort, 'usort_dir': usort_dir,
        'base_qs': _base_qs,
        'filter_qs': _filter_qs,
    })


@login_required
@transaction.atomic
def booking_declaration(request, club_slug, booking_id):
    from .models import DepartureDeclaration, DeclarationPassenger, FrequentPassenger
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    booking = get_object_or_404(Booking, club=club, id=booking_id)
    # Allow the member themselves or staff
    is_own = (booking.member.user == request.user)
    if not (is_own or actor.is_admin or actor.is_instructor):
        return render(request, 'core/no_access.html', {'club': club}, status=403)

    # Completed flights: declaration is always read-only for everyone
    # Departed flights: read-only for non-staff
    readonly = (booking.status == 'completed') or \
               (not is_own and not actor.is_admin and not actor.is_instructor) or \
               (booking.status == 'departed' and not actor.is_admin and not actor.is_instructor)

    decl, _ = DepartureDeclaration.objects.get_or_create(
        booking=booking,
        defaults={'submitted_by': request.user, 'is_draft': True}
    )
    error = None
    _is_app = request.path.startswith('/app/')

    _decl_url = f'{request.path}{"?inline=1" if request.GET.get("inline") == "1" else ""}'
    _saved_url = request.path + '?saved=1'

    if request.method == 'POST' and not readonly:
        action = request.POST.get('action', 'save_draft')

        # Passenger operations — handled independently, don't touch declaration fields
        if action == 'add_passenger':
            pax_name = request.POST.get('pax_name', '').strip()
            pax_phone = request.POST.get('pax_phone', '').strip()
            pax_nok_name = request.POST.get('pax_nok_name', '').strip()
            pax_nok_phone = request.POST.get('pax_nok_phone', '').strip()
            if pax_name:
                decl.save()
                DeclarationPassenger.objects.create(
                    declaration=decl, name=pax_name, phone=pax_phone,
                    next_of_kin_name=pax_nok_name, next_of_kin_phone=pax_nok_phone,
                )
            return redirect(_decl_url)

        elif action == 'remove_passenger':
            DeclarationPassenger.objects.filter(
                declaration=decl, id=request.POST.get('pax_id')
            ).delete()
            return redirect(_decl_url)

        # Declaration content fields
        from django.utils.dateparse import parse_datetime as _pdatetime
        decl.authorising_instructor_id = request.POST.get('authorising_instructor_id') or None
        decl.route_intentions = request.POST.get('route_intentions', '').strip()
        decl.destination = request.POST.get('destination', '').strip()
        decl.next_of_kin_name = request.POST.get('next_of_kin_name', '').strip()
        decl.next_of_kin_phone = request.POST.get('next_of_kin_phone', '').strip()
        for field in ['confirm_aip', 'confirm_weather', 'confirm_fuel',
                      'confirm_maps', 'confirm_fuel_card', 'confirm_afm', 'confirm_flight_plan']:
            setattr(decl, field, request.POST.get(field) == 'on')
        _er = request.POST.get('estimated_return', '').strip()
        decl.estimated_return = _pdatetime(_er.replace('T', ' ')) if _er else None
        _st = request.POST.get('sar_time', '').strip()
        decl.sar_time = _pdatetime(_st.replace('T', ' ')) if _st else None

        if action == 'submit':
            missing = []
            if not decl.route_intentions:
                missing.append('route intentions')
            if not decl.destination:
                missing.append('destination')
            if missing:
                error = 'Please complete: ' + ', '.join(missing)
            else:
                from django.utils import timezone as tz
                decl.is_draft = False
                decl.submitted_at = tz.now()
                decl.save()
                if _is_app:
                    return redirect('core:app_bookings', club_slug=club_slug)
                return redirect('core:booking_detail', club_slug=club_slug, booking_id=booking_id)
        else:
            decl.save()

        if not error:
            return redirect(_saved_url)

    instructors = ClubMember.objects.filter(club=club, is_on_instructor_roster=True).select_related('user')
    passengers = decl.passengers.all()
    frequent_passengers = FrequentPassenger.objects.filter(club_member=booking.member)
    member_nok_name = booking.member.next_of_kin_name
    member_nok_phone = booking.member.next_of_kin_phone

    from django.utils import timezone as _tz
    _now = _tz.now()
    _delta = (booking.scheduled_start - _now).total_seconds()
    hours_to_departure = max(0, _delta / 3600)

    _is_inline = request.GET.get('inline') == '1'
    if _is_app:
        _base = 'core/app/base.html'
    elif _is_inline:
        _base = 'core/base_inline.html'
    else:
        _base = 'core/base.html'
    return render(request, 'core/booking_declaration.html', {
        'club': club, 'club_member': actor,
        'booking': booking, 'decl': decl, 'error': error,
        'instructors': instructors,
        'passengers': passengers,
        'frequent_passengers': frequent_passengers,
        'member_nok_name': member_nok_name,
        'member_nok_phone': member_nok_phone,
        'readonly': readonly,
        'hours_to_departure': hours_to_departure,
        'is_inline': _is_inline,
        'is_app': _is_app,
        'base_template': _base,
        'inline_title': f'Bookings <span class="crumb-sep">›</span> {booking.aircraft.registration} {booking.scheduled_start.strftime("%j %b")} <span class="crumb-sep">›</span> <span class="crumb-cur">Declaration</span>',
    })


@login_required
def manage_aerodromes(request, club_slug):
    from .models import AerodromeFeeType
    club = get_object_or_404(Club, slug=club_slug)
    try:
        member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if not member.is_staff:
        return render(request, 'core/no_access.html', {'club': club}, status=403)

    error = None
    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add_aerodrome':
            icao = request.POST.get('icao', '').strip().upper()
            name = request.POST.get('name', '').strip()
            notes = request.POST.get('notes', '').strip()
            if icao and name:
                Aerodrome.objects.get_or_create(
                    club=club, icao_code=icao,
                    defaults={'name': name, 'notes': notes}
                )
            else:
                error = "ICAO code and name are required."

        elif action == 'edit_aerodrome':
            ae = Aerodrome.objects.filter(club=club, id=request.POST.get('ae_id')).first()
            if ae:
                ae.name = request.POST.get('name', ae.name).strip()
                ae.notes = request.POST.get('notes', '').strip()
                make_home = request.POST.get('is_home') == '1'
                if make_home and not ae.is_home:
                    Aerodrome.objects.filter(club=club).update(is_home=False)
                    ae.is_home = True
                elif not make_home:
                    ae.is_home = False
                ae.save(update_fields=['name', 'notes', 'is_home'])

        elif action in ('toggle_aerodrome', 'set_aerodrome_status'):
            ae = Aerodrome.objects.filter(club=club, id=request.POST.get('ae_id')).first()
            if ae:
                if action == 'set_aerodrome_status':
                    ae.is_active = request.POST.get('is_active') == '1'
                else:
                    ae.is_active = not ae.is_active
                ae.save(update_fields=['is_active'])

        elif action == 'delete_aerodrome':
            Aerodrome.objects.filter(club=club, id=request.POST.get('ae_id')).delete()

        elif action == 'add_fee_type':
            ae = Aerodrome.objects.filter(club=club, id=request.POST.get('ae_id')).first()
            ft_name = request.POST.get('ft_name', '').strip()
            ft_amount = request.POST.get('ft_amount', '0').strip()
            if ae and ft_name:
                AerodromeFeeType.objects.get_or_create(
                    aerodrome=ae, name=ft_name,
                    defaults={'default_amount': ft_amount or 0}
                )

        elif action == 'edit_fee_type':
            ft = AerodromeFeeType.objects.filter(
                aerodrome__club=club, id=request.POST.get('ft_id')
            ).first()
            if ft:
                ft.name = request.POST.get('ft_name', ft.name).strip()
                ft.default_amount = request.POST.get('ft_amount', ft.default_amount)
                ft.save()

        elif action == 'delete_fee_type':
            AerodromeFeeType.objects.filter(
                aerodrome__club=club, id=request.POST.get('ft_id')
            ).delete()

        elif action == 'save_fees':
            from decimal import Decimal, InvalidOperation
            ae = Aerodrome.objects.filter(club=club, id=request.POST.get('ae_id')).first()
            if ae:
                keep_ids    = [i for i in request.POST.getlist('fee_id') if i]
                fee_names   = request.POST.getlist('fee_name')
                fee_amounts = request.POST.getlist('fee_amount')
                for ft_id, name, amount in zip(keep_ids, fee_names, fee_amounts):
                    ft = AerodromeFeeType.objects.filter(aerodrome=ae, id=ft_id).first()
                    if ft and name.strip():
                        ft.name = name.strip()
                        try:
                            ft.default_amount = Decimal(str(amount).strip())
                        except (InvalidOperation, ValueError):
                            pass
                        ft.save()
                int_ids = []
                for i in keep_ids:
                    try:
                        int_ids.append(int(i))
                    except ValueError:
                        pass
                AerodromeFeeType.objects.filter(aerodrome=ae).exclude(id__in=int_ids).delete()
                for name, amount in zip(
                    request.POST.getlist('new_fee_name'),
                    request.POST.getlist('new_fee_amount'),
                ):
                    name = name.strip()
                    if name:
                        try:
                            amt = Decimal(str(amount).strip())
                        except (InvalidOperation, ValueError):
                            amt = Decimal('0')
                        AerodromeFeeType.objects.create(aerodrome=ae, name=name, default_amount=amt)

        return redirect('core:manage_aerodromes', club_slug=club_slug)

    aerodromes = Aerodrome.objects.filter(club=club).prefetch_related('fee_types').order_by('icao_code')
    return render(request, 'core/manage_aerodromes.html', {
        'club': club, 'club_member': member, 'aerodromes': aerodromes, 'error': error,
    })


@login_required
def manage_rates(request, club_slug):
    """Fuel levy rates moved to individual aircraft detail pages."""
    return redirect('core:manage_aircraft', club_slug=club_slug)


@login_required
def club_rates(request, club_slug):
    """Rates page dissolved — hire rates on aircraft pages, grade/surcharge in Settings."""
    return redirect('core:club_settings', club_slug=club_slug)


# ============================================================================
# INVOICING
# ============================================================================

@login_required
@transaction.atomic
def generate_invoice(request, club_slug, booking_id):
    """Create per-payee invoices from a completed booking's charge items."""
    club = get_object_or_404(Club, slug=club_slug)
    booking = get_object_or_404(Booking, club=club, id=booking_id)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    fc = getattr(booking, 'flight_completion', None)
    if not fc:
        return redirect('core:booking_detail', club_slug=club_slug, booking_id=booking_id)

    from django.urls import reverse as _rev
    import urllib.parse as _up
    _is_inline = request.GET.get('inline') == '1'
    _inline_sfx = '?inline=1' if _is_inline else ''
    _booking_url = _rev('core:booking_detail', kwargs={'club_slug': club_slug, 'booking_id': booking_id})
    def _inv_qs(inv_id):
        _qs_data = {'back': _booking_url}
        if _is_inline:
            _qs_data = {'inline': '1', 'back': _booking_url}
        return '?' + _up.urlencode(_qs_data)

    if not fc.charge_items.exists() or not fc.total_charge:
        from django.contrib import messages as _msg
        _msg.error(request, 'Cannot generate a $0 invoice — add charge items first.')
        return redirect('core:booking_detail', club_slug=club_slug, booking_id=booking_id)

    config = get_config(club)
    from datetime import date as _date, timedelta as _td
    from django.db.models import F
    from django.db import IntegrityError as _IErr
    ClubConfig = config.__class__

    today = timezone.localdate()
    due   = today + _td(days=config.payment_terms_days)
    description = booking.flight_type.name if booking.flight_type else ''

    UNIT_MAP = {'hire': 'Hr', 'instructor': 'Hr', 'fuel': 'Hr', 'landing': 'Ldg',
                'surcharge': 'Ea', 'one_off': 'Ea'}

    # Determine payees: use FlightPayments if split, else booking member
    # Tuple: (member, invoice_total, amount_paid, paid_at, fp_method)
    _FP_METHOD_MAP = {'cash': 'cash', 'eftpos': 'eftpos', 'credit': 'account_credit'}
    fp_list = list(fc.payments.all())
    existing_member_ids = set(fc.invoices.values_list('member_id', flat=True))

    if fp_list:
        payees = [
            (fp.member, fp.amount, fp.amount if fp.paid_at else 0, fp.paid_at,
             _FP_METHOD_MAP.get(fp.method, 'bank_transfer'))
            for fp in fp_list
            if fp.amount > 0 and fp.member_id not in existing_member_ids
        ]
        is_split = len(fp_list) > 1
    else:
        if booking.member.id not in existing_member_ids:
            _fp0 = fc.payments.filter(paid_at__isnull=False).first()
            _m0 = _FP_METHOD_MAP.get(_fp0.method, 'bank_transfer') if _fp0 else 'bank_transfer'
            payees = [(booking.member, fc.total_charge, fc.amount_paid or 0, fc.paid_at, _m0)]
        else:
            payees = []
        is_split = False

    if not payees:
        # All payees already invoiced — redirect to first invoice
        first = fc.invoices.order_by('invoice_number').first()
        if first:
            return redirect(_rev('core:invoice_detail', kwargs={'club_slug': club_slug, 'invoice_id': first.id}) + _inv_qs(first.id))
        return redirect('core:booking_detail', club_slug=club_slug, booking_id=booking_id)

    from .models import InvoicePayment as _IP
    from django.db.models import Max as _Max
    created = []
    for member, amount, paid, fp_paid_at, inv_method in payees:
        with transaction.atomic():
            _existing_max = Invoice.objects.filter(club=club).aggregate(m=_Max('invoice_number'))['m'] or 0
            inv_number = max(config.invoice_number_next, _existing_max + 1)
            ClubConfig.objects.filter(pk=config.pk).update(invoice_number_next=inv_number + 1)
            config.invoice_number_next = inv_number + 1  # keep in-memory copy in sync for multi-payee loops
            invoice = Invoice.objects.create(
                club=club,
                member=member,
                flight_completion=fc,
                invoice_number=inv_number,
                issue_date=today,
                due_date=due,
                description=description,
                gst_rate=config.gst_rate,
                created_by=request.user,
            )

        # Line items: for split, one summary line per payee; for single, snapshot all charge items
        if is_split:
            InvoiceLineItem.objects.create(
                invoice=invoice,
                description=f'Flight hire share — {booking.aircraft.registration} {today.strftime("%-d %b %Y")}',
                quantity=1,
                unit='Ea',
                rate=amount,
                amount=amount,
                sort_order=0,
            )
        else:
            for order, ci in enumerate(fc.charge_items.all()):
                InvoiceLineItem.objects.create(
                    invoice=invoice,
                    description=ci.description or ci.get_item_type_display(),
                    quantity=1,
                    unit=UNIT_MAP.get(ci.item_type, 'Ea'),
                    rate=ci.amount,
                    amount=ci.amount,
                    sort_order=order,
                    charge_item=ci,
                )

        # Auto-mark as paid if payment was already recorded — this is a receipt.
        # Create an InvoicePayment row so the ledger is complete, then sync the cache.
        if paid >= amount > 0:
            _IP.objects.create(
                invoice=invoice,
                amount=paid,
                method=inv_method,
                paid_at=fp_paid_at or timezone.now(),
                recorded_by=request.user,
            )
            invoice._sync_payment_cache()

        created.append(invoice)

    if len(created) == 1:
        return redirect(_rev('core:invoice_detail', kwargs={'club_slug': club_slug, 'invoice_id': created[0].id}) + _inv_qs(created[0].id))

    from django.contrib import messages as _msg
    _msg.success(request, f'{len(created)} invoice(s) generated.')
    _back = _rev('core:booking_detail', kwargs={'club_slug': club_slug, 'booking_id': booking_id})
    return redirect(_back + '?saved=1')


@login_required
def invoice_detail(request, club_slug, invoice_id):
    """View, edit status, and print an invoice."""
    club    = get_object_or_404(Club, slug=club_slug)
    invoice = get_object_or_404(Invoice, club=club, id=invoice_id)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    config = get_config(club)
    error = success = ''

    _is_inline = request.GET.get('inline') == '1'
    _stay_url  = request.path + ('?inline=1&saved=1' if _is_inline else '?saved=1')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        # Description only editable while draft; notes always editable
        _desc = request.POST.get('description', None)
        _notes = request.POST.get('notes', None)
        _fields = []
        if _desc is not None and invoice.status == 'draft':
            invoice.description = _desc.strip()
            _fields.append('description')
        if _notes is not None and invoice.status != 'void':
            invoice.notes = _notes.strip()
            _fields.append('notes')
        if _fields:
            invoice.save(update_fields=_fields)

        if action in ('update_details', 'save_notes'):
            return redirect(_stay_url)

        if action == 'mark_sent' and invoice.status == 'draft':
            invoice.status = 'sent'
            invoice.sent_at = timezone.now()
            invoice.save(update_fields=['status', 'sent_at'])
            from .services import notification_service as _ns
            _ns.notify_invoice_issued(invoice)
            from .email_notifications import invoice_sent as _email_invoice
            _email_invoice(invoice)
            return redirect(_stay_url)

        elif action == 'mark_paid' and invoice.status == 'sent':
            from decimal import Decimal as _D
            from .models import Account as _Acct, AccountTransaction as _AT, InvoicePayment as _IP
            _pay_method = request.POST.get('payment_method', '').strip()
            _valid_methods = {c[0] for c in _IP.PAYMENT_METHOD_CHOICES}
            if _pay_method not in _valid_methods:
                error = 'Select a payment method.'
            _reference = request.POST.get('reference', '').strip()
            pay_str = request.POST.get('payment_amount', '').strip()
            try:
                pay_amt = _D(pay_str)
                if pay_amt <= 0:
                    raise ValueError
            except (ValueError, Exception):
                if not error:
                    error = 'Enter a valid payment amount.'
                pay_amt = None
            if pay_amt and not error:
                if pay_amt > invoice.balance_due:
                    error = f'Payment ${pay_amt:.2f} exceeds outstanding balance ${invoice.balance_due:.2f}.'
                    pay_amt = None
            if pay_amt and not error and _pay_method == 'account_credit':
                if not invoice.member:
                    error = 'Cannot use account credit — invoice has no member.'
                    pay_amt = None
                else:
                    _acct, _ = _Acct.objects.get_or_create(club_member=invoice.member, defaults={'balance': 0})
                    from .services.charging_service import _check_credit_headroom
                    _credit_err = _check_credit_headroom(_acct, pay_amt)
                    if _credit_err:
                        error = _credit_err
                        pay_amt = None
            if pay_amt and not error:
                _IP.objects.create(
                    invoice=invoice,
                    amount=pay_amt,
                    method=_pay_method,
                    paid_at=timezone.now(),
                    reference=_reference,
                    recorded_by=request.user,
                )
                if _pay_method == 'account_credit' and invoice.member:
                    _acct, _ = _Acct.objects.get_or_create(club_member=invoice.member, defaults={'balance': 0})
                    _AT.objects.create(
                        account=_acct,
                        transaction_type='flight',
                        direction='debit',
                        amount=pay_amt,
                        description=f'Invoice {invoice.display_number}',
                        flight_completion=invoice.flight_completion,
                        payment_method='account',
                        created_by=request.user,
                    )
                    _acct.apply_transaction(pay_amt, 'debit')
                _now_fully_paid = invoice._sync_payment_cache()
                if _now_fully_paid:
                    _fc_linked = invoice.flight_completion
                    if _fc_linked is not None:
                        _fc_linked.refresh_from_db(fields=['amount_paid', 'total_charge', 'paid_at'])
                        if _fc_linked.balance_owing > _D('0'):
                            _fc_pay = min(pay_amt, _D(str(_fc_linked.balance_owing)))
                            from .services import charging_service as _cs
                            _cs.record_payment(
                                fc=_fc_linked,
                                booking=_fc_linked.booking,
                                user=request.user,
                                amount=_fc_pay,
                                method='invoice',
                            )
                    if invoice.subscription_expiry_date and invoice.member:
                        _sub_member = invoice.member
                        _old_exp = _sub_member.subscription_expires
                        _sub_member.subscription_expires = invoice.subscription_expiry_date
                        _upd = ['subscription_expires']
                        _was_pending = _sub_member.standing == 'pending'
                        if _was_pending:
                            _sub_member.standing = 'active'
                            _upd.append('standing')
                        _sub_member.save(update_fields=_upd)
                        MembershipHistoryEntry.objects.create(
                            club_member=_sub_member,
                            event_type='subscription_renewed',
                            changed_by=request.user,
                            old_value=str(_old_exp) if _old_exp else '—',
                            new_value=str(invoice.subscription_expiry_date),
                        )
                        if _was_pending:
                            MembershipHistoryEntry.objects.create(
                                club_member=_sub_member,
                                event_type='standing_change',
                                changed_by=request.user,
                                old_value='pending', new_value='active',
                                note='Standing set to Active — initial subscription invoice paid',
                            )
                            if _sub_member.user and not _sub_member.user.is_active:
                                _sub_member.user.is_active = True
                                _sub_member.user.save(update_fields=['is_active'])
                return redirect(_stay_url)

        elif action == 'void' and invoice.status in ('draft', 'sent'):
            invoice.status = 'void'
            invoice.save(update_fields=['status'])
            return redirect(_stay_url)

    line_items = invoice.line_items.all()
    from .models import Account as _Acct
    _member_acct = None
    if invoice.member:
        _member_acct, _ = _Acct.objects.get_or_create(club_member=invoice.member, defaults={'balance': 0})
    return render(request, 'core/invoice_detail.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'invoice': invoice, 'line_items': line_items, 'config': config,
        'error': error, 'success': success,
        'member_account_balance': _member_acct.balance if _member_acct else None,
        'base_template': 'core/base_inline.html' if _is_inline else 'core/base.html',
    })


@login_required
def invoice_print(request, club_slug, invoice_id):
    """Print-optimised view — browser Ctrl+P / Save as PDF."""
    club    = get_object_or_404(Club, slug=club_slug)
    invoice = get_object_or_404(Invoice, club=club, id=invoice_id)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err
    config  = get_config(club)
    return render(request, 'core/invoice_print.html', {
        'club': club, 'invoice': invoice,
        'line_items': invoice.line_items.all(), 'config': config,
    })


@login_required
def reports(request, club_slug):
    from django.db.models import Sum
    from django.db.models.functions import TruncMonth
    import json as _json

    def _sjson(obj):
        """JSON-encode and escape <, >, & so strings can't break out of <script> blocks."""
        return (_json.dumps(obj)
                .replace('&', '\\u0026')
                .replace('<', '\\u003c')
                .replace('>', '\\u003e'))

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    config, _ = ClubConfig.objects.get_or_create(club=club)
    fy_start = config.fy_start_month

    today = timezone.localdate()
    import calendar as _cal
    current_fy_year = today.year if today.month >= fy_start else today.year - 1

    # Selected FY — may be any past year when ?fy= param is set
    try:
        _req_fy = int(request.GET.get('fy', 0))
        fy_year = _req_fy if 2000 <= _req_fy <= current_fy_year else current_fy_year
    except (ValueError, TypeError):
        fy_year = current_fy_year

    fy_start_date = date(fy_year, fy_start, 1)
    fy_end_month = ((fy_start - 2) % 12) + 1
    fy_end_year = fy_year + 1 if fy_end_month < fy_start else fy_year
    fy_end_date = date(fy_end_year, fy_end_month, _cal.monthrange(fy_end_year, fy_end_month)[1])

    # Available FYs for selector dropdown — from first flight data to current
    _first_arr = (FlightCompletion.objects
                  .filter(booking__club=club, booking__arrived_at__isnull=False)
                  .order_by('booking__arrived_at')
                  .values_list('booking__arrived_at', flat=True)
                  .first())
    if _first_arr:
        _first_fy_avail = _first_arr.year if _first_arr.month >= fy_start else _first_arr.year - 1
    else:
        _first_fy_avail = current_fy_year
    available_fys = list(range(_first_fy_avail, current_fy_year + 1))

    leased_filter = request.GET.get('leased', 'all')  # 'all' | 'owned' | 'leased'

    qs = (FlightCompletion.objects
          .filter(booking__club=club,
                  booking__arrived_at__date__gte=fy_start_date,
                  booking__arrived_at__date__lte=fy_end_date,
                  actual_flight_hours__isnull=False)
          .select_related('booking__aircraft'))
    if leased_filter == 'owned':
        qs = qs.filter(booking__aircraft__is_leased=False)
    elif leased_filter == 'leased':
        qs = qs.filter(booking__aircraft__is_leased=True)

    # Build months list for the FY
    months = []
    m, y = fy_start, fy_year
    for _ in range(12):
        months.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1; y += 1
    month_labels = [d.strftime('%b %y') for d in months]

    aircraft_list = (Aircraft.objects.filter(club=club)
                     .exclude(status='retired')
                     .order_by('registration'))

    # hours[aircraft_id][month_index] = hours
    data = {ac.id: [0.0] * 12 for ac in aircraft_list}
    for fc in qs:
        ac_id = fc.booking.aircraft_id
        if ac_id not in data:
            continue
        idx = next((i for i, d in enumerate(months)
                    if d.year == fc.booking.arrived_at.year and d.month == fc.booking.arrived_at.month), None)
        if idx is not None:
            data[ac_id][idx] += float(fc.actual_flight_hours)

    datasets = []
    for i, ac in enumerate(aircraft_list):
        hours = [round(h, 1) for h in data[ac.id]]
        if any(h > 0 for h in hours):
            datasets.append({'label': ac.registration, 'data': hours})

    # ── Instructor hours chart ────────────────────────────────────────────────
    # One bar per instructor, stacked by month — flights where an instructor was assigned.
    instr_qs = (FlightCompletion.objects
                .filter(booking__club=club,
                        booking__arrived_at__date__gte=fy_start_date,
                        booking__arrived_at__date__lte=fy_end_date,
                        booking__instructor__isnull=False,
                        actual_flight_hours__isnull=False)
                .select_related('booking__instructor'))
    from django.contrib.auth import get_user_model as _gum
    _User = _gum()
    instr_ids = list(instr_qs.values_list('booking__instructor_id', flat=True).distinct())
    instr_users = {u.id: u for u in _User.objects.filter(id__in=instr_ids)}
    instr_data = {uid: [0.0] * 12 for uid in instr_ids}
    for fc in instr_qs:
        uid = fc.booking.instructor_id
        idx = next((i for i, d in enumerate(months)
                    if d.year == fc.booking.arrived_at.year and d.month == fc.booking.arrived_at.month), None)
        if idx is not None:
            instr_data[uid][idx] += float(fc.actual_flight_hours)
    instr_datasets = []
    for i, uid in enumerate(instr_ids):
        u = instr_users.get(uid)
        label = u.get_full_name() if u else f'Instructor {uid}'
        hours = [round(h, 1) for h in instr_data[uid]]
        if any(h > 0 for h in hours):
            instr_datasets.append({'label': label, 'data': hours})

    # ── Member activity table ─────────────────────────────────────────────────
    all_qs = (FlightCompletion.objects
              .filter(booking__club=club,
                      booking__arrived_at__date__gte=fy_start_date,
                      booking__arrived_at__date__lte=fy_end_date,
                      actual_flight_hours__isnull=False)
              .select_related('booking__member__user', 'booking__instructor'))

    # Optional month filter — scopes member activity table only
    selected_month_str = request.GET.get('month', '')
    member_filter_label = None
    if selected_month_str:
        try:
            _sm_year, _sm_mon = [int(x) for x in selected_month_str.split('-')]
            _sm_date = date(_sm_year, _sm_mon, 1)
            if fy_start_date <= _sm_date <= fy_end_date:
                _sm_end = date(_sm_year, _sm_mon, _cal.monthrange(_sm_year, _sm_mon)[1])
                member_qs = all_qs.filter(booking__arrived_at__date__gte=_sm_date,
                                          booking__arrived_at__date__lte=_sm_end)
                member_filter_label = _sm_date.strftime('%B %Y')
            else:
                selected_month_str = ''
                member_qs = all_qs
        except (ValueError, AttributeError):
            selected_month_str = ''
            member_qs = all_qs
    else:
        member_qs = all_qs

    member_stats = {}
    for fc in member_qs:
        m = fc.booking.member
        if m not in member_stats:
            member_stats[m] = {'count': 0, 'hours': 0.0, 'solo': 0.0, 'dual': 0.0}
        member_stats[m]['count'] += 1
        _fh = float(fc.actual_flight_hours)
        member_stats[m]['hours'] += _fh
        if fc.booking.instructor_id:
            member_stats[m]['dual'] += _fh
        else:
            member_stats[m]['solo'] += _fh
    member_rows = sorted(
        [{'name': m.user.get_full_name(), 'flight_count': v['count'],
          'total_hours': round(v['hours'], 1),
          'solo_hours': round(v['solo'], 1), 'dual_hours': round(v['dual'], 1)}
         for m, v in member_stats.items()],
        key=lambda r: -r['total_hours']
    )

    # ── Payment summary by month ───────────────────────────────────────────────
    from django.db.models import Sum as _Sum, Count as _Count
    payment_rows = []
    pay_total_count = pay_total_charged = pay_total_collected = 0
    for i, m_date in enumerate(months):
        month_fcs = [fc for fc in all_qs
                     if fc.booking.arrived_at.year == m_date.year
                     and fc.booking.arrived_at.month == m_date.month]
        charged   = sum(float(fc.total_charge) for fc in month_fcs)
        collected = sum(float(fc.amount_paid) for fc in month_fcs)
        if charged or collected:
            payment_rows.append({
                'label': m_date.strftime('%b %y'),
                'count': len(month_fcs),
                'charged':   round(charged, 2),
                'collected': round(collected, 2),
            })
            pay_total_count     += len(month_fcs)
            pay_total_charged   += charged
            pay_total_collected += collected

    # ── Dashboard KPIs ─────────────────────────────────────────────────────────
    from .models import Account as _DAcct
    from django.db.models import Sum as _DSum, Count as _DCnt

    dash_total_hours       = round(sum(float(fc.actual_flight_hours) for fc in all_qs), 1)
    dash_revenue_charged   = round(sum(float(fc.total_charge or 0)   for fc in all_qs), 2)
    dash_revenue_collected = round(sum(float(fc.amount_paid or 0)    for fc in all_qs), 2)
    dash_flights_count     = sum(1 for _ in all_qs)
    dash_solo_count        = sum(1 for fc in all_qs if not fc.booking.instructor_id)
    dash_dual_count        = sum(1 for fc in all_qs if fc.booking.instructor_id)
    active_members_count   = ClubMember.objects.filter(club=club, standing='active').count()

    # Members card extras
    from django.db.models import Count as _MCount
    dash_new_members_count = ClubMember.objects.filter(club=club, join_date__gte=fy_start_date).count()
    _sex_active = ClubMember.objects.filter(club=club, standing='active').exclude(sex='')
    dash_sex_m   = _sex_active.filter(sex='M').count()
    dash_sex_f   = _sex_active.filter(sex='F').count()
    dash_sex_set = dash_sex_m + dash_sex_f
    _cat_qs = (ClubMember.objects
               .filter(club=club, standing='active', membership_category__isnull=False)
               .values('membership_category__name')
               .annotate(cnt=_MCount('id'))
               .order_by('-cnt')[:4])
    dash_top_categories = [{'name': r['membership_category__name'], 'count': r['cnt']} for r in _cat_qs]

    # Trial flights (FY)
    _trial_qs = (FlightCompletion.objects
                 .filter(booking__club=club,
                         booking__arrived_at__date__gte=fy_start_date,
                         booking__arrived_at__date__lte=fy_end_date,
                         booking__flight_type__is_trial=True)
                 .select_related('booking'))
    dash_trial_count = _trial_qs.count()
    _dash_trial_flights = [0] * 12
    for _fc in _trial_qs:
        if _fc.booking.arrived_at:
            _tidx = next((i for i, d in enumerate(months)
                          if d.year == _fc.booking.arrived_at.year
                          and d.month == _fc.booking.arrived_at.month), None)
            if _tidx is not None:
                _dash_trial_flights[_tidx] += 1

    _debt_agg = (_DAcct.objects.filter(club_member__club=club, balance__lt=0)
                 .aggregate(total=_DSum('balance'), cnt=_DCnt('id')))
    dash_outstanding_debt = abs(round(float(_debt_agg['total'] or 0), 2))
    dash_debtor_count     = _debt_agg['cnt'] or 0

    _inv_agg = (Invoice.objects.filter(club=club, status='sent')
                .aggregate(total=_DSum('line_items__amount'), cnt=_DCnt('id', distinct=True)))
    dash_unpaid_invoices       = round(float(_inv_agg['total'] or 0), 2)
    dash_overdue_invoice_count = (Invoice.objects.filter(
        club=club, status='sent', due_date__lt=today).count())

    # Current-month hours KPI
    cm_start = date(today.year, today.month, 1)
    import calendar as _cal3
    cm_end = date(today.year, today.month, _cal3.monthrange(today.year, today.month)[1])
    _cm_qs = (FlightCompletion.objects
              .filter(booking__club=club,
                      booking__arrived_at__date__gte=cm_start,
                      booking__arrived_at__date__lte=cm_end,
                      actual_flight_hours__isnull=False)
              .select_related('booking'))
    dash_month_label = today.strftime('%B %Y')
    dash_month_hours = round(sum(float(fc.actual_flight_hours) for fc in _cm_qs), 1)
    dash_month_solo  = round(sum(float(fc.actual_flight_hours) for fc in _cm_qs if not fc.booking.instructor_id), 1)
    dash_month_dual  = round(sum(float(fc.actual_flight_hours) for fc in _cm_qs if fc.booking.instructor_id), 1)

    # Monthly arrays for dashboard charts (all flights, no leased filter)
    _dash_billed = [0.0] * 12
    _dash_collected = [0.0] * 12
    for fc in all_qs:
        if fc.booking.arrived_at:
            _idx = next((i for i, d in enumerate(months)
                         if d.year == fc.booking.arrived_at.year and d.month == fc.booking.arrived_at.month), None)
            if _idx is not None:
                _dash_billed[_idx]    += float(fc.total_charge or 0)
                _dash_collected[_idx] += float(fc.amount_paid or 0)
    dash_monthly_billed    = _json.dumps([round(v, 0) for v in _dash_billed])
    dash_monthly_collected = _json.dumps([round(v, 0) for v in _dash_collected])

    # Top aircraft/members for dashboard tables
    _ac_hrs_db = (FlightCompletion.objects
                  .filter(booking__club=club,
                          booking__arrived_at__date__gte=fy_start_date,
                          booking__arrived_at__date__lte=fy_end_date,
                          actual_flight_hours__isnull=False)
                  .values('booking__aircraft__registration')
                  .annotate(hrs=_DSum('actual_flight_hours'), cnt=_DCnt('id'))
                  .order_by('-hrs')[:6])
    top_aircraft_dash = [{'reg': r['booking__aircraft__registration'],
                          'hours': round(float(r['hrs']), 1),
                          'count': r['cnt']} for r in _ac_hrs_db]
    top_members_dash  = member_rows[:6]

    # ── Occurrence analytics ──────────────────────────────────────────────────
    from django.db.models import Count as _OCount
    from django.db.models.functions import TruncMonth as _TrMonth

    # All-time totals
    occ_qs = OccurrenceReport.objects.filter(club=club)
    occ_total      = occ_qs.count()
    occ_open       = occ_qs.filter(status=OccurrenceReport.STATUS_SUBMITTED).count()
    occ_reviewed   = occ_qs.filter(status=OccurrenceReport.STATUS_REVIEWED).count()
    occ_closed     = occ_qs.filter(status=OccurrenceReport.STATUS_CLOSED).count()

    # FY-scoped occurrence counts for dashboard KPI
    occ_fy_qs   = occ_qs.filter(date_of_occurrence__gte=fy_start_date,
                                 date_of_occurrence__lte=fy_end_date)
    occ_fy_total = occ_fy_qs.count()
    occ_fy_open  = occ_fy_qs.filter(status=OccurrenceReport.STATUS_SUBMITTED).count()

    # By type (all time)
    by_type = list(
        occ_qs.values('occurrence_type__name')
        .annotate(count=_OCount('id'))
        .order_by('-count')
    )
    occ_type_labels = _sjson([r['occurrence_type__name'] for r in by_type])
    occ_type_data   = _json.dumps([r['count'] for r in by_type])

    # By month (last 12 months, rolling)
    from datetime import timedelta as _td
    twelve_months_ago = today - _td(days=365)
    monthly_qs = (occ_qs
                  .filter(date_of_occurrence__gte=twelve_months_ago)
                  .annotate(m=_TrMonth('date_of_occurrence'))
                  .values('m')
                  .annotate(count=_OCount('id'))
                  .order_by('m'))
    monthly_map = {r['m'].strftime('%b %y'): r['count'] for r in monthly_qs}
    # Build last 12 month labels (oldest first)
    occ_month_labels_list = []
    _my, _mm = today.year, today.month
    for _back in range(11, -1, -1):
        _m2 = _mm - _back
        _y2 = _my
        while _m2 <= 0:
            _m2 += 12; _y2 -= 1
        occ_month_labels_list.append(date(_y2, _m2, 1).strftime('%b %y'))
    occ_month_labels = _json.dumps(occ_month_labels_list)
    occ_month_data   = _json.dumps([monthly_map.get(lbl, 0) for lbl in occ_month_labels_list])

    # By aircraft (top 10, all time)
    by_aircraft = list(
        occ_qs.filter(aircraft__isnull=False)
        .values('aircraft__registration')
        .annotate(count=_OCount('id'))
        .order_by('-count')[:10]
    )
    occ_ac_labels = _sjson([r['aircraft__registration'] for r in by_aircraft])
    occ_ac_data   = _json.dumps([r['count'] for r in by_aircraft])

    # Safety risk events & open actions
    occ_safety_risk_total = occ_qs.filter(is_safety_risk=True).count()
    occ_safety_risk_open  = occ_qs.filter(is_safety_risk=True,
                                          status=OccurrenceReport.STATUS_SUBMITTED).count()
    _action_qs_open = OccurrenceAction.objects.filter(report__club=club,
                                                      status=OccurrenceAction.STATUS_OPEN)
    occ_open_actions_count      = _action_qs_open.count()
    occ_safety_risk_open_actions = _action_qs_open.filter(report__is_safety_risk=True).count()
    # Person with most open actions
    _top_owner_row = (_action_qs_open
                      .filter(assigned_to__isnull=False)
                      .values('assigned_to__user__first_name', 'assigned_to__user__last_name')
                      .annotate(cnt=_OCount('id'))
                      .order_by('-cnt')
                      .first())
    occ_top_owner = None
    if _top_owner_row:
        _fn = _top_owner_row['assigned_to__user__first_name']
        _ln = _top_owner_row['assigned_to__user__last_name']
        occ_top_owner = {'name': f'{_fn} {_ln}'.strip(), 'count': _top_owner_row['cnt']}
    # Monthly actions raised (last 12 months, reuse occ_month_labels_list)
    _action_month_qs = (OccurrenceAction.objects
                        .filter(report__club=club, created_at__date__gte=twelve_months_ago)
                        .annotate(m=_TrMonth('created_at'))
                        .values('m')
                        .annotate(count=_OCount('id'))
                        .order_by('m'))
    _action_monthly_map = {r['m'].strftime('%b %y'): r['count'] for r in _action_month_qs}
    occ_action_month_data = _json.dumps([_action_monthly_map.get(lbl, 0)
                                         for lbl in occ_month_labels_list])

    # ── Budget vs actual (monthly, FY scope) ──────────────────────────────────
    _bfy_year = fy_year
    _budget_qs = FlyingBudget.objects.filter(club=club, fy_year=_bfy_year)
    _budget_by_month = [0.0] * 12
    for b in _budget_qs:
        idx = next((i for i, d in enumerate(months) if d.month == b.month and d.year == (
            _bfy_year if b.month >= fy_start else _bfy_year + 1)), None)
        if idx is not None:
            _budget_by_month[idx] += float(b.budgeted_hours)
    # actual monthly totals (re-use aircraft data already computed)
    _actual_by_month = [round(sum(data[ac.id][i] for ac in aircraft_list if ac.id in data), 1)
                        for i in range(12)]
    # cumulative variance: sum(actual - budget) month by month
    _cum_var = []
    _running = 0.0
    for i in range(12):
        _running += _actual_by_month[i] - _budget_by_month[i]
        _cum_var.append(round(_running, 1))
    has_budget = any(v > 0 for v in _budget_by_month)

    # Current-month budget
    _cm_idx = next((i for i, d in enumerate(months) if d.year == today.year and d.month == today.month), None)
    dash_month_budget = round(_budget_by_month[_cm_idx], 1) if _cm_idx is not None else 0

    # Monthly solo/dual flight counts for dashboard stacked flights chart
    _dash_solo_flights = [0] * 12
    _dash_dual_flights = [0] * 12
    for fc in all_qs:
        if fc.booking.arrived_at:
            _fidx = next((i for i, d in enumerate(months)
                          if d.year == fc.booking.arrived_at.year and d.month == fc.booking.arrived_at.month), None)
            if _fidx is not None:
                if fc.booking.instructor_id:
                    _dash_dual_flights[_fidx] += 1
                else:
                    _dash_solo_flights[_fidx] += 1

    from django.urls import reverse as _rev
    _occ_url     = _rev('core:occurrence_list',  kwargs={'club_slug': club.slug})
    _mbr_url     = _rev('core:manage_members',   kwargs={'club_slug': club.slug})
    _charges_url = _rev('core:manage_charges',   kwargs={'club_slug': club.slug})
    fy_label = f"{fy_start_date.strftime('%b %Y')} – {fy_end_date.strftime('%b %Y')}"

    # ── Historical monthly comparison (last 4 FYs) ────────────────────────────
    _hist_fy_years = [current_fy_year - 3, current_fy_year - 2, current_fy_year - 1, current_fy_year]
    # Months in FY order (e.g. [4,5,6,7,8,9,10,11,12,1,2,3] for Apr start)
    _hist_month_order = [(fy_start + i - 1) % 12 + 1 for i in range(12)]

    def _hist_fy_bounds(yr):
        _s = date(yr, fy_start, 1)
        _em = ((fy_start - 2) % 12) + 1
        _ey = yr + 1 if _em < fy_start else yr
        _e = date(_ey, _em, _cal.monthrange(_ey, _em)[1])
        return _s, _e

    _hist_all_data    = []
    _hist_leased_data = []
    _hist_owned_data  = []
    for _hyr in _hist_fy_years:
        _hs, _he = _hist_fy_bounds(_hyr)
        _h_qs = (FlightCompletion.objects
                 .filter(booking__club=club,
                         booking__arrived_at__date__gte=_hs,
                         booking__arrived_at__date__lte=_he,
                         actual_flight_hours__isnull=False)
                 .select_related('booking__aircraft'))
        _all_m = {m: 0.0 for m in range(1, 13)}
        _lsd_m = {m: 0.0 for m in range(1, 13)}
        _own_m = {m: 0.0 for m in range(1, 13)}
        for _hfc in _h_qs:
            _mo = _hfc.booking.arrived_at.month
            _hh = float(_hfc.actual_flight_hours)
            _all_m[_mo] += _hh
            if _hfc.booking.aircraft.is_leased:
                _lsd_m[_mo] += _hh
            else:
                _own_m[_mo] += _hh
        _hist_all_data.append([round(_all_m[m], 1) for m in _hist_month_order])
        _hist_leased_data.append([round(_lsd_m[m], 1) for m in _hist_month_order])
        _hist_owned_data.append([round(_own_m[m], 1) for m in _hist_month_order])

    # FY label: "FY25" based on end year of each FY
    def _fy_short_label(yr):
        end_yr = yr + 1 if fy_start > 1 else yr
        return f"FY{str(end_yr)[-2:]}"

    hist_month_labels = _json.dumps([date(2000, m, 1).strftime('%b') for m in _hist_month_order])
    hist_fy_labels    = _json.dumps([_fy_short_label(yr) for yr in _hist_fy_years])
    hist_all_data     = _json.dumps(_hist_all_data)
    hist_leased_data  = _json.dumps(_hist_leased_data)
    hist_owned_data   = _json.dumps(_hist_owned_data)

    # ── Annual totals across all FYs with data ────────────────────────────────
    _first_fc = (FlightCompletion.objects
                 .filter(booking__club=club, actual_flight_hours__isnull=False,
                         booking__arrived_at__isnull=False)
                 .order_by('booking__arrived_at').first())
    _first_fy = current_fy_year  # fallback: current FY only
    if _first_fc:
        _fd = _first_fc.booking.arrived_at.date()
        _first_fy = _fd.year if _fd.month >= fy_start else _fd.year - 1

    _all_fy_years = list(range(_first_fy, current_fy_year + 1))

    def _date_to_fy(d):
        return d.year if d.month >= fy_start else d.year - 1

    _ann_hours_all    = {yr: 0.0 for yr in _all_fy_years}
    _ann_hours_leased = {yr: 0.0 for yr in _all_fy_years}
    _ann_type_counts  = {yr: {} for yr in _all_fy_years}

    _ann_qs = (FlightCompletion.objects
               .filter(booking__club=club, actual_flight_hours__isnull=False,
                       booking__arrived_at__isnull=False)
               .select_related('booking__aircraft', 'booking__flight_type'))
    for _afc in _ann_qs:
        _afyr = _date_to_fy(_afc.booking.arrived_at.date())
        if _afyr not in _ann_hours_all:
            continue
        _ah = float(_afc.actual_flight_hours)
        _ann_hours_all[_afyr]    += _ah
        if _afc.booking.aircraft.is_leased:
            _ann_hours_leased[_afyr] += _ah
        _aft_id = _afc.booking.flight_type_id
        if _aft_id is not None:
            _ann_type_counts[_afyr][_aft_id] = _ann_type_counts[_afyr].get(_aft_id, 0) + 1

    _ann_labels    = [_fy_short_label(yr) for yr in _all_fy_years]
    _ann_all_hrs   = [round(_ann_hours_all[yr], 1) for yr in _all_fy_years]
    _ann_leased_hrs= [round(_ann_hours_leased[yr], 1) for yr in _all_fy_years]
    _ann_owned_hrs = [round(_ann_hours_all[yr] - _ann_hours_leased[yr], 1) for yr in _all_fy_years]

    # Flight type stacked chart — all types with any flight data
    _ft_ids_used = set()
    for _yc in _ann_type_counts.values():
        _ft_ids_used.update(_yc.keys())
    _ann_ft_list = list(FlightType.objects.filter(club=club, id__in=_ft_ids_used).order_by('name'))
    _ann_ft_datasets = []
    for _ft in _ann_ft_list:
        _counts = [_ann_type_counts[yr].get(_ft.id, 0) for yr in _all_fy_years]
        if any(c > 0 for c in _counts):
            _ann_ft_datasets.append({'label': _ft.name, 'data': _counts})

    ann_labels      = _json.dumps(_ann_labels)
    ann_all_hrs     = _json.dumps(_ann_all_hrs)
    ann_leased_hrs  = _json.dumps(_ann_leased_hrs)
    ann_owned_hrs   = _json.dumps(_ann_owned_hrs)
    ann_ft_datasets = _sjson(_ann_ft_datasets)

    # FY short-label helper (already defined above as _fy_short_label)
    _fy_months_for_tmpl = [(d.strftime('%Y-%m'), d.strftime('%B %Y')) for d in months]

    return render(request, 'core/reports.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'month_labels': _json.dumps(month_labels),
        'datasets': _sjson(datasets),
        'instr_datasets': _sjson(instr_datasets),
        'leased_filter': leased_filter,
        'fy_label': fy_label,
        'fy_start_date': fy_start_date,
        'fy_year': fy_year,
        'current_fy_year': current_fy_year,
        'available_fys': [(yr, _fy_short_label(yr)) for yr in available_fys],
        'selected_month_str': selected_month_str,
        'member_filter_label': member_filter_label,
        'fy_months': _fy_months_for_tmpl,
        'member_rows': member_rows,
        'occ_total': occ_total,
        'occ_stats': [
            ('Total',    occ_total,    '#1a1f2e', _occ_url),
            ('Open',     occ_open,     '#c0392b', _occ_url + '?status=submitted'),
            ('Reviewed', occ_reviewed, '#216c2a', _occ_url + '?status=reviewed'),
            ('Closed',   occ_closed,   '#8a93a0', _occ_url + '?status=closed'),
        ],
        'occ_type_labels': occ_type_labels, 'occ_type_data': occ_type_data,
        'occ_month_labels': occ_month_labels, 'occ_month_data': occ_month_data,
        'occ_ac_labels': occ_ac_labels, 'occ_ac_data': occ_ac_data,
        # Dashboard KPIs
        'dash_total_hours': dash_total_hours,
        'dash_month_label': dash_month_label,
        'dash_month_hours': dash_month_hours,
        'dash_month_solo':  dash_month_solo,
        'dash_month_dual':  dash_month_dual,
        'dash_month_budget': dash_month_budget,
        'dash_revenue_charged':   dash_revenue_charged,
        'dash_revenue_collected': dash_revenue_collected,
        'dash_flights_count': dash_flights_count,
        'dash_solo_count':    dash_solo_count,
        'dash_dual_count':    dash_dual_count,
        'active_members_count': active_members_count,
        'dash_new_members_count': dash_new_members_count,
        'dash_sex_m': dash_sex_m, 'dash_sex_f': dash_sex_f, 'dash_sex_set': dash_sex_set,
        'dash_top_categories': dash_top_categories,
        'dash_trial_count': dash_trial_count,
        'dash_trial_flights': _json.dumps(_dash_trial_flights),
        'dash_charges_url': _charges_url,
        'dash_outstanding_debt': dash_outstanding_debt,
        'dash_debtor_count':     dash_debtor_count,
        'dash_members_url': _mbr_url,
        'dash_debtor_url': _mbr_url + '?debt=1',
        'dash_unpaid_invoices':       dash_unpaid_invoices,
        'dash_overdue_invoice_count': dash_overdue_invoice_count,
        'dash_monthly_charged':   _json.dumps([round(v, 0) for v in _dash_billed]),
        'dash_monthly_collected': dash_monthly_collected,
        'dash_solo_flights':  _json.dumps(_dash_solo_flights),
        'dash_dual_flights':  _json.dumps(_dash_dual_flights),
        'top_aircraft_dash': top_aircraft_dash,
        'top_members_dash':  top_members_dash,
        'occ_fy_total': occ_fy_total,
        'occ_fy_open':  occ_fy_open,
        'occ_open':     occ_open,
        'occ_open_url': _occ_url + '?status=submitted',
        'occ_open_actions_count':       occ_open_actions_count,
        'occ_safety_risk_total':        occ_safety_risk_total,
        'occ_safety_risk_open':         occ_safety_risk_open,
        'occ_safety_risk_open_actions': occ_safety_risk_open_actions,
        'occ_top_owner':                occ_top_owner,
        'occ_action_month_data':        occ_action_month_data,
        'occ_actions_url': _rev('core:occurrence_actions', kwargs={'club_slug': club.slug}),
        'dash_budget_monthly': _json.dumps(_budget_by_month),
        'dash_actual_monthly': _json.dumps(_actual_by_month),
        'dash_cum_variance':   _json.dumps(_cum_var),
        'dash_has_budget': has_budget,
        'hist_month_labels': hist_month_labels,
        'hist_fy_labels':    hist_fy_labels,
        'hist_all_data':     hist_all_data,
        'hist_leased_data':  hist_leased_data,
        'hist_owned_data':   hist_owned_data,
        'ann_labels':        ann_labels,
        'ann_all_hrs':       ann_all_hrs,
        'ann_leased_hrs':    ann_leased_hrs,
        'ann_owned_hrs':     ann_owned_hrs,
        'ann_ft_datasets':   ann_ft_datasets,
        'ai_suggestions': [
            'Which aircraft flew the most hours this year?',
            'How much revenue did we collect last month?',
            'Who are the most active members this year?',
            'Which instructor has the most hours this year?',
            'Are any aircraft overdue for maintenance?',
            'How many active members do we have?',
        ],
    })


@login_required
def reports_members_xlsx(request, club_slug):
    """Download member activity for the current FY as an Excel file."""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    config = get_config(club)
    today = timezone.localdate()
    fy_start = config.fy_start_month
    fy_year = today.year if today.month >= fy_start else today.year - 1
    fy_start_date = date(fy_year, fy_start, 1)
    import calendar as _xcal
    fy_end_month = ((fy_start - 2) % 12) + 1
    fy_end_year = fy_year + 1 if fy_end_month < fy_start else fy_year
    fy_end_date = date(fy_end_year, fy_end_month, _xcal.monthrange(fy_end_year, fy_end_month)[1])

    all_qs = (FlightCompletion.objects
              .filter(booking__club=club,
                      booking__arrived_at__date__gte=fy_start_date,
                      booking__arrived_at__date__lte=fy_end_date,
                      actual_flight_hours__isnull=False)
              .select_related('booking__member__user'))
    member_stats = {}
    for fc in all_qs:
        m = fc.booking.member
        if m not in member_stats:
            member_stats[m] = {'count': 0, 'hours': 0.0}
        member_stats[m]['count'] += 1
        member_stats[m]['hours'] += float(fc.actual_flight_hours)
    rows = sorted(
        [(m.user.get_full_name(), v['count'], round(v['hours'], 1)) for m, v in member_stats.items()],
        key=lambda r: -r[2]
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Member activity'
    header_fill = PatternFill('solid', fgColor='2B2B2B')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    ws.append(['Member', 'Flights', 'Hours'])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    for name, count, hours in rows:
        ws.append([name, count, hours])

    fy_label = f"{fy_start_date.strftime('%b_%Y')}-{fy_end_date.strftime('%b_%Y')}"
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="member_activity_{fy_label}.xlsx"'
    wb.save(resp)
    return resp


@login_required
def reports_pivot(request, club_slug):
    """AJAX pivot-table builder — returns aggregated flight data as JSON."""
    import json as _json
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Access denied'}, status=403)
    if not (actor.is_admin or actor.is_instructor):
        return JsonResponse({'error': 'Access denied'}, status=403)

    rows   = request.GET.getlist('rows')
    values = request.GET.getlist('values')
    aggs   = request.GET.getlist('aggs')
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()

    VALID_DIMS    = {'aircraft', 'aircraft_type', 'flight_type', 'member',
                     'instructor', 'month', 'quarter', 'year', 'outcome'}
    VALID_METRICS = {'hours', 'charge', 'paid', 'flights', 'outstanding', 'budgeted_hours'}

    if not rows or not values:
        return JsonResponse({'error': 'rows and values required'}, status=400)
    for r in rows:
        if r not in VALID_DIMS:
            return JsonResponse({'error': f'Unknown dimension: {r}'}, status=400)
    for v in values:
        if v not in VALID_METRICS:
            return JsonResponse({'error': f'Unknown metric: {v}'}, status=400)

    qs = (FlightCompletion.objects
          .filter(booking__club=club, actual_flight_hours__isnull=False)
          .select_related('booking__aircraft__aircraft_type',
                          'booking__member__user',
                          'booking__instructor',
                          'booking__flight_type'))
    if date_from_str:
        try:
            qs = qs.filter(booking__arrived_at__date__gte=date_from_str)
        except Exception:
            pass
    if date_to_str:
        try:
            qs = qs.filter(booking__arrived_at__date__lte=date_to_str)
        except Exception:
            pass

    def get_dim(fc, dim):
        b = fc.booking
        if dim == 'aircraft':
            return b.aircraft.registration if b.aircraft_id else '—'
        if dim == 'aircraft_type':
            ac = b.aircraft
            return ac.aircraft_type.name if (ac and ac.aircraft_type_id) else '—'
        if dim == 'flight_type':
            return b.flight_type.name if b.flight_type_id else '—'
        if dim == 'member':
            m = b.member
            return m.user.get_full_name() if m else '—'
        if dim == 'instructor':
            i = b.instructor
            return i.get_full_name() if i else '(solo)'
        if dim == 'month':
            return b.arrived_at.strftime('%b %Y') if b.arrived_at else '—'
        if dim == 'quarter':
            if not b.arrived_at:
                return '—'
            q = ((b.arrived_at.month - 1) // 3) + 1
            return f'Q{q} {b.arrived_at.year}'
        if dim == 'year':
            return str(b.arrived_at.year) if b.arrived_at else '—'
        if dim == 'outcome':
            return fc.get_outcome_display()
        return '—'

    def get_metric(fc, metric):
        if metric == 'hours':
            return float(fc.actual_flight_hours or 0)
        if metric == 'charge':
            return float(fc.total_charge or 0)
        if metric == 'paid':
            return float(fc.amount_paid or 0)
        if metric == 'flights':
            return 1
        if metric == 'outstanding':
            return float(max(0, (fc.total_charge or 0) - (fc.amount_paid or 0)))
        return 0

    flight_values = [v for v in values if v != 'budgeted_hours']
    buckets = {}
    for fc in qs.iterator():
        key = tuple(get_dim(fc, d) for d in rows)
        if key not in buckets:
            buckets[key] = {v: [] for v in flight_values}
        for v in flight_values:
            buckets[key][v].append(get_metric(fc, v))

    # Build budget lookup if needed: (aircraft_reg, month_num, cal_year) → hours
    _budget_lookup = {}
    if 'budgeted_hours' in values:
        _bcfg = ClubConfig.objects.get(club=club)
        for _b in FlyingBudget.objects.filter(club=club).select_related('aircraft'):
            _cal_year = _b.fy_year if _b.month >= _bcfg.fy_start_month else _b.fy_year + 1
            _budget_lookup[(_b.aircraft.registration, _b.month, _cal_year)] = float(_b.budgeted_hours)

    def _resolve_budget(key):
        """Return budgeted hours for a pivot row given its dimension values."""
        from datetime import datetime as _dt2
        ac_reg = None
        month_num = year_num = None
        for i, dim in enumerate(rows):
            if dim == 'aircraft':
                ac_reg = key[i] if key[i] != '—' else None
            elif dim == 'month':
                ms = key[i]
                if ms and ms != '—':
                    try:
                        _parsed = _dt2.strptime(ms, '%b %Y')
                        month_num, year_num = _parsed.month, _parsed.year
                    except ValueError:
                        pass
        if month_num is None:
            return 0
        if ac_reg:
            return round(_budget_lookup.get((ac_reg, month_num, year_num), 0), 1)
        return round(sum(v for (r, m, y), v in _budget_lookup.items()
                         if m == month_num and y == year_num), 1)

    agg_map = {}
    for i, v in enumerate(values):
        if v == 'flights':
            agg_map[v] = 'count'
        elif v == 'budgeted_hours':
            agg_map[v] = 'sum'
        else:
            agg_map[v] = aggs[i] if i < len(aggs) and aggs[i] in ('sum', 'avg', 'count') else 'sum'

    def do_agg(vals, agg):
        if not vals:
            return 0
        if agg == 'avg':
            return round(sum(vals) / len(vals), 2)
        if agg == 'count':
            return int(sum(vals))
        return round(sum(vals), 2)

    def sort_key(item):
        key = item[0]
        from datetime import datetime as _dt
        try:
            parts = key[0].split()
            if len(parts) == 2:
                mn = _dt.strptime(parts[0], '%b').month
                return (int(parts[1]), mn, '')
        except Exception:
            pass
        try:
            parts = key[0].split()
            if len(parts) == 2 and parts[0].startswith('Q'):
                return (int(parts[1]), int(parts[0][1]), '')
        except Exception:
            pass
        return (0, 0, key[0])

    result_rows = []
    for key, mdata in sorted(buckets.items(), key=sort_key):
        vals_out = []
        for v in values:
            if v == 'budgeted_hours':
                vals_out.append(_resolve_budget(key))
            else:
                vals_out.append(do_agg(mdata[v], agg_map[v]))
        result_rows.append(list(key) + vals_out)

    n_dims = len(rows)
    totals = ['Total'] + [''] * (n_dims - 1)
    for i, v in enumerate(values):
        col_vals = [r[n_dims + i] for r in result_rows]
        agg = agg_map[v]
        if agg == 'avg':
            totals.append(round(sum(col_vals) / len(col_vals), 2) if col_vals else 0)
        elif agg == 'count':
            totals.append(int(sum(col_vals)))
        else:
            totals.append(round(sum(col_vals), 2))

    LABELS = {
        'aircraft': 'Aircraft', 'aircraft_type': 'A/C type', 'flight_type': 'Flight type',
        'member': 'Member', 'instructor': 'Instructor',
        'month': 'Month', 'quarter': 'Quarter', 'year': 'Year', 'outcome': 'Outcome',
        'hours': 'Hours', 'charge': 'Charge ($)', 'paid': 'Paid ($)',
        'flights': 'Flights', 'outstanding': 'Outstanding ($)',
        'budgeted_hours': 'Budgeted hrs',
    }
    headers = [LABELS.get(f, f) for f in rows + values]

    return JsonResponse({'headers': headers, 'rows': result_rows,
                         'totals': totals, 'count': len(result_rows)})


@login_required
def ai_ask(request, club_slug):
    """Natural-language query endpoint — fetches club data upfront and passes as context to Groq."""
    import json as _json
    from django.conf import settings as _settings
    from groq import Groq

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Not a member'}, status=403)
    if not actor.is_admin:
        return JsonResponse({'error': 'Admin access required'}, status=403)

    question = request.POST.get('question', '').strip()
    if not question:
        return JsonResponse({'error': 'No question provided'}, status=400)

    api_key = _settings.GROQ_API_KEY
    if not api_key:
        return JsonResponse({'error': 'AI not configured — add GROQ_API_KEY to .env and restart the server.'}, status=503)

    # ── Build club data snapshot ──────────────────────────────────────────────
    # Members
    members = ClubMember.objects.filter(club=club).select_related('role', 'user')
    member_rows = []
    for m in members:
        member_rows.append({
            'name': m.user.get_full_name(),
            'role': m.role.name if m.role else 'No role',
            'standing': m.standing or 'unknown',
            'subscription_expires': str(m.subscription_expires) if m.subscription_expires else None,
        })

    # Completed flights — all time
    fcs = (FlightCompletion.objects
           .filter(booking__club=club, actual_flight_hours__isnull=False)
           .select_related('booking__aircraft', 'booking__instructor', 'booking__member__user')
           .prefetch_related('charge_items')
           .order_by('booking__arrived_at'))
    flight_rows = []
    for fc in fcs:
        b = fc.booking
        charged = sum(float(ci.amount) for ci in fc.charge_items.all())
        flight_rows.append({
            'date': b.arrived_at.strftime('%Y-%m-%d') if b.arrived_at else None,
            'aircraft': b.aircraft.registration if b.aircraft else None,
            'member': b.member.user.get_full_name() if b.member else None,
            'instructor': b.instructor.get_full_name() if b.instructor else None,
            'hours': round(float(fc.actual_flight_hours), 1),
            'charged': round(charged, 2),
            'paid': round(float(fc.amount_paid or 0), 2),
        })

    # Aircraft
    aircraft_rows = []
    for ac in Aircraft.objects.filter(club=club).exclude(status='retired').prefetch_related('maintenance_items'):
        items = list(ac.maintenance_items.all())
        aircraft_rows.append({
            'registration': ac.registration,
            'type': ac.aircraft_type.name if ac.aircraft_type else 'Unknown',
            'status': ac.status,
            'is_leased': ac.is_leased,
            'overdue_maintenance': [m.name for m in items if getattr(m, 'urgency', '') == 'red'],
            'warning_maintenance': [m.name for m in items if getattr(m, 'urgency', '') == 'amber'],
        })

    # Account balances
    from .models import Account as _Acct
    account_rows = []
    for acc in _Acct.objects.filter(club_member__club=club).select_related('club_member__user'):
        account_rows.append({
            'member': acc.club_member.user.get_full_name(),
            'balance': round(float(acc.balance), 2),
            'credit_limit': float(acc.credit_limit) if acc.credit_limit is not None else None,
        })

    # Occurrences (last 2 years)
    from .models import OccurrenceReport as _OR
    occ_rows = []
    for o in _OR.objects.filter(club=club).select_related('occurrence_type', 'reported_by__user').order_by('-date_of_occurrence')[:100]:
        occ_rows.append({
            'date': str(o.date_of_occurrence),
            'type': o.occurrence_type.name,
            'status': o.status,
            'is_safety_risk': o.is_safety_risk,
            'reported_by': o.reported_by.user.get_full_name() if o.reported_by else None,
        })

    # Budget vs actual (current FY)
    _ai_cfg = ClubConfig.objects.get(club=club)
    _ai_today = timezone.localdate()
    _ai_fy = _ai_today.year if _ai_today.month >= _ai_cfg.fy_start_month else _ai_today.year - 1
    _ai_budgets = list(FlyingBudget.objects.filter(club=club, fy_year=_ai_fy).select_related('aircraft'))
    budget_rows = [
        {'aircraft': b.aircraft.registration, 'month': b.month, 'fy_year': b.fy_year,
         'budgeted_hours': float(b.budgeted_hours)}
        for b in _ai_budgets
    ]

    data_context = _json.dumps({
        'members': member_rows,
        'completed_flights': flight_rows,
        'aircraft': aircraft_rows,
        'account_balances': account_rows,
        'occurrences': occ_rows,
        'flying_budget': {
            'fy_year': _ai_fy,
            'fy_start_month': _ai_cfg.fy_start_month,
            'entries': budget_rows,
        },
    }, default=str)

    # ── Groq call ─────────────────────────────────────────────────────────────
    try:
        client = Groq(api_key=api_key)
        system = (
            f"You are a helpful data assistant for {club.name}, an aviation club. "
            f"Today is {timezone.localdate().strftime('%d %B %Y')}. "
            "The following JSON contains club data — members, completed flights, aircraft, "
            "account balances, and safety occurrences. "
            "Answer questions using only this data. Be concise. "
            "Format currency as $X,XXX.XX and hours as X.X hrs.\n\n"
            f"CLUB DATA:\n{data_context}"
        )
        resp = client.chat.completions.create(
            model='llama-3.3-70b-versatile',
            messages=[
                {'role': 'system', 'content': system},
                {'role': 'user',   'content': question},
            ],
            max_tokens=1024,
        )
        return JsonResponse({'answer': resp.choices[0].message.content})
    except Exception as exc:
        return JsonResponse({'error': f'AI error: {exc}'}, status=500)


@login_required
def notifications(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    # Auto-prune: delete notifications older than 90 days (keeps the list bounded)
    from django.utils import timezone as _tz
    cutoff = _tz.now() - __import__('datetime').timedelta(days=90)
    member.notifications.filter(created_at__lt=cutoff).delete()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'mark_read':
            member.notifications.filter(id=request.POST.get('id')).update(is_read=True)
        elif action == 'mark_all_read':
            member.notifications.filter(is_read=False).update(is_read=True)
        elif action == 'delete':
            member.notifications.filter(id=request.POST.get('id')).delete()
        elif action == 'delete_all_read':
            member.notifications.filter(is_read=True).delete()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'ok': True})
        _profile_url = redirect('core:notifications', club_slug=club_slug).url
        return redirect(_profile_url + '?saved=1')

    tab = request.GET.get('tab', 'unread')
    qs = member.notifications.all()
    if tab == 'unread':
        qs = qs.filter(is_read=False)
    items = list(qs[:100])
    unread_count = member.notifications.filter(is_read=False).count()
    return render(request, 'core/notifications.html', {
        'club': club, 'club_member': member, 'is_instructor': member.is_instructor,
        'notifications': items,
        'tab': tab,
        'unread_count': unread_count,
    })


@login_required
def manage_invoices(request, club_slug):
    """List all invoices with overdue indicators."""
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request): return err

    f_status = request.GET.get('status', '')
    f_member = request.GET.get('member', '')

    from urllib.parse import urlencode as _ue
    from django.core.paginator import Paginator as _Pag

    sort     = request.GET.get('sort', 'number')
    sort_dir = request.GET.get('dir', 'desc')
    _INV_SORT = {
        'number': ('invoice_number',),
        'member': ('member__user__last_name', 'member__user__first_name'),
        'issued': ('issue_date', 'invoice_number'),
        'due':    ('due_date', 'invoice_number'),
        'amount': ('total', 'invoice_number'),
        'status': ('status', '-invoice_number'),
    }
    if sort not in _INV_SORT:
        sort = 'number'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'desc'
    _inv_order = _INV_SORT[sort]
    if sort_dir == 'desc':
        _inv_order = tuple(f[1:] if f.startswith('-') else '-'+f for f in _inv_order)

    qs = (Invoice.objects.filter(club=club)
          .select_related('member__user', 'flight_completion__booking')
          .prefetch_related('line_items')
          .order_by(*_inv_order))

    today = timezone.localdate()

    if f_status == 'overdue':
        qs = qs.filter(status='sent', due_date__lt=today)
    elif f_status:
        qs = qs.filter(status=f_status)
    if f_member:
        qs = qs.filter(member__user_id=f_member)

    members_qs = ClubMember.objects.filter(club=club).select_related('user').order_by('user__last_name')
    overdue_count = Invoice.objects.filter(club=club, status='sent', due_date__lt=today).count()
    _paginator  = _Pag(qs, 50)
    invoices_page = _paginator.get_page(request.GET.get('page'))
    _filter_qs  = _ue({k: v for k, v in request.GET.items() if k != 'page' and v})
    _base_qs    = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir') and v})

    return render(request, 'core/manage_invoices.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'invoices_page': invoices_page, 'f_status': f_status, 'f_member': f_member,
        'members_qs': members_qs, 'overdue_count': overdue_count,
        'status_choices': Invoice.STATUS_CHOICES,
        'sort': sort, 'sort_dir': sort_dir, 'base_qs': _base_qs,
        'filter_qs': _filter_qs, 'total_count': _paginator.count,
    })


def _instructor_cred_issues(club, on_date):
    """
    Return instructors with active (non-cancelled, non-completed) bookings on
    `on_date` who have at least one critical lapsed credential:
    instructor certificate, medical, or flight review.

    Returns list of dicts: {cm, lapsed: [str], bookings: [Booking]}
    """
    from .models import MemberCredential, ClubConfig

    INSTR_CERT_CODES = {'instr_a', 'instr_b', 'instr_c', 'instr_d', 'instr_e', 'examiner'}
    MEDICAL_CODES    = {'medical_c1', 'medical_c2', 'medical_c3', 'dlr9'}

    instructor_ids = list(
        Booking.objects
        .filter(club=club, scheduled_start__date=on_date, instructor__isnull=False)
        .exclude(status__in=['cancelled', 'completed'])
        .values_list('instructor_id', flat=True)
        .distinct()
    )
    if not instructor_ids:
        return []

    config, _ = ClubConfig.objects.get_or_create(club=club)

    instructors = (
        ClubMember.objects
        .filter(club=club, user_id__in=instructor_ids, is_on_instructor_roster=True)
        .select_related('user')
        .prefetch_related('user__credentials__credential_type')
    )

    issues = []
    for instr in instructors:
        creds = list(instr.user.credentials.all())
        lapsed = []

        # Instructor certificate
        ic = [c for c in creds if c.credential_type.code in INSTR_CERT_CODES]
        if not ic:
            lapsed.append('No instructor certificate recorded')
        else:
            current = [c for c in ic if not c.expiry_date or c.expiry_date >= on_date]
            if not current:
                latest = max(ic, key=lambda c: c.expiry_date or date.min)
                lapsed.append(f'Instructor certificate expired {latest.expiry_date:%d %b %Y}')

        # Medical
        med = [c for c in creds if c.credential_type.code in MEDICAL_CODES]
        if not med:
            lapsed.append('No medical certificate recorded')
        else:
            current = [c for c in med if not c.expiry_date or c.expiry_date >= on_date]
            if not current:
                latest = max(med, key=lambda c: c.expiry_date or date.min)
                lapsed.append(f'Medical expired {latest.expiry_date:%d %b %Y}')

        # Flight review
        fr = [c for c in creds if c.credential_type.code == 'fr']
        if not fr:
            lapsed.append('No Flight Review recorded')
        else:
            current = []
            for c in fr:
                if c.expiry_date:
                    if c.expiry_date >= on_date:
                        current.append(c)
                elif c.issue_date:
                    due = c.issue_date + timedelta(days=config.bfr_interval_months * 30)
                    if due >= on_date:
                        current.append(c)
                else:
                    current.append(c)  # no dates at all — can't determine, don't flag
            if not current:
                lapsed.append('Flight Review overdue')

        if lapsed:
            day_bookings = list(
                Booking.objects
                .filter(club=club, instructor=instr.user, scheduled_start__date=on_date)
                .exclude(status__in=['cancelled', 'completed'])
                .select_related('aircraft', 'member__user')
                .order_by('scheduled_start')
            )
            issues.append({'cm': instr, 'lapsed': lapsed, 'bookings': day_bookings})

    return issues


@login_required
def manage_exceptions(request, club_slug):
    from django.db.models import Q as _Q
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err

    today = timezone.localdate()

    # 1. Unpaid completed flights
    unpaid_flights = (
        Booking.objects
        .filter(club=club, status='completed',
                flight_completion__paid_at__isnull=True,
                flight_completion__total_charge__gt=0,
                flight_completion__invoices__isnull=True)
        .select_related('member__user', 'aircraft', 'flight_type', 'flight_completion')
        .order_by('arrived_at')
    )

    # Pre-load instructor availability for off-roster detection
    from .models import InstructorAvailability as _IA
    _roster_uids = set(
        ClubMember.objects.filter(club=club, is_on_instructor_roster=True).values_list('user_id', flat=True)
    )
    _av_wins = {}  # user_id -> list[InstructorAvailability]
    for av in _IA.objects.filter(club_member__club=club).select_related('club_member'):
        _av_wins.setdefault(av.club_member.user_id, []).append(av)

    def _instr_off_roster(booking):
        uid = booking.instructor_id
        if not uid or uid not in _roster_uids:
            return False
        windows = _av_wins.get(uid, [])
        if not windows:
            return True  # no schedule = not available
        return not any(w.applies_on(booking.scheduled_start.date()) for w in windows)

    # 2. Booking conflicts (future non-cancelled bookings with a flagged issue)
    from django.db.models import Exists as _Exists, OuterRef as _OuterRef
    _ac_clash_sub = Booking.objects.filter(
        club=club, aircraft_id=_OuterRef('aircraft_id'),
        status__in=['pending', 'confirmed', 'departed'],
        scheduled_start__lt=_OuterRef('scheduled_end'),
        scheduled_end__gt=_OuterRef('scheduled_start'),
    ).exclude(pk=_OuterRef('pk'))
    _in_clash_sub = Booking.objects.filter(
        club=club, instructor_id=_OuterRef('instructor_id'),
        status__in=['pending', 'confirmed', 'departed'],
        scheduled_start__lt=_OuterRef('scheduled_end'),
        scheduled_end__gt=_OuterRef('scheduled_start'),
    ).exclude(pk=_OuterRef('pk'))
    _future_bks = (Booking.objects
        .filter(club=club, scheduled_start__date__gte=today)
        .exclude(status__in=['cancelled', 'completed']))
    _ac_clash_ids = set(
        _future_bks.filter(aircraft__isnull=False)
        .annotate(_cl=_Exists(_ac_clash_sub)).filter(_cl=True)
        .values_list('id', flat=True)
    )
    _in_clash_ids = set(
        _future_bks.filter(instructor__isnull=False)
        .annotate(_cl=_Exists(_in_clash_sub)).filter(_cl=True)
        .values_list('id', flat=True)
    )
    _clashing_ids = _ac_clash_ids | _in_clash_ids

    _db_conflicts = list(
        Booking.objects
        .filter(club=club, scheduled_start__date__gte=today)
        .exclude(status__in=['cancelled', 'completed'])
        .filter(
            _Q(blockout_conflict=True) |
            _Q(member__standing__in=['suspended', 'lapsed', 'resigned']) |
            _Q(member__standing='active',
               member__subscription_expires__isnull=False,
               member__subscription_expires__lt=today) |
            _Q(aircraft__status='retired') |
            _Q(id__in=_clashing_ids)
        )
        .select_related('member__user', 'aircraft', 'flight_type', 'instructor')
        .order_by('scheduled_start')
    )
    _seen_ids = {b.id for b in _db_conflicts}
    # Add instructor-off-roster bookings not caught by the DB query
    _instr_conflicts = [
        b for b in (
            Booking.objects
            .filter(club=club, scheduled_start__date__gte=today, instructor__isnull=False)
            .exclude(status__in=['cancelled', 'completed'])
            .exclude(id__in=_seen_ids)
            .select_related('member__user', 'aircraft', 'flight_type', 'instructor')
            .order_by('scheduled_start')
        )
        if _instr_off_roster(b)
    ]
    conflicts = sorted(_db_conflicts + _instr_conflicts, key=lambda b: b.scheduled_start)

    def _conflict_labels(b):
        labels = []
        if b.id in _ac_clash_ids:
            labels.append('Aircraft double-booked')
        if b.id in _in_clash_ids:
            labels.append('Instructor double-booked')
        if b.blockout_conflict:
            labels.append(b.blockout_conflict_reason or 'Block-out conflict')
        if b.member:
            if b.member.standing in ('suspended', 'lapsed', 'resigned'):
                labels.append(f'Member {b.member.get_standing_display()}')
            elif (b.member.standing == 'active' and b.member.subscription_expires
                  and b.member.subscription_expires < today):
                labels.append('Subscription expired')
        if b.aircraft and b.aircraft.status == 'retired':
            labels.append('Aircraft retired')
        if _instr_off_roster(b):
            labels.append('Instructor off roster')
        return labels

    conflicts_data = [{'b': b, 'labels': _conflict_labels(b)} for b in conflicts]

    # 3. Lapsed/expired credentials — members with upcoming bookings whose
    #    credentials have expired (any credential with an expiry_date in the past)
    from .models import MemberCredential
    future_member_ids = (
        Booking.objects
        .filter(club=club, scheduled_start__date__gte=today)
        .exclude(status__in=['cancelled', 'completed'])
        .values_list('member_id', flat=True)
        .distinct()
    )
    future_user_ids = (
        ClubMember.objects
        .filter(id__in=future_member_ids)
        .values_list('user_id', flat=True)
    )
    expired_cred_user_ids = (
        MemberCredential.objects
        .filter(member_id__in=future_user_ids, expiry_date__lt=today)
        .values_list('member_id', flat=True)
        .distinct()
    )
    members_lapsed_creds = (
        ClubMember.objects
        .filter(club=club, user_id__in=expired_cred_user_ids)
        .select_related('user')
        .order_by('user__last_name')
    )
    lapsed_creds_data = []
    for cm in members_lapsed_creds:
        expired = list(
            MemberCredential.objects
            .filter(member=cm.user, expiry_date__lt=today)
            .select_related('credential_type')
        )
        upcoming = (Booking.objects
                    .filter(club=club, member=cm, scheduled_start__date__gte=today)
                    .exclude(status__in=['cancelled', 'completed'])
                    .order_by('scheduled_start').first())
        lapsed_creds_data.append({'cm': cm, 'expired': expired, 'next_booking': upcoming})

    # 4. Maintenance — amber and red items on online aircraft
    from .models import AircraftMaintenanceItem, MaintenanceUrgency
    from django.db.models import Case, When, IntegerField as _IntF
    maint_items = (
        AircraftMaintenanceItem.objects
        .filter(aircraft__club=club, aircraft__status='online',
                urgency__in=[MaintenanceUrgency.AMBER, MaintenanceUrgency.RED])
        .select_related('aircraft')
        .order_by(
            Case(
                When(urgency=MaintenanceUrgency.RED, then=0),
                When(urgency=MaintenanceUrgency.AMBER, then=1),
                default=2, output_field=_IntF()
            ),
            'aircraft__registration', 'name'
        )
    )

    # 5. Unpaid invoices — draft or sent (not yet paid)
    from .models import Invoice as _Inv
    unpaid_invoices = (
        _Inv.objects
        .filter(club=club, status__in=(_Inv.STATUS_DRAFT, _Inv.STATUS_SENT))
        .select_related('member__user', 'flight_completion__booking__aircraft')
        .order_by('due_date', 'invoice_number')
    )

    # Overdue returns — departed >24h with no check-in (potential safety issue)
    _overdue_cutoff = timezone.now() - timedelta(hours=24)
    overdue_departures = list(
        Booking.objects
        .filter(club=club, status='departed', departed_at__lt=_overdue_cutoff)
        .select_related('member__user', 'aircraft')
        .order_by('departed_at')
    )

    # 6. Instructor credential issues for today
    instructor_cred_issues = _instructor_cred_issues(club, today)

    # ── Sorting for each Attention tab ────────────────────────────────────
    from urllib.parse import urlencode as _ue
    _exc_sort_keys = ('uf_sort','uf_dir','cf_sort','cf_dir','ui_sort','ui_dir',
                      'cr_sort','cr_dir','mt_sort','mt_dir','tab')
    exc_base_qs = _ue({k: v for k, v in request.GET.items()
                       if k not in _exc_sort_keys and v})

    # Unpaid flights
    uf_sort = request.GET.get('uf_sort', 'date')
    uf_dir  = request.GET.get('uf_dir',  'asc')
    _UF_SORT = {
        'member':   ('member__user__last_name', 'member__user__first_name'),
        'aircraft': ('aircraft__registration',),
        'type':     ('flight_type__name',),
        'date':     ('arrived_at',),
        'total':    ('flight_completion__total_charge',),
    }
    if uf_sort not in _UF_SORT: uf_sort = 'date'
    if uf_dir not in ('asc', 'desc'): uf_dir = 'asc'
    _uf_order = _UF_SORT[uf_sort]
    if uf_dir == 'desc': _uf_order = tuple('-'+f for f in _uf_order)
    unpaid_flights = list(unpaid_flights.order_by(*_uf_order))

    # Booking conflicts
    cf_sort = request.GET.get('cf_sort', 'date')
    cf_dir  = request.GET.get('cf_dir',  'asc')
    _CF_KEYS = {
        'member':   lambda d: d['b'].member.user.last_name.lower() if d['b'].member and d['b'].member.user else '',
        'aircraft': lambda d: d['b'].aircraft.registration if d['b'].aircraft else '',
        'date':     lambda d: d['b'].scheduled_start,
        'status':   lambda d: d['b'].status,
    }
    if cf_sort not in _CF_KEYS: cf_sort = 'date'
    if cf_dir not in ('asc', 'desc'): cf_dir = 'asc'
    conflicts_data.sort(key=_CF_KEYS[cf_sort], reverse=(cf_dir == 'desc'))

    # Unpaid invoices
    ui_sort = request.GET.get('ui_sort', 'due')
    ui_dir  = request.GET.get('ui_dir',  'asc')
    _UI_SORT = {
        'number':  ('invoice_number',),
        'member':  ('member__user__last_name', 'member__user__first_name'),
        'issued':  ('issue_date',),
        'due':     ('due_date',),
        'total':   ('total',),
        'status':  ('status',),
    }
    if ui_sort not in _UI_SORT: ui_sort = 'due'
    if ui_dir not in ('asc', 'desc'): ui_dir = 'asc'
    _ui_order = _UI_SORT[ui_sort]
    if ui_dir == 'desc': _ui_order = tuple('-'+f for f in _ui_order)
    unpaid_invoices = list(unpaid_invoices.order_by(*_ui_order))

    # Credentials (both sub-tables share cr_sort/cr_dir)
    cr_sort = request.GET.get('cr_sort', 'member')
    cr_dir  = request.GET.get('cr_dir',  'asc')
    if cr_sort not in ('member', 'next_booking'): cr_sort = 'member'
    if cr_dir not in ('asc', 'desc'): cr_dir = 'asc'
    from datetime import date as _date
    _CR_KEYS = {
        'member':       lambda d: d['cm'].user.last_name.lower() if d['cm'].user else '',
        'next_booking': lambda d: d['next_booking'].scheduled_start if d['next_booking'] else _date.max,
    }
    lapsed_creds_data.sort(key=_CR_KEYS[cr_sort], reverse=(cr_dir == 'desc'))
    instructor_cred_issues.sort(
        key=lambda d: d['cm'].user.last_name.lower() if d['cm'].user else '',
        reverse=(cr_dir == 'desc')
    )

    # Maintenance
    mt_sort = request.GET.get('mt_sort', 'urgency')
    mt_dir  = request.GET.get('mt_dir',  'asc')
    _URGENCY_ORDER = {'red': 0, 'amber': 1, 'green': 2}
    _MT_KEYS = {
        'aircraft': lambda d: d.aircraft.registration,
        'name':     lambda d: d.name.lower(),
        'urgency':  lambda d: _URGENCY_ORDER.get(d.urgency, 9),
        'due_hours':lambda d: d.due_hours or 999999,
        'due_date': lambda d: d.due_date or _date.max,
    }
    if mt_sort not in _MT_KEYS: mt_sort = 'urgency'
    if mt_dir not in ('asc', 'desc'): mt_dir = 'asc'
    maint_list = sorted(list(maint_items), key=_MT_KEYS[mt_sort], reverse=(mt_dir == 'desc'))

    total_issues = (
        len(overdue_departures) +
        len(unpaid_flights) + len(conflicts_data) +
        len(lapsed_creds_data) + len(maint_list) +
        len(unpaid_invoices) +
        len(instructor_cred_issues)
    )

    cred_count = len(lapsed_creds_data) + len(instructor_cred_issues)

    return render(request, 'core/manage_exceptions.html', {
        'club': club,
        'club_member': actor,
        'unpaid_flights': unpaid_flights,
        'conflicts_data': conflicts_data,
        'lapsed_creds_data': lapsed_creds_data,
        'maint_items': maint_list,
        'unpaid_invoices': unpaid_invoices,
        'overdue_departures': overdue_departures,
        'instructor_cred_issues': instructor_cred_issues,
        'cred_count': cred_count,
        'total_issues': total_issues,
        'today': today,
        'exc_base_qs': exc_base_qs,
        'uf_sort': uf_sort, 'uf_dir': uf_dir,
        'cf_sort': cf_sort, 'cf_dir': cf_dir,
        'ui_sort': ui_sort, 'ui_dir': ui_dir,
        'cr_sort': cr_sort, 'cr_dir': cr_dir,
        'mt_sort': mt_sort, 'mt_dir': mt_dir,
    })


@login_required
def manage_vouchers(request, club_slug):
    from .models import Voucher, Account, AccountTransaction
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'create_voucher':
            code = request.POST.get('code', '').strip().upper()
            value = request.POST.get('value', '').strip()
            desc = request.POST.get('description', '').strip()
            notes = request.POST.get('notes', '').strip()
            if code and value:
                try:
                    Voucher.objects.create(
                        club=club, code=code, value=value,
                        description=desc, notes=notes, created_by=request.user,
                    )
                    from django.contrib import messages as _messages
                    _messages.success(request, f'Voucher {code} created (${value}).')
                except Exception:
                    from django.contrib import messages as _messages
                    _messages.info(request, f'Voucher code {code} already exists.')

        elif action == 'redeem_voucher':
            voucher_id = request.POST.get('voucher_id')
            member_id  = request.POST.get('member_id')
            v = Voucher.objects.filter(club=club, id=voucher_id, is_redeemed=False).first()
            member = ClubMember.objects.filter(club=club, id=member_id).first()
            if v and member:
                from django.utils import timezone as _tz
                from django.contrib import messages as _messages
                acct, _ = Account.objects.get_or_create(
                    club_member=member, defaults={'balance': 0}
                )
                AccountTransaction.objects.create(
                    account=acct,
                    transaction_type='top_up',
                    direction='credit',
                    amount=v.value,
                    description=f'Voucher {v.code} redeemed — {v.description or "credit"}',
                    payment_method='other',
                    created_by=request.user,
                )
                from django.db.models import Sum as _Sum
                acct.balance = acct.transactions.filter(direction='credit').aggregate(
                    s=_Sum('amount'))['s'] or 0
                acct.balance -= acct.transactions.filter(direction='debit').aggregate(
                    s=_Sum('amount'))['s'] or 0
                acct.save(update_fields=['balance'])
                v.is_redeemed = True
                v.redeemed_by = member
                v.redeemed_at = _tz.now()
                v.save()
                _messages.success(request, f'Voucher {v.code} redeemed — ${v.value} credited to {member.user.get_full_name()}.')

        return redirect('core:manage_vouchers', club_slug=club_slug)

    f_show = request.GET.get('show', 'pending')  # pending | redeemed | all
    qs = Voucher.objects.filter(club=club).select_related('redeemed_by__user', 'created_by')
    if f_show == 'pending':
        qs = qs.filter(is_redeemed=False)
    elif f_show == 'redeemed':
        qs = qs.filter(is_redeemed=True)
    members = ClubMember.objects.filter(club=club).select_related('user').order_by('user__last_name')

    return render(request, 'core/manage_vouchers.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'vouchers': qs, 'f_show': f_show, 'members': members,
        'voucher_types': VoucherType.objects.filter(club=club, is_active=True),
    })


@login_required
def manage_sundry(request, club_slug):
    from .models import Account, AccountTransaction as _AT
    from decimal import Decimal as _D
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err

    error = None
    if request.method == 'POST':
        member_id  = request.POST.get('member_id', '').strip()
        description = request.POST.get('description', '').strip()
        amount_str  = request.POST.get('amount', '').strip()
        reference   = request.POST.get('reference', '').strip()
        member = ClubMember.objects.filter(club=club, id=member_id).select_related('user').first()
        try:
            amount = _D(amount_str)
            if amount <= 0:
                raise ValueError
        except (ValueError, Exception):
            amount = None
        if not member:
            error = 'Select a member.'
        elif not description:
            error = 'Description is required.'
        elif not amount:
            error = 'Enter a positive amount.'
        else:
            account, _ = Account.objects.get_or_create(club_member=member)
            _AT.objects.create(
                account=account,
                transaction_type='sale',
                direction='debit',
                amount=amount,
                description=description,
                reference=reference,
                created_by=request.user,
            )
            account.balance = account.balance - amount
            account.save(update_fields=['balance'])
            return redirect(f'{request.path}?saved=1')

    members = (ClubMember.objects
               .filter(club=club)
               .exclude(standing='resigned')
               .select_related('user')
               .order_by('user__last_name', 'user__first_name'))

    from urllib.parse import urlencode as _ue
    q_sales  = request.GET.get('q', '').strip()
    sort     = request.GET.get('sort', 'date')
    sort_dir = request.GET.get('dir', 'desc')
    _SD_SORT = {
        'member':      ('account__club_member__user__last_name', 'account__club_member__user__first_name'),
        'description': ('description', '-created_at'),
        'date':        ('created_at',),
        'amount':      ('amount', '-created_at'),
    }
    if sort not in _SD_SORT:
        sort = 'date'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'desc'
    _sd_order = _SD_SORT[sort]
    if sort_dir == 'desc':
        _sd_order = tuple(f[1:] if f.startswith('-') else '-'+f for f in _sd_order)
    sales_qs = (_AT.objects
                .filter(account__club_member__club=club, transaction_type='sale')
                .select_related('account__club_member__user', 'created_by')
                .order_by(*_sd_order))
    if q_sales:
        from django.db.models import Q as _Q
        sales_qs = sales_qs.filter(
            _Q(account__club_member__user__first_name__icontains=q_sales) |
            _Q(account__club_member__user__last_name__icontains=q_sales) |
            _Q(description__icontains=q_sales) |
            _Q(reference__icontains=q_sales)
        )
    from django.core.paginator import Paginator
    sales_page = Paginator(sales_qs, 30).get_page(request.GET.get('page'))
    _base_qs = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir') and v})

    return render(request, 'core/manage_sundry.html', {
        'club': club, 'club_member': actor,
        'members': members,
        'sales_page': sales_page,
        'q_sales': q_sales,
        'sort': sort, 'sort_dir': sort_dir, 'base_qs': _base_qs,
        'modal_error': error,
        'modal_error_id': 'new-sale-modal' if error else None,
        'modal_error_values': {
            'member_id': request.POST.get('member_id', ''),
            'description': request.POST.get('description', ''),
            'amount': request.POST.get('amount', ''),
            'reference': request.POST.get('reference', ''),
        } if error else {},
    })


@login_required
def health_check(request, club_slug):
    from decimal import Decimal as _D
    from django.db.models import Sum, F
    from django.utils import timezone as _tz

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    # ── Fix-drift actions ─────────────────────────────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'fix_balance_drift':
        from .models import Account as _Acct
        fixed = 0
        for _acc in _Acct.objects.filter(club_member__club=club):
            computed = _acc.recompute_balance()
            if abs(computed - _acc.balance) > _D('0.01'):
                _acc.balance = computed
                _acc.save(update_fields=['balance', 'updated_at'])
                fixed += 1
        return redirect(request.path + '?fixed=' + str(fixed))

    if request.method == 'POST' and request.POST.get('action') == 'fix_payment_drift':
        from .models import FlightPayment as _FP2
        fixed = 0
        for _fc in FlightCompletion.objects.filter(booking__club=club):
            _computed = (_FP2.objects.filter(completion=_fc, paid_at__isnull=False)
                         .aggregate(t=Sum('amount'))['t'] or _D('0'))
            if abs(_computed - _fc.amount_paid) > _D('0.01'):
                _fc._sync_payment_cache()
                fixed += 1
        return redirect(request.path + '?fixed=' + str(fixed))

    if request.method == 'POST' and request.POST.get('action') == 'mark_lapsed':
        _today = timezone.localdate()
        lapsed_count = ClubMember.objects.filter(
            club=club, standing='active',
            subscription_expires__isnull=False,
            subscription_expires__lt=_today,
        ).update(standing='lapsed')
        return redirect(request.path + '?lapsed=' + str(lapsed_count))

    from django.urls import reverse as _rev

    # group -> {label, description, issues: [{severity, message, rows: [{text, url}]}], ok: [str]}
    groups = {
        'financial': {
            'label': 'Financial',
            'description': (
                'Checks that stored totals match their underlying ledger entries. '
                'A drift means a balance or charge figure shown in the system does not match '
                'the sum of its transaction records — the transactions are always the source of truth.'
            ),
            'issues': [], 'ok': [],
        },
        'operations': {
            'label': 'Operations',
            'description': (
                'Checks for gaps or inconsistencies in aircraft Hobbs meter readings between '
                'consecutive flights. A positive gap means unaccounted flight time; '
                'a negative gap means overlapping or out-of-order entries.'
            ),
            'issues': [], 'ok': [],
        },
        'members': {
            'label': 'Members',
            'description': (
                'Checks membership standing against current subscription records. '
                'Active members with expired subscriptions may need to renew or have their standing updated.'
            ),
            'issues': [], 'ok': [],
        },
    }

    def _issue(group, sev, msg, rows=None):
        groups[group]['issues'].append({'severity': sev, 'message': msg, 'rows': rows or []})

    def _ok(group, name):
        groups[group]['ok'].append(name)

    # ── 1. Account balance drift ──────────────────────────────────────────────
    from .models import Account
    balance_drifts = []
    for acc in Account.objects.filter(club_member__club=club).select_related('club_member__user'):
        computed = acc.recompute_balance()
        if abs(_D(str(computed)) - acc.balance) > _D('0.01'):
            balance_drifts.append({
                'text': f"{acc.club_member.user.get_full_name()}: stored ${acc.balance}, computed ${computed:.2f}",
                'url': _rev('core:manage_member_detail', kwargs={'club_slug': club_slug, 'member_id': acc.club_member.id}) + '#account',
            })
    if balance_drifts:
        _issue('financial', 'err',
               f"{len(balance_drifts)} account balance(s) don't match ledger sum",
               rows=balance_drifts)
    else:
        _ok('financial', 'Account balances')

    # ── 2. FlightCompletion charge drift ─────────────────────────────────────
    charge_drifts = []
    for fc in (FlightCompletion.objects
               .filter(booking__club=club)
               .prefetch_related('charge_items')
               .select_related('booking__member__user', 'booking__aircraft')):
        computed = sum(ci.amount for ci in fc.charge_items.all()) or _D('0')
        if abs(computed - fc.total_charge) > _D('0.01'):
            charge_drifts.append({
                'text': (
                    f"FC #{fc.id} ({fc.booking.aircraft.registration if fc.booking.aircraft else '?'} "
                    f"{fc.booking.scheduled_start.strftime('%d %b %y') if fc.booking.scheduled_start else '?'}): "
                    f"stored ${fc.total_charge}, items sum ${computed:.2f}"
                ),
                'url': _rev('core:booking_detail', kwargs={'club_slug': club_slug, 'booking_id': fc.booking.id}),
            })
    if charge_drifts:
        _issue('financial', 'err',
               f"{len(charge_drifts)} flight completion(s) have charge total mismatches",
               rows=charge_drifts)
    else:
        _ok('financial', 'Flight charges')

    # ── 3. FlightCompletion payment drift ────────────────────────────────────
    # Only count payments where paid_at IS NOT NULL (money actually received),
    # matching _sync_payment_cache() semantics. Allocated-but-not-collected
    # invoice payments (paid_at=None) do not count as received.
    from .models import FlightPayment as _FP
    payment_drifts = []
    for fc in (FlightCompletion.objects
               .filter(booking__club=club)
               .select_related('booking__aircraft')):
        computed = (_FP.objects.filter(completion=fc, paid_at__isnull=False)
                    .aggregate(t=Sum('amount'))['t'] or _D('0'))
        if abs(computed - fc.amount_paid) > _D('0.01'):
            payment_drifts.append({
                'text': (
                    f"FC #{fc.id} ({fc.booking.aircraft.registration if fc.booking.aircraft else '?'} "
                    f"{fc.booking.scheduled_start.strftime('%d %b %y') if fc.booking.scheduled_start else '?'}): "
                    f"stored ${fc.amount_paid}, received payments sum ${computed:.2f}"
                ),
                'url': _rev('core:booking_detail', kwargs={'club_slug': club_slug, 'booking_id': fc.booking.id}),
            })
    if payment_drifts:
        _issue('financial', 'err',
               f"{len(payment_drifts)} flight completion(s) have payment total mismatches",
               rows=payment_drifts)
    else:
        _ok('financial', 'Flight payments')

    # ── 4. Invoice paid but linked FC still has balance owing ────────────────
    from .models import Invoice as _InvH
    inv_fc_drifts = []
    for _inv in (_InvH.objects
                 .filter(club=club, status='paid')
                 .select_related('flight_completion__booking__aircraft',
                                 'flight_completion__booking__member__user')
                 .exclude(flight_completion__isnull=True)):
        _ifc = _inv.flight_completion
        if _ifc and _D(str(_ifc.balance_owing or 0)) > _D('0.01'):
            _mn = _ifc.booking.member.user.get_full_name() if (_ifc.booking and _ifc.booking.member) else '?'
            inv_fc_drifts.append({
                'text': (
                    f"Invoice {_inv.display_number} paid (${_inv.total}) "
                    f"but FC #{_ifc.id} still shows ${_ifc.balance_owing:.2f} owing — {_mn}"
                ),
                'url': _rev('core:invoice_detail', kwargs={'club_slug': club_slug, 'invoice_id': _inv.id}),
            })
    if inv_fc_drifts:
        _issue('financial', 'warn',
               f"{len(inv_fc_drifts)} invoice(s) marked paid but linked flight still shows balance owing",
               rows=inv_fc_drifts)
    else:
        _ok('financial', 'Invoice/FC payment sync')

    # ── 5. Meter hour gaps ───────────────────────────────────────────────────
    from .models import MaintenanceLogEntry
    meter_gaps = []
    for ac in Aircraft.objects.filter(club=club).exclude(status='retired'):
        fcs = list(
            FlightCompletion.objects
            .filter(booking__aircraft=ac, hobbs_end__isnull=False, hobbs_start__isnull=False)
            .order_by('booking__departed_at')
            .values('id', 'hobbs_start', 'hobbs_end', 'booking__departed_at', 'booking__id')
        )
        for i in range(1, len(fcs)):
            prev_end = _D(str(fcs[i-1]['hobbs_end']))
            curr_start = _D(str(fcs[i]['hobbs_start']))
            gap = curr_start - prev_end
            if abs(gap) > _D('0.05'):
                meter_gaps.append({
                    'text': (
                        f"{ac.registration}: gap of {gap:+.2f} hrs before FC #{fcs[i]['id']} "
                        f"({fcs[i]['booking__departed_at'].strftime('%d %b %y') if fcs[i]['booking__departed_at'] else '?'})"
                    ),
                    'url': _rev('core:booking_detail', kwargs={'club_slug': club_slug, 'booking_id': fcs[i]['booking__id']}),
                    'link_label': 'Edit readings →',
                })
    if meter_gaps:
        _issue('operations', 'warn',
               f"{len(meter_gaps)} hobbs gap(s) detected between consecutive flights",
               rows=meter_gaps)
    else:
        _ok('operations', 'Meter readings')


    # ── 6. Active members with expired subscriptions ─────────────────────────
    today = timezone.localdate()
    lapsed_active = list(
        ClubMember.objects.filter(
            club=club, standing='active',
            subscription_expires__isnull=False,
            subscription_expires__lt=today,
        ).select_related('user').order_by('subscription_expires')
    )
    if lapsed_active:
        rows = [
            {
                'text': f"{m.user.get_full_name()} — expired {m.subscription_expires}",
                'url': _rev('core:manage_member_detail', kwargs={'club_slug': club_slug, 'member_id': m.id}) + '#membership',
            }
            for m in lapsed_active
        ]
        _issue('members', 'warn',
               f"{len(lapsed_active)} active member(s) have expired subscriptions", rows=rows)
    else:
        _ok('members', 'Subscription standing')

    # Invoices: open invoices on completed bookings is normal (unpaid bill),
    # tracked via Attention Items — not a data integrity issue.
    _ok('financial', 'Invoices')

    total_issues = sum(len(g['issues']) for g in groups.values())
    return render(request, 'core/health_check.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'groups': groups,
        'total_issues': total_issues,
        'lapsed_active_count': len(lapsed_active),
    })


@login_required
def manage_guide(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_staff(actor, club, request):
        return err
    config = get_config(club)
    return render(request, 'core/manage_guide.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'config': config,
    })


@login_required
def member_guide(request, club_slug):
    """Member-facing help guide for the web app — available to any club member."""
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    config = get_config(club)
    return render(request, 'core/member_guide.html', {
        'club': club, 'club_member': actor, 'config': config,
    })


@login_required
def app_guide(request, club_slug):
    """Short help guide for the mobile (PWA) app — available to any club member."""
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    config = get_config(club)
    return render(request, 'core/app/guide.html', {
        'club': club, 'club_member': actor, 'config': config,
    })


@login_required
def data_page(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err
    _ico = 'width="20" height="20" viewBox="0 0 15 15" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"'
    export_types = [
        ('members',     f'<svg {_ico}><circle cx="7.5" cy="5" r="3"/><path d="M1.5 14c0-3.3 2.7-6 6-6s6 2.7 6 6"/></svg>',
                        'Members',        'Name, email, role, standing, account balance, subscription expiry'),
        ('credentials', f'<svg {_ico}><rect x="2" y="2" width="11" height="11" rx="1.5"/><line x1="4.5" y1="5.5" x2="10.5" y2="5.5"/><line x1="4.5" y1="7.5" x2="10.5" y2="7.5"/><line x1="4.5" y1="9.5" x2="7.5" y2="9.5"/></svg>',
                        'Credentials',    'Member licences, ratings, medicals, flight reviews — type, issue/expiry dates'),
        ('flights',     f'<svg {_ico}><path d="M13.5 1.5 1.5 6.5l4.5 2 1.5 4.5 2.5-2.5z"/><line x1="6" y1="8.5" x2="13.5" y2="1.5"/></svg>',
                        'Flight history', 'All completed flights — dates, aircraft, pilot, instructor, hours, charges'),
        ('aircraft',    f'<svg {_ico}><path d="M7.5 1.5 6 5.5 1 9l1 1 5-2 -.5 4-2 1 .5 1 3-1.5 3 1.5.5-1-2-1L9.5 9l5 2 1-1-5-3.5z"/></svg>',
                        'Aircraft',       'Fleet list with type, serial, seats, billing method, initial meters'),
        ('financial',   f'<svg {_ico}><rect x="1.5" y="3.5" width="12" height="8.5" rx="1.5"/><line x1="1.5" y1="7" x2="13.5" y2="7"/><line x1="4" y1="10.5" x2="6.5" y2="10.5"/></svg>',
                        'Financial',      'Account transactions, payments, outstanding balances'),
        ('invoices',    f'<svg {_ico}><rect x="2.5" y="1.5" width="10" height="12" rx="1"/><line x1="5" y1="5" x2="10" y2="5"/><line x1="5" y1="7.5" x2="10" y2="7.5"/><line x1="5" y1="10" x2="8" y2="10"/></svg>',
                        'Invoices',       'All invoices with status, amounts, and line items'),
        ('maintenance', f'<svg {_ico}><path d="M13 2.5a3 3 0 0 0-4.2 4.2L4 11.5a1 1 0 1 0 1.4 1.4l4.8-4.8A3 3 0 0 0 13 2.5z"/><line x1="11" y1="4" x2="12.5" y2="5.5"/></svg>',
                        'Maintenance',    'Per-aircraft maintenance hour log and scheduled items'),
        ('occurrences', f'<svg {_ico}><path d="M7.5 2 2 12.5h11z"/><line x1="7.5" y1="6" x2="7.5" y2="9"/><circle cx="7.5" cy="11" r=".5" fill="currentColor"/></svg>',
                        'Occurrences',    'Safety occurrence reports — type, date, description, status, review notes'),
    ]
    aircraft_list = Aircraft.objects.filter(club=club).exclude(status='retired').order_by('registration')
    return render(request, 'core/data_page.html', {
        'club': club, 'club_member': actor, 'is_instructor': actor.is_instructor,
        'export_types': export_types,
        'aircraft_list': aircraft_list,
    })


@login_required
def export_data(request, club_slug, export_type):
    import csv, io, zipfile
    from django.http import HttpResponse, StreamingHttpResponse
    from .models import (Voucher, Account, AccountTransaction,
                         AircraftMaintenanceItem, MaintenanceLogEntry, MemberCredential)

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    ALLOWED = {'members', 'flights', 'aircraft', 'maintenance', 'financial',
               'credentials', 'invoices', 'occurrences', 'all'}
    if export_type not in ALLOWED:
        return HttpResponse('Unknown export type', status=400)

    logger.info('export club=%s type=%s by=%s', club.slug, export_type, request.user.email)

    def csv_bytes(rows, headers):
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        return buf.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility

    # ── Members ──────────────────────────────────────────────────────────────
    def members_csv():
        hdrs = ['last_name', 'first_name', 'email', 'role', 'standing',
                'caa_number', 'phone_mobile', 'phone_home',
                'account_balance', 'subscription_expires']
        rows = []
        for m in (ClubMember.objects.filter(club=club)
                  .select_related('user', 'role')
                  .prefetch_related('account')
                  .order_by('user__last_name')):
            bal = ''
            try:
                bal = str(m.account.balance)
            except Exception:
                pass
            rows.append([
                m.user.last_name, m.user.first_name, m.user.email,
                m.role.name if m.role else '', m.standing,
                m.caa_number, m.phone_mobile, m.phone_home,
                bal, m.subscription_expires or '',
            ])
        return csv_bytes(rows, hdrs)

    # ── Flights ───────────────────────────────────────────────────────────────
    def flights_csv():
        hdrs = ['date', 'member', 'aircraft', 'flight_type', 'instructor',
                'scheduled_start', 'scheduled_end', 'actual_hours',
                'total_charge', 'amount_paid', 'balance_owing', 'outcome']
        rows = []
        qs = (FlightCompletion.objects.filter(booking__club=club)
              .select_related('booking__member__user', 'booking__aircraft',
                              'booking__flight_type', 'booking__instructor')
              .order_by('booking__scheduled_start'))
        for fc in qs:
            b = fc.booking
            instr = b.instructor.get_full_name() if b.instructor else ''
            rows.append([
                b.scheduled_start.date(),
                b.member.user.get_full_name(),
                b.aircraft.registration,
                str(b.flight_type),
                instr,
                b.scheduled_start.strftime('%H:%M'),
                b.scheduled_end.strftime('%H:%M'),
                fc.actual_flight_hours,
                fc.total_charge,
                fc.amount_paid,
                fc.balance_owing,
                fc.get_outcome_display(),
            ])
        return csv_bytes(rows, hdrs)

    # ── Aircraft ──────────────────────────────────────────────────────────────
    def aircraft_csv():
        hdrs = ['registration', 'type', 'serial', 'seats', 'engines',
                'status', 'is_leased', 'total_time_method',
                'maint_time_source', 'maint_time_fraction', 'maint_hours_initial',
                'hobbs_initial', 'tacho_initial', 'airswitch_initial']
        rows = []
        for ac in Aircraft.objects.filter(club=club).select_related('aircraft_type').order_by('registration'):
            rows.append([
                ac.registration,
                ac.aircraft_type.name if ac.aircraft_type else '',
                ac.serial_number,
                ac.seats, ac.engine_count,
                ac.get_status_display(),
                'Yes' if ac.is_leased else 'No',
                ac.get_total_time_method_display(),
                ac.maint_time_source, ac.maint_time_fraction,
                ac.maint_hours_initial or '',
                ac.hobbs_initial or '', ac.tacho_initial or '', ac.airswitch_initial or '',
            ])
        return csv_bytes(rows, hdrs)

    # ── Maintenance ───────────────────────────────────────────────────────────
    def maintenance_csv():
        # Items
        item_hdrs = ['aircraft', 'item_name', 'due_date', 'due_hours',
                     'interval_hours', 'interval_days', 'warn_hours', 'alert_hours',
                     'last_completed_date', 'last_completed_hours', 'urgency']
        item_rows = []
        for item in (AircraftMaintenanceItem.objects
                     .filter(aircraft__club=club)
                     .select_related('aircraft')
                     .order_by('aircraft__registration', 'name')):
            item_rows.append([
                item.aircraft.registration, item.name,
                item.due_date or '', item.due_hours or '',
                item.interval_hours or '', item.interval_days or '',
                item.warn_hours or '', item.alert_hours or '',
                item.last_completed_date or '', item.last_completed_hours or '',
                item.get_urgency_display(),
            ])
        # Log
        log_hdrs = ['aircraft', 'date', 'hobbs', 'tacho', 'airswitch',
                    'maint_hours_this_flight', 'maint_hours_cumulative', 'notes']
        log_rows = []
        for entry in (MaintenanceLogEntry.objects
                      .filter(aircraft__club=club)
                      .select_related('aircraft')
                      .order_by('aircraft__registration', 'date')):
            log_rows.append([
                entry.aircraft.registration, entry.date,
                entry.hobbs_reading or '', entry.tacho_reading or '',
                entry.airswitch_reading or '',
                entry.maint_hours_flight, entry.maint_hours_total,
                entry.notes,
            ])
        # Combine as two sections in one CSV
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['=== MAINTENANCE ITEMS ==='])
        w.writerow(item_hdrs)
        w.writerows(item_rows)
        w.writerow([])
        w.writerow(['=== MAINTENANCE LOG ==='])
        w.writerow(log_hdrs)
        w.writerows(log_rows)
        return buf.getvalue().encode('utf-8-sig')

    # ── Financial ─────────────────────────────────────────────────────────────
    def financial_csv():
        hdrs = ['date', 'member', 'type', 'direction', 'amount',
                'description', 'payment_method', 'reference']
        rows = []
        for tx in (AccountTransaction.objects
                   .filter(account__club_member__club=club)
                   .select_related('account__club_member__user')
                   .order_by('created_at')):
            rows.append([
                tx.created_at.date(),
                tx.account.club_member.user.get_full_name(),
                tx.get_transaction_type_display(),
                tx.direction,
                tx.amount,
                tx.description,
                tx.payment_method,
                tx.reference,
            ])
        return csv_bytes(rows, hdrs)

    # ── Excel helper: build a single .xlsx with named Excel Tables ──────────────
    def make_xlsx(sheets):
        """sheets = list of (sheet_name, headers, rows). Returns bytes.
        Each sheet becomes a named Excel Table so Power Query can load them
        straight into the Data Model."""
        import openpyxl, decimal as _dec, datetime as _dt, re as _re
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        HDR_FILL = PatternFill('solid', fgColor='1E3A5F')
        HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
        TBL_STYLE = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True)
        DATE_FMT  = 'yyyy-mm-dd'
        NUM_FMT   = '#,##0.00'

        def coerce(v):
            """Preserve dates/numbers; convert everything else to string."""
            if v is None or v == '':
                return None
            if isinstance(v, bool):
                return 'Yes' if v else 'No'
            if isinstance(v, (int, float)):
                return v
            if isinstance(v, _dec.Decimal):
                return float(v)
            if isinstance(v, _dt.datetime):
                return v.replace(tzinfo=None)   # Excel doesn't handle tz-aware
            if isinstance(v, _dt.date):
                return v
            return str(v)

        def tbl_name(sheet_name):
            n = _re.sub(r'[^A-Za-z0-9_]', '_', sheet_name)
            return ('T' + n) if n[:1].isdigit() else n

        # ── Instructions sheet ──────────────────────────────────────────────────
        ws_i = wb.create_sheet('_Instructions', 0)
        ws_i.sheet_view.showGridLines = False
        ws_i.column_dimensions['A'].width = 4
        ws_i.column_dimensions['B'].width = 28
        ws_i.column_dimensions['C'].width = 68

        def irow(row_num, b_val='', c_val='', b_bold=False, b_fill=None, c_fill=None, height=None):
            ws_i.row_dimensions[row_num].height = height or 15
            cb = ws_i.cell(row_num, 2, b_val)
            cc = ws_i.cell(row_num, 3, c_val)
            if b_bold: cb.font = Font(bold=True, size=11)
            if b_fill: cb.fill = c_fill = PatternFill('solid', fgColor=b_fill)
            if c_fill: cc.fill = PatternFill('solid', fgColor=c_fill)
            cc.alignment = Alignment(wrap_text=True)

        irow(1,  'ClubHangar Data Export',  '', b_bold=True); ws_i.row_dimensions[1].height = 22
        ws_i.cell(1, 2).font = Font(bold=True, size=14, color='1E3A5F')
        irow(2,  'Each tab in this workbook is a named Excel Table.', '')
        irow(3,  'The tables mirror the ClubHangar database and can be linked into an Excel Data Model', '')
        irow(4,  'for pivot tables, charts, and Power BI reports.', '')
        irow(5)
        irow(6,  'HOW TO BUILD THE DATA MODEL IN EXCEL', '', b_bold=True)
        ws_i.cell(6, 2).font = Font(bold=True, size=11, color='1E3A5F')
        irow(7,  'Step 1', 'Open this workbook in Microsoft Excel (2016 or later).')
        irow(8,  'Step 2', 'Go to the Data tab → Get Data → From Table/Range.')
        irow(9,  '',       'Select each table one at a time and choose "Load to… → Only create connection" + tick "Add to Data Model".')
        irow(10, 'Step 3', 'Repeat Step 2 for every table (Members, Flights, Aircraft, Financial, Invoices, etc.).')
        irow(11, 'Step 4', 'Once all tables are in the Data Model, go to Data → Relationships (or Power Pivot → Manage).')
        irow(12, 'Step 5', 'Create the relationships listed in the table below, then build pivot tables or Power BI reports.')
        irow(13)
        irow(14, 'TIP',    'On a Mac you may need to use Power Query (Data → Get Data) rather than the Power Pivot add-in.')
        ws_i.cell(14, 2).font = Font(bold=True, color='C07A1C')
        irow(15, 'TIP',    'If you use Google Sheets, import each CSV file as a separate sheet and use VLOOKUP/QUERY to join them.')
        ws_i.cell(15, 2).font = Font(bold=True, color='C07A1C')
        irow(16)
        irow(17, 'RELATIONSHIPS', '', b_bold=True)
        ws_i.cell(17, 2).font = Font(bold=True, size=11, color='1E3A5F')

        rel_hdrs = ['Table', 'Column', '→  relates to  →', 'Table', 'Column', 'Notes']
        rels = [
            ('Members',            'Email',          'Flights',             'Member email',    'Primary join for flight records'),
            ('Members',            'Email',          'Financial',           'Member',          'Use Email in Members; match to member full name or add email to Financial'),
            ('Members',            'Email',          'Invoices',            'Member',          'Match by full name or email'),
            ('Members',            'Email',          'Credentials',         'Member',          'Full name match'),
            ('Aircraft',           'Registration',   'Flights',             'Aircraft',        'Direct registration match'),
            ('Aircraft',           'Registration',   'Maintenance Items',   'Aircraft',        'Direct registration match'),
            ('Aircraft',           'Registration',   'Maintenance Log',     'Aircraft',        'Direct registration match'),
            ('Invoices',           'Invoice #',      'Invoice Line Items',  'Invoice #',       'Direct invoice number match'),
        ]
        ws_i.cell(18, 2).value = rel_hdrs[0]; ws_i.cell(18, 2).font = Font(bold=True)
        ws_i.cell(18, 3).value = rel_hdrs[1]; ws_i.cell(18, 3).font = Font(bold=True)
        ws_i.cell(18, 4).value = rel_hdrs[2]; ws_i.cell(18, 4).font = Font(bold=True, color='888888')
        ws_i.cell(18, 5).value = rel_hdrs[3]; ws_i.cell(18, 5).font = Font(bold=True)
        ws_i.cell(18, 6).value = rel_hdrs[4]; ws_i.cell(18, 6).font = Font(bold=True)
        ws_i.cell(18, 7).value = rel_hdrs[5]; ws_i.cell(18, 7).font = Font(bold=True, color='888888')
        ws_i.column_dimensions['D'].width = 8
        ws_i.column_dimensions['E'].width = 22
        ws_i.column_dimensions['F'].width = 16
        ws_i.column_dimensions['G'].width = 48
        for ri, (t1, c1, t2, c2, note) in enumerate(rels, 19):
            ws_i.cell(ri, 2, t1)
            ws_i.cell(ri, 3, c1)
            ws_i.cell(ri, 4, '→')
            ws_i.cell(ri, 4).font = Font(color='888888')
            ws_i.cell(ri, 4).alignment = Alignment(horizontal='center')
            ws_i.cell(ri, 5, t2)
            ws_i.cell(ri, 6, c2)
            ws_i.cell(ri, 7, note)
            ws_i.cell(ri, 7).font = Font(color='888888', italic=True)

        # ── Data sheets ─────────────────────────────────────────────────────────
        for sheet_name, headers, rows in sheets:
            ws = wb.create_sheet(sheet_name[:31])
            ws.append(headers)
            for cell in ws[1]:
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal='left')
            for row in rows:
                coerced = [coerce(v) for v in row]
                ws.append(coerced)
            # Format date/number columns
            for col_idx, hdr in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                hdr_lower = hdr.lower()
                if any(k in hdr_lower for k in ('date', 'expires', 'issued')):
                    for cell in ws[col_letter][1:]:
                        if isinstance(cell.value, _dt.date):
                            cell.number_format = DATE_FMT
                elif any(k in hdr_lower for k in ('amount', 'balance', 'charge', 'paid', 'price', 'total', 'rate')):
                    for cell in ws[col_letter][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = NUM_FMT
            # Auto-fit columns
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
            # Add Excel Table (requires at least 1 data row for openpyxl)
            if ws.max_row > 1:
                n_cols = len(headers)
                tbl = Table(displayName=tbl_name(sheet_name),
                            ref=f'A1:{get_column_letter(n_cols)}{ws.max_row}')
                tbl.tableStyleInfo = TBL_STYLE
                ws.add_table(tbl)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def members_data():
        hdrs = ['Last name', 'First name', 'Email', 'Role', 'Standing',
                'CAA number', 'Mobile', 'Home phone', 'Account balance', 'Subscription expires']
        rows = []
        for m in (ClubMember.objects.filter(club=club)
                  .select_related('user', 'role').prefetch_related('account')
                  .order_by('user__last_name')):
            bal = ''
            try: bal = m.account.balance
            except Exception: pass
            rows.append([m.user.last_name, m.user.first_name, m.user.email,
                         m.role.name if m.role else '', m.standing,
                         m.caa_number, m.phone_mobile, m.phone_home, bal,
                         m.subscription_expires or ''])
        return hdrs, rows

    def flights_data():
        hdrs = ['Date', 'Member', 'Aircraft', 'Flight type', 'Instructor',
                'Start', 'End', 'Actual hours', 'Total charge', 'Paid', 'Balance owing', 'Outcome']
        rows = []
        qs = (FlightCompletion.objects.filter(booking__club=club)
              .select_related('booking__member__user', 'booking__aircraft',
                              'booking__flight_type', 'booking__instructor')
              .order_by('booking__scheduled_start'))
        for fc in qs:
            b = fc.booking
            rows.append([b.scheduled_start.date(), b.member.user.get_full_name(),
                         b.aircraft.registration, str(b.flight_type),
                         b.instructor.get_full_name() if b.instructor else '',
                         b.scheduled_start.strftime('%H:%M'), b.scheduled_end.strftime('%H:%M'),
                         fc.actual_flight_hours, fc.total_charge, fc.amount_paid,
                         fc.balance_owing, fc.get_outcome_display()])
        return hdrs, rows

    def aircraft_data():
        hdrs = ['Registration', 'Type', 'Serial', 'Seats', 'Engines', 'Status',
                'Leased', 'Billing method', 'Maint source', 'Maint fraction', 'Maint hrs initial',
                'Hobbs initial', 'Tacho initial', 'Airswitch initial']
        rows = []
        for ac in Aircraft.objects.filter(club=club).select_related('aircraft_type').order_by('registration'):
            rows.append([ac.registration, ac.aircraft_type.name if ac.aircraft_type else '',
                         ac.serial_number, ac.seats, ac.engine_count, ac.get_status_display(),
                         'Yes' if ac.is_leased else 'No', ac.get_total_time_method_display(),
                         ac.maint_time_source, ac.maint_time_fraction, ac.maint_hours_initial or '',
                         ac.hobbs_initial or '', ac.tacho_initial or '', ac.airswitch_initial or ''])
        return hdrs, rows

    def maintenance_data():
        from .models import AircraftMaintenanceItem, MaintenanceLogEntry
        item_hdrs = ['Aircraft', 'Item', 'Due date', 'Due hours', 'Interval hrs',
                     'Interval days', 'Warn hrs', 'Alert hrs',
                     'Last done date', 'Last done hours', 'Urgency']
        item_rows = []
        for item in (AircraftMaintenanceItem.objects.filter(aircraft__club=club)
                     .select_related('aircraft').order_by('aircraft__registration', 'name')):
            item_rows.append([item.aircraft.registration, item.name,
                               item.due_date or '', item.due_hours or '',
                               item.interval_hours or '', item.interval_days or '',
                               item.warn_hours or '', item.alert_hours or '',
                               item.last_completed_date or '', item.last_completed_hours or '',
                               item.get_urgency_display()])
        log_hdrs = ['Aircraft', 'Date', 'Hobbs', 'Tacho', 'Airswitch',
                    'Maint hrs this flight', 'Maint hrs cumulative', 'Notes']
        log_rows = []
        for e in (MaintenanceLogEntry.objects.filter(aircraft__club=club)
                  .select_related('aircraft').order_by('aircraft__registration', 'date')):
            log_rows.append([e.aircraft.registration, e.date,
                              e.hobbs_reading or '', e.tacho_reading or '', e.airswitch_reading or '',
                              e.maint_hours_flight, e.maint_hours_total, e.notes])
        return (item_hdrs, item_rows), (log_hdrs, log_rows)

    def financial_data():
        from .models import AccountTransaction
        hdrs = ['Date', 'Member', 'Type', 'Direction', 'Amount',
                'Description', 'Payment method', 'Reference']
        rows = []
        for tx in (AccountTransaction.objects.filter(account__club_member__club=club)
                   .select_related('account__club_member__user').order_by('created_at')):
            rows.append([tx.created_at.date(), tx.account.club_member.user.get_full_name(),
                         tx.get_transaction_type_display(), tx.direction, tx.amount,
                         tx.description, tx.payment_method, tx.reference])
        return hdrs, rows

    def credentials_data():
        from .models import MemberCredential
        club_user_ids = ClubMember.objects.filter(club=club).values_list('user_id', flat=True)
        hdrs = ['Member', 'Type', 'Name / Aircraft type', 'Issue date',
                'Expiry date', 'Certificate number', 'Notes']
        rows = []
        for c in (MemberCredential.objects
                  .filter(member_id__in=club_user_ids)
                  .select_related('member', 'credential_type', 'aircraft_type')
                  .order_by('member__last_name', 'credential_type__display_order')):
            name = c.name or (c.aircraft_type.name if c.aircraft_type else '')
            rows.append([c.member.get_full_name(),
                         c.credential_type.name if c.credential_type_id else '', name,
                         c.issue_date or '', c.expiry_date or '',
                         c.certificate_number, c.notes])
        return hdrs, rows

    def invoices_data():
        from .models import Invoice, InvoiceLineItem
        inv_hdrs = ['Invoice #', 'Member', 'Description', 'Status',
                    'Issue date', 'Due date', 'GST rate', 'Total', 'Paid', 'Notes']
        inv_rows = []
        for inv in (Invoice.objects.filter(club=club)
                    .select_related('member__user')
                    .order_by('invoice_number')):
            inv_rows.append([inv.display_number,
                             inv.member.user.get_full_name() if inv.member else '',
                             inv.description, inv.get_status_display(),
                             inv.issue_date, inv.due_date, inv.gst_rate,
                             inv.total, inv.amount_paid, inv.notes])
        line_hdrs = ['Invoice #', 'Description', 'Quantity', 'Unit price', 'Total']
        line_rows = []
        for li in (InvoiceLineItem.objects.filter(invoice__club=club)
                   .select_related('invoice')
                   .order_by('invoice__invoice_number', 'id')):
            line_rows.append([li.invoice.display_number, li.description,
                              li.quantity, li.unit_price, li.total_price])
        return (inv_hdrs, inv_rows), (line_hdrs, line_rows)

    def occurrences_data():
        hdrs = ['ID', 'Type', 'Date', 'Time', 'Location', 'Aircraft',
                'Reported by', 'Status', 'Reported at', 'Description',
                'Immediate action', 'Review notes']
        rows = []
        for r in (OccurrenceReport.objects.filter(club=club)
                  .select_related('occurrence_type', 'reported_by__user', 'aircraft')
                  .order_by('date_of_occurrence')):
            rows.append([r.id, r.occurrence_type.name,
                         r.date_of_occurrence.isoformat(),
                         r.time_of_occurrence.strftime('%H:%M') if r.time_of_occurrence else '',
                         r.location,
                         r.aircraft.registration if r.aircraft else '',
                         r.reported_by.user.get_full_name(),
                         r.get_status_display(),
                         r.reported_at.strftime('%Y-%m-%d %H:%M'),
                         r.description, r.immediate_action, r.review_notes])
        return hdrs, rows

    def membership_history_data():
        from .models import MembershipHistoryEntry
        hdrs = ['Member', 'Event', 'Date', 'Changed by', 'From', 'To', 'Note']
        rows = []
        for e in (MembershipHistoryEntry.objects
                  .filter(club_member__club=club)
                  .select_related('club_member__user', 'changed_by')
                  .order_by('club_member__user__last_name', 'changed_at')):
            rows.append([e.club_member.user.get_full_name(),
                         e.get_event_type_display(),
                         e.changed_at.strftime('%Y-%m-%d %H:%M'),
                         e.changed_by.get_full_name() if e.changed_by else '',
                         e.old_value, e.new_value, e.note])
        return hdrs, rows

    # ── Dispatch ──────────────────────────────────────────────────────────────
    slug = club.slug
    ts = timezone.localdate().strftime('%Y%m%d')
    fmt = request.GET.get('fmt', 'csv')  # ?fmt=xlsx or default csv

    if export_type == 'all':
        if fmt == 'xlsx':
            m_hdrs, m_rows       = members_data()
            f_hdrs, f_rows       = flights_data()
            a_hdrs, a_rows       = aircraft_data()
            (mi_hdrs, mi_rows), (ml_hdrs, ml_rows) = maintenance_data()
            fi_hdrs, fi_rows     = financial_data()
            cr_hdrs, cr_rows     = credentials_data()
            (iv_hdrs, iv_rows), (li_hdrs, li_rows) = invoices_data()
            oc_hdrs, oc_rows     = occurrences_data()
            mh_hdrs, mh_rows     = membership_history_data()
            xlsx = make_xlsx([
                ('Members',            m_hdrs,  m_rows),
                ('Membership History', mh_hdrs, mh_rows),
                ('Credentials',        cr_hdrs, cr_rows),
                ('Flights',            f_hdrs,  f_rows),
                ('Aircraft',           a_hdrs,  a_rows),
                ('Maintenance Items',  mi_hdrs, mi_rows),
                ('Maintenance Log',    ml_hdrs, ml_rows),
                ('Financial',          fi_hdrs, fi_rows),
                ('Invoices',           iv_hdrs, iv_rows),
                ('Invoice Line Items', li_hdrs, li_rows),
                ('Occurrences',        oc_hdrs, oc_rows),
            ])
            resp = HttpResponse(xlsx,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{slug}_export_{ts}.xlsx"'
            return resp
        # CSV ZIP
        buf = io.BytesIO()
        def _csv(hdrs, rows):
            b = io.StringIO()
            w = csv.writer(b)
            w.writerow(hdrs)
            w.writerows(rows)
            return b.getvalue().encode('utf-8-sig')
        (mi_hdrs, mi_rows), (ml_hdrs, ml_rows) = maintenance_data()
        (iv_hdrs, iv_rows), (li_hdrs, li_rows) = invoices_data()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f'{slug}_members_{ts}.csv',            _csv(*members_data()))
            zf.writestr(f'{slug}_membership_history_{ts}.csv', _csv(*membership_history_data()))
            zf.writestr(f'{slug}_credentials_{ts}.csv',        _csv(*credentials_data()))
            zf.writestr(f'{slug}_flights_{ts}.csv',            _csv(*flights_data()))
            zf.writestr(f'{slug}_aircraft_{ts}.csv',           _csv(*aircraft_data()))
            zf.writestr(f'{slug}_maint_items_{ts}.csv',        _csv(mi_hdrs, mi_rows))
            zf.writestr(f'{slug}_maint_log_{ts}.csv',          _csv(ml_hdrs, ml_rows))
            zf.writestr(f'{slug}_financial_{ts}.csv',          _csv(*financial_data()))
            zf.writestr(f'{slug}_invoices_{ts}.csv',           _csv(iv_hdrs, iv_rows))
            zf.writestr(f'{slug}_invoice_lines_{ts}.csv',      _csv(li_hdrs, li_rows))
            zf.writestr(f'{slug}_occurrences_{ts}.csv',        _csv(*occurrences_data()))
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/zip')
        resp['Content-Disposition'] = f'attachment; filename="{slug}_export_{ts}.zip"'
        return resp

    # Single-table export
    def _csv(hdrs, rows):
        b = io.StringIO()
        w = csv.writer(b)
        w.writerow(hdrs)
        w.writerows(rows)
        return b.getvalue().encode('utf-8-sig')

    data_map = {
        'members':     (members_data,     f'{slug}_members_{ts}'),
        'flights':     (flights_data,     f'{slug}_flights_{ts}'),
        'aircraft':    (aircraft_data,    f'{slug}_aircraft_{ts}'),
        'financial':   (financial_data,   f'{slug}_financial_{ts}'),
        'credentials': (credentials_data, f'{slug}_credentials_{ts}'),
        'occurrences': (occurrences_data, f'{slug}_occurrences_{ts}'),
    }
    if export_type == 'invoices':
        (iv_hdrs, iv_rows), (li_hdrs, li_rows) = invoices_data()
        if fmt == 'xlsx':
            xlsx = make_xlsx([('Invoices', iv_hdrs, iv_rows),
                              ('Invoice Line Items', li_hdrs, li_rows)])
            resp = HttpResponse(xlsx,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{slug}_invoices_{ts}.xlsx"'
            return resp
        buf = io.StringIO()
        w2 = csv.writer(buf)
        w2.writerow(['=== INVOICES ===']); w2.writerow(iv_hdrs); w2.writerows(iv_rows)
        w2.writerow([]); w2.writerow(['=== LINE ITEMS ===']); w2.writerow(li_hdrs); w2.writerows(li_rows)
        resp = HttpResponse(buf.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{slug}_invoices_{ts}.csv"'
        return resp
    if export_type == 'maintenance':
        (mi_hdrs, mi_rows), (ml_hdrs, ml_rows) = maintenance_data()
        if fmt == 'xlsx':
            xlsx = make_xlsx([('Maintenance Items', mi_hdrs, mi_rows),
                              ('Maintenance Log', ml_hdrs, ml_rows)])
            resp = HttpResponse(xlsx,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{slug}_maintenance_{ts}.xlsx"'
            return resp
        buf = io.StringIO()
        w2 = csv.writer(buf)
        w2.writerow(['=== MAINTENANCE ITEMS ===']); w2.writerow(mi_hdrs); w2.writerows(mi_rows)
        w2.writerow([]); w2.writerow(['=== MAINTENANCE LOG ===']); w2.writerow(ml_hdrs); w2.writerows(ml_rows)
        resp = HttpResponse(buf.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{slug}_maintenance_{ts}.csv"'
        return resp

    fn, base = data_map[export_type]
    hdrs, rows = fn()
    if fmt == 'xlsx':
        sheet_name = export_type.capitalize()
        xlsx = make_xlsx([(sheet_name, hdrs, rows)])
        resp = HttpResponse(xlsx,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{base}.xlsx"'
        return resp
    resp = HttpResponse(_csv(hdrs, rows), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{base}.csv"'
    return resp


@login_required
def export_import_template(request, club_slug):
    """
    Generate a blank Excel import template: one worksheet per object type,
    with headers, format notes, and a few greyed-out example rows so the
    user knows exactly what columns and values to provide.
    A reference sheet lists the club's current flight types and aircraft types.
    """
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_admin(actor, club, request): return err

    from .models import FlightType, AircraftType

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Styles ────────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill('solid', fgColor='1E3A5F')
    HDR_FONT  = Font(color='FFFFFF', bold=True, size=10)
    HDR_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

    NOTE_FILL = PatternFill('solid', fgColor='FFF3CD')   # amber — required field note
    EX_FILL   = PatternFill('solid', fgColor='F2F4F7')   # grey  — example rows
    EX_FONT   = Font(italic=True, color='9AA3AD', size=9)
    EX_ALIGN  = Alignment(horizontal='left')

    REF_FILL  = PatternFill('solid', fgColor='E8F0FB')   # blue-tint — reference rows
    REF_FONT  = Font(color='2B4DA0', size=9)

    THIN      = Side(style='thin', color='D6DAE0')
    BORDER    = Border(bottom=Side(style='thin', color='E3E6EA'))

    def add_sheet(name, headers, notes, examples, col_widths=None):
        """
        headers: list of strings (use * suffix for required fields)
        notes:   list of per-column notes (same length as headers, or empty string)
        examples: list of lists (rows of example data, rendered greyed-out)
        col_widths: optional list of explicit widths; otherwise auto-sized
        """
        ws = wb.create_sheet(name[:31])

        # Header row
        ws.append(headers)
        ws.row_dimensions[1].height = 28
        for i, cell in enumerate(ws[1], 1):
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = HDR_ALIGN
            if col_widths and i <= len(col_widths):
                ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

        # Notes row (per-column hints, shown in amber)
        if any(notes):
            ws.append(notes)
            ws.row_dimensions[2].height = 14
            for cell in ws[2]:
                cell.font  = Font(italic=True, color='7A5800', size=8)
                cell.fill  = NOTE_FILL
                cell.alignment = Alignment(horizontal='left', wrap_text=True)

        # Example rows
        for ex in examples:
            ws.append(ex)
            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 13
            for cell in ws[row_idx]:
                cell.font  = EX_FONT
                cell.fill  = EX_FILL
                cell.alignment = EX_ALIGN

        # Auto-width if not explicitly specified
        if not col_widths:
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = 'A2'   # freeze header row
        return ws

    # ── Sheet 1: Instructions ─────────────────────────────────────────────────
    ws_info = wb.create_sheet('Instructions', 0)
    instructions = [
        ('ClubHangar Import Template', True),
        (f'Generated for: {club.name}', False),
        ('', False),
        ('HOW TO USE THIS FILE', True),
        ('1. Fill in the coloured sheets (Members, Aircraft, Flights).', False),
        ('2. Delete the grey example rows before importing — they are just to show the format.', False),
        ('3. Columns marked * are required. Leave optional columns blank if not known.', False),
        ('4. Dates must be in YYYY-MM-DD format (e.g. 2024-06-15).', False),
        ('5. Emails must exactly match existing members when cross-referencing.', False),
        ('6. Aircraft registrations must exactly match aircraft already added to the system.', False),
        ('7. Flight types and aircraft types must exactly match what is configured in Settings.', False),
        ('', False),
        ('SHEET SUMMARY', True),
        ('Members  — one row per club member. Fill this in first.', False),
        ('Aircraft — one row per aircraft in the fleet.', False),
        ('Flights  — one row per completed historical flight.', False),
        ('Ref: Flight Types  — read-only reference, do not edit.', False),
        ('Ref: Aircraft Types — read-only reference, do not edit.', False),
    ]
    TITLE_FONT = Font(bold=True, size=11, color='1E3A5F')
    BODY_FONT  = Font(size=10, color='2B333D')
    ws_info.column_dimensions['A'].width = 72
    for i, (text, is_title) in enumerate(instructions, 1):
        cell = ws_info.cell(row=i, column=1, value=text)
        cell.font  = TITLE_FONT if is_title else BODY_FONT
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        ws_info.row_dimensions[i].height = 18 if is_title else 14
    ws_info.sheet_view.showGridLines = False

    # ── Sheet 2: Members ──────────────────────────────────────────────────────
    add_sheet(
        name='Members',
        headers=['Last name *', 'First name *', 'Email *',
                 'Role', 'CAA number', 'Mobile', 'Home phone',
                 'Date of birth', 'Medical expiry', 'BFR expiry',
                 'Opening balance ($)'],
        notes=['e.g. Smith', 'e.g. John', 'e.g. john@example.com',
               'Member / Instructor / Admin', 'e.g. 123456',
               'e.g. +64 21 123 4567', 'e.g. +64 9 555 1234',
               'YYYY-MM-DD', 'YYYY-MM-DD', 'YYYY-MM-DD',
               '0.00 for no balance'],
        examples=[
            ['Smith',   'John',   'john.smith@example.com',   'Member',     '123456', '+64 21 123 4567', '',              '1985-03-15', '2026-06-30', '2025-12-01', '0.00'],
            ['Nguyen',  'Sarah',  'sarah.nguyen@example.com', 'Instructor', '654321', '+64 27 987 6543', '+64 9 555 0000', '1978-07-22', '2027-03-15', '2026-04-01', '-45.00'],
        ],
        col_widths=[14, 14, 28, 12, 13, 18, 18, 14, 14, 14, 14],
    )

    # ── Sheet 3: Aircraft ─────────────────────────────────────────────────────
    ac_type_names = ', '.join(
        AircraftType.objects.filter(club=club).values_list('name', flat=True)
    ) or 'e.g. Cessna 172'
    add_sheet(
        name='Aircraft',
        headers=['Registration *', 'Aircraft type *', 'Seats', 'Engines',
                 'Serial number', 'Hobbs initial', 'Tacho initial',
                 'Air switch initial', 'Fuel consumption (L/hr)', 'Leased'],
        notes=['e.g. ZK-ABC', f'One of: {ac_type_names}', '1–20', '1–4',
               'Manufacturer serial', 'Hours at system entry', 'Hours at system entry',
               'Hours at system entry', 'e.g. 28.0', 'Yes / No'],
        examples=[
            ['ZK-ABC', 'Cessna 172', '4', '1', '17267890', '2340.5', '1987.2', '',       '28.0', 'No'],
            ['ZK-DEF', 'Piper PA-28', '4', '1', 'PA28-4512', '1250.0', '',      '980.0', '26.0', 'No'],
        ],
        col_widths=[15, 20, 8, 9, 16, 14, 14, 16, 20, 9],
    )

    # ── Sheet 4: Flights ──────────────────────────────────────────────────────
    ft_names = ', '.join(
        FlightType.objects.filter(club=club).values_list('name', flat=True)
    ) or 'e.g. Training, Solo'
    add_sheet(
        name='Flights',
        headers=['Date *', 'Member email *', 'Aircraft registration *', 'Flight type *',
                 'Instructor email', 'Hobbs start', 'Hobbs end',
                 'Tacho start', 'Tacho end', 'Air switch start', 'Air switch end',
                 'Charge ($)', 'Notes'],
        notes=['YYYY-MM-DD', 'Must match a member', 'Must match aircraft in system',
               f'One of: {ft_names}',
               'Leave blank for solo', 'Leave blank if not used', 'Leave blank if not used',
               'Leave blank if not used', 'Leave blank if not used',
               'Leave blank if not used', 'Leave blank if not used',
               '0.00 or leave blank', 'Outcome notes'],
        examples=[
            ['2024-01-15', 'john.smith@example.com',   'ZK-ABC', 'Training', 'sarah.nguyen@example.com', '2340.5', '2342.3', '', '', '', '', '125.00', 'First solo circuit'],
            ['2024-01-16', 'sarah.nguyen@example.com', 'ZK-DEF', 'Solo',     '',                         '1250.0', '1251.5', '', '', '', '', '78.00',  ''],
        ],
        col_widths=[12, 28, 22, 16, 28, 12, 12, 12, 12, 14, 14, 10, 28],
    )

    # ── Sheet 5: Ref — Flight Types ───────────────────────────────────────────
    ws_ft = wb.create_sheet('Ref: Flight Types')
    ws_ft.append(['Flight type name', 'Billable', 'Requires declaration'])
    for cell in ws_ft[1]:
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = HDR_ALIGN
    for ft in FlightType.objects.filter(club=club).order_by('name'):
        row = ws_ft.max_row + 1
        ws_ft.append([ft.name,
                       'Yes' if ft.is_billable else 'No',
                       'Yes' if getattr(ft, 'requires_declaration', False) else 'No'])
        for cell in ws_ft[row]:
            cell.font = REF_FONT; cell.fill = REF_FILL
    for col in ws_ft.columns:
        ws_ft.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value or '')) for c in col), default=8) + 4

    # ── Sheet 6: Ref — Aircraft Types ─────────────────────────────────────────
    ws_at = wb.create_sheet('Ref: Aircraft Types')
    ws_at.append(['Aircraft type name', 'ICAO designator'])
    for cell in ws_at[1]:
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = HDR_ALIGN
    for at in AircraftType.objects.filter(club=club).order_by('name'):
        row = ws_at.max_row + 1
        ws_at.append([at.name, getattr(at, 'icao_designator', '') or ''])
        for cell in ws_at[row]:
            cell.font = REF_FONT; cell.fill = REF_FILL
    for col in ws_at.columns:
        ws_at.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value or '')) for c in col), default=8) + 4

    # ── Respond ───────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = club.slug
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{slug}_import_template.xlsx"'
    return resp


def _fetch_avwx(icao, endpoint='metar', api_key=''):
    """Fetch METAR or TAF from AVWX REST API. Cached 5 minutes."""
    from django.core.cache import cache
    import urllib.request, urllib.error, json as _json, ssl as _ssl, certifi
    cache_key = f'avwx_{endpoint}_{icao}'
    cached = cache.get(cache_key)
    if cached is not None:
        return cached
    url = f'https://avwx.rest/api/{endpoint}/{icao}?options=summary,translate'
    ctx = _ssl.create_default_context(cafile=certifi.where())
    try:
        req = urllib.request.Request(url, headers={
            'Authorization': f'TOKEN {api_key}',
            'User-Agent': 'ClubHangar/1.0',
        })
        with urllib.request.urlopen(req, timeout=8, context=ctx) as r:
            data = _json.loads(r.read().decode())
        cache.set(cache_key, data, 300)
        return data
    except Exception as exc:
        return {'error': str(exc)}


def _decode_wx_conditions(conditions):
    """
    Turn a list of AVWX weather_condition dicts into human-readable strings.
    Each dict has keys: intensity, descriptor, precipitation, obscuration, other —
    each is either None or {"repr": "XX", "value": "English word"}.
    """
    def _val(field):
        if not field:
            return ''
        if isinstance(field, dict):
            return field.get('value', '') or ''
        return str(field)

    out = []
    for wx in (conditions or []):
        intensity  = _val(wx.get('intensity'))
        descriptor = _val(wx.get('descriptor'))
        precip     = _val(wx.get('precipitation'))
        obscur     = _val(wx.get('obscuration'))
        other      = _val(wx.get('other'))

        parts = []
        if intensity:
            parts.append(intensity)

        if descriptor.lower() == 'shower' and precip:
            parts.append(f'showers of {precip.lower()}')
        elif descriptor.lower() == 'thunderstorm':
            parts.append('thunderstorm')
            if precip:
                parts.append(f'with {precip.lower()}')
        elif descriptor.lower() == 'freezing' and precip:
            parts.append(f'freezing {precip.lower()}')
        elif descriptor.lower() == 'blowing' and precip:
            parts.append(f'blowing {precip.lower()}')
        else:
            if descriptor:
                parts.append(descriptor.lower())
            if precip:
                parts.append(precip.lower())

        if obscur:
            parts.append(obscur.lower())
        if other:
            parts.append(other.lower())

        if parts:
            s = ' '.join(parts)
            out.append(s[0].upper() + s[1:])
    return out


def _parse_avwx_taf(data):
    """Transform raw AVWX TAF response — multiply cloud altitudes ×100 and convert times to NZ local."""
    import copy
    from datetime import datetime, timezone as _tz
    from zoneinfo import ZoneInfo
    NZ = ZoneInfo('Pacific/Auckland')
    result = copy.deepcopy(data)

    def _localise(dt_str):
        if not dt_str:
            return None
        try:
            return datetime.fromisoformat(dt_str.replace('Z', '+00:00')).astimezone(NZ)
        except Exception:
            return None

    # Issue time + age
    issue_str = (result.get('time') or {}).get('dt', '')
    if issue_str:
        issued = _localise(issue_str)
        result['issued_local'] = issued
        try:
            age_min = int((datetime.now(_tz.utc) -
                           datetime.fromisoformat(issue_str.replace('Z', '+00:00'))).total_seconds() / 60)
            result['age_minutes'] = age_min
        except Exception:
            result['age_minutes'] = None
    else:
        result['issued_local'] = None
        result['age_minutes'] = None

    # Validity
    result['start_local'] = _localise((result.get('start_time') or {}).get('dt', ''))
    result['end_local']   = _localise((result.get('end_time')   or {}).get('dt', ''))

    # Cloud altitudes ×100 and weather decoding in each forecast period
    for period in (result.get('forecast') or []):
        period['start_local'] = _localise((period.get('start_time') or {}).get('dt', ''))
        period['end_local']   = _localise((period.get('end_time')   or {}).get('dt', ''))
        for cloud in (period.get('clouds') or []):
            if cloud.get('altitude') is not None:
                cloud['altitude'] = cloud['altitude'] * 100
        wx_conds = period.get('weather_conditions') or []
        if wx_conds:
            period['wx_human'] = _decode_wx_conditions(wx_conds)
        else:
            # AVWX doesn't always populate weather_conditions for simple codes
            period['wx_human'] = [
                c.get('value') or c.get('repr', '')
                for c in (period.get('wx_codes') or [])
                if c.get('value') or c.get('repr')
            ]
    return result


def _parse_avwx_metar(data):
    """Return display-friendly dict from a raw AVWX METAR response."""
    from datetime import datetime, timezone as _tz
    from zoneinfo import ZoneInfo
    NZ = ZoneInfo('Pacific/Auckland')
    result = {
        'raw':            data.get('raw', ''),
        'flight_rules':   data.get('flight_rules', ''),
        'wind_dir':       (data.get('wind_direction') or {}).get('value'),
        'wind_dir_repr':  (data.get('wind_direction') or {}).get('repr', ''),
        'wind_speed':     (data.get('wind_speed') or {}).get('value'),
        'wind_gust':      (data.get('wind_gust') or {}).get('value'),
        'visibility':     (data.get('visibility') or {}).get('value'),
        'vis_repr':       (data.get('visibility') or {}).get('repr', ''),
        'temp':           (data.get('temperature') or {}).get('value'),
        'dew':            (data.get('dewpoint') or {}).get('value'),
        'qnh':            (data.get('altimeter') or {}).get('value'),
        'qnh_repr':       (data.get('altimeter') or {}).get('repr', ''),
        'clouds':         [],
        'ceiling':        None,
        'issued_local':   None,
        'age_minutes':    None,
    }
    for c in (data.get('clouds') or []):
        alt_raw = c.get('altitude')
        alt_ft  = alt_raw * 100 if alt_raw is not None else None
        result['clouds'].append({'type': c.get('type', ''), 'altitude': alt_ft})
        if c.get('type') in ('BKN', 'OVC') and result['ceiling'] is None:
            result['ceiling'] = alt_ft
    dt_str = (data.get('time') or {}).get('dt', '')
    if dt_str:
        try:
            issued_utc = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
            result['issued_local'] = issued_utc.astimezone(NZ)
            result['age_minutes'] = int(
                (datetime.now(_tz.utc) - issued_utc).total_seconds() / 60)
        except Exception:
            pass
    return result


@login_required
def avwx_metar_api(request, club_slug):
    """JSON endpoint — METAR summary for the club's home aerodrome (for the Gantt pill)."""
    from django.conf import settings as _s
    club = get_object_or_404(Club, slug=club_slug)
    if not ClubMember.objects.filter(user=request.user, club=club).exists():
        return JsonResponse({'error': 'forbidden'}, status=403)
    api_key = getattr(_s, 'AVWX_API_KEY', '')
    if not api_key:
        return JsonResponse({'error': 'AVWX API key not configured'})
    home_ae = Aerodrome.objects.filter(club=club, is_home=True).first()
    if not home_ae:
        return JsonResponse({'error': 'No home aerodrome set'})
    raw = _fetch_avwx(home_ae.icao_code, 'metar', api_key)
    if 'raw' not in raw:
        return JsonResponse({'error': raw.get('error', 'No data')})
    m = _parse_avwx_metar(raw)
    raw_t = _fetch_avwx(home_ae.icao_code, 'taf', api_key)
    taf_raw = raw_t.get('raw', '') if 'raw' in raw_t else ''

    return JsonResponse({
        'icao':         home_ae.icao_code,
        'raw':          m['raw'],
        'flight_rules': m['flight_rules'],
        'wind_dir':     m['wind_dir_repr'],
        'wind_speed':   m['wind_speed'],
        'wind_gust':    m['wind_gust'],
        'qnh':          m['qnh_repr'],
        'vis':          m['vis_repr'],
        'age_minutes':  m['age_minutes'],
        'issued_local': m['issued_local'].strftime('%H:%M %Z') if m['issued_local'] else None,
        'taf_raw':      taf_raw,
    })


@login_required
def club_weather(request, club_slug):
    from django.conf import settings as _s
    from .models import WeatherWebcam
    club = get_object_or_404(Club, slug=club_slug)
    try:
        member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    if request.method == 'POST' and member.is_staff:
        action = request.POST.get('action', '')
        if action == 'add_webcam':
            name  = request.POST.get('name', '').strip()
            url   = request.POST.get('url', '').strip()
            desc  = request.POST.get('description', '').strip()
            embed = request.POST.get('embed_code', '').strip()
            if name and url:
                WeatherWebcam.objects.create(
                    club=club, name=name, url=url, description=desc, embed_code=embed,
                    display_order=WeatherWebcam.objects.filter(club=club).count(),
                )
        elif action == 'edit_webcam':
            wc = WeatherWebcam.objects.filter(club=club, id=request.POST.get('wc_id')).first()
            if wc:
                wc.name        = request.POST.get('name', wc.name).strip()
                wc.url         = request.POST.get('url', wc.url).strip()
                wc.description = request.POST.get('description', '').strip()
                wc.embed_code  = request.POST.get('embed_code', '').strip()
                wc.save(update_fields=['name', 'url', 'description', 'embed_code'])
        elif action == 'delete_webcam':
            WeatherWebcam.objects.filter(club=club, id=request.POST.get('wc_id')).delete()
        return redirect('core:club_weather', club_slug=club_slug)

    home_ae   = Aerodrome.objects.filter(club=club, is_home=True).first()
    api_key   = getattr(_s, 'AVWX_API_KEY', '')
    metar_data= taf_data = None
    metar_err = taf_err  = None

    if not api_key:
        metar_err = 'AVWX API key not configured — add AVWX_API_KEY to your .env'
    elif not home_ae:
        metar_err = 'No home aerodrome set — tick "Home aerodrome" in Settings › Aerodromes'
    else:
        raw_m = _fetch_avwx(home_ae.icao_code, 'metar', api_key)
        if 'raw' in raw_m:
            metar_data = _parse_avwx_metar(raw_m)
        else:
            metar_err = raw_m.get('error', 'No METAR available')

        raw_t = _fetch_avwx(home_ae.icao_code, 'taf', api_key)
        if 'raw' in raw_t:
            taf_data = _parse_avwx_taf(raw_t)
        else:
            taf_err = raw_t.get('error', 'No TAF available for this aerodrome')

    webcams = WeatherWebcam.objects.filter(club=club, is_active=True)
    return render(request, 'core/weather.html', {
        'club': club, 'club_member': member,
        'home_ae': home_ae,
        'metar_data': metar_data, 'metar_err': metar_err,
        'taf_data':   taf_data,   'taf_err':   taf_err,
        'webcams': webcams,
    })


@login_required
def adsb_proxy(request, club_slug, aircraft_id):
    """
    Live position proxy. Sources tried in order:
      1. OpenSky Network — free, no auth, NZ bounding-box query (callsign match)
      2. adsb.fi         — free, no auth, /v2/registration/[reg] (ADSBExchange v2)
      3. ADSB.one        — free, no auth, /v2/reg/[reg] (ADSBExchange v2)
    All sources fall through on any error (429, timeout, network failure).
    Responses cached server-side 45s. 'tried' list returned for client diagnostics.
    """
    import urllib.request, urllib.error, ssl as _ssl, json as _json
    from django.core.cache import cache
    from django.http import JsonResponse

    club = get_object_or_404(Club, slug=club_slug)
    try:
        ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Access denied'}, status=403)
    ac = get_object_or_404(Aircraft, club=club, id=aircraft_id)
    # Mode S callsign format: strip dash, uppercase  (ZK-ABC → ZKABC)
    callsign = ac.registration.replace('-', '').upper()
    # Official registration kept as-is for database-lookup APIs
    reg = ac.registration.upper()

    cache_key = f'adsb_{club_slug}_{aircraft_id}'
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse(cached)

    ctx = _ssl.create_default_context()
    try:
        import certifi as _certifi
        ctx = _ssl.create_default_context(cafile=_certifi.where())
    except ImportError:
        ctx.check_hostname = False
        ctx.verify_mode = _ssl.CERT_NONE

    def _get(url):
        req = urllib.request.Request(url, headers={'User-Agent': 'ClubHangar/1.0'})
        with urllib.request.urlopen(req, timeout=8, context=ctx) as r:
            return _json.loads(r.read())

    def _parse_v2(ac_list, source_name):
        """Parse an ADSBExchange v2 'ac' array into our response dict."""
        for a in ac_list or []:
            alt = a.get('alt_baro')
            try:
                alt_ft = float(alt)
                if alt_ft <= 0:
                    continue  # on ground or invalid reading
            except (TypeError, ValueError):
                continue  # 'ground', None, or any non-numeric string
            lat, lon = a.get('lat'), a.get('lon')
            if lat is None or lon is None:
                continue
            return {
                'found': True,
                'lat': lat, 'lon': lon,
                'alt_ft': round(alt_ft),
                'speed_kt': round(float(a['gs'])) if a.get('gs') else None,
                'track': round(float(a['track'])) if a.get('track') else None,
                'squawk': a.get('squawk'),
                'seen': round(a.get('seen', 0)),
                'registration': ac.registration,
                'source': source_name,
            }
        return None

    tried = []
    result = None

    # ── Source 1: adsb.fi (/v2/registration/[reg], ADSBExchange v2) ───────────
    try:
        data = _get(f'https://opendata.adsb.fi/api/v2/registration/{reg}')
        tried.append('adsb.fi')
        result = _parse_v2(data.get('ac'), 'adsb.fi')
    except Exception:
        tried.append('adsb.fi (failed)')

    # ── Source 2: adsb.lol (/v2/registration/[reg], ADSBExchange community) ───
    if result is None:
        try:
            data = _get(f'https://api.adsb.lol/v2/registration/{reg}')
            tried.append('adsb.lol')
            result = _parse_v2(data.get('ac'), 'adsb.lol')
        except Exception:
            tried.append('adsb.lol (failed)')

    # ── Source 3: airplanes.live (/v2/registration/[reg], ADSBExchange v2) ────
    if result is None:
        try:
            data = _get(f'https://api.airplanes.live/v2/registration/{reg}')
            tried.append('airplanes.live')
            result = _parse_v2(data.get('ac'), 'airplanes.live')
        except Exception:
            tried.append('airplanes.live (failed)')

    # ── Source 4: ADSB.one (/v2/reg/[callsign], ADSBExchange v2) ─────────────
    if result is None:
        try:
            data = _get(f'https://api.adsb.one/v2/reg/{callsign}')
            tried.append('ADSB.one')
            result = _parse_v2(data.get('ac'), 'ADSB.one')
        except Exception:
            tried.append('ADSB.one (failed)')

    if result is None:
        result = {
            'found': False,
            'registration': ac.registration,
            'note': f'{ac.registration} not detected in any ADS-B source',
        }
    result['tried'] = tried
    cache.set(cache_key, result, 90)
    return JsonResponse(result)


@login_required
def live_positions(request, club_slug):
    """
    Returns ADS-B positions for every active club aircraft currently
    discoverable on ADS-B (not just departed ones). Aircraft with no signal are
    omitted. Fetches all aircraft in parallel (threads); per-aircraft 45s cache.
    """
    import urllib.request, ssl as _ssl, json as _json, threading
    from django.core.cache import cache

    club = get_object_or_404(Club, slug=club_slug)
    try:
        ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return JsonResponse({'error': 'Access denied'}, status=403)

    today = timezone.localdate()
    # Pilot names (when known) come from today's departed bookings.
    pilot_by_aircraft = {}
    for b in (Booking.objects
              .filter(club=club, status='departed', scheduled_start__date=today)
              .select_related('member', 'member__user')
              .order_by('scheduled_start')):
        if b.member and b.member.user:
            pilot_by_aircraft[b.aircraft_id] = (
                b.member.user.get_full_name() or b.member.user.username)

    # Try to discover ALL active club aircraft on ADS-B, not just departed ones.
    aircraft_objs = list(
        Aircraft.objects.filter(club=club).exclude(status='retired')
        .select_related('aircraft_type')
    )

    ctx = _ssl.create_default_context()
    try:
        import certifi as _certifi
        ctx = _ssl.create_default_context(cafile=_certifi.where())
    except ImportError:
        ctx.check_hostname = False
        ctx.verify_mode = _ssl.CERT_NONE

    def _get(url):
        req = urllib.request.Request(url, headers={'User-Agent': 'ClubHangar/1.0'})
        with urllib.request.urlopen(req, timeout=8, context=ctx) as r:
            return _json.loads(r.read())

    def _parse_v2(ac_list, reg):
        for a in ac_list or []:
            alt = a.get('alt_baro')
            try:
                alt_ft = float(alt)
                if alt_ft <= 0:
                    continue
            except (TypeError, ValueError):
                continue  # 'ground', None, or non-numeric
            lat, lon = a.get('lat'), a.get('lon')
            if lat is None or lon is None:
                continue
            return {
                'found': True,
                'lat': lat, 'lon': lon,
                'alt_ft': round(alt_ft),
                'speed_kt': round(float(a['gs'])) if a.get('gs') else None,
                'track': round(float(a['track'])) if a.get('track') else None,
                'registration': reg,
            }
        return None

    def fetch_one(ac):
        reg = ac.registration.upper()
        callsign = reg.replace('-', '')
        cache_key = f'adsb_{club_slug}_{ac.id}'
        cached = cache.get(cache_key)
        if cached is not None:
            return cached
        result = None
        sources = [
            f'https://opendata.adsb.fi/api/v2/registration/{reg}',
            f'https://api.adsb.lol/v2/registration/{reg}',
            f'https://api.airplanes.live/v2/registration/{reg}',
            f'https://api.adsb.one/v2/reg/{callsign}',
        ]
        for url in sources:
            try:
                result = _parse_v2(_get(url).get('ac'), reg)
                if result:
                    break
            except Exception:
                pass
        if result is None:
            result = {'found': False, 'registration': reg}
        cache.set(cache_key, result, 90)
        return result

    positions = {}

    def worker(ac):
        positions[ac.id] = fetch_one(ac)

    threads = [threading.Thread(target=worker, args=(ac,), daemon=True)
               for ac in aircraft_objs]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=10)

    aircraft_list = []
    for ac in aircraft_objs:
        pos = positions.get(ac.id) or {'found': False, 'registration': ac.registration}
        at = ac.aircraft_type
        engine_count = ac.engine_count or 1
        # Icon is decided purely by engine count: 2+ = twin, otherwise single.
        icon = 'twin' if engine_count >= 2 else 'high_wing'
        aircraft_list.append({
            'registration': ac.registration,
            'aircraft_type': at.name if at else '',
            'icao_designator': (at.icao_designator or '').upper() if at else '',
            'engine_count': engine_count,
            'icon': icon,
            'pilot_name': pilot_by_aircraft.get(ac.id, ''),
            **pos,
        })

    return JsonResponse({'aircraft': aircraft_list})


@login_required
def live_map(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    try:
        member = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('core:club_list')
    home_ae = Aerodrome.objects.filter(club=club, is_home=True).first()
    return render(request, 'core/live_map.html', {
        'club': club,
        'club_member': member,
        'home_ae': home_ae,
    })


# ─────────────────────────────────────────────────────────────────────────────
# OCCURRENCE / SAFETY REPORTING
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def occurrence_submit(request, club_slug):
    """Any member can submit an occurrence report."""
    from .models import OccurrenceType, OccurrenceReport, Aircraft as _Aircraft
    from datetime import date as _date
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    error = success = ''
    occ_types = OccurrenceType.objects.filter(club=club, is_active=True)
    aircraft_list = _Aircraft.objects.filter(club=club, status='online').order_by('registration')

    # Recent bookings for this member (last 30 days) for optional link
    from datetime import timedelta
    recent_bookings = (Booking.objects
                       .filter(club=club, member=actor,
                               scheduled_start__date__gte=timezone.localdate() - timedelta(days=30))
                       .exclude(status='cancelled')
                       .select_related('aircraft', 'flight_type')
                       .order_by('-scheduled_start')[:10])

    if request.method == 'POST':
        from decimal import InvalidOperation
        ot_id = request.POST.get('occurrence_type')
        date_str = request.POST.get('date_of_occurrence', '').strip()
        description = request.POST.get('description', '').strip()
        if not ot_id or not date_str or not description:
            error = 'Occurrence type, date, and description are required.'
        else:
            try:
                ot = OccurrenceType.objects.get(id=ot_id, club=club)
                from datetime import date as _d
                occ_date = _d.fromisoformat(date_str)
                time_str = request.POST.get('time_of_occurrence', '').strip()
                occ_time = None
                if time_str:
                    from datetime import time as _t
                    h, m = time_str.split(':')
                    occ_time = _t(int(h), int(m))
                ac_id = request.POST.get('aircraft')
                bk_id = request.POST.get('related_booking')
                ac = _Aircraft.objects.filter(id=ac_id, club=club).first() if ac_id else None
                bk = Booking.objects.filter(id=bk_id, club=club, member=actor).first() if bk_id else None
                report = OccurrenceReport.objects.create(
                    club=club,
                    occurrence_type=ot,
                    reported_by=actor,
                    status=OccurrenceReport.STATUS_SUBMITTED,
                    date_of_occurrence=occ_date,
                    time_of_occurrence=occ_time,
                    location=request.POST.get('location', '').strip(),
                    aircraft=ac,
                    related_booking=bk,
                    description=description,
                    immediate_action=request.POST.get('immediate_action', '').strip(),
                    is_safety_risk=request.POST.get('is_safety_risk') == 'on',
                )
                OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Submitted')
                from .email_notifications import occurrence_submitted as _email_occ
                _email_occ(report)
                if request.path.startswith('/app/'):
                    from django.urls import reverse as _rev
                    return redirect(_rev('core:app_home', kwargs={'club_slug': club_slug}) + '?event_reported=1')
                return redirect(f"{request.path}?saved=1")
            except Exception as e:
                error = f'Error saving report: {e}'

    is_app    = request.GET.get('app') == '1' or request.POST.get('_from_app') == '1' or request.path.startswith('/app/')
    is_inline = request.GET.get('inline') == '1'
    if is_app:
        base_template = 'core/app/base.html'
        template = 'core/app/occurrence_submit.html'
    elif is_inline:
        base_template = 'core/base_inline.html'
        template = 'core/occurrence_submit.html'
    else:
        base_template = 'core/base.html'
        template = 'core/occurrence_submit.html'
    return render(request, template, {
        'club': club, 'club_member': actor,
        'occ_types': occ_types,
        'aircraft_list': aircraft_list,
        'recent_bookings': recent_bookings,
        'error': error,
        'today': timezone.localdate().isoformat(),
        'base_template': base_template,
        'is_inline': is_inline,
        'is_app': is_app,
        'back': request.GET.get('back', ''),
    })


@login_required
def occurrence_list(request, club_slug):
    """Admin/instructor view — all reports with filter and review."""
    from .models import OccurrenceType, OccurrenceReport
    from django.core.paginator import Paginator
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err

    from urllib.parse import urlencode as _ue
    f_status      = request.GET.get('status', '')
    f_type        = request.GET.get('type', '')
    f_safety_risk = request.GET.get('safety_risk', '')
    sort          = request.GET.get('sort', 'date')
    sort_dir      = request.GET.get('dir', 'desc')
    _OCC_SORT = {
        'date':     ('reported_at',),
        'type':     ('occurrence_type__name', '-reported_at'),
        'aircraft': ('aircraft__registration', '-reported_at'),
        'status':   ('status', '-reported_at'),
    }
    if sort not in _OCC_SORT:
        sort = 'date'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'desc'
    _occ_order = _OCC_SORT[sort]
    if sort_dir == 'desc':
        _occ_order = tuple(f[1:] if f.startswith('-') else '-'+f for f in _occ_order)

    qs = (OccurrenceReport.objects
          .filter(club=club)
          .select_related('occurrence_type', 'reported_by__user', 'aircraft', 'reviewed_by')
          .order_by(*_occ_order))
    if f_status == 'all':
        pass
    elif f_status:
        qs = qs.filter(status=f_status)
    else:
        qs = qs.filter(status__in=[OccurrenceReport.STATUS_SUBMITTED, OccurrenceReport.STATUS_REVIEWED])
    if f_type:
        qs = qs.filter(occurrence_type_id=f_type)
    if f_safety_risk == '1':
        qs = qs.filter(is_safety_risk=True)

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page'))
    _base_qs  = _ue({k: v for k, v in request.GET.items() if k not in ('page', 'sort', 'dir') and v})

    occ_types = OccurrenceType.objects.filter(club=club, is_active=True)
    return render(request, 'core/occurrence_list.html', {
        'club': club, 'club_member': actor,
        'page_obj': page_obj,
        'occ_types': occ_types,
        'f_status': f_status, 'f_type': f_type, 'f_safety_risk': f_safety_risk,
        'status_choices': OccurrenceReport.STATUS_CHOICES,
        'open_count': OccurrenceReport.objects.filter(club=club, status=OccurrenceReport.STATUS_SUBMITTED).count(),
        'sort': sort, 'sort_dir': sort_dir, 'base_qs': _base_qs,
    })


@login_required
def occurrence_detail(request, club_slug, report_id):
    """View + review a single occurrence report."""
    from .models import OccurrenceReport
    from django.utils import timezone as _tz
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')

    report = get_object_or_404(OccurrenceReport, club=club, id=report_id)

    # Members can only view their own; manage-access users see all
    if report.reported_by != actor and not actor.can_access_manage:
        return render(request, 'core/no_access.html', {'club': club}, status=403)

    is_inline = request.GET.get('inline') == '1' or request.POST.get('inline') == '1'
    now = _tz.now()

    if request.method == 'POST' and actor.can_access_manage:
        act = request.POST.get('action')
        notes = request.POST.get('review_notes', '').strip()

        if act == 'save_notes':
            if notes != report.review_notes:
                report.review_notes = notes
                if not report.reviewed_by:
                    report.reviewed_by = request.user
                    report.reviewed_at = now
                report.save(update_fields=['review_notes', 'reviewed_by', 'reviewed_at'])
                OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Notes updated', note=notes[:200])

        elif act == 'set_safety_risk':
            report.is_safety_risk = request.POST.get('is_safety_risk') == '1'
            report.save(update_fields=['is_safety_risk'])
            OccurrenceAuditEntry.objects.create(report=report, actor=actor,
                verb='Flagged as safety risk' if report.is_safety_risk else 'Safety risk flag removed')

        elif act == 'mark_reviewed':
            report.status = OccurrenceReport.STATUS_REVIEWED
            report.reviewed_by = request.user
            report.reviewed_at = now
            report.review_notes = notes
            report.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'review_notes'])
            OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Marked reviewed', note=notes[:200])

        elif act == 'close_no_action':
            report.status = OccurrenceReport.STATUS_CLOSED
            report.reviewed_by = request.user
            report.reviewed_at = now
            report.review_notes = notes
            report.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'review_notes'])
            OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Closed — no action required', note=notes[:200])

        elif act == 'close':
            if report.all_actions_resolved:
                report.status = OccurrenceReport.STATUS_CLOSED
                report.save(update_fields=['status'])
                OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Closed')

        elif act == 'reopen':
            report.status = OccurrenceReport.STATUS_SUBMITTED
            report.save(update_fields=['status'])
            OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Reopened')

        elif act == 'add_action':
            desc = request.POST.get('action_desc', '').strip()
            if desc:
                assigned_id = request.POST.get('assigned_to') or None
                assigned = ClubMember.objects.filter(club=club, id=assigned_id).first() if assigned_id else None
                due_raw = request.POST.get('due_date', '').strip()
                due = None
                if due_raw:
                    try:
                        from datetime import date as _d; due = _d.fromisoformat(due_raw)
                    except ValueError:
                        pass
                oa = OccurrenceAction.objects.create(
                    report=report, description=desc,
                    assigned_to=assigned, due_date=due, created_by=actor,
                )
                verb_note = f"Action added: {desc[:80]}"
                if assigned:
                    verb_note += f" — assigned to {assigned.user.get_full_name()}"
                OccurrenceAuditEntry.objects.create(report=report, actor=actor, verb='Action added', note=verb_note)
                # Notify assigned instructor
                if assigned:
                    from .models import Notification as _Notif
                    from django.urls import reverse as _rev
                    _Notif.objects.create(
                        club_member=assigned,
                        notification_type='safety_action',
                        subject='Safety action assigned to you',
                        body=f"Action: {desc[:120]}",
                        action_url=_rev('core:occurrence_detail', args=[club.slug, report.id]),
                    )

        elif act == 'complete_action':
            oa = OccurrenceAction.objects.filter(id=request.POST.get('action_id'), report=report).first()
            if oa and oa.status == OccurrenceAction.STATUS_OPEN:
                oa.status = OccurrenceAction.STATUS_COMPLETE
                oa.completed_by = request.user
                oa.completed_at = now
                oa.save()
                OccurrenceAuditEntry.objects.create(report=report, actor=actor,
                    verb='Action completed', note=oa.description[:80])

        elif act == 'override_action':
            oa = OccurrenceAction.objects.filter(id=request.POST.get('action_id'), report=report).first()
            if oa and oa.status == OccurrenceAction.STATUS_OPEN:
                oa.status = OccurrenceAction.STATUS_OVERRIDDEN
                oa.override_note = request.POST.get('override_note', '').strip()
                oa.save()
                OccurrenceAuditEntry.objects.create(report=report, actor=actor,
                    verb='Action overridden', note=oa.override_note[:200])

        # Terminal actions (close, close_no_action) drop inline=1 so the overlay closes.
        # Non-terminal actions (save_notes, mark_reviewed, etc.) keep inline=1 to stay open.
        _terminal = act in ('close', 'close_no_action', 'reopen')
        return redirect(f"{request.path}?{'inline=1&' if (is_inline and not _terminal) else ''}saved=1")

    instructors = ClubMember.objects.filter(club=club, is_on_instructor_roster=True).select_related('user')
    base_template = 'core/base_inline.html' if is_inline else 'core/base.html'
    return render(request, 'core/occurrence_detail.html', {
        'club': club, 'club_member': actor,
        'report': report,
        'actions': report.actions.select_related('assigned_to__user', 'completed_by').all(),
        'audit': report.audit_entries.select_related('actor__user').all(),
        'instructors': instructors,
        'back': request.GET.get('back', ''),
        'base_template': base_template,
        'is_inline': is_inline,
    })


@login_required
def occurrence_export(request, club_slug):
    """CSV export of all occurrence reports for this club."""
    import csv as _csv
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err
    logger.info('export club=%s type=occurrences by=%s', club.slug, request.user.email)

    from django.http import StreamingHttpResponse

    qs = (OccurrenceReport.objects
          .filter(club=club)
          .select_related('occurrence_type', 'reported_by__user', 'aircraft', 'reviewed_by')
          .order_by('-date_of_occurrence'))

    def _rows():
        yield ['ID', 'Type', 'Date', 'Time', 'Location', 'Aircraft',
               'Reported by', 'Status', 'Reported at',
               'Description', 'Immediate action', 'Review notes']
        for r in qs:
            yield [
                r.id,
                r.occurrence_type.name,
                r.date_of_occurrence.isoformat(),
                r.time_of_occurrence.strftime('%H:%M') if r.time_of_occurrence else '',
                r.location,
                r.aircraft.registration if r.aircraft else '',
                r.reported_by.user.get_full_name(),
                r.get_status_display(),
                r.reported_at.strftime('%Y-%m-%d %H:%M'),
                r.description,
                r.immediate_action,
                r.review_notes,
            ]

    class EchoWriter:
        def write(self, value): return value

    writer = _csv.writer(EchoWriter())
    response = StreamingHttpResponse(
        (writer.writerow(row) for row in _rows()),
        content_type='text/csv',
    )
    response['Content-Disposition'] = f'attachment; filename="occurrences-{club.slug}.csv"'
    return response


@login_required
def occurrence_actions(request, club_slug):
    """Safety action items — open OccurrenceActions assigned to anyone (manage) or just this user."""
    from .models import OccurrenceAction as _OA
    from django.utils import timezone as _tz
    club = get_object_or_404(Club, slug=club_slug)
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        return redirect('login')
    if err := require_manage(actor, club, request): return err

    if request.method == 'POST':
        act = request.POST.get('action')
        oa = _OA.objects.filter(id=request.POST.get('action_id'), report__club=club).first()
        if oa and oa.status == _OA.STATUS_OPEN:
            if act == 'complete_action':
                oa.status = _OA.STATUS_COMPLETE
                oa.completed_by = request.user
                oa.completed_at = _tz.now()
                oa.save()
                from .models import OccurrenceAuditEntry
                OccurrenceAuditEntry.objects.create(report=oa.report, actor=actor,
                    verb='Action completed', note=oa.description[:80])
            elif act == 'override_action':
                note = request.POST.get('override_note', '').strip()
                if note:
                    oa.status = _OA.STATUS_OVERRIDDEN
                    oa.override_note = note
                    oa.save()
                    from .models import OccurrenceAuditEntry
                    OccurrenceAuditEntry.objects.create(report=oa.report, actor=actor,
                        verb='Action overridden', note=note[:200])
        return redirect('core:occurrence_actions', club_slug=club_slug)

    f_assigned = request.GET.get('assigned', '')
    qs = (_OA.objects
          .filter(report__club=club, status=_OA.STATUS_OPEN)
          .select_related('report__occurrence_type', 'report__reported_by__user',
                          'assigned_to__user', 'report__aircraft')
          .order_by('due_date', 'created_at'))
    if f_assigned:
        qs = qs.filter(assigned_to_id=f_assigned)

    from datetime import date as _date
    instructors = ClubMember.objects.filter(club=club, is_on_instructor_roster=True).select_related('user')
    return render(request, 'core/occurrence_actions.html', {
        'club': club, 'club_member': actor,
        'actions': qs,
        'f_assigned': f_assigned,
        'instructors': instructors,
        'today': timezone.localdate(),
    })


# ── Mobile PWA views ──────────────────────────────────────────────────────────

def _app_actor(request, club_slug):
    """Return (club, actor) or raise redirect. Shared by all app views."""
    club = get_object_or_404(Club, slug=club_slug)
    if not request.user.is_authenticated:
        return club, None
    try:
        actor = ClubMember.objects.get(user=request.user, club=club)
    except ClubMember.DoesNotExist:
        actor = None
    return club, actor


@login_required
@login_required
def app_root(request):
    """Mobile app entry point (start_url). Routes a multi-club member to a
    club picker; a single-club member straight into their club's app home."""
    memberships = (ClubMember.objects
                   .filter(user=request.user)
                   .select_related('club', 'role')
                   .order_by('club__name'))
    if not memberships.exists():
        return render(request, 'core/no_access.html')
    if memberships.count() == 1:
        return redirect('core:app_home', club_slug=memberships.first().club.slug)
    return render(request, 'core/app/club_select.html', {'memberships': memberships})


def app_home(request, club_slug):
    from datetime import date as _d, timedelta as _td
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')
    today = timezone.localdate()

    next_booking = (Booking.objects
                    .filter(member=actor, club=club, scheduled_start__date__gte=today)
                    .exclude(status='cancelled')
                    .select_related('aircraft', 'flight_type', 'instructor')
                    .order_by('scheduled_start')
                    .first())

    pending_count = 0
    if actor.is_instructor:
        pending_count = (Booking.objects
                         .filter(club=club, status='pending',
                                 instructor=request.user,
                                 scheduled_start__date__gte=today)
                         .count())

    try:
        balance = actor.account.balance
    except Exception:
        balance = None

    # Debt summary for home banner
    from .models import FlightCompletion as _FC, Invoice as _Inv, Announcement as _Ann
    from .models import MemberCredential as _Cred
    from django.db.models import Sum as _Sum, Count as _Count
    _fc_agg = (
        _FC.objects
        .filter(booking__member=actor, booking__club=club,
                total_charge__gt=0, paid_at__isnull=True)
        .exclude(payment_method='invoice')
        .aggregate(total=_Sum('total_charge'), paid=_Sum('amount_paid'), cnt=_Count('id'))
    )
    from decimal import Decimal as _D
    flight_debt  = max(_D('0'), (_fc_agg['total'] or _D('0')) - (_fc_agg['paid'] or _D('0')))
    flight_count = _fc_agg['cnt'] or 0
    invoice_count = _Inv.objects.filter(member=actor, club=club, status='sent').count()

    # Last completed flight
    last_flight = (Booking.objects
                   .filter(member=actor, club=club, status='completed')
                   .select_related('aircraft', 'flight_completion')
                   .order_by('-scheduled_start')
                   .first())

    # Credentials expiring within 60 days
    expiry_threshold = today + _td(days=60)
    expiring_creds = list(_Cred.objects.filter(
        member=actor.user,
        expiry_date__isnull=False,
        expiry_date__lte=expiry_threshold,
    ).select_related('credential_type').order_by('expiry_date'))

    # Active announcements (pinned first, then newest)
    from django.db.models import Q as _Q
    announcements = list(_Ann.objects.filter(
        club=club
    ).filter(
        _Q(expires_at__isnull=True) | _Q(expires_at__gte=today)
    ).order_by('-is_pinned', '-created_at')[:10])

    # Fleet / instructor summary
    online_aircraft = list(Aircraft.objects.filter(club=club, status='online')
                           .select_related('aircraft_type'))
    ac_by_type = {}
    for ac in online_aircraft:
        label = ac.aircraft_type.name if ac.aircraft_type else 'Aircraft'
        ac_by_type[label] = ac_by_type.get(label, 0) + 1
    fleet_summary = ', '.join(
        f"{cnt} {typ}" for typ, cnt in sorted(ac_by_type.items())
    ) if ac_by_type else None

    instr_on_roster = ClubMember.objects.filter(
        club=club, is_on_instructor_roster=True
    ).count()

    # Quick slots check: any aircraft with ≥60 min free today after now
    from .availability import find_free_spans
    _spans = find_free_spans(club, today, today, min_minutes=60)
    slots_available_today = bool(_spans)
    first_slot_time = None
    available_ac_regs = []
    if _spans:
        _all_starts = [sp['start'] for row in _spans for sp in row['spans']]
        if _all_starts:
            first_slot_time = timezone.localtime(min(_all_starts)).strftime('%H:%M')
        available_ac_regs = [row['aircraft'].registration for row in _spans]

    # Greeting based on local hour
    _local_hour = timezone.localtime(timezone.now()).hour
    if _local_hour < 12:
        greeting_word = 'Good morning'
    elif _local_hour < 17:
        greeting_word = 'Good afternoon'
    else:
        greeting_word = 'Good evening'

    return render(request, 'core/app/home.html', {
        'club': club, 'club_member': actor,
        'next_booking': next_booking,
        'today': today,
        'pending_count': pending_count,
        'balance': balance,
        'flight_debt': flight_debt,
        'flight_count': flight_count,
        'invoice_count': invoice_count,
        'last_flight': last_flight,
        'expiring_creds': expiring_creds,
        'announcements': announcements,
        'fleet_summary': fleet_summary,
        'instr_on_roster': instr_on_roster,
        'slots_available_today': slots_available_today,
        'first_slot_time': first_slot_time,
        'available_ac_regs': available_ac_regs,
        'greeting_word': greeting_word,
    })


@login_required
def app_schedule(request, club_slug, year=None, month=None, day=None):
    from datetime import date as _d, timedelta as _td
    from django.urls import reverse
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    today = timezone.localdate()
    selected = _d(year, month, day) if year else today

    # Week strip — Mon–Sun of selected week
    week_start = selected - _td(days=selected.weekday())
    week_days = []
    for i in range(7):
        d = week_start + _td(days=i)
        week_days.append({
            'date': d,
            'name': d.strftime('%a'),
            'num': d.day,
            'is_today': d == today,
            'is_selected': d == selected,
            'url': reverse('core:app_schedule_date', args=[club_slug, d.year, d.month, d.day]),
        })

    # Prev / next day
    prev_d = selected - _td(days=1)
    next_d = selected + _td(days=1)

    # Aircraft (online)
    aircraft_qs = (Aircraft.objects
                   .filter(club=club, status='online')
                   .select_related('aircraft_type')
                   .order_by('registration'))

    # Bookings for the day
    day_bookings = (Booking.objects
                    .filter(club=club, scheduled_start__date=selected)
                    .exclude(status='cancelled')
                    .select_related('member__user', 'aircraft', 'flight_type', 'instructor')
                    .order_by('scheduled_start'))

    # Timeline geometry
    DAY_START = 7
    DAY_END   = 20
    PX_PER_HR = 60
    TL_HEIGHT = (DAY_END - DAY_START) * PX_PER_HR  # 780px

    def to_px(dt):
        loc = timezone.localtime(dt)
        hrs = loc.hour + loc.minute / 60
        return max(0, min(TL_HEIGHT, round((hrs - DAY_START) * PX_PER_HR)))

    config = get_config(club)
    def _blend(hex_color, pct):
        r1,g1,b1 = int(hex_color[1:3],16),int(hex_color[3:5],16),int(hex_color[5:7],16)
        f = pct/100
        return '#{:02x}{:02x}{:02x}'.format(round(r1*f+255*(1-f)),round(g1*f+255*(1-f)),round(b1*f+255*(1-f)))
    _p = config.theme_primary
    status_color = {
        'pending':   _blend(_p, 18),
        'confirmed': _blend(_p, 42),
        'departed':  _p,
        'completed': '#f1efe8',
    }

    from .models import SlotWatch as _SlotWatch
    watched_ids = set(
        _SlotWatch.objects
        .filter(club_member=actor)
        .values_list('booking_id', flat=True)
    )

    aircraft_data = []
    for ac in aircraft_qs:
        blocks = []
        for b in day_bookings:
            if b.aircraft_id != ac.id:
                continue
            top    = to_px(b.scheduled_start)
            height = max(22, to_px(b.scheduled_end) - top)
            blocks.append({
                'id':            b.id,
                'top':           top,
                'height':        height,
                'is_mine':       b.member == actor,
                'name':          'You' if b.member == actor else b.member.user.get_short_name(),
                'member_full':   b.member.user.get_full_name(),
                'time':          '{}–{}'.format(
                    timezone.localtime(b.scheduled_start).strftime('%H:%M'),
                    timezone.localtime(b.scheduled_end).strftime('%H:%M')),
                'status':        b.status,
                'color':         status_color.get(b.status, _blend(_p, 42)),
                'solo':          not bool(b.instructor),
                'instructor':    b.instructor.get_short_name() if b.instructor else None,
                'aircraft_reg':  ac.registration,
                'flight_type':   b.flight_type.name if b.flight_type else '',
                'watched':       b.id in watched_ids,
            })
        aircraft_data.append({'ac': ac, 'blocks': blocks})

    time_labels = [
        {'label': '{:02d}:00'.format(h), 'top': (h - DAY_START) * PX_PER_HR}
        for h in range(DAY_START, DAY_END + 1)
    ]

    # Block-outs for the selected day (prefetch instructors M2M for affects_instructor())
    from .models import BlockOut as _AppBlockOut
    from django.db.models import Q as _Q
    _bo_candidates = list(_AppBlockOut.objects.filter(club=club).filter(
        _Q(recurrence='one_off', date=selected) |
        _Q(recurrence='weekly') |
        _Q(recurrence='daily')
    ).select_related('blockout_type').prefetch_related('instructors'))
    full_blockouts = [
        bo for bo in _bo_candidates
        if bo.scope == 'all' and bo.applies_on(selected)
    ]

    # Instructor availability bars
    from .models import InstructorAvailability as _InstrAv
    from datetime import datetime as _dt2, time as _t2
    _INSTR_COLORS = [
        '#2563eb', '#16a34a', '#dc2626', '#d97706',
        '#7c3aed', '#0891b2', '#be185d', '#065f46', '#92400e', '#1e40af',
    ]
    _day_s_dt = timezone.make_aware(_dt2.combine(selected, _t2(DAY_START, 0)))
    _day_e_dt = timezone.make_aware(_dt2.combine(selected, _t2(DAY_END,   0)))
    _roster = list(
        ClubMember.objects
        .filter(club=club, is_on_instructor_roster=True)
        .select_related('user')
        .prefetch_related('availability_windows')
        .order_by('user__last_name')
    )

    def _subtract_blockout_from_bars(bars, bo_top_px, bo_bot_px):
        """Remove a block-out pixel range from a list of {top, height} bars."""
        result = []
        for bar in bars:
            bar_top = bar['top']
            bar_bot = bar['top'] + bar['height']
            if bo_top_px >= bar_bot or bo_bot_px <= bar_top:
                result.append(bar)
            else:
                if bar_top < bo_top_px:
                    result.append({'top': bar_top, 'height': bo_top_px - bar_top})
                if bo_bot_px < bar_bot:
                    result.append({'top': bo_bot_px, 'height': bar_bot - bo_bot_px})
        return result

    instructor_bars = []
    for _i, _instr in enumerate(_roster):
        _color    = _INSTR_COLORS[_i % len(_INSTR_COLORS)]
        _initials = (_instr.user.first_name[:1] + _instr.user.last_name[:1]).upper() or '?'
        _av_wins  = list(_instr.availability_windows.all())
        if not _av_wins:
            _bars = [{'top': 0, 'height': TL_HEIGHT}]
        else:
            _bars = []
            for _w in _av_wins:
                _iv = _w.interval_on(selected, _day_s_dt, _day_e_dt)
                if _iv:
                    _top = to_px(_iv[0])
                    _h   = max(4, to_px(_iv[1]) - _top)
                    _bars.append({'top': _top, 'height': _h})
        # Subtract block-outs that affect this instructor
        for _bo in _bo_candidates:
            if not _bo.applies_on(selected):
                continue
            if not _bo.affects_instructor(_instr.user):
                continue
            _bo_iv = _bo.interval_on(selected)
            if _bo_iv:
                _bars = _subtract_blockout_from_bars(_bars, to_px(_bo_iv[0]), to_px(_bo_iv[1]))
        # Subtract existing bookings assigned to this instructor
        for _bk in day_bookings:
            if _bk.instructor_id == _instr.user_id:
                _bars = _subtract_blockout_from_bars(_bars, to_px(_bk.scheduled_start), to_px(_bk.scheduled_end))
        instructor_bars.append({
            'initials': _initials,
            'name':     _instr.user.get_short_name(),
            'color':    _color,
            'bars':     _bars,
        })

    from .permissions import check_booking_block
    bb_blocked, bb_msg = check_booking_block(actor, get_config(club))

    return render(request, 'core/app/schedule.html', {
        'club': club, 'club_member': actor,
        'selected': selected,
        'today': today,
        'week_days': week_days,
        'aircraft_data': aircraft_data,
        'time_labels': time_labels,
        'tl_height': TL_HEIGHT,
        'px_per_hr': PX_PER_HR,
        'day_start': DAY_START,
        'prev_url': reverse('core:app_schedule_date', args=[club.slug, prev_d.year, prev_d.month, prev_d.day]),
        'next_url': reverse('core:app_schedule_date', args=[club.slug, next_d.year, next_d.month, next_d.day]),
        'prev_date': prev_d,
        'next_date': next_d,
        'full_blockouts': full_blockouts,
        'instructor_bars': instructor_bars,
        'booking_blocked': bb_blocked,
        'booking_block_msg': bb_msg,
    })


@login_required
def app_bookings(request, club_slug):
    from datetime import date as _d
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')
    today = timezone.localdate()

    if request.method == 'POST' and request.POST.get('action') == 'cancel_booking':
        booking_id = request.POST.get('booking_id')
        booking = get_object_or_404(Booking, id=booking_id, club=club, member=actor)
        if booking.status in ('pending', 'confirmed'):
            reason = request.POST.get('cancellation_reason', 'no_longer_required')
            result = booking_service.cancel(booking, request.user, reason=reason)
            if result.ok:
                _audit(booking, request.user, 'cancelled')
                from .email_notifications import booking_cancelled as _email_cancelled
                _email_cancelled(booking, reason=reason)
        return redirect('core:app_bookings', club_slug=club_slug)

    upcoming = (Booking.objects
                .filter(member=actor, club=club, scheduled_start__date__gte=today)
                .exclude(status='cancelled')
                .select_related('aircraft', 'flight_type', 'instructor', 'declaration')
                .order_by('scheduled_start')[:20])

    past = (Booking.objects
            .filter(member=actor, club=club, scheduled_start__date__lt=today)
            .exclude(status='cancelled')
            .select_related('aircraft', 'flight_type')
            .order_by('-scheduled_start')[:20])

    return render(request, 'core/app/bookings.html', {
        'club': club, 'club_member': actor,
        'upcoming': upcoming,
        'past': past,
        'today': today,
    })


@login_required
def app_profile(request, club_slug):
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    from .models import MemberCredential, NotificationPreference as _NP
    _profile_url = redirect('core:app_profile', club_slug=club_slug).url

    if request.method == 'POST':
        action = request.POST.get('action', '')

    credentials = list(MemberCredential.objects
                       .filter(member=actor.user)
                       .select_related('credential_type', 'aircraft_type')
                       .order_by('expiry_date'))
    if any(c.is_expired for c in credentials):
        cred_status = 'red'
    elif any(c.is_expiring_soon for c in credentials):
        cred_status = 'amber'
    elif credentials:
        cred_status = 'green'
    else:
        cred_status = 'grey'

    today = timezone.localdate()

    sub_exp = actor.subscription_expires
    sub_status = None
    sub_days   = None
    if sub_exp:
        delta = (sub_exp - today).days
        if delta < 0:
            sub_status = 'expired'
        elif delta <= 60:
            sub_status = 'expiring'
            sub_days = delta
        else:
            sub_status = 'ok'

    return render(request, 'core/app/profile.html', {
        'club': club, 'club_member': actor,
        'credentials': credentials,
        'cred_status': cred_status,
        'today': today,
        'saved': request.GET.get('saved') == '1',
        'sub_status': sub_status,
        'sub_days': sub_days,
    })


@login_required
def app_notifications(request, club_slug):
    from .models import NotificationPreference as _NP
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')
    from django.urls import reverse as _rev
    _notif_url = _rev('core:app_notifications', args=[club_slug])
    if request.method == 'POST':
        pref, _ = _NP.objects.get_or_create(club_member=actor)
        _toggle_fields = [
            'booking_confirmed', 'booking_cancelled', 'booking_reminder',
            'credential_expiring', 'subscription_expiring',
            'instructor_booking_urgent', 'instructor_booking_upcoming',
            'maintenance_alert', 'lapsed_credentials', 'slot_released',
            'payment_reminder', 'invoice_sent',
        ]
        for f in _toggle_fields:
            setattr(pref, f, request.POST.get(f) == 'on')
        pref.save()
        return redirect(_notif_url + '?saved=1')
    try:
        notification_pref = actor.notification_prefs
    except _NP.DoesNotExist:
        notification_pref = None
    _raw_toggles = [
        ('booking_confirmed',     'Booking confirmed',                        True),
        ('booking_cancelled',     'Booking cancelled',                        True),
        ('booking_reminder',      'Booking reminder (day before)',             True),
        ('credential_expiring',   'Credential expiring soon',                 True),
        ('subscription_expiring', 'Subscription expiring',                    True),
        ('payment_reminder',      'Account balance reminder',                 True),
        ('invoice_sent',          'Invoice issued to me',                     True),
        ('slot_released',         'Slot freed up by another member (opt-in)', False),
    ]
    if actor.is_instructor:
        _raw_toggles += [
            ('instructor_booking_urgent',   'New booking assigned — urgent (≤2 days)',    True),
            ('instructor_booking_upcoming', 'New booking assigned — upcoming (≤10 days)', True),
            ('maintenance_alert',           'Maintenance alert (amber/red items)',         True),
            ('lapsed_credentials',          'Member lapsed credentials — flying today',   True),
        ]
    notification_toggle_fields = [
        {'field': f, 'label': l,
         'enabled': getattr(notification_pref, f, default) if notification_pref else default}
        for f, l, default in _raw_toggles
    ]
    return render(request, 'core/app/notifications.html', {
        'club': club, 'club_member': actor,
        'notification_toggle_fields': notification_toggle_fields,
        'saved': request.GET.get('saved') == '1',
    })


@login_required
def app_account(request, club_slug):
    """Mobile app — account balance, transactions, outstanding amounts."""
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    from .models import AccountTransaction, FlightCompletion as _FC, Invoice as _Inv
    from datetime import timedelta as _timedelta
    from django.core.paginator import Paginator as _Paginator

    transactions_page = None
    balance = None
    try:
        acct = actor.account
        _tx_qs = AccountTransaction.objects.filter(account=acct).order_by('-created_at')
        _paginator = _Paginator(_tx_qs, 20)
        _page_num = request.GET.get('page', 1)
        transactions_page = _paginator.get_page(_page_num)
        balance = acct.balance
    except Exception:
        pass

    today = timezone.localdate()
    _thirty_days_ago = today - _timedelta(days=30)

    unpaid_flights = list(
        _FC.objects
        .filter(booking__member=actor, booking__club=club,
                total_charge__gt=0, paid_at__isnull=True)
        .exclude(payment_method='invoice')
        .select_related('booking__aircraft__aircraft_type', 'booking__instructor',
                        'booking__flight_type', 'booking')
        .order_by('-booking__scheduled_start')
    )

    unpaid_invoices = list(
        _Inv.objects
        .filter(member=actor, club=club, status='sent')
        .select_related('flight_completion__booking__aircraft__aircraft_type',
                        'flight_completion__booking__instructor',
                        'flight_completion__booking__flight_type')
        .prefetch_related('line_items')
        .order_by('-issue_date')
    )
    for _inv in unpaid_invoices:
        _inv.needs_attention = _inv.is_overdue or _inv.issue_date <= _thirty_days_ago

    from decimal import Decimal as _D
    unpaid_flights_total = sum((fc.balance_owing or _D(0)) for fc in unpaid_flights)
    unpaid_invoices_total = sum((inv.balance_due or _D(0)) for inv in unpaid_invoices)

    return render(request, 'core/app/account.html', {
        'club': club, 'club_member': actor,
        'balance': balance,
        'transactions_page': transactions_page,
        'unpaid_flights': unpaid_flights,
        'unpaid_invoices': unpaid_invoices,
        'unpaid_flights_total': unpaid_flights_total,
        'unpaid_invoices_total': unpaid_invoices_total,
    })


@login_required
def app_credential_add(request, club_slug):
    """Mobile app — add a credential with optional evidence upload."""
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    from .models import MemberCredential, CredentialType, AircraftType
    from django.urls import reverse as _rev

    if request.method == 'POST':
        ct_id = request.POST.get('credential_type', '').strip()
        ct_obj = CredentialType.objects.filter(id=ct_id).first() if ct_id else None
        ac_type_id = request.POST.get('cred_aircraft_type_id', '').strip() or None
        name = request.POST.get('cred_name', '').strip()
        cert_num = request.POST.get('certificate_number', '').strip()
        issue_str = request.POST.get('issue_date', '').strip()
        expiry_str = request.POST.get('expiry_date', '').strip()
        notes = request.POST.get('notes', '').strip()

        from datetime import date as _d
        def _pd(s):
            try:
                return _d.fromisoformat(s) if s else None
            except ValueError:
                return None

        if ct_obj:
            cred = MemberCredential(
                member=actor.user,
                credential_type=ct_obj,
                name=name,
                certificate_number=cert_num,
                issue_date=_pd(issue_str),
                expiry_date=_pd(expiry_str),
                notes=notes,
            )
            if ac_type_id and ct_obj.category == 'type_rating':
                try:
                    cred.aircraft_type_id = int(ac_type_id)
                except (ValueError, TypeError):
                    pass
            if 'evidence' in request.FILES:
                cred.evidence = request.FILES['evidence']
            cred.save()
        return redirect(_rev('core:app_credentials', args=[club.slug]) + '?saved=1')

    aircraft_types = AircraftType.objects.filter(club=club).order_by('name')
    return render(request, 'core/app/credential_add.html', {
        'club': club, 'club_member': actor,
        'credential_types': CredentialType.objects.filter(region='NZ-CAA').order_by('display_order'),
        'aircraft_type_list': aircraft_types,
    })


def app_credential_edit(request, club_slug, cred_id):
    """Mobile app — edit an existing credential (must belong to the current member)."""
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    from .models import MemberCredential, CredentialType, AircraftType
    from django.urls import reverse as _rev

    cred = get_object_or_404(MemberCredential, id=cred_id, member=actor.user)

    if request.method == 'POST':
        ct_id = request.POST.get('credential_type', '').strip()
        ct_obj = CredentialType.objects.filter(id=ct_id).first() if ct_id else None
        ac_type_id = request.POST.get('cred_aircraft_type_id', '').strip() or None
        name = request.POST.get('cred_name', '').strip()
        cert_num = request.POST.get('certificate_number', '').strip()
        issue_str = request.POST.get('issue_date', '').strip()
        expiry_str = request.POST.get('expiry_date', '').strip()
        notes = request.POST.get('notes', '').strip()

        from datetime import date as _d
        def _pd(s):
            try:
                return _d.fromisoformat(s) if s else None
            except ValueError:
                return None

        if ct_obj:
            cred.credential_type = ct_obj
        cred.name = name
        cred.certificate_number = cert_num
        cred.issue_date = _pd(issue_str)
        cred.expiry_date = _pd(expiry_str)
        cred.notes = notes
        cred.aircraft_type_id = None
        if ac_type_id and ct_obj and ct_obj.category == 'type_rating':
            try:
                cred.aircraft_type_id = int(ac_type_id)
            except (ValueError, TypeError):
                pass
        if 'evidence' in request.FILES:
            cred.evidence = request.FILES['evidence']
        elif request.POST.get('remove_evidence') and cred.evidence:
            cred.evidence.delete(save=False)
            cred.evidence = None
        cred.save()
        return redirect(_rev('core:app_credentials', args=[club.slug]) + '?saved=1')

    aircraft_types = AircraftType.objects.filter(club=club).order_by('name')
    return render(request, 'core/app/credential_edit.html', {
        'club': club, 'club_member': actor, 'cred': cred,
        'credential_types': CredentialType.objects.filter(region='NZ-CAA').order_by('display_order'),
        'aircraft_type_list': aircraft_types,
    })


def app_credential_delete(request, club_slug, cred_id):
    """Mobile app — delete a credential (POST only, must belong to current member)."""
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    from .models import MemberCredential
    from django.urls import reverse as _rev

    cred = get_object_or_404(MemberCredential, id=cred_id, member=actor.user)
    if request.method == 'POST':
        cred.delete()
    return redirect(_rev('core:app_credentials', args=[club.slug]) + '?saved=1')


@login_required
def app_profile_edit(request, club_slug):
    """Mobile app — edit own contact details and avatar."""
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    if request.method == 'POST':
        if request.POST.get('action') == 'remove_avatar':
            if actor.avatar:
                actor.avatar.delete(save=True)
            from django.urls import reverse
            return redirect(reverse('core:app_profile', args=[club.slug]) + '?saved=1')
        actor.phone_mobile  = request.POST.get('phone_mobile', '').strip()
        actor.phone_home    = request.POST.get('phone_home', '').strip()
        actor.address_line1 = request.POST.get('address_line1', '').strip()
        actor.address_line2 = request.POST.get('address_line2', '').strip()
        actor.suburb        = request.POST.get('suburb', '').strip()
        actor.postcode      = request.POST.get('postcode', '').strip()
        actor.next_of_kin_name  = request.POST.get('next_of_kin_name', '').strip()
        actor.next_of_kin_phone = request.POST.get('next_of_kin_phone', '').strip()
        if 'avatar' in request.FILES:
            if actor.avatar:
                actor.avatar.delete(save=False)
            actor.avatar = request.FILES['avatar']
        actor.save()
        from django.urls import reverse
        return redirect(reverse('core:app_profile', args=[club.slug]) + '?saved=1')

    return render(request, 'core/app/profile_edit.html', {
        'club': club, 'club_member': actor,
    })


@login_required
def app_book_availability(request, club_slug):
    """
    Mobile booking creation.

    GET with no meaningful params → show picker form (mode + date).
    GET ?mode=solo|dual&date=YYYY-MM-DD → show slot list.
    """
    from datetime import date as _d, timedelta as _td, datetime as _dt
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    today = timezone.localdate()
    mode  = request.GET.get('mode', '')
    ac_param = request.GET.get('ac', '')
    date_str  = request.GET.get('date', '')

    from .permissions import check_booking_block
    _config = get_config(club)
    booking_blocked, booking_block_msg = check_booking_block(actor, _config)

    # ── Step 1: picker ───────────────────────────────────────────────────────
    if mode not in ('solo', 'dual'):
        # Build a list of the next 14 days for the date chips
        date_options = [today + _td(days=i) for i in range(14)]
        return render(request, 'core/app/book_pick.html', {
            'club': club, 'club_member': actor,
            'date_options': date_options,
            'today': today,
            'booking_blocked': booking_blocked,
            'booking_block_msg': booking_block_msg,
        })

    # ── Step 2: slot list ────────────────────────────────────────────────────
    try:
        selected = _d.fromisoformat(date_str) if date_str else today
    except ValueError:
        selected = today

    # Scrollable date band: 28 days from today, with month labels
    week_days = []
    for i in range(28):
        d = today + _td(days=i)
        week_days.append({
            'date': d, 'name': d.strftime('%a'), 'num': d.day,
            'is_today': d == today, 'is_selected': d == selected,
            'month_label': d.strftime('%b') if (d.day == 1 or i == 0) else '',
        })

    # Member type ratings → set of AircraftType IDs they hold
    from .models import MemberCredential
    rated_type_ids = set(
        MemberCredential.objects
        .filter(member=actor.user, credential_type__code='type')
        .values_list('aircraft_type_id', flat=True)
    )
    rated_type_ids.discard(None)  # credentials without an aircraft_type are non-restrictive

    # Aircraft (online, for hire). Filter to rated types if member has any type ratings.
    ac_qs = (Aircraft.objects
             .filter(club=club, status='online', is_available_for_hire=True)
             .select_related('aircraft_type')
             .order_by('registration'))
    if rated_type_ids:
        ac_qs = ac_qs.filter(aircraft_type_id__in=rated_type_ids)
    all_aircraft = list(ac_qs)

    # Day bookings
    day_bookings = list(
        Booking.objects
        .filter(club=club, scheduled_start__date=selected)
        .exclude(status='cancelled')
        .select_related('aircraft')
    )

    # Instructors (for dual)
    instructors = []
    if mode == 'dual':
        instructors = list(
            ClubMember.objects
            .filter(club=club, is_on_instructor_roster=True)
            .select_related('user')
            .order_by('user__last_name')
        )

    # ── Slot specs from ClubConfig ───────────────────────────────────────────
    from .models import ClubConfig as _CC
    _cfg = _CC.objects.filter(club=club).first()
    _slot_specs = _cfg.parsed_booking_slots() if _cfg else []
    if not _slot_specs:
        # Fallback: compute from duration + interval + operating hours
        _dur = _cfg.default_booking_duration if _cfg else 90
        _iv  = _cfg.time_slot_interval       if _cfg else 30
        _ds  = (_cfg.operating_hours_start.hour * 60 + _cfg.operating_hours_start.minute) if _cfg else 7 * 60
        _de  = (_cfg.operating_hours_end.hour   * 60 + _cfg.operating_hours_end.minute)   if _cfg else 21 * 60
        _s = _ds
        while _s + _dur <= _de:
            _e = _s + _dur
            _slot_specs.append((
                '{:02d}:{:02d}'.format(_s // 60, _s % 60),
                '{:02d}:{:02d}'.format(_e // 60, _e % 60),
                _s // 60, _s % 60, _e // 60, _e % 60,
            ))
            _s += _iv

    def _overlaps(bk, sh, sm, eh, em):
        """True if booking bk overlaps the half-open interval [sh:sm, eh:em)."""
        loc_s = timezone.localtime(bk.scheduled_start)
        loc_e = timezone.localtime(bk.scheduled_end)
        slot_start_min = sh * 60 + sm
        slot_end_min   = eh * 60 + em
        bk_start_min   = loc_s.hour * 60 + loc_s.minute
        bk_end_min     = loc_e.hour * 60 + loc_e.minute
        return bk_start_min < slot_end_min and bk_end_min > slot_start_min

    def _free_ac_for_slot(sh, sm, eh, em):
        return [
            ac for ac in all_aircraft
            if not any(_overlaps(bk, sh, sm, eh, em) for bk in day_bookings if bk.aircraft_id == ac.id)
        ]

    def _type_pills(free_ac):
        """One pill per aircraft (reg + type) — used for both solo and dual."""
        return _ac_pills(free_ac)

    def _ac_pills(free_ac):
        """One pill per AIRCRAFT — used for solo (member books a specific tail)."""
        return [
            {'reg': ac.registration,
             'type_name': ac.aircraft_type.name if ac.aircraft_type_id else ac.registration,
             'aircraft_id': ac.id}
            for ac in free_ac
        ]

    slots = []
    instructor_slot_groups = []

    if mode == 'solo':
        for start_str, end_str, sh, sm, eh, em in _slot_specs:
            free_ac = _free_ac_for_slot(sh, sm, eh, em)
            slots.append({
                'start': start_str, 'end': end_str, 'h': sh, 'm': sm,
                'available': bool(free_ac),
                'type_pills': _ac_pills(free_ac),
                'free_ac': free_ac,
            })
    else:
        for inst in instructors:
            inst_slots = []
            for start_str, end_str, sh, sm, eh, em in _slot_specs:
                inst_busy = any(
                    _overlaps(bk, sh, sm, eh, em)
                    for bk in day_bookings if bk.instructor_id == inst.user_id
                )
                if inst_busy:
                    continue
                free_ac = _free_ac_for_slot(sh, sm, eh, em)
                if not free_ac:
                    continue
                inst_slots.append({
                    'start': start_str, 'end': end_str,
                    'type_pills': _type_pills(free_ac),
                })
            if inst_slots:
                instructor_slot_groups.append({
                    'instructor_name': f"{inst.user.first_name} {inst.user.last_name}".strip(),
                    'instructor_id': inst.id,
                    'slot_rows': inst_slots,
                })

    prev_d = selected - _td(days=1)
    next_d = selected + _td(days=1)

    return render(request, 'core/app/book_slots.html', {
        'club': club, 'club_member': actor,
        'mode': mode,
        'selected': selected,
        'today': today,
        'week_days': week_days,
        'slots': slots,
        'instructor_slot_groups': instructor_slot_groups,
        'all_aircraft': all_aircraft,
        'prev_date': prev_d,
        'next_date': next_d,
    })


@login_required
def app_book_confirm(request, club_slug):
    """
    Mobile booking creation — step 3 (confirm + POST).

    GET: show confirm form with summary.
    POST: create booking → redirect to bookings list.
    """
    from datetime import date as _d, time as _t, datetime as _dt
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return redirect('login')

    today = timezone.localdate()

    from .permissions import check_booking_block as _chk_block
    _bb, _bb_msg = _chk_block(actor, get_config(club))

    if request.method == 'POST':
        if _bb:
            return redirect('core:app_book', club_slug=club_slug)
        mode      = request.POST.get('mode', 'solo')
        date_str  = request.POST.get('date', '')
        ac_id     = request.POST.get('ac_id', '')
        start_str = request.POST.get('start', '')
        end_str   = request.POST.get('end', '')
        ft_id     = request.POST.get('flight_type', '')
        inst_id   = request.POST.get('instructor', '')
        note      = request.POST.get('note', '')

        errors = []
        try:
            sel_date = _d.fromisoformat(date_str)
        except (ValueError, TypeError):
            errors.append('Invalid date.')
            sel_date = today

        try:
            ac = Aircraft.objects.get(id=int(ac_id), club=club)
        except (Aircraft.DoesNotExist, ValueError):
            errors.append('Aircraft not found.')
            ac = None

        try:
            ft = FlightType.objects.get(id=int(ft_id), club=club)
        except (FlightType.DoesNotExist, ValueError):
            errors.append('Flight type not found.')
            ft = None

        inst_user = None
        if mode == 'dual' and not inst_id:
            errors.append('Please select an instructor.')
        elif mode == 'dual' and inst_id:
            try:
                inst_member = ClubMember.objects.get(id=int(inst_id), club=club)
                inst_user = inst_member.user
            except (ClubMember.DoesNotExist, ValueError):
                errors.append('Instructor not found.')

        try:
            sh, sm = [int(x) for x in start_str.split(':')]
            eh, em = [int(x) for x in end_str.split(':')]
            sched_start = timezone.make_aware(_dt.combine(sel_date, _t(sh, sm)))
            sched_end   = timezone.make_aware(_dt.combine(sel_date, _t(eh, em)))
            if sched_end <= sched_start:
                errors.append('End time must be after start.')
        except (ValueError, TypeError, AttributeError):
            errors.append('Invalid times.')
            sched_start = sched_end = None

        if not errors and ac and ft and sched_start and sched_end:
            try:
                booking = Booking.objects.create(
                    club=club,
                    member=actor,
                    aircraft=ac,
                    flight_type=ft,
                    instructor=inst_user,
                    scheduled_start=sched_start,
                    scheduled_end=sched_end,
                    status='pending',
                    description=note,
                    created_by=request.user,
                )
                from django.urls import reverse
                return redirect(reverse('core:app_bookings', args=[club.slug]) + '?saved=1')
            except Exception as e:
                errors.append(str(e))

        # Fall through to re-render confirm with errors
        _all_ac = Aircraft.objects.filter(club=club, status='online').select_related('aircraft_type').order_by('registration')
        _ac_obj = Aircraft.objects.filter(club=club, id=int(ac_id)).select_related('aircraft_type').first() if ac_id else None
        _ac_list = _all_ac.filter(aircraft_type_id=_ac_obj.aircraft_type_id) if (_ac_obj and _ac_obj.aircraft_type_id) else _all_ac
        return render(request, 'core/app/book_confirm.html', {
            'club': club, 'club_member': actor,
            'errors': errors,
            'mode': mode,
            'date_str': date_str,
            'ac_id': ac_id,
            'aircraft_list': _ac_list,
            'start': start_str,
            'end': end_str,
            'ft_id': ft_id,
            'inst_id': inst_id,
            'note': note,
            'flight_types': FlightType.objects.filter(club=club).order_by('name'),
            'instructors': ClubMember.objects.filter(club=club, is_on_instructor_roster=True).select_related('user').order_by('user__last_name'),
        })

    # GET
    mode      = request.GET.get('mode', 'solo')
    date_str  = request.GET.get('date', str(today))
    ac_id     = request.GET.get('ac', '')
    ac_type   = request.GET.get('ac_type', '')
    start_str = request.GET.get('start', '')
    end_str   = request.GET.get('end', '')
    inst_id   = request.GET.get('instructor', '')

    try:
        sel_date = _d.fromisoformat(date_str)
    except (ValueError, TypeError):
        sel_date = today

    ac = None
    if ac_id:
        try:
            ac = Aircraft.objects.select_related('aircraft_type').get(id=int(ac_id), club=club)
        except (Aircraft.DoesNotExist, ValueError):
            pass

    # Aircraft list for the dropdown — same type as the pill tapped, or all if unknown
    all_ac = Aircraft.objects.filter(club=club, status='online').select_related('aircraft_type').order_by('registration')
    if ac and ac.aircraft_type_id:
        aircraft_list = all_ac.filter(aircraft_type_id=ac.aircraft_type_id)
    elif ac_type:
        aircraft_list = all_ac.filter(aircraft_type__name=ac_type)
    else:
        aircraft_list = all_ac

    flight_types = FlightType.objects.filter(club=club).order_by('name')

    # Filter instructors to those available at the requested slot start time.
    # "Available" = on roster for that day (or no schedule defined) AND no block-out.
    from .models import BlockOut as _ConfBO
    from django.db.models import Q as _ConfQ
    from datetime import datetime as _cdt, time as _ct
    _instr_all = list(
        ClubMember.objects
        .filter(club=club, is_on_instructor_roster=True)
        .select_related('user')
        .prefetch_related('availability_windows')
        .order_by('user__last_name')
    )
    try:
        _conf_date = _d.fromisoformat(date_str)
        _csh, _csm = [int(x) for x in start_str.split(':')]
        _sched_s   = timezone.make_aware(_cdt.combine(_conf_date, _ct(_csh, _csm)))
        _conf_day_s = timezone.make_aware(_cdt.combine(_conf_date, _ct(7,  0)))
        _conf_day_e = timezone.make_aware(_cdt.combine(_conf_date, _ct(20, 0)))
        _conf_bos = [
            bo for bo in _ConfBO.objects.filter(club=club).filter(
                _ConfQ(recurrence='one_off', date=_conf_date) |
                _ConfQ(recurrence='weekly') |
                _ConfQ(recurrence='daily')
            ).prefetch_related('instructors')
            if bo.applies_on(_conf_date)
        ]
        _conf_bookings = list(
            Booking.objects
            .filter(club=club, scheduled_start__date=_conf_date, instructor__isnull=False)
            .exclude(status='cancelled')
            .values_list('instructor_id', 'scheduled_start', 'scheduled_end')
        )
        def _instr_ok(cm):
            wins = list(cm.availability_windows.all())
            if wins:
                on_roster = False
                for w in wins:
                    iv = w.interval_on(_conf_date, _conf_day_s, _conf_day_e)
                    if iv and iv[0] <= _sched_s < iv[1]:
                        on_roster = True
                        break
                if not on_roster:
                    return False
            for bo in _conf_bos:
                if not bo.affects_instructor(cm.user):
                    continue
                biv = bo.interval_on(_conf_date)
                if biv and biv[0] <= _sched_s < biv[1]:
                    return False
            for instr_id, bk_s, bk_e in _conf_bookings:
                if instr_id == cm.user_id and bk_s <= _sched_s < bk_e:
                    return False
            return True
        instructors = [i for i in _instr_all if _instr_ok(i)]
    except (ValueError, TypeError, AttributeError):
        instructors = _instr_all  # fall back to all if times can't be parsed

    default_ft = None
    if mode == 'solo':
        default_ft = flight_types.filter(is_solo=True).first()
    elif mode == 'dual':
        default_ft = flight_types.filter(is_training=True, is_solo=False).first()
    if not default_ft:
        default_ft = flight_types.first()

    return render(request, 'core/app/book_confirm.html', {
        'club': club, 'club_member': actor,
        'mode': mode,
        'sel_date': sel_date,
        'date_str': date_str,
        'ac': ac,
        'ac_id': ac_id,
        'aircraft_list': aircraft_list,
        'start': start_str,
        'end': end_str,
        'inst_id': inst_id,
        'flight_types': flight_types,
        'instructors': instructors,
        'default_ft': default_ft,
        'errors': [],
    })


# ── Web Push ──────────────────────────────────────────────────────────────────

def pwa_manifest(request, club_slug):
    """Dynamic PWA manifest — uses club's uploaded logo as icon, falling back to ClubHangar mark."""
    import json as _json
    from django.conf import settings
    club = get_object_or_404(Club, slug=club_slug)
    cfg  = get_config(club)
    site = getattr(settings, 'SITE_URL', '').rstrip('/')

    from django.contrib.staticfiles.storage import staticfiles_storage
    mark_svg = request.build_absolute_uri(staticfiles_storage.url('core/img/clubhangar-mark.svg'))
    mark_png = request.build_absolute_uri(staticfiles_storage.url('core/img/clubhangar-icon-512.png'))

    if cfg.logo:
        logo_url = (site + cfg.logo.url) if site else cfg.logo.url
        icons = [
            {'src': logo_url, 'sizes': 'any', 'type': 'image/png', 'purpose': 'any maskable'},
            {'src': mark_png, 'sizes': '512x512', 'type': 'image/png', 'purpose': 'any'},
        ]
    else:
        icons = [
            {'src': mark_png, 'sizes': '512x512', 'type': 'image/png', 'purpose': 'any maskable'},
            {'src': mark_svg, 'sizes': 'any', 'type': 'image/svg+xml', 'purpose': 'any'},
        ]

    manifest = {
        'name': 'ClubHangar',
        'short_name': 'ClubHangar',
        'description': f'{club.name} — powered by ClubHangar',
        # Open to the club picker so multi-club members choose each time;
        # single-club members are redirected straight into their club.
        'start_url': '/app/',
        'scope': '/app/',
        'display': 'standalone',
        'orientation': 'portrait',
        'background_color': cfg.theme_banner or '#1d3a5f',
        'theme_color': cfg.theme_banner or '#1d3a5f',
        'icons': icons,
    }
    resp = JsonResponse(manifest)
    resp['Content-Type'] = 'application/manifest+json'
    resp['Cache-Control'] = 'public, max-age=3600'
    return resp


def sw_js(request):
    """Serve the service worker from /sw.js with scope header."""
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    path = finders.find('core/sw.js')
    if path:
        with open(path) as f:
            content = f.read()
    else:
        content = '// service worker not found'
    resp = HttpResponse(content, content_type='application/javascript')
    resp['Service-Worker-Allowed'] = '/'
    resp['Cache-Control'] = 'no-store'
    return resp


@login_required
@require_POST
def push_subscribe(request, club_slug):
    """Save or update a push subscription for the current member."""
    import json
    from .models import PushSubscription
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return JsonResponse({'ok': False}, status=403)
    try:
        data = json.loads(request.body)
        endpoint = data['endpoint']
        p256dh   = data['keys']['p256dh']
        auth     = data['keys']['auth']
    except (KeyError, ValueError):
        return JsonResponse({'ok': False, 'error': 'bad payload'}, status=400)

    PushSubscription.objects.update_or_create(
        endpoint=endpoint,
        defaults={'club_member': actor, 'p256dh': p256dh, 'auth': auth},
    )
    return JsonResponse({'ok': True})


@login_required
@require_POST
def push_unsubscribe(request, club_slug):
    """Remove all push subscriptions for the current member on this device."""
    import json
    from .models import PushSubscription
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return JsonResponse({'ok': False}, status=403)
    try:
        data = json.loads(request.body)
        endpoint = data.get('endpoint')
    except ValueError:
        return JsonResponse({'ok': False}, status=400)
    if endpoint:
        PushSubscription.objects.filter(club_member=actor, endpoint=endpoint).delete()
    return JsonResponse({'ok': True})


@login_required
@require_POST
def push_test(request, club_slug):
    """Send a test push to all of the current member's subscriptions so they can verify delivery."""
    import logging as _log
    from .models import PushSubscription
    club, actor = _app_actor(request, club_slug)
    if not actor:
        return JsonResponse({'ok': False}, status=403)
    if not PushSubscription.objects.filter(club_member=actor).exists():
        return JsonResponse({'ok': False, 'error': 'No push subscription found. Turn push on first.'})
    try:
        from .push import notify_member
        from django.urls import reverse as _rev
        app_url = _rev('core:app_home', kwargs={'club_slug': club_slug})
        notify_member(actor, 'ClubHangar test', 'Push notifications are working ✓', url=app_url)
        return JsonResponse({'ok': True})
    except Exception as exc:
        _log.getLogger(__name__).exception('push_test failed for %s', actor)
        return JsonResponse({'ok': False, 'error': str(exc)})


@login_required
def submit_feedback(request, club_slug):
    club = get_object_or_404(Club, slug=club_slug)
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    msg_type = request.POST.get('message_type', FeedbackMessage.TYPE_FEEDBACK)
    if msg_type not in {FeedbackMessage.TYPE_FEEDBACK, FeedbackMessage.TYPE_FEATURE, FeedbackMessage.TYPE_BUG}:
        msg_type = FeedbackMessage.TYPE_FEEDBACK
    message = request.POST.get('message', '').strip()
    if message:
        FeedbackMessage.objects.create(
            club=club,
            sender=request.user,
            message_type=msg_type,
            message=message,
        )
    return redirect(request.META.get('HTTP_REFERER', '/') + '?saved=1')
